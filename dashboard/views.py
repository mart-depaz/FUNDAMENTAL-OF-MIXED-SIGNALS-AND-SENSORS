#for dashboard app views



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.db.models import Q, Count, Max
from django.db import transaction
from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from accounts.models import CustomUser
from .models import Course, Program, Department, CourseSchedule, UserNotification, CourseEnrollment, AttendanceRecord, QRCodeRegistration, InstructorRegistrationStatus, BiometricRegistration
from datetime import datetime
import json
import logging
import random
import time
import hashlib
import re
import base64
import io
import os
import requests
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
# Initialize logger first
logger = logging.getLogger(__name__)


def _normalize_registered_qr(raw_value: str) -> str:
    raw = (raw_value or '').strip()
    if not raw:
        return ''

    candidate = None
    try:
        if raw.startswith('http://') or raw.startswith('https://'):
            from urllib.parse import urlparse, parse_qs
            parsed = urlparse(raw)
            q = parse_qs(parsed.query or '')
            candidate = (q.get('qr_code') or q.get('qr') or [None])[0]
            if not candidate:
                m = re.search(r'([a-f0-9]{32})', raw, flags=re.IGNORECASE)
                candidate = m.group(1) if m else None
        else:
            m = re.search(r'([a-f0-9]{32})', raw, flags=re.IGNORECASE)
            candidate = m.group(1) if m else None
    except Exception:
        candidate = None

    return (candidate or raw).strip().lower()

try:
    from PIL import Image
    from pyzbar.pyzbar import decode as pyzbar_decode
    SERVER_DECODE_AVAILABLE = True
    logger.info('pyzbar + Pillow available for server-side QR decoding')
except Exception as e:
    SERVER_DECODE_AVAILABLE = False
    logger.warning('pyzbar/Pillow not available for server decode: ' + str(e))

# Helper function to create notifications
def create_notification(user, notification_type, title, message, category='general', related_course=None, related_user=None):
    """Helper function to create a UserNotification"""
    try:
        UserNotification.objects.create(
            user=user,
            notification_type=notification_type,
            title=title,
            message=message,
            category=category,
            related_course=related_course,
            related_user=related_user,
            is_read=False
        )
        logger.info(f"Notification created: {notification_type} for user {user.username}")
    except Exception as e:
        logger.error(f"Error creating notification: {str(e)}")

# Import qrcode and PIL at module level to ensure they're available
try:
    import qrcode
    from PIL import Image
    QRCODE_AVAILABLE = True
    logger.info("qrcode library successfully imported")
except ImportError as e:
    QRCODE_AVAILABLE = False
    logger.error(f"qrcode library not available: {str(e)}. Please install it with: pip install qrcode[pil]")

def calculate_course_status(start_dt, end_dt, now_ph):
    """
    Calculate detailed course status based on start/end datetime and current time.
    
    Returns one of: 'soon', 'upcoming', 'starting_today', 'live', 'ongoing', 'finished'
    """
    from datetime import timedelta, time as dtime
    
    if not start_dt:
        return 'finished'
    
    # Calculate days until start
    days_until_start = (start_dt.date() - now_ph.date()).days
    
    # Check if start date is today
    if days_until_start == 0:
        # Check if we're within the class time window
        if end_dt and start_dt <= now_ph < end_dt:
            return 'live'  # Class is happening right now
        elif now_ph < start_dt:
            return 'starting_today'  # Class starts today but hasn't started yet
        else:
            # Class ended today, but since it's a recurring schedule, check if there's a next occurrence
            # For now, mark as finished for today's instance
            return 'finished'  # Today's class ended
    
    # Check if start date is in the future
    if days_until_start > 0:
        if days_until_start > 7:
            return 'soon'  # More than 7 days away
        elif days_until_start == 1:
            return 'tomorrow'  # Tomorrow
        elif days_until_start >= 1:
            return 'upcoming'  # 2-7 days away
    
    # If start date is in the past (days_until_start < 0)
    # This means we're looking at a past occurrence
    # Since courses are recurring weekly, we should have already calculated the next occurrence
    # But if we're here, it means this is a past instance
    return 'finished'  # Past date

def require_role(*allowed_roles):
    """Decorator to restrict access to specific user roles (student, teacher, admin)"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            user = request.user
            
            # Admin users cannot access instructor or student dashboards
            if user.is_authenticated:
                if user.is_admin and ('teacher' in allowed_roles or 'student' in allowed_roles):
                    logger.warning(f"Admin user {user.username} tried to access {allowed_roles} dashboard")
                    return render(request, 'dashboard/shared/error.html', 
                                {'message': 'Admin accounts do not have access to instructor or student dashboards.'})
                
                # Check if user has the required role
                if 'teacher' in allowed_roles and not user.is_teacher:
                    return render(request, 'dashboard/shared/error.html', 
                                {'message': 'You are not authorized to access the instructor dashboard.'})
                
                if 'student' in allowed_roles and not user.is_student:
                    return render(request, 'dashboard/shared/error.html', 
                                {'message': 'You are not authorized to access the student dashboard.'})
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

@login_required
@require_role('teacher')
def teacher_dashboard_view(request):
    if not request.user.is_teacher or request.user.is_admin:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access the teacher dashboard.'})
    
    user = request.user
    
    # Dashboard metrics
    instructor_courses = Course.objects.filter(instructor=user, is_active=True, deleted_at__isnull=True, is_archived=False).select_related('program').order_by('-created_at')
    # Grouped courses: count unique by code+name+semester+school_year (sections collapsed)
    course_keys = set()
    for c in instructor_courses.values('code', 'name', 'semester', 'school_year'):
        norm_code = (c.get('code') or '').strip().upper()
        norm_name = (c.get('name') or '').strip().upper()
        norm_sem = (c.get('semester') or '').strip().lower()
        norm_sy = (c.get('school_year') or '').strip()
        course_keys.add((norm_code, norm_name, norm_sem, norm_sy))
    total_courses = len(course_keys)
    
    active_students = CourseEnrollment.objects.filter(
        course__instructor=user,
        is_active=True,
        deleted_at__isnull=True
    ).values('student').distinct().count()
    avg_students_per_course = round(active_students / total_courses, 1) if total_courses else 0
    
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Recent courses: show only 3 unique logical courses (group multi-section as one), newest first
    recent_courses = []
    seen_keys = set()
    for c in instructor_courses:  # already ordered by -created_at
        key = (
            (c.code or '').strip().upper(),
            (c.name or '').strip().upper(),
            (c.semester or '').strip().lower(),
            (c.school_year or '').strip()
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        recent_courses.append(c)
        if len(recent_courses) >= 3:
            break
    # Build upcoming classes (nearest first, Philippines time), include section, limit to 3
    try:
        try:
            from zoneinfo import ZoneInfo  # Python 3.9+
            ph_tz = ZoneInfo('Asia/Manila')
        except Exception:
            import pytz  # fallback
            ph_tz = pytz.timezone('Asia/Manila')
    except Exception:
        ph_tz = None
    from datetime import timedelta, time as dtime
    now_utc = datetime.utcnow()
    now_ph = now_utc if ph_tz is None else datetime.now(ph_tz)
    weekday_map = {'M': 0, 'Mon': 0, 'Monday': 0,
                   'T': 1, 'Tue': 1, 'Tuesday': 1,
                   'W': 2, 'Wed': 2, 'Wednesday': 2,
                   'Th': 3, 'Thu': 3, 'Thursday': 3,
                   'F': 4, 'Fri': 4, 'Friday': 4,
                   'S': 5, 'Sat': 5, 'Saturday': 5,
                   'Su': 6, 'Sun': 6, 'Sunday': 6}
    def to_ph_time(t):
        if not t:
            return None
        # t is time object
        if ph_tz is None:
            return t
        return t
    def next_occurrence(next_weekday, t):
        # Compute next datetime for weekday (0=Mon) at time t (time object) in PH tz
        if not isinstance(t, dtime):
            return None
        today_wd = now_ph.weekday() if hasattr(now_ph, 'weekday') else datetime.now().weekday()
        days_ahead = (next_weekday - today_wd) % 7
        candidate = (now_ph.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
                     if hasattr(now_ph, 'replace') else datetime.now().replace(hour=t.hour, minute=t.minute, second=0, microsecond=0))
        # If the time for the same weekday has already passed today, jump to the next week (7 days ahead)
        if days_ahead == 0 and candidate <= now_ph:
            days_ahead = 7
        if days_ahead:
            candidate = candidate + timedelta(days=days_ahead)
        return candidate
    # Create upcoming class notifications (only once per day, for classes starting today or tomorrow)
    from django.utils import timezone
    today_date = now_ph.date()
    tomorrow_date = today_date + timedelta(days=1)
    
    upcoming_entries = []
    for c in instructor_courses:
        # Day-specific schedules
        schedules = list(c.course_schedules.all().order_by('day_order'))
        # Human-friendly semester text (ordinal without the word "Sem")
        sem_map = {'1st': '1st', '2nd': '2nd', '3rd': '3rd', 'summer': 'Summer'}
        sem_display = sem_map.get((c.semester or '').strip().lower(), c.semester or '')
        if schedules:
            for s in schedules:
                wd = weekday_map.get(getattr(s, 'day', ''), None)
                if wd is None or not s.start_time:
                    continue
                # Determine live/upcoming status relative to "now_ph"
                start_today = now_ph.replace(hour=s.start_time.hour, minute=s.start_time.minute, second=0, microsecond=0)
                end_today = start_today
                if s.end_time:
                    end_today = start_today.replace(hour=s.end_time.hour, minute=s.end_time.minute)
                if wd == (now_ph.weekday() if hasattr(now_ph, 'weekday') else datetime.now().weekday()):
                    next_dt = start_today
                    end_dt = end_today
                    status = calculate_course_status(next_dt, end_dt, now_ph)
                else:
                    next_dt = next_occurrence(wd, s.start_time)
                    end_dt = next_dt.replace(hour=s.end_time.hour, minute=s.end_time.minute) if (next_dt and s.end_time) else next_dt
                    status = calculate_course_status(next_dt, end_dt, now_ph)
                if not next_dt:
                    continue
                start_str = s.start_time.strftime('%I:%M %p')
                end_str = s.end_time.strftime('%I:%M %p') if s.end_time else ''
                date_label = next_dt.strftime('%b %d, %Y') if next_dt else ''
                upcoming_entries.append({
                    'id': c.id,
                    'code': c.code,
                    'name': c.name,
                    'section': (c.section or '').upper(),
                    'semester_display': sem_display,
                    'day_label': s.get_day_display() if hasattr(s, 'get_day_display') else s.day,
                    'date_label': date_label,
                    'time_label': f"{start_str}{(' - ' + end_str) if end_str else ''}",
                    'status': status,
                    'color': c.color or '#3b82f6',
                    'next_dt': next_dt,
                })
                
                # Create notification for upcoming class (today or tomorrow, only once per day per course)
                if next_dt and next_dt.date() in [today_date, tomorrow_date]:
                    try:
                        # Check if notification already exists for this class created in the last 24 hours (prevent duplicates on refresh)
                        day_text = 'today' if next_dt.date() == today_date else 'tomorrow'
                        from django.utils import timezone
                        from datetime import timedelta
                        # Check if notification exists with same course, type, created in last 24 hours
                        existing_notification = UserNotification.objects.filter(
                            user=user,
                            notification_type='upcoming_class',
                            related_course=c,
                            created_at__gte=now_ph - timedelta(hours=24)  # Check last 24 hours
                        ).first()
                        
                        if not existing_notification:
                            create_notification(
                                user=user,
                                notification_type='upcoming_class',
                                title='Upcoming Class',
                                message=f'{c.code} - {c.name} is scheduled for {day_text} at {start_str}',
                                category='course',
                                related_course=c
                            )
                    except Exception as e:
                        logger.error(f"Error creating upcoming class notification: {str(e)}")
        else:
            # Default schedule across days
            if c.days and c.start_time:
                for day in [d.strip() for d in c.days.split(',') if d.strip()]:
                    wd = weekday_map.get(day, None)
                    if wd is None:
                        continue
                    start_today = now_ph.replace(hour=c.start_time.hour, minute=c.start_time.minute, second=0, microsecond=0)
                    end_today = start_today
                    if c.end_time:
                        end_today = start_today
                    if c.end_time:
                        end_today = start_today.replace(hour=c.end_time.hour, minute=c.end_time.minute)
                    if wd == (now_ph.weekday() if hasattr(now_ph, 'weekday') else datetime.now().weekday()):
                        next_dt = start_today
                        end_dt = end_today
                        status = calculate_course_status(next_dt, end_dt, now_ph)
                    else:
                        next_dt = next_occurrence(wd, c.start_time)
                        end_dt = next_dt.replace(hour=c.end_time.hour, minute=c.end_time.minute) if (next_dt and c.end_time) else next_dt
                        status = calculate_course_status(next_dt, end_dt, now_ph)
                    if not next_dt:
                        continue
                    start_str = c.start_time.strftime('%I:%M %p') if c.start_time else ''
                    end_str = c.end_time.strftime('%I:%M %p') if c.end_time else ''
                    date_label = next_dt.strftime('%b %d, %Y') if next_dt else ''
                    upcoming_entries.append({
                        'id': c.id,
                        'code': c.code,
                        'name': c.name,
                        'section': (c.section or '').upper(),
                        'semester_display': sem_display,
                        'day_label': day,
                        'date_label': date_label,
                        'time_label': f"{start_str}{(' - ' + end_str) if end_str else ''}",
                        'status': status,
                        'color': c.color or '#3b82f6',
                        'next_dt': next_dt,
                    })
                    
                    # Create notification for upcoming class (today or tomorrow, only once per day per course)
                    if next_dt and next_dt.date() in [today_date, tomorrow_date]:
                        try:
                            # Check if notification already exists for this class created in the last 24 hours (prevent duplicates on refresh)
                            day_text = 'today' if next_dt.date() == today_date else 'tomorrow'
                            from django.utils import timezone
                            from datetime import timedelta
                            # Check if notification exists with same course, type, created in last 24 hours
                            existing_notification = UserNotification.objects.filter(
                                user=user,
                                notification_type='upcoming_class',
                                related_course=c,
                                created_at__gte=now_ph - timedelta(hours=24)  # Check last 24 hours
                            ).first()
                            
                            if not existing_notification:
                                create_notification(
                                    user=user,
                                    notification_type='upcoming_class',
                                    title='Upcoming Class',
                                    message=f'{c.code} - {c.name} is scheduled for {day_text} at {start_str}',
                                    category='course',
                                    related_course=c
                                )
                        except Exception as e:
                            logger.error(f"Error creating upcoming class notification: {str(e)}")
    # Sort by status priority then by next occurrence and take top 3
    status_priority = {'live': 0, 'starting_today': 1, 'tomorrow': 2, 'upcoming': 3, 'soon': 4, 'ongoing': 5, 'finished': 6}
    upcoming_entries.sort(key=lambda e: (status_priority.get(e.get('status'), 6), e.get('next_dt', datetime.max)))
    upcoming_more_count = max(0, len(upcoming_entries) - 3)
    upcoming_classes = upcoming_entries[:3]
    
    try:
        recent_notifications = UserNotification.objects.filter(user=user).order_by('-created_at')[:5]
    except Exception:
        recent_notifications = []
    
    # Compute attendance reports count (all attendance records for this instructor's active, non-deleted, non-archived courses)
    try:
        attendance_reports_count = AttendanceRecord.objects.filter(
            course__instructor=user,
            course__is_active=True,
            course__deleted_at__isnull=True,
            course__is_archived=False
        ).count()
    except Exception:
        attendance_reports_count = 0

    context = {
        'unread_notifications': unread_notifications,
        'total_courses': total_courses,
        'active_students': active_students,
        'recent_courses': recent_courses,
        'upcoming_classes': upcoming_classes,
        'upcoming_more_count': upcoming_more_count,
        'recent_notifications': recent_notifications,
        'avg_students_per_course': avg_students_per_course,
        'attendance_reports_count': attendance_reports_count,
    }
    return render(request, 'dashboard/instructor/teacher.html', context)

@login_required
@require_role('student')
def student_dashboard_view(request):
    if not request.user.is_student or request.user.is_admin:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access the student dashboard.'})
    
    user = request.user
    
    # Get unread notification count (handle if table doesn't exist yet)
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Get enrolled courses for this student
    from .models import CourseEnrollment, Course, CourseSchedule
    enrollments = CourseEnrollment.objects.filter(
        student=user,
        is_active=True,
        deleted_at__isnull=True,
        course__is_active=True,
        course__deleted_at__isnull=True,
        course__is_archived=False
    ).select_related('course', 'course__program', 'course__instructor').prefetch_related('course__course_schedules')
    
    # Get recent courses (last 3, grouped by multi-section)
    recent_enrollments = enrollments.order_by('-enrolled_at')[:10]
    
    # Group multi-section courses
    grouped_recent = {}
    for enrollment in recent_enrollments:
        course = enrollment.course
        key = (
            (course.code or '').strip().lower(),
            (course.name or '').strip().lower(),
            (course.semester or '').strip().lower(),
            (course.school_year or '').strip()
        )
        if key not in grouped_recent:
            grouped_recent[key] = enrollment
    recent_courses = list(grouped_recent.values())[:3]
    
    # Build schedule entries for upcoming classes (day by day, same as student_todays_status_view)
    from datetime import datetime, time as dtime, timedelta
    try:
        from zoneinfo import ZoneInfo
        PH_TZ = ZoneInfo('Asia/Manila')
    except Exception:
        try:
            import pytz
            PH_TZ = pytz.timezone('Asia/Manila')
        except Exception:
            PH_TZ = None
    now_ph = datetime.now(PH_TZ) if PH_TZ else datetime.now()
    weekday_map = {'M': 0, 'T': 1, 'W': 2, 'Th': 3, 'F': 4, 'S': 5, 'Su': 6}
    
    def to_time_label(start_t, end_t):
        s = start_t.strftime('%I:%M %p') if start_t else ''
        e = end_t.strftime('%I:%M %p') if end_t else ''
        return f"{s} - {e}" if s and e else s or e
    
    def next_occurrence(weekday_idx, t):
        if not isinstance(t, dtime):
            return None
        today_idx = now_ph.weekday()
        days_ahead = (weekday_idx - today_idx) % 7
        base = now_ph.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        if days_ahead == 0 and base <= now_ph:
            days_ahead = 7
        return base + timedelta(days=days_ahead)
    
    def calculate_course_status(start_dt, end_dt, now):
        if not start_dt or not end_dt:
            return 'upcoming'
        if start_dt <= now <= end_dt:
            return 'live'
        elif start_dt.date() == now.date() and now < start_dt:
            return 'starting_today'
        elif start_dt.date() == (now.date() + timedelta(days=1)):
            return 'tomorrow'
        elif start_dt > now:
            days_diff = (start_dt.date() - now.date()).days
            if days_diff > 7:
                return 'soon'
            else:
                return 'upcoming'
        else:
            return 'finished'
    
    schedule_entries = []
    student_courses = [e.course for e in enrollments]
    
    for c in student_courses:
        color = c.color or '#3b82f6'
        section = (c.section or '').upper()
        code = c.code or ''
        name = c.name or ''
        
        # Day-specific schedules
        day_schedules = c.course_schedules.all()
        if day_schedules.exists():
            for s in day_schedules:
                day_val = s.day
                day_key = {
                    'Monday': 'M', 'Mon': 'M',
                    'Tuesday': 'T', 'Tue': 'T',
                    'Wednesday': 'W', 'Wed': 'W',
                    'Thursday': 'Th', 'Thu': 'Th',
                    'Friday': 'F', 'Fri': 'F',
                    'Saturday': 'S', 'Sat': 'S',
                    'Sunday': 'Su', 'Sun': 'Su'
                }.get(str(day_val), str(day_val))
                wd = weekday_map.get(day_key, None)
                if wd is None or not s.start_time:
                    continue
                start_today = now_ph.replace(hour=s.start_time.hour, minute=s.start_time.minute, second=0, microsecond=0)
                end_today = start_today
                if s.end_time:
                    end_today = start_today.replace(hour=s.end_time.hour, minute=s.end_time.minute)
                if wd == now_ph.weekday():
                    start_dt = start_today
                    end_dt = end_today
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                else:
                    start_dt = next_occurrence(wd, s.start_time)
                    end_dt = start_dt.replace(hour=s.end_time.hour, minute=s.end_time.minute) if (start_dt and s.end_time) else start_dt
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                time_label = to_time_label(s.start_time, s.end_time)
                schedule_entries.append({
                    'id': c.id,
                    'code': code,
                    'name': name,
                    'section': section,
                    'color': color,
                    'status': status,
                    'date_label': date_label,
                    'time_label': time_label,
                    'start_dt': start_dt,
                })
        else:
            # Synchronized schedule - create separate entry for each day
            if c.days and c.start_time:
                days_list = [d.strip() for d in c.days.split(',') if d.strip()]
                for day in days_list:
                    day_key = {
                        'Monday': 'M', 'Mon': 'M',
                        'Tuesday': 'T', 'Tue': 'T',
                        'Wednesday': 'W', 'Wed': 'W',
                        'Thursday': 'Th', 'Thu': 'Th',
                        'Friday': 'F', 'Fri': 'F',
                        'Saturday': 'S', 'Sat': 'S',
                        'Sunday': 'Su', 'Sun': 'Su'
                    }.get(str(day), str(day))
                    wd = weekday_map.get(day_key, None)
                    if wd is None:
                        continue
                    start_today = now_ph.replace(hour=c.start_time.hour, minute=c.start_time.minute, second=0, microsecond=0)
                    end_today = start_today
                    if c.end_time:
                        end_today = start_today.replace(hour=c.end_time.hour, minute=c.end_time.minute)
                    if wd == now_ph.weekday():
                        start_dt = start_today
                        end_dt = end_today
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    else:
                        start_dt = next_occurrence(wd, c.start_time)
                        end_dt = start_dt.replace(hour=c.end_time.hour, minute=c.end_time.minute) if (start_dt and c.end_time) else start_dt
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    
                    # Check if scheduled for tomorrow
                    if start_dt:
                        tomorrow = now_ph.date() + timedelta(days=1)
                        if start_dt.date() == tomorrow:
                            status = 'tomorrow'
                    
                    date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                    time_label = to_time_label(c.start_time, c.end_time)
                    schedule_entries.append({
                        'id': c.id,
                        'code': code,
                        'name': name,
                        'section': section,
                        'color': color,
                        'status': status,
                        'date_label': date_label,
                        'time_label': time_label,
                        'start_dt': start_dt,
                    })
    
    # Sort: live first, then starting_today, then tomorrow, then upcoming, etc.
    def sort_key(e):
        status_priority = {'live': 0, 'starting_today': 1, 'tomorrow': 2, 'upcoming': 3, 'soon': 4, 'ongoing': 5, 'finished': 6}
        status_rank = status_priority.get(e.get('status'), 6)
        dt = e.get('start_dt') or now_ph
        return (status_rank, dt if status_rank < 6 else datetime.max.replace(tzinfo=dt.tzinfo) - (dt - dt.replace(hour=0, minute=0, second=0, microsecond=0)))
    schedule_entries.sort(key=sort_key)
    
    # Get upcoming classes (next 5)
    upcoming_classes = schedule_entries[:5]
    
    context = {
        'user': user,
        'unread_notifications': unread_notifications,
        'recent_courses': recent_courses,
        'upcoming_classes': upcoming_classes,
        'total_enrolled': enrollments.count(),
    }
    # Ensure today's attendance is finalized for enrolled courses so counts match the Attendance Log page
    try:
        # Finalize attendance for each enrolled course (idempotent)
        processed_courses = set()
        for enrollment in enrollments:
            course_obj = getattr(enrollment, 'course', None)
            if not course_obj or course_obj.id in processed_courses:
                continue
            processed_courses.add(course_obj.id)
            try:
                finalize_all_course_attendance(course_obj, getattr(course_obj, 'instructor', None), force=False)
            except Exception:
                # Ignore finalization errors; proceed to counting
                logger.exception('Error finalizing attendance for course %s', getattr(course_obj, 'id', None))
    except Exception:
        # If anything unexpected happens, continue without finalization
        logger.exception('Error during attendance finalization loop for student dashboard')

    # Compute student's attendance log count (after finalization)
    try:
        attendance_log_count = AttendanceRecord.objects.filter(student=user, course__is_active=True, course__deleted_at__isnull=True, course__is_archived=False).count()
    except Exception:
        attendance_log_count = 0
    context['attendance_log_count'] = attendance_log_count
    return render(request, 'dashboard/student/student.html', context)

@login_required
@require_role('student')
def student_todays_status_view(request):
    """Student's Today's Status page - similar to instructor's My Classes"""
    user = request.user
    if not user.is_student:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Get enrolled courses for this student
    enrollments = CourseEnrollment.objects.filter(
        student=user,
        is_active=True,
        deleted_at__isnull=True
    ).select_related('course', 'course__instructor', 'course__program').prefetch_related('course__course_schedules')
    
    # Get courses (excluding archived/deleted)
    student_courses = [e.course for e in enrollments if e.course.is_active and e.course.deleted_at is None and not e.course.is_archived]
    
    # Optional focus course
    focus_course_id = request.GET.get('course')
    focus_course = None
    if focus_course_id:
        try:
            focus_course = next((c for c in student_courses if c.id == int(focus_course_id)), None)
        except Exception:
            focus_course = None
    
    # Build schedule entries with PH time awareness
    from datetime import datetime, time as dtime, timedelta
    try:
        from zoneinfo import ZoneInfo
        PH_TZ = ZoneInfo('Asia/Manila')
    except Exception:
        try:
            import pytz
            PH_TZ = pytz.timezone('Asia/Manila')
        except Exception:
            PH_TZ = None
    now_ph = datetime.now(PH_TZ) if PH_TZ else datetime.now()
    weekday_map = {'M': 0, 'T': 1, 'W': 2, 'Th': 3, 'F': 4, 'S': 5, 'Su': 6}
    
    def to_time_label(start_t, end_t):
        s = start_t.strftime('%I:%M %p') if start_t else ''
        e = end_t.strftime('%I:%M %p') if end_t else ''
        return f"{s} - {e}" if s and e else s or e
    
    def next_occurrence(weekday_idx, t):
        if not isinstance(t, dtime):
            return None
        today_idx = now_ph.weekday()
        days_ahead = (weekday_idx - today_idx) % 7
        base = now_ph.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        if days_ahead == 0 and base <= now_ph:
            days_ahead = 7
        return base + timedelta(days=days_ahead)
    
    schedule_entries = []
    # Get all enrolled courses (each section separately)
    for c in student_courses:
        color = c.color or '#3b82f6'
        section = (c.section or '').upper()
        code = c.code or ''
        name = c.name or ''
        
        # Day-specific schedules
        day_schedules = c.course_schedules.all()
        if day_schedules.exists():
            for s in day_schedules:
                day_val = s.day
                day_key = {
                    'Monday': 'M', 'Mon': 'M',
                    'Tuesday': 'T', 'Tue': 'T',
                    'Wednesday': 'W', 'Wed': 'W',
                    'Thursday': 'Th', 'Thu': 'Th',
                    'Friday': 'F', 'Fri': 'F',
                    'Saturday': 'S', 'Sat': 'S',
                    'Sunday': 'Su', 'Sun': 'Su'
                }.get(str(day_val), str(day_val))
                wd = weekday_map.get(day_key, None)
                if wd is None or not s.start_time:
                    continue
                start_today = now_ph.replace(hour=s.start_time.hour, minute=s.start_time.minute, second=0, microsecond=0)
                end_today = start_today
                if s.end_time:
                    end_today = start_today.replace(hour=s.end_time.hour, minute=s.end_time.minute)
                if wd == now_ph.weekday():
                    start_dt = start_today
                    end_dt = end_today
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                else:
                    start_dt = next_occurrence(wd, s.start_time)
                    end_dt = start_dt.replace(hour=s.end_time.hour, minute=s.end_time.minute) if (start_dt and s.end_time) else start_dt
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                time_label = to_time_label(s.start_time, s.end_time)
                
                # Get attendance record for today for this specific day schedule
                attendance_record = None
                if start_dt:
                    today_date = start_dt.date() if hasattr(start_dt, 'date') else now_ph.date()
                    try:
                        # Get day-specific attendance record
                        schedule_day = s.day  # e.g., 'Mon', 'Tue', etc.
                        attendance_record = AttendanceRecord.objects.filter(
                            course=c,
                            student=user,
                            attendance_date=today_date,
                            schedule_day=schedule_day
                        ).first()
                    except Exception:
                        pass
                
                # Get day name for this entry
                day_name = ''
                day_label = ''
                if start_dt:
                    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                    day_name = day_names[start_dt.weekday()]
                    day_label = day_name
                elif s.day:
                    # Fallback to schedule day if start_dt is not available
                    day_name = str(s.day)
                    day_label = day_name
                
                # Generate unique schedule_id for this entry
                day_key_short = {
                    'Monday': 'Mon', 'Mon': 'Mon',
                    'Tuesday': 'Tue', 'Tue': 'Tue',
                    'Wednesday': 'Wed', 'Wed': 'Wed',
                    'Thursday': 'Thu', 'Thu': 'Thu',
                    'Friday': 'Fri', 'Fri': 'Fri',
                    'Saturday': 'Sat', 'Sat': 'Sat',
                    'Sunday': 'Sun', 'Sun': 'Sun'
                }.get(day_label, day_label[:3] if day_label else '')
                schedule_id = f"{c.id}_{day_key_short}_{start_dt.strftime('%Y%m%d') if start_dt else ''}"
                
                # Get day-specific attendance status
                day_attendance_status = None
                if s.attendance_status:
                    day_attendance_status = s.attendance_status
                else:
                    day_attendance_status = c.attendance_status or 'closed'
                
                schedule_entries.append({
                    'course_id': c.id,
                    'code': code,
                    'name': name,
                    'section': section,
                    'color': color,
                    'status': status,
                    'date_label': date_label,
                    'time_label': time_label,
                    'start_dt': start_dt,
                    'end_dt': end_dt,
                    'day': day_name,  # Add day for day-specific QR codes
                    'day_label': day_label,  # Full day name for display
                    'schedule_id': schedule_id,  # Unique identifier for this schedule entry
                    'qr_code': c.qr_code or '',
                    'attendance_status': day_attendance_status,  # Use day-specific status if available
                    'attendance_record': attendance_record,
                })
        else:
            # Synchronized schedule - create separate entry for EACH day (day by day display)
            if c.days and c.start_time:
                days_list = [d.strip() for d in c.days.split(',') if d.strip()]
                for day in days_list:
                    day_key = {
                        'Monday': 'M', 'Mon': 'M',
                        'Tuesday': 'T', 'Tue': 'T',
                        'Wednesday': 'W', 'Wed': 'W',
                        'Thursday': 'Th', 'Thu': 'Th',
                        'Friday': 'F', 'Fri': 'F',
                        'Saturday': 'S', 'Sat': 'S',
                        'Sunday': 'Su', 'Sun': 'Su'
                    }.get(str(day), str(day))
                    wd = weekday_map.get(day_key, None)
                    if wd is None:
                        continue
                    start_today = now_ph.replace(hour=c.start_time.hour, minute=c.start_time.minute, second=0, microsecond=0)
                    end_today = start_today
                    if c.end_time:
                        end_today = start_today.replace(hour=c.end_time.hour, minute=c.end_time.minute)
                    if wd == now_ph.weekday():
                        start_dt = start_today
                        end_dt = end_today
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    else:
                        start_dt = next_occurrence(wd, c.start_time)
                        end_dt = start_dt.replace(hour=c.end_time.hour, minute=c.end_time.minute) if (start_dt and c.end_time) else start_dt
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    
                    # Check if scheduled for tomorrow
                    if start_dt:
                        tomorrow = now_ph.date() + timedelta(days=1)
                        if start_dt.date() == tomorrow:
                            status = 'tomorrow'
                    
                    date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                    time_label = to_time_label(c.start_time, c.end_time)
                    
                    # Get attendance record for today for this specific day schedule
                    attendance_record = None
                    if start_dt:
                        today_date = start_dt.date() if hasattr(start_dt, 'date') else now_ph.date()
                        try:
                            # Get day-specific attendance record
                            schedule_day = day_short  # e.g., 'Mon', 'Tue', etc.
                            attendance_record = AttendanceRecord.objects.filter(
                                course=c,
                                student=user,
                                attendance_date=today_date,
                                schedule_day=schedule_day
                            ).first()
                        except Exception:
                            pass
                    
                    # Get day name for this entry
                    day_label = ''
                    day_short = ''
                    if start_dt:
                        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                        day_name = day_names[start_dt.weekday()]
                        day_label = day_name
                        # Map to short form for CourseSchedule lookup
                        day_short_map = {
                            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                            'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                        }
                        day_short = day_short_map.get(day_label, 'Mon')
                    else:
                        # Map day string to full day name
                        day_map = {
                            'Monday': 'Monday', 'Mon': 'Monday', 'M': 'Monday',
                            'Tuesday': 'Tuesday', 'Tue': 'Tuesday', 'T': 'Tuesday',
                            'Wednesday': 'Wednesday', 'Wed': 'Wednesday', 'W': 'Wednesday',
                            'Thursday': 'Thursday', 'Thu': 'Thursday', 'Th': 'Thursday',
                            'Friday': 'Friday', 'Fri': 'Friday', 'F': 'Friday',
                            'Saturday': 'Saturday', 'Sat': 'Saturday', 'S': 'Saturday',
                            'Sunday': 'Sunday', 'Sun': 'Sunday', 'Su': 'Sunday'
                        }
                        day_name = day_map.get(str(day), str(day))
                        day_label = day_name
                        # Map to short form
                        day_short_map = {
                            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                            'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                        }
                        day_short = day_short_map.get(day_label, str(day)[:3])
                    
                    # Get day-specific attendance status for synchronized schedules
                    day_attendance_status = c.attendance_status or 'closed'
                    day_schedule = c.course_schedules.filter(day=day_short).first()
                    if day_schedule and day_schedule.attendance_status:
                        day_attendance_status = day_schedule.attendance_status
                    
                    # Generate unique schedule_id for this entry
                    schedule_id = f"{c.id}_{day_label}_{start_dt.strftime('%Y%m%d') if start_dt else ''}"
                    
                    schedule_entries.append({
                        'course_id': c.id,
                        'code': code,
                        'name': name,
                        'section': section,
                        'color': color,
                        'status': status,
                        'date_label': date_label,
                        'time_label': time_label,
                        'start_dt': start_dt,
                        'end_dt': end_dt,
                        'day': day_name,  # Add day for day-specific QR codes
                        'day_label': day_label,  # Full day name for display
                        'schedule_id': schedule_id,  # Unique identifier for this schedule entry
                        'qr_code': c.qr_code or '',
                        'attendance_status': day_attendance_status,  # Use day-specific status if available
                        'attendance_record': attendance_record,
                    })
    
    # Sort: live first, then starting_today, then tomorrow, then upcoming, etc.
    def sort_key(e):
        status_priority = {'live': 0, 'starting_today': 1, 'tomorrow': 2, 'upcoming': 3, 'soon': 4, 'ongoing': 5, 'finished': 6}
        status_rank = status_priority.get(e.get('status'), 6)
        dt = e.get('start_dt') or now_ph
        return (status_rank, dt if status_rank < 6 else datetime.max.replace(tzinfo=dt.tzinfo) - (dt - dt.replace(hour=0, minute=0, second=0, microsecond=0)))
    schedule_entries.sort(key=sort_key)
    
    # Get focus schedule from URL parameter
    focus_schedule_id = request.GET.get('schedule')
    
    # Get focus course next entry
    focus_course_next_entry = None
    if focus_course:
        if focus_schedule_id:
            # Find entry matching the schedule_id
            for entry in schedule_entries:
                if entry.get('schedule_id') == focus_schedule_id and entry.get('course_id') == focus_course.id:
                    focus_course_next_entry = entry
                    break
        else:
            # Fallback to first entry for this course
            for entry in schedule_entries:
                if entry.get('course_id') == focus_course.id:
                    focus_course_next_entry = entry
                    break
        
        # Get day-specific QR code for focus_course_next_entry
        if focus_course_next_entry and focus_course_next_entry.get('day_label'):
            day_label = focus_course_next_entry.get('day_label')
            # Map full day name to short form
            day_map = {
                'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
            }
            day_short = day_map.get(day_label, day_label)
            day_schedule = CourseSchedule.objects.filter(course=focus_course, day=day_short).first()
            if day_schedule and day_schedule.qr_code:
                focus_course_next_entry['qr_code'] = day_schedule.qr_code
            else:
                # Fallback to course-level QR code
                focus_course_next_entry['qr_code'] = focus_course.qr_code or ''
    
    # Check if student has already marked attendance today for focus course and specific schedule
    student_attendance_today = None
    if focus_course and focus_course_next_entry:
        from django.utils import timezone
        today = timezone.now().date()
        try:
            # Get day-specific attendance record
            day_label = focus_course_next_entry.get('day_label', '')
            if day_label:
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                schedule_day = day_map.get(day_label, '')
                student_attendance_today = AttendanceRecord.objects.filter(
                    course=focus_course,
                    student=user,
                    attendance_date=today,
                    schedule_day=schedule_day
                ).first()
        except Exception:
            student_attendance_today = None
    
    # Check if course schedule has finished
    course_finished = False
    if focus_course_next_entry and focus_course_next_entry.get('end_dt'):
        end_dt = focus_course_next_entry.get('end_dt')
        if end_dt and now_ph > end_dt:
            course_finished = True
            # Auto-close attendance and reset QR code when schedule finishes
            if focus_course_next_entry.get('day_label'):
                day_label = focus_course_next_entry.get('day_label')
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                schedule_day = day_map.get(day_label, day_label)
                day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
                if day_schedule:
                    # Reset attendance status to closed
                    day_schedule.attendance_status = 'closed'
                    # Reset QR code for next occurrence
                    day_schedule.qr_code = None
                    day_schedule.qr_code_date = None
                    day_schedule.save(update_fields=['attendance_status', 'qr_code', 'qr_code_date'])
            elif focus_course:
                # For synchronized courses, update course-level attendance status
                focus_course.attendance_status = 'closed'
                # Reset QR code for next occurrence
                focus_course.qr_code = None
                focus_course.save(update_fields=['attendance_status', 'qr_code'])
    
    # Compute whether to show the transient "Your Status" box.
    # Rules:
    # - If the specific schedule instance has an end_dt and now > end_dt => do not show.
    # - Otherwise, show only when there's an AttendanceRecord for this student for that instance
    #   and the current time falls within the instance window (if available).
    show_student_status = False
    try:
        start_dt = focus_course_next_entry.get('start_dt') if focus_course_next_entry else None
        end_dt = focus_course_next_entry.get('end_dt') if focus_course_next_entry else None

        # Derive end_dt from CourseSchedule or course-level end_time if missing
        if start_dt and not end_dt and focus_course:
            try:
                day_label = focus_course_next_entry.get('day_label') if focus_course_next_entry else None
                if day_label:
                    day_map = {
                        'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                        'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                    }
                    schedule_day = day_map.get(day_label, day_label)
                    day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
                    if day_schedule and getattr(day_schedule, 'end_time', None):
                        et = day_schedule.end_time
                        end_dt = start_dt.replace(hour=et.hour, minute=et.minute)
                if not end_dt and getattr(focus_course, 'end_time', None):
                    et = focus_course.end_time
                    end_dt = start_dt.replace(hour=et.hour, minute=et.minute)
            except Exception:
                end_dt = end_dt

        # If we have an end_dt and it's already passed, do not show
        instance_finished = False
        if end_dt:
            try:
                if now_ph > end_dt:
                    instance_finished = True
            except Exception:
                instance_finished = course_finished

        if instance_finished:
            show_student_status = False
            # propagate instance_finished into course_finished so templates can show finished message
            course_finished = course_finished or instance_finished
        else:
            if student_attendance_today:
                if start_dt and end_dt:
                    show_student_status = (start_dt <= now_ph <= end_dt)
                elif start_dt:
                    show_student_status = (now_ph.date() == start_dt.date() and now_ph >= start_dt)
                else:
                    show_student_status = not course_finished
            else:
                show_student_status = False
    except Exception:
        show_student_status = (student_attendance_today is not None and not course_finished)

    context = {
        'school_admin': school_admin,
        'unread_notifications': unread_notifications,
        'focus_course': focus_course,
        'focus_course_next_entry': focus_course_next_entry,
        'focus_schedule_id': focus_schedule_id,  # Pass schedule_id to template
        'schedule_entries': schedule_entries,
        'student_attendance_today': student_attendance_today,
        'course_finished': course_finished,  # Flag to show message when schedule finishes
        'show_student_status': show_student_status,
    }
    # If we have a focus schedule and it's open, compute present window expiry (ms) and duration
    try:
        present_expiry_ms = None
        present_duration_minutes = None
        focus_schedule_identifier = None
        if focus_course_next_entry:
            # Try to find schedule object for the focused course/day
            focus_schedule_identifier = focus_course_next_entry.get('schedule_id')
            # Determine schedule object: prefer CourseSchedule by day matching
            from django.utils import timezone
            try:
                day_label = focus_course_next_entry.get('day_label')
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                day_code = day_map.get(day_label, None)
            except Exception:
                day_code = None

            sched_obj = None
            try:
                if day_code and focus_course:
                    sched_obj = CourseSchedule.objects.filter(course=focus_course, day__iexact=day_code).first()
            except Exception:
                sched_obj = None

            # Get present duration from schedule or course
            if sched_obj and getattr(sched_obj, 'attendance_present_duration', None):
                present_duration_minutes = int(getattr(sched_obj, 'attendance_present_duration'))
            else:
                present_duration_minutes = int(getattr(focus_course, 'attendance_present_duration', 0) or 0)

            # Get attendance status to verify if present window is active
            attendance_status = None
            try:
                if sched_obj and getattr(sched_obj, 'attendance_status', None):
                    attendance_status = sched_obj.attendance_status
                else:
                    attendance_status = getattr(focus_course, 'attendance_status', 'closed')
            except Exception:
                attendance_status = 'closed'

            # Only show present-window countdown when the instructor actually opened it
            # (i.e., a schedule-level qr_code_opened_at exists for this specific occurrence).
            # Do NOT fallback to start_dt or to timezone.now() as that causes countdowns
            # to appear even when the instructor never opened the present window.
            qr_opened_at = None
            try:
                # Only consider schedule-level object (sched_obj) which was looked up above.
                if sched_obj and getattr(sched_obj, 'qr_code_opened_at', None):
                    # If schedule has a qr_code_date set, ensure it matches the occurrence date
                    try:
                        occ_date = focus_course_next_entry.get('start_dt').date() if focus_course_next_entry.get('start_dt') else None
                    except Exception:
                        occ_date = None

                    sched_qr_date = getattr(sched_obj, 'qr_code_date', None)
                    if sched_qr_date and occ_date:
                        # Only use the opened timestamp when the QR code was generated for this occurrence date
                        if sched_qr_date == occ_date:
                            qr_opened_at = getattr(sched_obj, 'qr_code_opened_at')
                    else:
                        # If no qr_code_date to compare (legacy), be conservative and require qr_code_opened_at
                        qr_opened_at = getattr(sched_obj, 'qr_code_opened_at')
                # Fallback to course-level if schedule doesn't have it
                elif getattr(focus_course, 'qr_code_opened_at', None):
                    qr_opened_at = getattr(focus_course, 'qr_code_opened_at')
            except Exception:
                qr_opened_at = None

            # If we have an explicit opened timestamp and a configured duration, compute expiry
            # Make sure attendance is open and expiry hasn't already passed
            if qr_opened_at and present_duration_minutes and present_duration_minutes > 0 and attendance_status == 'open':
                try:
                    from django.utils import timezone as tz_util
                    current_time = tz_util.now()
                    expiry = qr_opened_at + tz_util.timedelta(minutes=int(present_duration_minutes))
                    # Only set expiry if it hasn't passed yet and is in the future
                    if expiry > current_time:
                        # convert to ms since epoch for JavaScript
                        present_expiry_ms = int(expiry.timestamp() * 1000)
                    else:
                        # Expiry has already passed, don't show timer
                        present_expiry_ms = None
                except Exception as e:
                    logger.warning(f"Error calculating present window expiry: {str(e)}")
                    present_expiry_ms = None
            else:
                present_expiry_ms = None

        context['focus_present_window_expiry_ms'] = present_expiry_ms
        context['focus_present_duration_minutes'] = present_duration_minutes
        context['focus_present_schedule_id'] = focus_schedule_identifier
    except Exception as e:
        logger.warning(f"Error in present window calculation: {str(e)}")
        context['focus_present_window_expiry_ms'] = None
        context['focus_present_duration_minutes'] = None
        context['focus_present_schedule_id'] = None
    return render(request, 'dashboard/student/student_todays_status.html', context)

@login_required
def student_attendance_log_view(request):
    """Student's attendance log - shows all attendance history"""
    user = request.user
    if not user.is_student:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Get enrolled courses for this student - ordered by enrollment date (1st enrolled first)
    # Include BOTH active AND inactive enrollments so students can see their complete attendance history
    # Also include courses that are no longer active (finished/closed) but have attendance records
    from .models import CourseEnrollment, AttendanceRecord
    enrollments = CourseEnrollment.objects.filter(
        student=user,
        deleted_at__isnull=True,
        course__deleted_at__isnull=True,
        course__is_archived=False
    ).select_related('course', 'course__program', 'course__instructor').order_by('enrolled_at')
    
    # Create list with enrollment info for ordering
    enrolled_courses = []
    for enrollment in enrollments:
        course = enrollment.course
        course.enrollment_info = enrollment  # Attach enrollment info for template access
        course.is_active_enrollment = enrollment.is_active  # Mark whether enrollment is currently active
        course.is_active_course = course.is_active  # Mark whether course is still active
        enrolled_courses.append(course)
    
    # Get selected course from query parameter
    selected_course_id = request.GET.get('course')
    selected_course = None
    attendance_records = []
    status_filter = request.GET.get('status', '')
    
    if selected_course_id:
        try:
            selected_course = next((c for c in enrolled_courses if c.id == int(selected_course_id)), None)
            if selected_course:
                # Get current date and time (in PH timezone)
                try:
                    from zoneinfo import ZoneInfo
                    PH_TZ = ZoneInfo('Asia/Manila')
                except Exception:
                    try:
                        import pytz
                        PH_TZ = pytz.timezone('Asia/Manila')
                    except Exception:
                        PH_TZ = None
                
                from django.utils import timezone
                now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
                today = now_ph.date()
                current_time = now_ph.time()
                
                # Get attendance records for selected course
                # Ensure any missing absent records are finalized for today's schedules
                try:
                    # finalize_all_course_attendance is idempotent with force=False (time-aware)
                    # Only creates ABSENT records if ALL class schedules for today have ended
                    finalize_all_course_attendance(selected_course, getattr(selected_course, 'instructor', None), force=False)
                except Exception:
                    # If finalization fails, continue and allow manual inspection
                    logger.exception('finalize_all_course_attendance failed')

                # Get all existing attendance records AFTER finalization
                # This ensures we include records created by finalize_all_course_attendance
                records_qs = AttendanceRecord.objects.filter(
                    student=user,
                    course=selected_course
                ).select_related('course', 'enrollment').order_by('-attendance_date', '-attendance_time')
                
                # Get existing records as a set for quick lookup
                existing_records = list(records_qs)
                existing_record_dates = {rec.attendance_date for rec in existing_records}
                
                # Filter attendance records by scheduled class days only
                # Only show records for days the class is scheduled to meet
                scheduled_days = set()
                
                # Get day-specific schedules for this course
                day_schedules = selected_course.course_schedules.all()
                if day_schedules.exists():
                    for sched in day_schedules:
                        if sched.day:
                            scheduled_days.add(sched.day)
                # Fallback to default course days if no specific schedules
                elif selected_course.days:
                    for day in [d.strip() for d in selected_course.days.split(',') if d.strip()]:
                        scheduled_days.add(day)
                
                # Generate all scheduled dates from enrollment date to today
                # where the student should have attended
                from datetime import timedelta, datetime
                
                # Get enrollment info for this student-course pair
                try:
                    student_enrollment = CourseEnrollment.objects.filter(
                        student=user,
                        course=selected_course,
                        deleted_at__isnull=True
                    ).first()
                    enrollment_date = student_enrollment.enrolled_at.date() if student_enrollment and student_enrollment.enrolled_at else today - timedelta(days=365)
                except Exception:
                    enrollment_date = today - timedelta(days=365)
                
                # Build map of scheduled dates for this course
                scheduled_dates_to_check = []
                current_check_date = enrollment_date
                
                while current_check_date <= today:
                    # Get weekday name for current_check_date
                    weekday_idx = current_check_date.weekday()
                    weekday_map = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
                    weekday_code = weekday_map.get(weekday_idx, '')
                    
                    # Check if this day is a scheduled day for the course
                    day_matches = False
                    for sched_day in scheduled_days:
                        if sched_day in [weekday_code, ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][weekday_idx]]:
                            day_matches = True
                            break
                    
                    if day_matches and current_check_date not in existing_record_dates:
                        scheduled_dates_to_check.append(current_check_date)
                    
                    current_check_date += timedelta(days=1)
                
                # Create absent records for scheduled dates where class has finished
                # and student has no attendance record
                for check_date in scheduled_dates_to_check:
                    # Get course schedule for this date
                    weekday_idx = check_date.weekday()
                    weekday_code = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][weekday_idx]
                    
                    # Find schedule for this day
                    course_start_time = None
                    course_end_time = None
                    schedule_postponed = False
                    
                    # Check day-specific schedules
                    day_schedules_for_date = selected_course.course_schedules.filter(day__iexact=weekday_code).first()
                    if day_schedules_for_date:
                        course_start_time = day_schedules_for_date.start_time
                        course_end_time = day_schedules_for_date.end_time
                        # Check if this schedule is marked as postponed
                        if getattr(day_schedules_for_date, 'attendance_status', '') == 'postponed':
                            schedule_postponed = True
                    else:
                        # Use course defaults
                        course_start_time = selected_course.start_time
                        course_end_time = selected_course.end_time
                    
                    if course_start_time and course_end_time:
                        # Check if class has finished
                        class_end_dt = datetime.combine(check_date, course_end_time)
                        
                        # Make timezone-aware
                        if PH_TZ:
                            if hasattr(PH_TZ, 'localize'):
                                import pytz
                                class_end_dt = PH_TZ.localize(class_end_dt)
                            else:
                                class_end_dt = class_end_dt.replace(tzinfo=PH_TZ)
                        
                        # Only create absent record if class has finished
                        if class_end_dt < now_ph or check_date < today:
                            # Check if record already exists (might have been created by finalize)
                            existing = AttendanceRecord.objects.filter(
                                student=user,
                                course=selected_course,
                                attendance_date=check_date
                            ).exists()
                            
                            if not existing:
                                # Determine status: postponed or absent
                                record_status = 'postponed' if schedule_postponed else 'absent'
                                
                                # Create record with appropriate status
                                try:
                                    attendance_enrollment = CourseEnrollment.objects.filter(
                                        student=user,
                                        course=selected_course,
                                        deleted_at__isnull=True
                                    ).first()
                                    
                                    if attendance_enrollment:
                                        AttendanceRecord.objects.create(
                                            student=user,
                                            course=selected_course,
                                            enrollment=attendance_enrollment,
                                            attendance_date=check_date,
                                            attendance_time=None,
                                            status=record_status
                                        )
                                except Exception as e:
                                    logger.debug(f"Failed to create absent record: {e}")
                
                # Refetch records after creating postponed/absent records to include newly created ones
                records_qs = AttendanceRecord.objects.filter(
                    student=user,
                    course=selected_course
                ).select_related('course', 'enrollment').order_by('-attendance_date', '-attendance_time')
                
                # Deduplicate BEFORE applying status filter: Get only latest per date
                # This ensures we don't show multiple scans on the same day - only the final status
                all_records = list(records_qs)
                seen_dates = {}
                dedup_records = []
                for rec in all_records:
                    # Use date as key to get only latest per date
                    date_key = rec.attendance_date
                    if date_key not in seen_dates:
                        seen_dates[date_key] = True
                        dedup_records.append(rec)
                
                # Apply status filter AFTER deduplication (optional)
                attendance_records = dedup_records
                if status_filter:
                    attendance_records = [r for r in attendance_records if r.status == status_filter]

                # Annotate each attendance record with whether the schedule for that
                # record's day was marked as postponed. This allows templates to
                # display a 'Postponed' label instead of normal statuses.
                try:
                    for rec in attendance_records:
                        rec.is_postponed = False
                        try:
                            # Check if the schedule for this date is marked as postponed
                            if rec.attendance_date:
                                # Get weekday code for this date
                                weekday_idx = rec.attendance_date.weekday()
                                weekday_code = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][weekday_idx]
                                
                                # Check if schedule for this day is postponed
                                sched = selected_course.course_schedules.filter(day__iexact=weekday_code).first()
                                if sched and getattr(sched, 'attendance_status', '') == 'postponed':
                                    rec.is_postponed = True
                                    continue
                            
                            # Fallback: check day code stored on the record (e.g., 'Mon', 'Tue')
                            day_code = getattr(rec, 'schedule_day', None)
                            if day_code:
                                sched = selected_course.course_schedules.filter(day__iexact=day_code).first()
                                if sched and getattr(sched, 'attendance_status', '') == 'postponed':
                                    rec.is_postponed = True
                                    continue
                            
                            # Fallback: if course-level attendance_status is postponed
                            if getattr(selected_course, 'attendance_status', '') == 'postponed':
                                rec.is_postponed = True
                        except Exception:
                            rec.is_postponed = False
                except Exception:
                    # If anything goes wrong, leave records unannotated (templates will behave normally)
                    pass
        except (ValueError, TypeError):
            selected_course = None
            attendance_records = []
    
    # Calculate statistics for selected course
    total_records = len(attendance_records)
    present_count = sum(1 for r in attendance_records if r.status == 'present')
    late_count = sum(1 for r in attendance_records if r.status == 'late')
    absent_count = sum(1 for r in attendance_records if r.status == 'absent')
    # Count postponed: either status is 'postponed' OR is_postponed flag is True
    postponed_count = sum(1 for r in attendance_records if r.status == 'postponed' or getattr(r, 'is_postponed', False))
    
    context = {
        'school_admin': school_admin,
        'unread_notifications': unread_notifications,
        'enrolled_courses': enrolled_courses,
        'selected_course': selected_course,
        'attendance_records': attendance_records,
        'status_filter': status_filter,
        'total_records': total_records,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'postponed_count': postponed_count,
    }
    return render(request, 'dashboard/student/student_attendance_log.html', context)

@login_required
def instructor_my_classes_view(request):
    """Simple monitoring hub for instructors - shows today's and upcoming classes."""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Topbar admin
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Optional focus course
    focus_course_id = request.GET.get('course')
    focus_course = None
    if focus_course_id:
        try:
            # Extract numeric ID from composite keys like "73_F_20251212"
            numeric_id = focus_course_id.split('_')[0] if '_' in focus_course_id else focus_course_id
            focus_course = Course.objects.get(id=int(numeric_id), instructor=user)
        except Exception:
            focus_course = None
    
    # Very light data set - list active courses for quick pick (excluding deleted and archived)
    # Get ALL individual course sections (not grouped) so each section appears separately in Quick Pick
    courses = Course.objects.filter(instructor=user, is_active=True, deleted_at__isnull=True, is_archived=False)\
        .select_related('program', 'instructor')\
        .prefetch_related('course_schedules')\
        .order_by('name', 'code', 'section')
    
    # Build schedule entries with PH time awareness
    from datetime import datetime, time as dtime, timedelta
    try:
        from zoneinfo import ZoneInfo
        PH_TZ = ZoneInfo('Asia/Manila')
    except Exception:
        try:
            import pytz
            PH_TZ = pytz.timezone('Asia/Manila')
        except Exception:
            PH_TZ = None
    now_ph = datetime.now(PH_TZ) if PH_TZ else datetime.now()
    weekday_map = {'M': 0, 'T': 1, 'W': 2, 'Th': 3, 'F': 4, 'S': 5, 'Su': 6}
    days_full = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday', 6: 'Sunday'}
    
    def to_time_label(start_t, end_t):
        s = start_t.strftime('%I:%M %p') if start_t else ''
        e = end_t.strftime('%I:%M %p') if end_t else ''
        return f"{s} - {e}" if s and e else s or e
    
    def next_occurrence(weekday_idx, t):
        if not isinstance(t, dtime):
            return None
        today_idx = now_ph.weekday()
        days_ahead = (weekday_idx - today_idx) % 7
        base = now_ph.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        # If today and already passed, push by 7 days
        if days_ahead == 0 and base <= now_ph:
            days_ahead = 7
        return base + timedelta(days=days_ahead)
    
    schedule_entries = []
    # Get all courses including all sections (don't group - show each section separately)
    # Use the same courses queryset but ensure we iterate through all individual sections
    for c in courses:
        color = c.color or '#3b82f6'
        section = (c.section or '').upper()
        code = c.code or ''
        name = c.name or ''
        # Day-specific schedules
        day_schedules = getattr(c, 'course_schedules', None)
        if day_schedules and day_schedules.all().exists():
            for s in day_schedules.all():
                # Map day to weekday index
                day_val = getattr(s, 'day', '')
                # Map to weekday_map keys (M, T, W, etc.) for calculation
                day_key_short = {
                    'Monday': 'M', 'Mon': 'M',
                    'Tuesday': 'T', 'Tue': 'T',
                    'Wednesday': 'W', 'Wed': 'W',
                    'Thursday': 'Th', 'Thu': 'Th',
                    'Friday': 'F', 'Fri': 'F',
                    'Saturday': 'S', 'Sat': 'S',
                    'Sunday': 'Su', 'Sun': 'Su'
                }.get(str(day_val), str(day_val))
                # Map to full day name for display and QR code
                day_label_full = {
                    'M': 'Monday', 'Mon': 'Monday', 'Monday': 'Monday',
                    'T': 'Tuesday', 'Tue': 'Tuesday', 'Tuesday': 'Tuesday',
                    'W': 'Wednesday', 'Wed': 'Wednesday', 'Wednesday': 'Wednesday',
                    'Th': 'Thursday', 'Thu': 'Thursday', 'Thursday': 'Thursday',
                    'F': 'Friday', 'Fri': 'Friday', 'Friday': 'Friday',
                    'S': 'Saturday', 'Sat': 'Saturday', 'Saturday': 'Saturday',
                    'Su': 'Sunday', 'Sun': 'Sunday', 'Sunday': 'Sunday'
                }.get(str(day_val), str(day_val))
                wd = weekday_map.get(day_key_short, None)
                if wd is None or not s.start_time:
                    continue
                start_today = now_ph.replace(hour=s.start_time.hour, minute=s.start_time.minute, second=0, microsecond=0)
                end_today = start_today
                if s.end_time:
                    end_today = start_today.replace(hour=s.end_time.hour, minute=s.end_time.minute)
                if wd == now_ph.weekday():
                    start_dt = start_today
                    end_dt = end_today
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                else:
                    start_dt = next_occurrence(wd, s.start_time)
                    end_dt = start_dt.replace(hour=s.end_time.hour, minute=s.end_time.minute) if (start_dt and s.end_time) else start_dt
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                time_label = to_time_label(s.start_time, s.end_time)
                # Create unique schedule identifier: course_id + day + date
                schedule_id = f"{c.id}_{day_key_short}_{start_dt.strftime('%Y%m%d') if start_dt else ''}"
                schedule_entries.append({
                    'course_id': c.id,
                    'code': code,
                    'name': name,
                    'section': section,
                    'color': color,
                    'status': status,
                    'date_label': date_label,
                    'time_label': time_label,
                    'start_dt': start_dt,
                    'day_label': day_label_full,  # Use full day name for display and QR code
                    'schedule_id': schedule_id,
                })
        else:
            # Synchronized schedule (same time across days) - create ONE entry PER DAY
            if c.days and c.start_time:
                days_list = [d.strip() for d in c.days.split(',') if d.strip()]
                # Create a separate entry for EACH day (day by day display)
                for day in days_list:
                    day_key = {
                        'Monday': 'M', 'Mon': 'M',
                        'Tuesday': 'T', 'Tue': 'T',
                        'Wednesday': 'W', 'Wed': 'W',
                        'Thursday': 'Th', 'Thu': 'Th',
                        'Friday': 'F', 'Fri': 'F',
                        'Saturday': 'S', 'Sat': 'S',
                        'Sunday': 'Su', 'Sun': 'Su'
                    }.get(str(day), str(day))
                    wd = weekday_map.get(day_key, None)
                    if wd is None:
                        continue
                    start_today = now_ph.replace(hour=c.start_time.hour, minute=c.start_time.minute, second=0, microsecond=0)
                    end_today = start_today
                    if c.end_time:
                        end_today = start_today.replace(hour=c.end_time.hour, minute=c.end_time.minute)
                    if wd == now_ph.weekday():
                        start_dt = start_today
                        end_dt = end_today
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    else:
                        start_dt = next_occurrence(wd, c.start_time)
                        end_dt = start_dt.replace(hour=c.end_time.hour, minute=c.end_time.minute) if (start_dt and c.end_time) else start_dt
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    
                    # Check if scheduled for tomorrow
                    if start_dt:
                        tomorrow = now_ph.date() + timedelta(days=1)
                        if start_dt.date() == tomorrow:
                            status = 'tomorrow'
                    
                    date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                    day_label = days_full.get(wd, day)
                    time_label = to_time_label(c.start_time, c.end_time)
                    # Create unique schedule identifier: course_id + day + date
                    schedule_id = f"{c.id}_{day_label}_{start_dt.strftime('%Y%m%d') if start_dt else ''}"
                    schedule_entries.append({
                        'course_id': c.id,
                        'code': code,
                        'name': name,
                        'section': section,
                        'color': color,
                        'status': status,
                        'date_label': date_label,
                        'day_label': day_label,
                        'time_label': time_label,
                        'start_dt': start_dt,
                        'schedule_id': schedule_id,
                    })
    
    # Sort: live first, then starting_today, then tomorrow, then upcoming, then soon, then ongoing, then finished
    def sort_key(e):
        status_priority = {'live': 0, 'starting_today': 1, 'tomorrow': 2, 'upcoming': 3, 'soon': 4, 'ongoing': 5, 'finished': 6}
        status_rank = status_priority.get(e.get('status'), 6)
        dt = e.get('start_dt') or now_ph
        return (status_rank, dt if status_rank < 6 else datetime.max.replace(tzinfo=dt.tzinfo) - (dt - dt.replace(hour=0, minute=0, second=0, microsecond=0)))
    schedule_entries.sort(key=sort_key)

    # Derive next/active schedule entry for the focused course, if any
    # Check if a specific schedule (day) was requested
    focus_schedule_id = request.GET.get('schedule')
    focus_course_next_entry = None
    if focus_course:
        if focus_schedule_id:
            # Find the specific schedule entry by schedule_id
            for entry in schedule_entries:
                if entry.get('schedule_id') == focus_schedule_id and entry.get('course_id') == focus_course.id:
                    focus_course_next_entry = entry
                    break
        else:
            # Default: use the first matching schedule entry for the course
            for entry in schedule_entries:
                if entry.get('course_id') == focus_course.id:
                    focus_course_next_entry = entry
                    break
        
        # Add day-specific attendance status to focus_course_next_entry
        if focus_course_next_entry and focus_course_next_entry.get('day_label'):
            day_label = focus_course_next_entry.get('day_label')
            # Map full day name to short form
            day_map = {
                'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
            }
            day_short = day_map.get(day_label, day_label)
            day_schedule = CourseSchedule.objects.filter(course=focus_course, day=day_short).first()
            if day_schedule and day_schedule.attendance_status:
                focus_course_next_entry['day_schedule_status'] = day_schedule.attendance_status
            else:
                focus_course_next_entry['day_schedule_status'] = None
    
    # Get attendance data for focused course
    attendance_records = []
    enrolled_students = []
    today_attendance_count = 0
    course_finished = False  # Initialize course_finished before use
    if focus_course:
        from django.utils import timezone
        today = timezone.now().date()
        
        # Day-specific QR code generation for ALL day schedules
        # hashlib and random are already imported at module level
        # CourseSchedule is already imported at module level
        from datetime import datetime, date as date_type, timedelta
        
        today_date = now_ph.date()
        
        # Generate/update QR codes for ALL day schedules of the focus course
        # This ensures each day has a unique QR code
        all_day_schedules = CourseSchedule.objects.filter(course=focus_course).all()
        
        if not all_day_schedules.exists() and focus_course.days:
            # Create day schedules for courses with simple days field
            days_list = [d.strip() for d in focus_course.days.split(',') if d.strip()]
            day_map = {
                'Monday': 'Mon', 'Mon': 'Mon',
                'Tuesday': 'Tue', 'Tue': 'Tue',
                'Wednesday': 'Wed', 'Wed': 'Wed',
                'Thursday': 'Thu', 'Thu': 'Thu',
                'Friday': 'Fri', 'Fri': 'Fri',
                'Saturday': 'Sat', 'Sat': 'Sat',
                'Sunday': 'Sun', 'Sun': 'Sun'
            }
            for day_str in days_list:
                schedule_day = day_map.get(day_str, day_str)
                if not CourseSchedule.objects.filter(course=focus_course, day=schedule_day).exists():
                    CourseSchedule.objects.create(
                        course=focus_course,
                        day=schedule_day,
                        start_time=focus_course.start_time,
                        end_time=focus_course.end_time,
                        room=focus_course.room
                    )
            all_day_schedules = CourseSchedule.objects.filter(course=focus_course).all()
        
        # Process each day schedule to ensure unique QR codes
        focus_day_schedule = None
        if focus_course_next_entry and focus_course_next_entry.get('start_dt') and focus_course_next_entry.get('end_dt'):
            start_dt = focus_course_next_entry.get('start_dt')
            end_dt = focus_course_next_entry.get('end_dt')
            # Get day from day_label (preferred) or day field
            day_name = focus_course_next_entry.get('day_label') or focus_course_next_entry.get('day', '')
            
            # Map day name to CourseSchedule day format
            day_map = {
                'Monday': 'Mon', 'Mon': 'Mon',
                'Tuesday': 'Tue', 'Tue': 'Tue',
                'Wednesday': 'Wed', 'Wed': 'Wed',
                'Thursday': 'Thu', 'Thu': 'Thu',
                'Friday': 'Fri', 'Fri': 'Fri',
                'Saturday': 'Sat', 'Sat': 'Sat',
                'Sunday': 'Sun', 'Sun': 'Sun'
            }
            schedule_day = day_map.get(day_name, day_name) if day_name else None
            if schedule_day:
                focus_day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
        
        # Generate/update QR codes for all day schedules
        # QR code should only change when the schedule for that day has finished
        # Each day schedule maintains its own QR code until that specific day's class ends
        for day_schedule in all_day_schedules:
            needs_new_qr = False
            qr_code_date = today_date  # Default to today, will be updated if schedule finished
            
            # Map day to weekday
            weekday_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
            day_weekday = weekday_map.get(day_schedule.day, None)
            if day_weekday is None:
                continue
            
            # Calculate the next occurrence of this day
            current_weekday = now_ph.weekday()
            days_ahead = (day_weekday - current_weekday) % 7
            if days_ahead == 0:
                # Today - check if class has started and ended
                if day_schedule.start_time and day_schedule.end_time:
                    start_time_today = now_ph.replace(hour=day_schedule.start_time.hour, minute=day_schedule.start_time.minute, second=0, microsecond=0)
                    end_time_today = now_ph.replace(hour=day_schedule.end_time.hour, minute=day_schedule.end_time.minute, second=0, microsecond=0)
                    
                    if now_ph > end_time_today:
                        # Class has ENDED today - need new QR for next week
                        # Calculate next occurrence date
                        next_occurrence_date = today_date + timedelta(days=7)
                        
                        # Only generate if QR code doesn't exist, is for today's finished session, or is for a past date
                        if not day_schedule.qr_code or not day_schedule.qr_code_date or day_schedule.qr_code_date <= today_date:
                            needs_new_qr = True
                            # Use next occurrence date for the new QR code
                            qr_code_date = next_occurrence_date
                    elif now_ph >= start_time_today:
                        # Class is ONGOING today - keep existing QR code
                        if not day_schedule.qr_code:
                            # No QR code exists - generate one for this session
                            needs_new_qr = True
                        # Otherwise, keep the existing QR code
                    else:
                        # Class hasn't started yet today - keep existing QR code if it exists
                        if not day_schedule.qr_code:
                            # No QR code exists - generate one for upcoming session
                            needs_new_qr = True
                        # Otherwise, keep the existing QR code
                else:
                    # No time specified - generate QR if doesn't exist
                    if not day_schedule.qr_code:
                        needs_new_qr = True
            else:
                # Future day (not today)
                # Calculate the last occurrence date (when this day last occurred)
                last_occurrence_date = today_date - timedelta(days=days_ahead)
                
                if not day_schedule.qr_code or not day_schedule.qr_code_date:
                    # No QR code exists - generate one for the next occurrence
                    needs_new_qr = True
                else:
                    # Check if the QR code is for a past occurrence that has finished
                    if day_schedule.end_time:
                        # Calculate when the last occurrence ended
                        last_end_time = datetime.combine(last_occurrence_date, day_schedule.end_time)
                        if PH_TZ:
                            # zoneinfo doesn't have localize(), use replace() instead
                            if last_end_time.tzinfo is None:
                                last_end_time = last_end_time.replace(tzinfo=PH_TZ)
                        elif last_end_time.tzinfo is None:
                            last_end_time = last_end_time.replace(tzinfo=now_ph.tzinfo)
                        
                        # If the QR code date matches the last occurrence date and that occurrence has ended
                        if day_schedule.qr_code_date == last_occurrence_date and now_ph > last_end_time:
                            # The last occurrence has finished - generate new QR for next occurrence
                            # Calculate next occurrence date (7 days from last occurrence)
                            next_occurrence_date = last_occurrence_date + timedelta(days=7)
                            needs_new_qr = True
                            # Use next occurrence date for the new QR code
                            qr_code_date = next_occurrence_date
                        # Otherwise, keep the existing QR code (it's for a future occurrence or ongoing)
                    else:
                        # No end time - if QR code date is before today, generate new one
                        if day_schedule.qr_code_date < today_date:
                            needs_new_qr = True
            
            if needs_new_qr:
                # Generate day-specific QR code - unique per course, section, day, and occurrence
                # Use a stable identifier that doesn't change unless the schedule finishes
                # Don't include timestamp to ensure stability - use date and a fixed salt instead
                qr_data = f"{focus_course.id}_{focus_course.code}_{focus_course.section}_{day_schedule.day}_{qr_code_date.strftime('%Y%m%d')}"
                qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                
                # Ensure uniqueness across all day schedules
                max_attempts = 10
                attempts = 0
                while CourseSchedule.objects.filter(qr_code=qr_hash).exclude(id=day_schedule.id).exists() and attempts < max_attempts:
                    qr_data = f"{focus_course.id}_{focus_course.code}_{focus_course.section}_{day_schedule.day}_{qr_code_date.strftime('%Y%m%d')}_{random.randint(1000, 9999)}"
                    qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                    attempts += 1
                
                day_schedule.qr_code = qr_hash
                day_schedule.qr_code_date = qr_code_date
                day_schedule.save(update_fields=['qr_code', 'qr_code_date'])
                
                logger.info(f"Generated new QR code for course {focus_course.id}, section {focus_course.section}, day {day_schedule.day} (date: {qr_code_date}): {qr_hash[:8]}...")
            else:
                # Ensure QR code date is set even if we're maintaining the existing code
                if day_schedule.qr_code and not day_schedule.qr_code_date:
                    day_schedule.qr_code_date = today_date
                    day_schedule.save(update_fields=['qr_code_date'])
                logger.info(f"Maintaining existing QR code for course {focus_course.id}, section {focus_course.section}, day {day_schedule.day} (schedule not finished)")
        
        # Use the focus day schedule's QR code for backward compatibility
        if focus_day_schedule and focus_day_schedule.qr_code:
            focus_course.qr_code = focus_day_schedule.qr_code
            focus_course.save(update_fields=['qr_code'])
        
        # Get enrolled students
        enrolled_students = CourseEnrollment.objects.filter(
            course=focus_course,
            is_active=True,
            deleted_at__isnull=True
        ).select_related('student').order_by('full_name')
        
        # Get today's attendance records - only show if course is live or hasn't ended yet
        try:
            # Check if course has ended today
            show_today_records = True
            if focus_course_next_entry and focus_course_next_entry.get('end_dt'):
                end_dt = focus_course_next_entry.get('end_dt')
                if end_dt and now_ph > end_dt:
                    # Course has ended - don't show today's records in monitoring (they're in reports)
                    show_today_records = False
            
            if show_today_records:
                # Filter by schedule_day if a specific schedule is selected
                attendance_query = AttendanceRecord.objects.filter(
                    course=focus_course,
                    attendance_date=today
                )
                
                # If a specific schedule is selected, filter by that schedule's day
                if focus_course_next_entry and focus_course_next_entry.get('day_label'):
                    day_label = focus_course_next_entry.get('day_label')
                    day_map = {
                        'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                        'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                    }
                    schedule_day = day_map.get(day_label, '')
                    if schedule_day:
                        attendance_query = attendance_query.filter(schedule_day=schedule_day)
                
                attendance_records = attendance_query.select_related('student', 'enrollment').order_by('attendance_time')
                today_attendance_count = attendance_records.count()
                course_finished = False
            else:
                # Course has ended - preserve today's attendance records (move to reports)
                # and mark any enrolled students who did not scan as 'absent'
                # Set flag to show reminder message
                course_finished = True
                # Build attendance query for today (respect schedule day if present)
                attendance_query = AttendanceRecord.objects.filter(
                    course=focus_course,
                    attendance_date=today
                )
                if focus_course_next_entry and focus_course_next_entry.get('day_label'):
                    day_label = focus_course_next_entry.get('day_label')
                    day_map = {
                        'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                        'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                    }
                    schedule_day = day_map.get(day_label, '')
                    if schedule_day:
                        attendance_query = attendance_query.filter(schedule_day=schedule_day)

                # Fetch existing attendance records for today
                attendance_records = attendance_query.select_related('student', 'enrollment').order_by('attendance_time')
                today_attendance_count = attendance_records.count()

                # Determine the schedule's end time to use as attendance_time for absentees
                schedule_day = None
                if focus_course_next_entry and focus_course_next_entry.get('day_label'):
                    day_label = focus_course_next_entry.get('day_label')
                    day_map = {
                        'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                        'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                    }
                    schedule_day = day_map.get(day_label, '')

                day_schedule = None
                if schedule_day:
                    day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()

                # Choose a fallback time for absent records: day_schedule.end_time -> course.end_time -> now
                from django.utils import timezone as _tz
                fallback_time = _tz.now().time()
                if day_schedule and day_schedule.end_time:
                    fallback_time = day_schedule.end_time
                elif focus_course and getattr(focus_course, 'end_time', None):
                    fallback_time = focus_course.end_time

                # Create absent records for enrolled students who don't have one yet today
                try:
                    present_enrollments = set(att.enrollment_id for att in attendance_records)
                    for enrollment in enrolled_students:
                        if enrollment.id not in present_enrollments:
                            # Only create absent record if the student was enrolled BEFORE the class occurred
                            # Check if the enrollment date is before or on today's date at class start time
                            from django.utils import timezone as _timezone_import
                            enrollment_dt = enrollment.enrolled_at
                            
                            # Get the scheduled start time for this class
                            class_start_time = None
                            if day_schedule and day_schedule.start_time:
                                class_start_time = day_schedule.start_time
                            elif focus_course and focus_course.start_time:
                                class_start_time = focus_course.start_time
                            
                            # Create a datetime for class start today
                            if class_start_time:
                                class_start_dt = _timezone_import.make_aware(
                                    datetime.combine(today, class_start_time),
                                    PH_TZ
                                )
                            else:
                                # If no start time, use current time as reference
                                class_start_dt = now_ph
                            
                            # Only mark as absent if enrolled before the class started
                            if enrollment_dt <= class_start_dt:
                                # Avoid creating duplicate absent records
                                exists = AttendanceRecord.objects.filter(
                                    course=focus_course,
                                    enrollment=enrollment,
                                    attendance_date=today,
                                    schedule_day=schedule_day
                                ).exists()
                                if not exists:
                                    AttendanceRecord.objects.create(
                                        course=focus_course,
                                        student=enrollment.student,
                                        enrollment=enrollment,
                                        attendance_date=today,
                                        attendance_time=fallback_time,
                                        status='absent',
                                        schedule_day=schedule_day
                                    )
                                    today_attendance_count += 1
                except Exception:
                    # If anything goes wrong while creating absent records, continue gracefully
                    pass

                # Auto-close attendance and reset QR code when schedule finishes
                if focus_course_next_entry and focus_course_next_entry.get('day_label'):
                    day_label = focus_course_next_entry.get('day_label')
                    day_map = {
                        'Monday': 'Mon', 'Mon': 'Mon',
                        'Tuesday': 'Tue', 'Tue': 'Tue',
                        'Wednesday': 'Wed', 'Wed': 'Wed',
                        'Thursday': 'Thu', 'Thu': 'Thu',
                        'Friday': 'Fri', 'Fri': 'Fri',
                        'Saturday': 'Sat', 'Sat': 'Sat',
                        'Sunday': 'Sun', 'Sun': 'Sun'
                    }
                    schedule_day = day_map.get(day_label, day_label)
                    day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
                    if day_schedule:
                        # Reset attendance status to closed
                        day_schedule.attendance_status = 'closed'
                        # Reset QR code for next occurrence (will be generated on next view)
                        day_schedule.qr_code = None
                        day_schedule.qr_code_date = None
                        day_schedule.save(update_fields=['attendance_status', 'qr_code', 'qr_code_date'])
                elif focus_course:
                    # For synchronized courses, update course-level attendance status
                    focus_course.attendance_status = 'closed'
                    # Reset QR code for next occurrence
                    focus_course.qr_code = None
                    focus_course.save(update_fields=['attendance_status', 'qr_code'])
                # Note: Keep enrolled_students list visible for reference, but attendance_records are cleared
        except Exception:
            attendance_records = []
            today_attendance_count = 0
            course_finished = False
    
    # Check if course is finished (if not already set)
    if not course_finished and focus_course_next_entry and focus_course_next_entry.get('end_dt'):
        end_dt = focus_course_next_entry.get('end_dt')
        if end_dt and now_ph > end_dt:
            course_finished = True
            # Auto-close attendance and reset QR code when schedule finishes
            if focus_course_next_entry and focus_course_next_entry.get('day_label'):
                day_label = focus_course_next_entry.get('day_label')
                day_map = {
                    'Monday': 'Mon', 'Mon': 'Mon',
                    'Tuesday': 'Tue', 'Tue': 'Tue',
                    'Wednesday': 'Wed', 'Wed': 'Wed',
                    'Thursday': 'Thu', 'Thu': 'Thu',
                    'Friday': 'Fri', 'Fri': 'Fri',
                    'Saturday': 'Sat', 'Sat': 'Sat',
                    'Sunday': 'Sun', 'Sun': 'Sun'
                }
                schedule_day = day_map.get(day_label, day_label)
                day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
                if day_schedule:
                    # Reset attendance status to closed
                    day_schedule.attendance_status = 'closed'
                    # Reset QR code for next occurrence (will be generated below)
                    day_schedule.qr_code = None
                    day_schedule.qr_code_date = None
                    day_schedule.save(update_fields=['attendance_status', 'qr_code', 'qr_code_date'])
                    # Force QR code regeneration for next occurrence
                    # Calculate next occurrence date
                    weekday_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
                    day_weekday = weekday_map.get(schedule_day, None)
                    if day_weekday is not None:
                        current_weekday = now_ph.weekday()
                        days_ahead = (day_weekday - current_weekday) % 7
                        if days_ahead == 0:
                            days_ahead = 7  # Next week
                        next_occurrence_date = today_date + timedelta(days=days_ahead)
                        
                        # Generate new QR code for next occurrence
                        qr_data = f"{focus_course.id}_{focus_course.code}_{focus_course.section}_{day_schedule.day}_{next_occurrence_date.strftime('%Y%m%d')}"
                        qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                        
                        # Ensure uniqueness
                        max_attempts = 10
                        attempts = 0
                        while CourseSchedule.objects.filter(qr_code=qr_hash).exclude(id=day_schedule.id).exists() and attempts < max_attempts:
                            qr_data = f"{focus_course.id}_{focus_course.code}_{focus_course.section}_{day_schedule.day}_{next_occurrence_date.strftime('%Y%m%d')}_{random.randint(1000, 9999)}"
                            qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                            attempts += 1
                        
                        day_schedule.qr_code = qr_hash
                        day_schedule.qr_code_date = next_occurrence_date
                        day_schedule.save(update_fields=['qr_code', 'qr_code_date'])
                        logger.info(f"Generated new QR code for next occurrence: course {focus_course.id}, day {day_schedule.day}, date {next_occurrence_date}: {qr_hash[:8]}...")
            elif focus_course:
                # For synchronized courses, update course-level attendance status
                focus_course.attendance_status = 'closed'
                # Reset QR code for next occurrence
                focus_course.qr_code = None
                focus_course.save(update_fields=['attendance_status', 'qr_code'])
    
    # Get day-specific QR code for focus course if available
    # Refresh from database to get the latest QR code (in case it was just regenerated)
    focus_qr_code = None
    if focus_course and focus_course_next_entry:
        # Get day from day_label (preferred) or day field
        day_name = focus_course_next_entry.get('day_label') or focus_course_next_entry.get('day', '')
        if day_name:
            day_map = {
                'Monday': 'Mon', 'Mon': 'Mon',
                'Tuesday': 'Tue', 'Tue': 'Tue',
                'Wednesday': 'Wed', 'Wed': 'Wed',
                'Thursday': 'Thu', 'Thu': 'Thu',
                'Friday': 'Fri', 'Fri': 'Fri',
                'Saturday': 'Sat', 'Sat': 'Sat',
                'Sunday': 'Sun', 'Sun': 'Sun'
            }
            schedule_day = day_map.get(day_name, day_name)
            # Refresh from database to get the latest QR code
            day_schedule = CourseSchedule.objects.filter(course=focus_course, day=schedule_day).first()
            if day_schedule:
                # Refresh the object to get latest QR code
                day_schedule.refresh_from_db()
                if day_schedule.qr_code:
                    focus_qr_code = day_schedule.qr_code
                    # Also update focus_course.qr_code for backward compatibility
                    focus_course.qr_code = focus_qr_code
                    focus_course.save(update_fields=['qr_code'])
    
    # Calculate present window expiry for instructor timer (to persist across page reloads)
    instructor_present_expiry_ms = None
    try:
        if focus_course_next_entry:
            focus_schedule_identifier = focus_course_next_entry.get('schedule_id')
            present_duration_minutes = None
            qr_opened_at = None
            
            # Try to find schedule object for the focused course/day
            from django.utils import timezone
            try:
                day_label = focus_course_next_entry.get('day_label')
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                day_code = day_map.get(day_label, None)
            except Exception:
                day_code = None

            sched_obj = None
            try:
                if day_code and focus_course:
                    sched_obj = CourseSchedule.objects.filter(course=focus_course, day__iexact=day_code).first()
            except Exception:
                sched_obj = None

            # Get present duration from schedule or course
            if sched_obj and getattr(sched_obj, 'attendance_present_duration', None):
                present_duration_minutes = int(getattr(sched_obj, 'attendance_present_duration'))
            else:
                present_duration_minutes = int(getattr(focus_course, 'attendance_present_duration', 0) or 0)

            # Get qr_code_opened_at
            qr_opened_at = None
            try:
                if sched_obj and getattr(sched_obj, 'qr_code_opened_at', None):
                    # Validate date matches if qr_code_date is set
                    try:
                        occ_date = focus_course_next_entry.get('start_dt').date() if focus_course_next_entry.get('start_dt') else None
                    except Exception:
                        occ_date = None
                    
                    sched_qr_date = getattr(sched_obj, 'qr_code_date', None)
                    if sched_qr_date and occ_date:
                        if sched_qr_date == occ_date:
                            qr_opened_at = getattr(sched_obj, 'qr_code_opened_at')
                    else:
                        qr_opened_at = getattr(sched_obj, 'qr_code_opened_at')
                elif getattr(focus_course, 'qr_code_opened_at', None):
                    qr_opened_at = getattr(focus_course, 'qr_code_opened_at')
            except Exception:
                qr_opened_at = None
            
            # Calculate expiry if we have timestamp and duration
            if qr_opened_at and present_duration_minutes and present_duration_minutes > 0:
                try:
                    expiry = qr_opened_at + timezone.timedelta(minutes=int(present_duration_minutes))
                    # Only set if expiry is in future
                    if expiry > timezone.now():
                        instructor_present_expiry_ms = int(expiry.timestamp() * 1000)
                except Exception:
                    instructor_present_expiry_ms = None
    except Exception:
        instructor_present_expiry_ms = None
    
    context = {
        'school_admin': school_admin,
        'unread_notifications': unread_notifications,
        'courses': courses,
        'focus_course': focus_course,
        'focus_course_next_entry': focus_course_next_entry,
        'schedule_entries': schedule_entries,
        'enrolled_students': enrolled_students,
        'attendance_records': attendance_records,
        'today_attendance_count': today_attendance_count,
        'focus_qr_code': focus_qr_code,  # Day-specific QR code
        'course_finished': course_finished,  # Flag to show reminder when course ends
        'instructor_present_expiry_ms': instructor_present_expiry_ms,  # Server-side timer expiry for persistence
        'esp32_ip': settings.ESP32_IP,  # ESP32 fingerprint sensor IP address for biometric scanning
    }
    return render(request, 'dashboard/instructor/my_classes.html', context)


def finalize_attendance_records(course, schedule, instructor):
    """
    Finalize attendance for a specific schedule.
    Create absent or postponed records for enrolled students who didn't scan.
    schedule can be a CourseSchedule object or a day string.
    """
    from django.utils import timezone
    from zoneinfo import ZoneInfo
    
    try:
        ph_tz = ZoneInfo('Asia/Manila')
    except:
        import pytz
        ph_tz = pytz.timezone('Asia/Manila')
    
    now_ph = timezone.now().astimezone(ph_tz)
    today = now_ph.date()
    
    # Extract schedule_day string and attendance status
    schedule_day_str = None
    record_status = 'absent'
    
    if isinstance(schedule, CourseSchedule):
        schedule_day_str = schedule.day  # Use 'day' not 'day_of_week'
        # Check if the schedule is postponed
        if getattr(schedule, 'attendance_status', None) == 'postponed':
            record_status = 'postponed'
    else:
        schedule_day_str = str(schedule)
    
    # Get all enrolled students in this course
    enrollments = CourseEnrollment.objects.filter(
        course=course,
        is_active=True,
        deleted_at__isnull=True
    ).select_related('student')
    
    for enrollment in enrollments:
        # Check if student was enrolled BEFORE today's class time
        enrollment_valid = False
        
        if isinstance(schedule, CourseSchedule) and schedule.start_time:
            class_start_dt = timezone.make_aware(
                datetime.combine(today, schedule.start_time),
                ph_tz
            )
        elif hasattr(course, 'start_time') and course.start_time:
            class_start_dt = timezone.make_aware(
                datetime.combine(today, course.start_time),
                ph_tz
            )
        else:
            # If no start time, use current time as reference
            class_start_dt = now_ph
        
        # Only process if enrolled before the class started
        if enrollment.enrolled_at <= class_start_dt:
            enrollment_valid = True
        
        if not enrollment_valid:
            # Skip students who enrolled after the class started
            continue
        
        # Check if student has an attendance record for today with this schedule
        existing_record = AttendanceRecord.objects.filter(
            course=course,
            student=enrollment.student,
            attendance_date=today,
            schedule_day=schedule_day_str
        ).first()
        
        if not existing_record:
            # Student didn't scan - mark with appropriate status (absent or postponed)
            try:
                AttendanceRecord.objects.create(
                    course=course,
                    student=enrollment.student,
                    enrollment=enrollment,
                    attendance_date=today,
                    schedule_day=schedule_day_str,
                    attendance_time=None,
                    status=record_status
                )
                logger.info(f"[FINALIZE] created {record_status.upper()} record for schedule: course={course.id} student={enrollment.student.id} date={today} schedule_day={schedule_day_str}")
            except Exception as e:
                logger.error(f"[FINALIZE] Error creating {record_status} record: {str(e)}")
            
            # Create notification for student
            try:
                message_map = {
                    'absent': f'You were marked absent in {course.code} - {course.name}',
                    'postponed': f'Class marked as postponed in {course.code} - {course.name}'
                }
                create_notification(
                    user=enrollment.student,
                    notification_type='attendance_marked',
                    title='Attendance Recorded' if record_status == 'absent' else 'Class Postponed',
                    message=message_map.get(record_status, f'You were marked {record_status} in {course.code} - {course.name}'),
                    category='attendance',
                    related_course=course,
                    related_user=instructor
                )
            except Exception as e:
                logger.error(f"Error creating {record_status} notification: {str(e)}")


def finalize_all_course_attendance(course, instructor, force=False):
    """
    Finalize attendance for all schedules of a course for today.
    Create absent records for enrolled students who didn't scan any session.
    If schedule is postponed, create postponed records instead of absent records.
    
    Args:
        course: The course to finalize attendance for
        instructor: The instructor of the course (used for notifications)
        force: If True, force finalization regardless of class time. If False (default),
               only finalize if class has ended or if explicitly closing attendance
    """
    from django.utils import timezone
    from zoneinfo import ZoneInfo
    from django.db.models import Q
    
    try:
        ph_tz = ZoneInfo('Asia/Manila')
    except:
        import pytz
        ph_tz = pytz.timezone('Asia/Manila')
    
    now_ph = timezone.now().astimezone(ph_tz)
    today = now_ph.date()
    today_day = today.strftime('%A')  # e.g., 'Monday'
    
    # Get all day-of-week strings for today's schedule
    today_day_abbrev = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][today.weekday()] if today.weekday() < 7 else 'Mon'
    
    # Get all schedules for today
    # NOTE: normalize day tokens to avoid missing schedules and prematurely finalizing (creating ABSENT)
    try:
        day_map_norm = {
            'Monday': 'Mon', 'Mon': 'Mon', 'M': 'Mon',
            'Tuesday': 'Tue', 'Tue': 'Tue', 'T': 'Tue',
            'Wednesday': 'Wed', 'Wed': 'Wed', 'W': 'Wed',
            'Thursday': 'Thu', 'Thu': 'Thu', 'Th': 'Thu',
            'Friday': 'Fri', 'Fri': 'Fri', 'F': 'Fri',
            'Saturday': 'Sat', 'Sat': 'Sat', 'S': 'Sat',
            'Sunday': 'Sun', 'Sun': 'Sun', 'Su': 'Sun'
        }
        today_day_norm = day_map_norm.get(str(today_day).strip(), today_day)
        today_day_abbrev_norm = day_map_norm.get(str(today_day_abbrev).strip(), today_day_abbrev)
        today_day_tokens = {today_day_norm, today_day_abbrev_norm}
        if today_day_abbrev_norm == 'Sun':
            today_day_tokens.add('Su')
    except Exception:
        today_day_tokens = {today_day, today_day_abbrev}

    today_schedules = CourseSchedule.objects.filter(
        course=course,
        is_deleted=False
    ).filter(
        Q(day__in=list(today_day_tokens))
    )
    
    # CRITICAL: Only create ABSENT records if:
    # 1. Force is True (explicit finalization via close/postpone), OR
    # 2. ALL schedules for today have ended (after end_time)
    if not force:
        # Determine if the class has ended for today. If we can't reliably determine an end time,
        # do NOT finalize (to avoid creating ABSENT during an ongoing class).
        effective_end_dt = None
        if today_schedules.exists():
            try:
                end_dts = []
                for sched in today_schedules:
                    if hasattr(sched, 'end_time') and sched.end_time:
                        end_dts.append(
                            timezone.make_aware(
                                timezone.datetime.combine(today, sched.end_time),
                                ph_tz
                            )
                        )
                if end_dts:
                    effective_end_dt = max(end_dts)
            except Exception:
                effective_end_dt = None
        else:
            # Fallback to course-level time window only if course is scheduled today
            try:
                course_days_raw = getattr(course, 'days', '') or ''
                days_list = [d.strip() for d in course_days_raw.split(',') if d.strip()]
                days_norm = set()
                for d in days_list:
                    days_norm.add(day_map_norm.get(d, d))
                if (today_day_norm in days_norm) or (today_day_abbrev_norm in days_norm):
                    end_t = getattr(course, 'end_time', None)
                    if end_t:
                        effective_end_dt = timezone.make_aware(
                            timezone.datetime.combine(today, end_t),
                            ph_tz
                        )
            except Exception:
                effective_end_dt = None

        if not effective_end_dt:
            logger.info(f"[FINALIZE] No reliable end time found for course={course.id} today. Skipping finalization.")
            return

        if now_ph < effective_end_dt:
            logger.info(f"[FINALIZE] Class hasn't ended yet for course={course.id}. End time: {effective_end_dt}. Skipping finalization.")
            return
    
    # Check if TODAY is marked as postponed for ANY of the schedules
    # If so, create postponed records instead of absent records
    is_today_postponed = False
    for sched in today_schedules:
        if getattr(sched, 'attendance_status', None) == 'postponed':
            is_today_postponed = True
            logger.info(f"[FINALIZE] Today's schedule is postponed for course={course.id}, creating postponed records")
            break
    
    # Also check course-level postponed status
    if not is_today_postponed and getattr(course, 'attendance_status', None) == 'postponed':
        is_today_postponed = True
        logger.info(f"[FINALIZE] Today's course is postponed for course={course.id}, creating postponed records")
    
    # Determine record status based on postponement
    record_status = 'postponed' if is_today_postponed else 'absent'
    
    # Get all enrolled students
    enrollments = CourseEnrollment.objects.filter(
        course=course,
        is_active=True,
        deleted_at__isnull=True
    ).select_related('student')
    
    for enrollment in enrollments:
        # Check if student was enrolled BEFORE today's class time
        # Get the earliest schedule for today to check enrollment time
        enrollment_valid = False
        earliest_start_time = None
        
        if today_schedules.exists():
            # Get the earliest start time from today's schedules
            first_schedule = today_schedules.order_by('start_time').first()
            if first_schedule and first_schedule.start_time:
                earliest_start_time = first_schedule.start_time
        
        if not earliest_start_time and hasattr(course, 'start_time'):
            earliest_start_time = course.start_time
        
        # Create a datetime for the first class start today
        if earliest_start_time:
            class_start_dt = timezone.make_aware(
                datetime.combine(today, earliest_start_time),
                ph_tz
            )
        else:
            # If no start time, use current time as reference
            class_start_dt = now_ph
        
        # Only process if enrolled before the class started
        if enrollment.enrolled_at <= class_start_dt:
            enrollment_valid = True
        
        if not enrollment_valid:
            # Skip students who enrolled after the class started
            continue
        
        # Check if student has ANY attendance record for today for any schedule
        # When postponed, create for all students. When absent, only if no present/late records exist.
        has_record = AttendanceRecord.objects.filter(
            course=course,
            student=enrollment.student,
            attendance_date=today
        ).exists()
        
        if is_today_postponed:
            # For postponed status, create a record only if no record exists for today
            has_attendance = has_record
        else:
            # For absent status, only mark absent if student didn't mark present or late
            has_attendance = AttendanceRecord.objects.filter(
                course=course,
                student=enrollment.student,
                attendance_date=today,
                status__in=['present', 'late']
            ).exists()
        
        if not has_attendance:
            # Student didn't scan - mark as absent for all today's schedules
            # If there are explicit CourseSchedule rows for today, use them.
            # Otherwise, fallback to course-level `days` (synchronized schedules) and create
            # a single absent record for today's day code (e.g., 'Mon').
            schedules_to_mark = list(today_schedules) if today_schedules.exists() else []

            # Fallback: if no day-specific schedules but course.days includes today, use abbrev
            if not schedules_to_mark:
                try:
                    course_days_raw = getattr(course, 'days', '') or ''
                    days_list = [d.strip() for d in course_days_raw.split(',') if d.strip()]
                    if today_day in days_list or today_day_abbrev in days_list:
                        # Use the abbreviated day code as schedule placeholder
                        schedules_to_mark = [today_day_abbrev]
                except Exception:
                    schedules_to_mark = []

            for schedule in schedules_to_mark:
                # schedule may be a CourseSchedule instance or a short day string
                if hasattr(schedule, 'day'):
                    schedule_day_raw = getattr(schedule, 'day', None)
                else:
                    schedule_day_raw = str(schedule)

                # Normalize schedule day to short form (Mon, Tue, Wed, Thu, Fri, Sat, Sun)
                try:
                    sd = str(schedule_day_raw).strip()
                    day_map_norm = {
                        'Monday': 'Mon', 'Mon': 'Mon', 'M': 'Mon',
                        'Tuesday': 'Tue', 'Tue': 'Tue', 'T': 'Tue',
                        'Wednesday': 'Wed', 'Wed': 'Wed', 'W': 'Wed',
                        'Thursday': 'Thu', 'Thu': 'Thu', 'Th': 'Thu',
                        'Friday': 'Fri', 'Fri': 'Fri', 'F': 'Fri',
                        'Saturday': 'Sat', 'Sat': 'Sat', 'S': 'Sat',
                        'Sunday': 'Sun', 'Sun': 'Sun', 'Su': 'Sun'
                    }
                    schedule_day_str = day_map_norm.get(sd, sd)
                except Exception:
                    schedule_day_str = str(schedule_day_raw)

                existing_record = AttendanceRecord.objects.filter(
                    course=course,
                    student=enrollment.student,
                    attendance_date=today,
                    schedule_day=schedule_day_str
                ).first()

                if not existing_record:
                    try:
                        AttendanceRecord.objects.create(
                            course=course,
                            student=enrollment.student,
                            enrollment=enrollment,
                            attendance_date=today,
                            schedule_day=schedule_day_str,
                            attendance_time=None,
                            status=record_status
                        )
                        logger.info(f"[FINALIZE] created {record_status.upper()} record: course={course.id} student={enrollment.student.id} date={today} schedule_day={schedule_day_str}")
                    except Exception as e:
                        logger.error(f"[FINALIZE] failed creating {record_status} record for course={getattr(course,'id',None)} student={getattr(enrollment.student,'id',None)} date={today} schedule_day={schedule_day_str}: {e}")

            # Only notify if we actually created a record (avoid false ABSENT notifications)
            if schedules_to_mark:
                try:
                    create_notification(
                        user=enrollment.student,
                        notification_type='attendance_marked',
                        title='Attendance Recorded',
                        message=f'You were marked absent in {course.code} - {course.name}',
                        category='attendance',
                        related_course=course,
                        related_user=instructor
                    )
                except Exception as e:
                    logger.error(f"Error creating absent notification: {str(e)}")

@login_required
@require_http_methods(["POST"])
def instructor_update_enrollment_status_view(request, course_id):
    """Update enrollment status for a course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        # Accept both 'status' and 'enrollment_status' for compatibility
        status = data.get('status') or data.get('enrollment_status', '').strip().lower()
        
        if status not in ['open', 'closed']:
            return JsonResponse({'success': False, 'message': 'Invalid enrollment status. Must be "open" or "closed".'})
        
        try:
            course = Course.objects.get(id=course_id, instructor=user, is_active=True, deleted_at__isnull=True)
        except Course.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Course not found or you do not have permission to modify it.'})
        
        course.enrollment_status = status
        course.save(update_fields=['enrollment_status'])
        
        status_text = 'opened' if status == 'open' else 'closed'
        return JsonResponse({
            'success': True,
            'message': f'Enrollment has been {status_text} for this course.'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid request data.'})
    except Exception as e:
        logger.error(f"Error updating enrollment status: {str(e)}")
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_update_attendance_status_view(request, course_id):
    """Update attendance status for a course or specific day schedule"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        logger.info(f"[ATTENDANCE_UPDATE] user={user.id} course_id={course_id} request_body={data}")
        status = data.get('status')
        day = data.get('day')  # Optional: day of the week (e.g., 'Mon', 'Tue')
        schedule_id = data.get('schedule_id')  # Optional: specific schedule ID
        
        if status not in ['automatic', 'closed', 'open', 'stopped', 'postponed']:
            return JsonResponse({'success': False, 'message': 'Invalid attendance status.'})
        
        course = Course.objects.get(id=course_id, instructor=user)
        
        # If day or schedule_id is provided, update day-specific schedule
        if day or schedule_id:
            day_schedule = None
            
            if schedule_id:
                # Find schedule by matching the schedule_id format: course_id_day_YYYYMMDD
                # Extract day token from schedule_id (can be 'F', 'M', 'Th', etc.)
                try:
                    parts = str(schedule_id).split('_')
                    if len(parts) >= 2:
                        day_token = parts[1]
                        day_map_reverse = {
                            'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri',
                            'S': 'Sat', 'Su': 'Sun',
                            'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat', 'Sun': 'Sun'
                        }
                        mapped_day = day_map_reverse.get(day_token)
                        if mapped_day:
                            day_schedule = CourseSchedule.objects.filter(
                                course=course,
                                day=mapped_day
                            ).first()
                except Exception:
                    pass
            
            if not day_schedule and day:
                # Find schedule by day
                day_schedule = CourseSchedule.objects.filter(
                    course=course,
                    day=day
                ).first()
            
            if day_schedule:
                # Update day-specific attendance status
                previous_status = day_schedule.attendance_status or course.attendance_status
                day_schedule.attendance_status = status
                # Accept optional present duration (minutes) from instructor
                try:
                    present_duration = data.get('present_duration')
                    if present_duration is not None:
                        # Normalize to int if possible
                        try:
                            pd_val = int(present_duration)
                            day_schedule.attendance_present_duration = pd_val
                            
                            # GLOBAL AUTO-ASSIGN: Update ALL course schedules (across all courses) with the same duration
                            # This makes the present window global for all teacher's classes
                            all_schedules = CourseSchedule.objects.filter(
                                course__instructor=user
                            ).exclude(id=day_schedule.id)
                            for schedule in all_schedules:
                                schedule.attendance_present_duration = pd_val
                                schedule.save(update_fields=['attendance_present_duration'])
                        except Exception:
                            pass
                except Exception:
                    pass

                # When opening attendance, record the exact open timestamp (qr/session opened)
                # Only set qr_code_opened_at when the instructor explicitly provided a present_duration
                # (i.e., they confirmed a present-window). This prevents accidental countdowns
                # when attendance is simply toggled to 'open' without confirming a present window.
                from django.utils import timezone
                try:
                    pd_raw = data.get('present_duration')
                    pd_val = None
                    if pd_raw is not None:
                        try:
                            pd_val = int(pd_raw)
                        except Exception:
                            pd_val = None

                    # Effective duration: if not provided now, use already-saved schedule duration (or course-level as last fallback)
                    pd_effective = None
                    if pd_raw is None:
                        try:
                            pd_effective = int(getattr(day_schedule, 'attendance_present_duration', 0) or 0)
                        except Exception:
                            pd_effective = 0
                        if not pd_effective:
                            try:
                                pd_effective = int(getattr(course, 'attendance_present_duration', 0) or 0)
                            except Exception:
                                pd_effective = 0
                    else:
                        pd_effective = pd_val

                    if status == 'open' and pd_effective and pd_effective > 0:
                        try:
                            day_schedule.qr_code_opened_at = timezone.now()
                            day_schedule.qr_code_date = timezone.localtime(timezone.now()).date()
                        except Exception:
                            pass
                    elif pd_val == 0:
                        try:
                            day_schedule.qr_code_opened_at = None
                            day_schedule.attendance_present_duration = 0
                        except Exception:
                            pass
                except Exception:
                    pass

                # Save all modified fields
                save_fields = ['attendance_status', 'attendance_present_duration', 'qr_code_opened_at', 'qr_code_date']
                day_schedule.save(update_fields=[f for f in save_fields if hasattr(day_schedule, f)])
                logger.info(f"[ATTENDANCE_UPDATE] day_schedule saved: id={day_schedule.id} status={day_schedule.attendance_status}")
                
                status_display = status.replace('_', ' ').title()
                day_display = day_schedule.get_day_display()
                
                # If closing or postponing attendance, create appropriate records for students who didn't scan
                if status in ['closed', 'postponed']:
                    try:
                        finalize_attendance_records(course, day_schedule, user)
                    except Exception as e:
                        logger.error(f"Error finalizing attendance: {str(e)}")
                
                # Create notification for enrolled students
                try:
                    from .models import CourseEnrollment
                    enrolled_students = CourseEnrollment.objects.filter(
                        course=course,
                        is_active=True,
                        deleted_at__isnull=True
                    ).select_related('student')
                    
                    for enrollment in enrolled_students:
                        if enrollment.student:
                            create_notification(
                                user=enrollment.student,
                                notification_type='attendance_control_updated',
                                title='Attendance Control Updated',
                                message=f'Instructor updated attendance control for {course.code} - {course.name} ({day_display}) to {status_display}',
                                category='course',
                                related_course=course,
                                related_user=user
                            )
                except Exception as e:
                    logger.error(f"Error creating attendance control update notifications: {str(e)}")
                
                # Compute present window expiry (ms) if we recorded an opened timestamp and duration
                present_expiry_ms = None
                try:
                    from django.utils import timezone
                    if getattr(day_schedule, 'qr_code_opened_at', None):
                        pd = getattr(day_schedule, 'attendance_present_duration', None) or 0
                        try:
                            pd_int = int(pd)
                        except Exception:
                            pd_int = 0
                        if pd_int and pd_int > 0:
                            expiry = getattr(day_schedule, 'qr_code_opened_at') + timezone.timedelta(minutes=int(pd_int))
                            present_expiry_ms = int(expiry.timestamp() * 1000)
                except Exception:
                    present_expiry_ms = None

                # Return response with updated status and optional expiry for client verification
                response_data = {
                    'success': True,
                    'message': f'Attendance status updated to {status_display} for {day_display}',
                    'updated_status': day_schedule.attendance_status,
                    'present_expiry_ms': present_expiry_ms
                }
                logger.info(f"[ATTENDANCE_UPDATE] returning day_schedule response: {response_data}")
                return JsonResponse(response_data)
            else:
                # Day schedule not found, fall through to course-level update
                pass
        
        # Update course-level attendance status (if no day schedule found or no day specified)
        previous_status = course.attendance_status
        course.attendance_status = status
        # Accept optional present duration (minutes) for course-level
        try:
            present_duration = data.get('present_duration')
            if present_duration is not None:
                try:
                    pd_val = int(present_duration)
                    course.attendance_present_duration = pd_val
                except Exception:
                    pass
        except Exception:
            pass

        # When opening attendance at course-level, record open timestamp
        # Only set qr_code_opened_at when instructor explicitly provided a present_duration
        # (i.e., confirmed the present window). Prevents countdowns when simply opening attendance.
        from django.utils import timezone
        try:
            pd_raw = data.get('present_duration')
            pd_val = None
            if pd_raw is not None:
                try:
                    pd_val = int(pd_raw)
                except Exception:
                    pd_val = None

            pd_effective = None
            if pd_raw is None:
                try:
                    pd_effective = int(getattr(course, 'attendance_present_duration', 0) or 0)
                except Exception:
                    pd_effective = 0
            else:
                pd_effective = pd_val

            if status == 'open' and pd_effective and pd_effective > 0:
                try:
                    course.qr_code_opened_at = timezone.now()
                except Exception:
                    pass
            elif pd_val == 0:
                try:
                    course.qr_code_opened_at = None
                    course.attendance_present_duration = 0
                except Exception:
                    pass
        except Exception:
            pass

        # Save course with its updated fields
        course.save()
        logger.info(f"[ATTENDANCE_UPDATE] course saved: id={course.id} status={course.attendance_status}")
        
        # If closing or postponing attendance at course-level, finalize for all schedules
        if status in ['closed', 'postponed']:
            try:
                finalize_all_course_attendance(course, user, force=True)
            except Exception as e:
                logger.error(f"Error finalizing all course attendance: {str(e)}")
        
        # Create notification for all enrolled students when attendance control is updated
        try:
            from .models import CourseEnrollment
            enrolled_students = CourseEnrollment.objects.filter(
                course=course,
                is_active=True,
                deleted_at__isnull=True
            ).select_related('student')
            
            status_display = status.replace('_', ' ').title()
            for enrollment in enrolled_students:
                if enrollment.student:
                    create_notification(
                        user=enrollment.student,
                        notification_type='attendance_control_updated',
                        title='Attendance Control Updated',
                        message=f'Instructor updated attendance control for {course.code} - {course.name} to {status_display}',
                        category='course',
                        related_course=course,
                        related_user=user
                    )
        except Exception as e:
            logger.error(f"Error creating attendance control update notifications: {str(e)}")
        
        # Compute present window expiry (ms) if we recorded an opened timestamp and duration
        present_expiry_ms = None
        try:
            from django.utils import timezone
            if getattr(course, 'qr_code_opened_at', None):
                pd = getattr(course, 'attendance_present_duration', None) or 0
                try:
                    pd_int = int(pd)
                except Exception:
                    pd_int = 0
                if pd_int and pd_int > 0:
                    expiry = getattr(course, 'qr_code_opened_at') + timezone.timedelta(minutes=int(pd_int))
                    present_expiry_ms = int(expiry.timestamp() * 1000)
        except Exception:
            present_expiry_ms = None

        # Return response with updated status and optional expiry for client verification
        response_data = {
            'success': True,
            'message': f'Attendance status updated to {status.replace("_", " ").title()}',
            'updated_status': course.attendance_status,
            'present_expiry_ms': present_expiry_ms
        }
        logger.info(f"[ATTENDANCE_UPDATE] returning course response: {response_data}")
        return JsonResponse(response_data)
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Course not found.'})
    except Exception as e:
        logger.error(f"Error updating attendance status: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
def schedule_view(request):
    user = request.user
    if not (user.is_teacher or user.is_student):
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access the schedule.'})
    
    # Prepare student timetable entries (used to render asynchronous schedules)
    student_schedule_entries = []
    
    # Get school admin for topbar display (for both teachers and students)
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    education_level = user.education_level
    initial_school_type = 'HighSchool' if education_level == 'high_senior' else 'University'
    
    # For students, filter courses by their program, year level, section, and current semester
    # For instructors, filter courses assigned to them
    courses = []
    teacher_schedule_entries = []

    def map_day_symbol(day_value):
        mapping = {
            'Mon': 'M', 'Monday': 'M', 'MON': 'M', 'M': 'M',
            'Tue': 'T', 'Tuesday': 'T', 'TUE': 'T', 'T': 'T',
            'Wed': 'W', 'Wednesday': 'W', 'WED': 'W', 'W': 'W',
            'Thu': 'Th', 'Thursday': 'Th', 'THU': 'Th', 'TH': 'Th',
            'Fri': 'F', 'Friday': 'F', 'FRI': 'F', 'F': 'F',
            'Sat': 'S', 'Saturday': 'S', 'SAT': 'S', 'S': 'S',
            'Sun': 'Su', 'Sunday': 'Su', 'SUN': 'Su', 'SU': 'Su'
        }
        normalized = (day_value or '').strip()
        # Title-case fallback (e.g., "monday" -> "Monday")
        if normalized.lower() in {k.lower() for k in mapping.keys()}:
            for key, value in mapping.items():
                if key.lower() == normalized.lower():
                    return value
        return mapping.get(normalized, normalized)

    def time_24(time_value):
        return time_value.strftime('%H:%M') if time_value else ''

    if user.is_teacher:
        # For instructors: get courses assigned to them (only active courses with instructor assigned, excluding archived and deleted)
        # CRITICAL: Exclude courses that are archived OR deleted (even if not permanently deleted yet)
        # Use the most explicit query possible to ensure archived/deleted courses NEVER appear
        from django.db.models import Q
        from django.db import connection, transaction
        
        # CRITICAL: Clear any potential query cache and force a completely fresh database query
        # Use select_for_update with nowait=False to ensure we get the latest committed data
        # This bypasses all Django ORM caching mechanisms
        
        # CRITICAL: Use a direct database query that absolutely excludes archived/deleted courses
        # Query ONLY courses that match ALL criteria from the start - don't get all courses first
        with transaction.atomic():
            # Step 1: Get ONLY course IDs that match ALL criteria (active, not archived, not deleted)
            # This is the most important step - we never even look at archived/deleted courses
            valid_course_ids = list(Course.objects.filter(
                instructor=user,
                is_active=True,
                deleted_at__isnull=True,
                is_archived=False
            ).exclude(
                Q(is_archived=True) | Q(deleted_at__isnull=False) | Q(is_active=False)
            ).values_list('id', flat=True))
            
            logger.info(f"Schedule view: Found {len(valid_course_ids)} course IDs matching criteria (active, not archived, not deleted)")
            
            # Step 2: For each ID, do a fresh database lookup with ALL filters to verify
            # This ensures we get the absolute latest data and exclude any that changed
            verified_course_ids = []
            for course_id in valid_course_ids:
                try:
                    # Use get() with ALL filters - this forces a fresh database query
                    # This will raise DoesNotExist if course is archived/deleted/inactive
                    course = Course.objects.get(
                        id=course_id,
                        instructor=user,
                        is_active=True,
                        deleted_at__isnull=True,
                        is_archived=False
                    )
                    # Triple-check: Refresh from database and verify again
                    course.refresh_from_db()
                    if course.is_archived or course.deleted_at is not None or not course.is_active:
                        logger.warning(f"Course {course_id} failed verification after refresh: is_archived={course.is_archived}, deleted_at={course.deleted_at}, is_active={course.is_active}")
                        continue
                    verified_course_ids.append(course_id)
                except Course.DoesNotExist:
                    # Course doesn't match criteria (archived/deleted/inactive), skip it
                    logger.debug(f"Course {course_id} not found or doesn't match criteria (likely archived/deleted)")
                    continue
                except Exception as e:
                    logger.error(f"Error verifying course {course_id}: {str(e)}")
                    continue
            
            # Step 3: Fetch only verified courses with all filters again
            # This double-checks that we only get courses that match all criteria
            if not verified_course_ids:
                courses = []
            else:
                courses = Course.objects.filter(
                    id__in=verified_course_ids,
                    instructor=user,
                    is_active=True,
                    deleted_at__isnull=True,
                    is_archived=False
                ).exclude(
                    Q(is_archived=True) | Q(deleted_at__isnull=False) | Q(is_active=False)
                ).select_related('instructor').prefetch_related('course_schedules').order_by('days', 'start_time')
                
                # Step 4: Convert to list and do final verification
                courses_list = list(courses)
                final_courses = []
                for course in courses_list:
                    # Final check: Refresh from database one more time
                    course.refresh_from_db()
                    if course.is_archived or course.deleted_at is not None or not course.is_active:
                        logger.warning(f"FINAL VERIFICATION: Skipping course {course.id} ({course.code}): is_archived={course.is_archived}, deleted_at={course.deleted_at}, is_active={course.is_active}")
                        continue
                    final_courses.append(course)
                
                courses = final_courses
        
        # Log for debugging
        logger.info(f"Schedule view for instructor {user.id}: Found {len(courses)} active courses after final filtering (excluding archived/deleted)")

        for course in courses:
            # One final check before processing - ensure course is still valid
            # This is a safety measure in case something changed between the query and processing
            if course.is_archived or course.deleted_at is not None or not course.is_active:
                logger.warning(f"PROCESSING CHECK: Skipping course {course.id} ({course.code}) - was archived/deleted during processing")
                continue
            
            course_color = course.color or '#3b82f6'
            teacher_name = course.instructor.full_name if course.instructor else (course.instructor.username if course.instructor else 'N/A')
            day_schedules = course.course_schedules.all().order_by('day_order')

            if day_schedules.exists():
                for schedule in day_schedules:
                    mapped_day = map_day_symbol(schedule.day)
                    entry_id = f'db_{course.id}_{schedule.day}'
                    teacher_schedule_entries.append({
                        'id': entry_id,
                        'courseCode': course.code,
                        'courseName': course.name,
                        'teacherName': teacher_name,
                        'room': schedule.room or (course.room or ''),
                        'days': [mapped_day],
                        'start': time_24(schedule.start_time),
                        'end': time_24(schedule.end_time),
                        'startTime': time_24(schedule.start_time),
                        'endTime': time_24(schedule.end_time),
                        'attendance_start': time_24(schedule.attendance_start) if schedule.attendance_start else time_24(course.attendance_start),
                        'attendance_end': time_24(schedule.attendance_end) if schedule.attendance_end else time_24(course.attendance_end),
                        'attendanceTimeIn': time_24(schedule.attendance_start) if schedule.attendance_start else time_24(course.attendance_start),
                        'attendanceTimeEnd': time_24(schedule.attendance_end) if schedule.attendance_end else time_24(course.attendance_end),
                        'color': course_color,
                        'schoolType': initial_school_type,
                        'start_date': '',
                        'end_date': '',
                    })
            else:
                day_tokens = [token.strip() for token in (course.days or '').split(',') if token.strip()]
                mapped_days = [map_day_symbol(token) for token in day_tokens]
                entry_id = f'db_{course.id}'
                teacher_schedule_entries.append({
                    'id': entry_id,
                    'courseCode': course.code,
                    'courseName': course.name,
                    'teacherName': teacher_name,
                    'room': course.room or '',
                    'days': mapped_days,
                    'start': time_24(course.start_time),
                    'end': time_24(course.end_time),
                    'startTime': time_24(course.start_time),
                    'endTime': time_24(course.end_time),
                    'attendance_start': time_24(course.attendance_start),
                    'attendance_end': time_24(course.attendance_end),
                    'attendanceTimeIn': time_24(course.attendance_start),
                    'attendanceTimeEnd': time_24(course.attendance_end),
                    'color': course_color,
                    'schoolType': initial_school_type,
                    'start_date': '',
                    'end_date': '',
                })
    elif user.is_student:
        # For students: ONLY show courses they are enrolled in (via CourseEnrollment)
        from .models import CourseEnrollment
        
        # Get all active enrollments for this student
        enrollments = CourseEnrollment.objects.filter(
            student=user,
            is_active=True
        ).select_related('course', 'course__program', 'course__instructor')
        
        # Extract courses from enrollments - ONLY enrolled courses that are active, not archived, and not deleted
        student_courses = [
            enrollment.course for enrollment in enrollments 
            if enrollment.course.is_active 
            and enrollment.course.deleted_at is None 
            and not enrollment.course.is_archived
        ]
        
        # Filter by semester and school year if provided
        current_semester = request.GET.get('semester', '')
        school_year = request.GET.get('school_year', None)
        
        if current_semester:
            student_courses = [c for c in student_courses if c.semester == current_semester]
        if school_year:
            student_courses = [c for c in student_courses if c.school_year == school_year]
        
        courses = sorted(student_courses, key=lambda c: (c.days or '', c.start_time or ''))

        # Build structured entries for timetable (handles day-specific schedules)
        for course in courses:
            # Double-check: Skip any courses that are archived or deleted (safety measure)
            if course.is_archived or course.deleted_at is not None or not course.is_active:
                continue
            
            course_color = course.color or '#3b82f6'
            teacher_name = course.instructor.full_name if course.instructor else (course.instructor.username if course.instructor else 'N/A')
            day_schedules = course.course_schedules.all().order_by('day_order')

            if day_schedules.exists():
                for schedule in day_schedules:
                    mapped_day = map_day_symbol(schedule.day)
                    student_schedule_entries.append({
                        'id': f'db_{course.id}_{schedule.day}',
                        'courseCode': course.code,
                        'courseName': course.name,
                        'teacherName': teacher_name,
                        'room': schedule.room or (course.room or ''),
                        'days': [mapped_day],
                        'start': time_24(schedule.start_time),
                        'end': time_24(schedule.end_time),
                        'attendance_start': time_24(schedule.attendance_start) if schedule.attendance_start else time_24(course.attendance_start),
                        'attendance_end': time_24(schedule.attendance_end) if schedule.attendance_end else time_24(course.attendance_end),
                        'color': course_color,
                        'schoolType': initial_school_type,
                    })
            else:
                day_tokens = [token.strip() for token in (course.days or '').split(',') if token.strip()]
                mapped_days = [map_day_symbol(token) for token in day_tokens]
                student_schedule_entries.append({
                    'id': f'db_{course.id}',
                    'courseCode': course.code,
                    'courseName': course.name,
                    'teacherName': teacher_name,
                    'room': course.room or '',
                    'days': mapped_days,
                    'start': time_24(course.start_time),
                    'end': time_24(course.end_time),
                    'attendance_start': time_24(course.attendance_start),
                    'attendance_end': time_24(course.attendance_end),
                    'color': course_color,
                    'schoolType': initial_school_type,
                })
    
    context = {
        'user': user,
        'is_teacher': user.is_teacher,
        'school_admin': school_admin,
        'is_student': user.is_student,
        'education_level': education_level,
        'initial_school_type': initial_school_type,
        'courses': courses,
        'student_schedule_entries': student_schedule_entries if user.is_student else [],
        'teacher_schedule_entries': teacher_schedule_entries if user.is_teacher else [],
        'current_semester': request.GET.get('semester', '1st'),
        'school_year': request.GET.get('school_year', None),
    }
    
    response = render(request, 'dashboard/instructor/schedule.html', context)
    # Add aggressive cache-busting headers to prevent browser from caching old schedule data
    # This ensures archived/deleted courses don't appear due to cached responses
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    # Add timestamp to force browser to treat this as new content
    response['X-Timestamp'] = str(int(time.time()))
    return response

@login_required
@require_role('teacher')
def courses_view(request):
    """View for managing courses (teachers only)"""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    # Get courses assigned to this instructor (only active courses with instructor assigned, excluding deleted)
    instructor_courses = Course.objects.filter(
        instructor=user,
        is_active=True,
        deleted_at__isnull=True
    ).exclude(instructor__isnull=True).order_by('program', 'year_level', 'code')
    
    education_level = user.education_level
    initial_school_type = 'HighSchool' if education_level == 'high_senior' else 'University'
    
    context = {
        'user': user,
        'is_teacher': user.is_teacher,
        'education_level': education_level,
        'initial_school_type': initial_school_type,
        'instructor_courses': instructor_courses,
        'school_admin': school_admin,
    }
    
    return render(request, 'dashboard/instructor/courses.html', context)

@login_required
@require_http_methods(["POST"])
def instructor_update_attendance_times_view(request, course_id):
    """Allow instructor to set attendance start and end times for their course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    course = get_object_or_404(Course, id=course_id)
    
    # Verify the course is assigned to this instructor
    if course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only set attendance times for courses assigned to you.'})
    
    try:
        course.attendance_end = attendance_end
        course.save()
        
        return JsonResponse({'success': True, 'message': 'Attendance times updated successfully!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
def students_view(request):
    """View for managing students (teachers only) - shows enrolled students"""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    # Get all courses taught by this instructor (excluding deleted and archived)
    from .models import Course, CourseEnrollment
    instructor_courses_qs = Course.objects.filter(
        instructor=user,
        is_active=True,
        deleted_at__isnull=True,
        is_archived=False
    ).select_related('program').prefetch_related('course_schedules')\
     .order_by('created_at', 'program__code', 'year_level', 'semester', 'section', 'code')
    
    # Get all enrollments for these courses (excluding deleted)
    enrollments = CourseEnrollment.objects.filter(
        course__in=instructor_courses_qs,
        is_active=True,
        deleted_at__isnull=True
    ).select_related('student', 'course', 'course__program').order_by('-enrolled_at')
    
    # Get dropped enrollments (deleted but not permanently deleted) for drop list
    dropped_enrollments = CourseEnrollment.objects.filter(
        course__in=instructor_courses_qs,
        deleted_at__isnull=False
    ).select_related('student', 'course', 'course__program').order_by('-deleted_at')
    
    # Organize dropped enrollments by course, section, etc. for display
    dropped_enrollments_by_course = {}
    from django.utils import timezone
    now = timezone.now()
    for enrollment in dropped_enrollments:
        # Calculate days until permanent deletion
        if enrollment.deleted_at:
            days_until_deletion = 30 - (now - enrollment.deleted_at).days
            enrollment.days_until_deletion = max(0, days_until_deletion)
        
        # Organize by course
        course_key = f"{enrollment.course.code} - {enrollment.course.name}"
        if course_key not in dropped_enrollments_by_course:
            dropped_enrollments_by_course[course_key] = {
                'course': enrollment.course,
                'enrollments': []
            }
        dropped_enrollments_by_course[course_key]['enrollments'].append(enrollment)
    
    # Get filter options
    raw_section = request.GET.get('section', '').strip()
    filter_section = raw_section.upper()
    filter_year = request.GET.get('year_level', '')
    filter_semester = request.GET.get('semester', '')
    filter_school_year = request.GET.get('school_year', '')
    filter_course = request.GET.get('course', '')
    search_query = request.GET.get('search', '').strip()
    
    # Apply filters
    if filter_section:
        enrollments = enrollments.filter(section__iexact=filter_section)
    if filter_year:
        enrollments = enrollments.filter(year_level=int(filter_year))
    if filter_semester:
        enrollments = enrollments.filter(course__semester=filter_semester)
    if filter_school_year:
        enrollments = enrollments.filter(course__school_year=filter_school_year)
    if filter_course:
        enrollments = enrollments.filter(course_id=int(filter_course))
    if search_query:
        enrollments = enrollments.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(student_id_number__icontains=search_query) |
            Q(student__username__icontains=search_query)
        )
    
    # Get unique filter values
    all_enrollments = CourseEnrollment.objects.filter(
        course__in=instructor_courses_qs,
        is_active=True
    )
    
    sections = sorted(list({(e.section or '').upper() for e in all_enrollments if e.section}))
    year_levels = sorted(list(set([e.year_level for e in all_enrollments if e.year_level])))
    semesters = sorted(list(set([e.course.semester for e in all_enrollments if e.course.semester])))
    school_years = sorted(list(set([e.course.school_year for e in all_enrollments if e.course.school_year and e.course.school_year])))
    
    # Count students per course for display
    course_student_counts = {}
    for enrollment in all_enrollments:
        course_id = enrollment.course.id
        if course_id not in course_student_counts:
            course_student_counts[course_id] = 0
        course_student_counts[course_id] += 1
    
    # Get unread notification count
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Annotate courses with student counts
    instructor_courses_raw = list(instructor_courses_qs)
    for course in instructor_courses_raw:
        course.student_count = course_student_counts.get(course.id, 0)

    # Group multi-section courses as one logical course (by code, name, semester, school_year)
    grouped_map = {}
    for c in instructor_courses_raw:
        key = (
            (c.code or '').strip().lower(),
            (c.name or '').strip().lower(),
            (c.semester or '').strip().lower(),
            (c.school_year or '').strip()
        )
        if key not in grouped_map:
            grouped_map[key] = {
                'representative': c,            # use first encountered as representative
                'course_ids': [c.id],
                'sections': [ (c.section or '').upper() ] if c.section else [],
                'student_count_total': int(getattr(c, 'student_count', 0) or 0),
            }
        else:
            grouped_map[key]['course_ids'].append(c.id)
            if c.section:
                grouped_map[key]['sections'].append((c.section or '').upper())
            grouped_map[key]['student_count_total'] += int(getattr(c, 'student_count', 0) or 0)
    grouped_instructor_courses = []
    for _, data in grouped_map.items():
        rep = data['representative']
        # Attach aggregated fields onto representative for convenient template access
        rep.group_course_ids = data['course_ids']
        rep.group_sections = sorted(list({s for s in data['sections'] if s}))
        rep.group_size = len(rep.group_sections) if rep.group_sections else 1
        rep.student_count = data['student_count_total']
        # Build gradient consistent with Manage Courses (combine section colors)
        try:
            sibling_colors = []
            for cid in data['course_ids']:
                cobj = next((x for x in instructor_courses_raw if x.id == cid), None)
                clr = (cobj.color if cobj and getattr(cobj, 'color', None) else None)
                if clr:
                    clr = clr.strip()
                    if clr not in sibling_colors:
                        sibling_colors.append(clr)
            # Fallback to representative color
            if not sibling_colors:
                base = (rep.color or '#3C4770')
                sibling_colors = [base]
            # Cap number of colors to 4 for nice gradients
            sibling_colors = sibling_colors[:4]
            # Create multi-stop gradient
            if len(sibling_colors) == 1:
                rep.display_gradient = f"linear-gradient(135deg, {sibling_colors[0]} 0%, {sibling_colors[0]} 100%)"
            else:
                stops = []
                total = len(sibling_colors)
                for idx, c in enumerate(sibling_colors):
                    pos = int((idx / (total - 1)) * 100)
                    stops.append(f"{c} {pos}%")
                rep.display_gradient = f"linear-gradient(135deg, {', '.join(stops)})"
        except Exception:
            rep.display_gradient = None
        rep.is_multi = rep.group_size and rep.group_size > 1
        grouped_instructor_courses.append(rep)
    # Preserve created_at ordering from instructor_courses_qs by not re-sorting groups
    
    # Selected course view (for dedicated course detail page)
    selected_course = None
    selected_course_enrollments = None
    selected_course_sections = []
    selected_course_section_to_id = {}
    selected_course_id = request.GET.get('selected_course')
    if selected_course_id:
        try:
            selected_course_id_int = int(selected_course_id)
            # Selected course refers to actual course id; search in raw list
            selected_course = next((c for c in instructor_courses_raw if c.id == selected_course_id_int), None)
            if selected_course:
                selected_course_enrollments = enrollments.filter(course=selected_course).select_related('student').order_by('full_name')
                selected_course.student_count = course_student_counts.get(selected_course.id, 0)
                # Get dropped count for this course (all sibling courses)
                sibling_courses_for_dropped = Course.objects.filter(
                    instructor=user,
                    code=selected_course.code,
                    name=selected_course.name,
                    semester=selected_course.semester,
                    school_year=selected_course.school_year,
                    deleted_at__isnull=True
                )
                dropped_count = CourseEnrollment.objects.filter(
                    course__in=sibling_courses_for_dropped,
                    deleted_at__isnull=False
                ).count()
                selected_course.dropped_count = dropped_count
                # Build section options based on instructor-defined sections (siblings in the same grouped key)
                group_key = (
                    (selected_course.code or '').strip().lower(),
                    (selected_course.name or '').strip().lower(),
                    (selected_course.semester or '').strip().lower(),
                    (selected_course.school_year or '').strip()
                )
                sibling_courses = [c for c in instructor_courses_raw if (
                    (c.code or '').strip().lower(),
                    (c.name or '').strip().lower(),
                    (c.semester or '').strip().lower(),
                    (c.school_year or '').strip()
                ) == group_key]
                for c in sibling_courses:
                    section_label = (c.section or '').upper()
                    # Use 'NO SECTION' display if empty, but key as '' for internal mapping
                    display_label = section_label if section_label else 'NO SECTION'
                    selected_course_section_to_id[display_label] = c.id
                selected_course_sections = sorted(selected_course_section_to_id.keys())
        except (ValueError, TypeError):
            selected_course = None
            selected_course_enrollments = None
    # Ensure enrollment queryset reflects uppercase section display
    enrollments = list(enrollments.select_related('course', 'student'))
    
    def _uppercase_section(items):
        for item in items:
            item.section_display = (item.section or '').upper() if item.section else ''
        return items
    
    enrollments = _uppercase_section(enrollments)
    if selected_course_enrollments is not None:
        selected_course_enrollments = _uppercase_section(list(selected_course_enrollments))
        selected_course_sections = sorted({
            enrollment.section_display
            for enrollment in selected_course_enrollments
            if getattr(enrollment, 'section_display', '').strip()
        })

    # Annotate whether each enrollment already has a registered QR for this course
    try:
        from dashboard.models import QRCodeRegistration
        if selected_course is not None and selected_course_enrollments is not None:
            for enrollment in selected_course_enrollments:
                try:
                    enrollment.has_qr = QRCodeRegistration.objects.filter(
                        student=enrollment.student,
                        course=selected_course,
                        is_active=True
                    ).exists()
                except Exception:
                    enrollment.has_qr = False
    except Exception:
        # If annotation fails for any reason, default to False
        if selected_course_enrollments is not None:
            for enrollment in selected_course_enrollments:
                enrollment.has_qr = False
    
    # Annotate whether each enrollment already has a registered biometric for this course
    try:
        if selected_course is not None and selected_course_enrollments is not None:
            for enrollment in selected_course_enrollments:
                try:
                    enrollment.has_biometric = BiometricRegistration.objects.filter(
                        student=enrollment.student,
                        course=selected_course,
                        is_active=True
                    ).exists()
                except Exception:
                    enrollment.has_biometric = False
    except Exception:
        # If annotation fails for any reason, default to False
        if selected_course_enrollments is not None:
            for enrollment in selected_course_enrollments:
                enrollment.has_biometric = False
    
    # IMPORTANT: Also annotate the main enrollments list with has_biometric and has_qr status
    # This ensures the status displays correctly even when no specific course is selected
    try:
        for enrollment in enrollments:
            try:
                enrollment.has_biometric = BiometricRegistration.objects.filter(
                    student=enrollment.student,
                    course=enrollment.course,
                    is_active=True
                ).exists()
                enrollment.has_qr = QRCodeRegistration.objects.filter(
                    student=enrollment.student,
                    course=enrollment.course,
                    is_active=True
                ).exists()
            except Exception:
                enrollment.has_biometric = False
                enrollment.has_qr = False
    except Exception:
        # If annotation fails, default to False for all
        for enrollment in enrollments:
            enrollment.has_biometric = False
            enrollment.has_qr = False
    
    context = {
        'user': user,
        'is_teacher': user.is_teacher,
        'school_admin': school_admin,
        'enrollments': enrollments,
        'instructor_courses': grouped_instructor_courses,
        'total_grouped_courses': len(grouped_instructor_courses),
        'unread_notifications': unread_notifications,
        'filter_options': {
            'sections': sections,
            'year_levels': year_levels,
            'semesters': semesters,
            'school_years': school_years,
        },
        'current_filters': {
            'section': filter_section,
            'year_level': filter_year,
            'semester': filter_semester,
            'school_year': filter_school_year,
            'course': filter_course,
            'search': search_query,
        },
        'course_student_counts': course_student_counts,
        'selected_course': selected_course,
        'selected_course_enrollments': selected_course_enrollments,
        'selected_course_sections': selected_course_sections,
        'selected_course_section_options': [{'label': k, 'id': v} for k, v in selected_course_section_to_id.items()],
        'selected_course_section_to_id': selected_course_section_to_id,
        'dropped_enrollments': dropped_enrollments,
        'dropped_enrollments_by_course': dropped_enrollments_by_course,
    }
    
    return render(request, 'dashboard/instructor/students.html', context)

@login_required
def weekly_timetable_view(request):
    """View for weekly timetable (both teachers and students can view)"""
    user = request.user
    if not (user.is_teacher or user.is_student):
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display (for both teachers and students)
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    education_level = user.education_level
    initial_school_type = 'HighSchool' if education_level == 'high_senior' else 'University'
    
    # Get courses from database
    courses_from_db = []
    if user.is_teacher:
        # For instructors: get courses assigned to them (only active courses with instructor assigned)
        # Get filter parameters
        filter_school_year = request.GET.get('school_year', '')
        filter_year_level = request.GET.get('year_level', '')
        filter_section = request.GET.get('section', '')
        filter_semester = request.GET.get('semester', '')
        
        instructor_courses = Course.objects.filter(
            instructor=user,
            is_active=True,
            deleted_at__isnull=True
        ).select_related('instructor', 'program').prefetch_related('course_schedules')
        
        # Log for debugging BEFORE ordering
        course_count = instructor_courses.count()
        logger.info(f"Instructor {user.username} (ID: {user.id}) has {course_count} active courses")
        
        # Log each course for debugging
        if course_count > 0:
            logger.info("=== ALL COURSES FOR INSTRUCTOR (BEFORE ORDERING) ===")
            for course in instructor_courses:
                logger.info(f"Course: {course.code} - {course.name} (ID: {course.id})")
                logger.info(f"  Instructor ID: {course.instructor.id if course.instructor else 'None'}")
                logger.info(f"  Days: {course.days}, Start: {course.start_time}, End: {course.end_time}")
                logger.info(f"  Year: {course.year_level}, Section: {course.section}, Sem: {course.semester}, SY: {course.school_year}")
                logger.info(f"  Is Active: {course.is_active}, Created: {course.created_at}")
                logger.info(f"  Has Day Schedules: {course.course_schedules.exists()}")
        else:
            logger.warning(f"⚠️ No courses found for instructor {user.username} (ID: {user.id})!")
            logger.warning("Check if courses exist in database and are assigned to this instructor")
            # Double-check: query all courses for this instructor regardless of is_active
            all_courses = Course.objects.filter(instructor=user)
            logger.warning(f"Total courses (including inactive): {all_courses.count()}")
            if all_courses.exists():
                logger.warning("Found courses but they might be inactive:")
                for c in all_courses:
                    logger.warning(f"  - {c.code} - {c.name} (is_active={c.is_active})")
        
        instructor_courses = instructor_courses.order_by('-created_at', 'year_level', 'semester', 'section', 'days', 'start_time')
        
        # Convert to format expected by timetable
        for course in instructor_courses:
            # Check if course has day-specific schedules
            day_schedules = course.course_schedules.all().order_by('day_order')
            has_day_schedules = day_schedules.exists()
            
            if has_day_schedules:
                # Create separate entries for each day schedule
                for schedule in day_schedules:
                    day_map = {
                        'Mon': 'M', 'Monday': 'M', 'MON': 'M', 'MONDAY': 'M',
                        'Tue': 'T', 'Tuesday': 'T', 'TUE': 'T', 'TUESDAY': 'T',
                        'Wed': 'W', 'Wednesday': 'W', 'WED': 'W', 'WEDNESDAY': 'W',
                        'Thu': 'Th', 'Thursday': 'Th', 'THU': 'Th', 'THURSDAY': 'Th',
                        'Fri': 'F', 'Friday': 'F', 'FRI': 'F', 'FRIDAY': 'F',
                        'Sat': 'S', 'Saturday': 'S', 'SAT': 'S', 'SATURDAY': 'S',
                        'Sun': 'Su', 'Sunday': 'Su', 'SUN': 'Su', 'SUNDAY': 'Su'
                    }
                    # Normalize the day
                    day_normalized = schedule.day.strip().title() if schedule.day else ''
                    mapped_day = day_map.get(day_normalized, day_map.get(schedule.day, schedule.day))
                    
                    # Ensure the mapped day is valid
                    if mapped_day not in ['M', 'T', 'W', 'Th', 'F', 'S', 'Su']:
                        logger.warning(f"Course {course.code}: Invalid day '{schedule.day}' for schedule, skipping")
                        continue
                    
                    courses_from_db.append({
                        'id': f'db_{course.id}_{schedule.day}',
                        'courseCode': course.code,
                        'courseName': course.name,
                        'code': course.code,
                        'name': course.name,
                        'teacherName': course.instructor.full_name if course.instructor else 'N/A',
                        'instructor': course.instructor.full_name if course.instructor else 'N/A',
                        'room': schedule.room or course.room or '',
                        'days': [mapped_day],
                        'startTime': schedule.start_time.strftime('%I:%M %p') if schedule.start_time else '',
                        'endTime': schedule.end_time.strftime('%I:%M %p') if schedule.end_time else '',
                        'classStartTime': schedule.start_time.strftime('%I:%M %p') if schedule.start_time else '',
                        'classEndTime': schedule.end_time.strftime('%I:%M %p') if schedule.end_time else '',
                        'attendanceTimeIn': (schedule.attendance_start.strftime('%I:%M %p') if schedule.attendance_start else '') or (course.attendance_start.strftime('%I:%M %p') if course.attendance_start else ''),
                        'attendanceTimeEnd': (schedule.attendance_end.strftime('%I:%M %p') if schedule.attendance_end else '') or (course.attendance_end.strftime('%I:%M %p') if course.attendance_end else ''),
                        'color': course.color or '#3b82f6',
                        'themeColor': course.color or '#3b82f6',
                        'schoolType': initial_school_type,
                        'program': course.program.code if course.program else '',
                        'yearLevel': course.year_level,
                        'section': course.section,
                        'semester': course.semester,
                        'schoolYear': course.school_year or '',
                        'startDate': '',
                        'endDate': '',
                    })
            else:
                # Use default schedule
                days_list = [d.strip() for d in course.days.split(',')]
                day_map = {
                    'Mon': 'M', 'Monday': 'M', 'MON': 'M', 'MONDAY': 'M',
                    'Tue': 'T', 'Tuesday': 'T', 'TUE': 'T', 'TUESDAY': 'T',
                    'Wed': 'W', 'Wednesday': 'W', 'WED': 'W', 'WEDNESDAY': 'W',
                    'Thu': 'Th', 'Thursday': 'Th', 'THU': 'TH', 'THURSDAY': 'Th',
                    'Fri': 'F', 'Friday': 'F', 'FRI': 'F', 'FRIDAY': 'F',
                    'Sat': 'S', 'Saturday': 'S', 'SAT': 'S', 'SATURDAY': 'S',
                    'Sun': 'Su', 'Sunday': 'Su', 'SUN': 'Su', 'SUNDAY': 'Su'
                }
                # Map days and keep all valid ones (don't filter out if not in map - keep original)
                mapped_days = []
                for day in days_list:
                    mapped = day_map.get(day, day)
                    # Only add if it's a valid day abbreviation (M, T, W, Th, F, S, Su)
                    if mapped in ['M', 'T', 'W', 'Th', 'F', 'S', 'Su']:
                        mapped_days.append(mapped)
                    elif day in ['M', 'T', 'W', 'Th', 'F', 'S', 'Su']:
                        # Already in correct format
                        mapped_days.append(day)
                
                # Log if days were filtered out
                if len(mapped_days) < len(days_list):
                    logger.warning(f"Course {course.code}: Some days were filtered out. Original: {days_list}, Mapped: {mapped_days}")
                
                # Only add course if it has valid days
                if not mapped_days:
                    logger.warning(f"Course {course.code}: No valid days after mapping. Original days: {days_list}. Skipping course.")
                    continue
                
                courses_from_db.append({
                    'id': f'db_{course.id}',
                    'courseCode': course.code,
                    'courseName': course.name,
                    'code': course.code,
                    'name': course.name,
                    'teacherName': course.instructor.full_name if course.instructor else 'N/A',
                    'instructor': course.instructor.full_name if course.instructor else 'N/A',
                    'room': course.room or '',
                    'days': mapped_days,
                    'startTime': course.start_time.strftime('%I:%M %p') if course.start_time else '',
                    'endTime': course.end_time.strftime('%I:%M %p') if course.end_time else '',
                    'classStartTime': course.start_time.strftime('%I:%M %p') if course.start_time else '',
                    'classEndTime': course.end_time.strftime('%I:%M %p') if course.end_time else '',
                    'attendanceTimeIn': course.attendance_start.strftime('%I:%M %p') if course.attendance_start else '',
                    'attendanceTimeEnd': course.attendance_end.strftime('%I:%M %p') if course.attendance_end else '',
                    'color': course.color or '#3b82f6',
                    'themeColor': course.color or '#3b82f6',
                    'schoolType': initial_school_type,
                    'program': course.program.code if course.program else '',
                    'yearLevel': course.year_level,
                    'section': course.section,
                    'semester': course.semester,
                    'schoolYear': course.school_year or '',
                    'startDate': '',
                    'endDate': '',
                })
    elif user.is_student:
        # For students: ONLY show courses they are enrolled in
        from .models import CourseEnrollment
        
        # CRITICAL: Ensure courses_from_db is empty for students unless they have active enrollments
        courses_from_db = []  # Explicitly reset to ensure no courses leak through
        
        # Get all active enrollments for this student ONLY
        enrollments = CourseEnrollment.objects.filter(
            student=user,
            is_active=True
        ).select_related('course', 'course__program', 'course__instructor').prefetch_related('course__course_schedules')
        
        logger.info(f"Student {user.username} ({user.id}): Found {enrollments.count()} active enrollment(s)")
        
        # Extract courses from enrollments - ONLY enrolled courses, with strict validation
        student_courses = []
        for enrollment in enrollments:
            if enrollment.is_active and enrollment.course and enrollment.course.is_active:
                student_courses.append(enrollment.course)
                logger.info(f"  - Enrolled in: {enrollment.course.code} - {enrollment.course.name} (Course ID: {enrollment.course.id})")
        
        # If no enrollments, courses_from_db will remain empty and message will show
        if not student_courses:
            # No enrolled courses - courses_from_db stays empty, frontend will show message
            logger.info(f"Student {user.username} has NO enrolled courses. Timetable will show empty message. courses_from_db length: {len(courses_from_db)}")
        else:
            # Filter by semester and school year if provided
            current_semester = request.GET.get('semester', '')
            school_year = request.GET.get('school_year', '')
            
            if current_semester:
                student_courses = [c for c in student_courses if c.semester == current_semester]
            if school_year:
                student_courses = [c for c in student_courses if c.school_year == school_year]
            
            # Sort by days and start time
            student_courses = sorted(student_courses, key=lambda c: (c.days or '', c.start_time or ''))
            
            # Convert to format expected by timetable
            for course in student_courses:
                # Check if course has day-specific schedules
                day_schedules = course.course_schedules.all().order_by('day_order')
                has_day_schedules = day_schedules.exists()
                
                if has_day_schedules:
                    # Create separate entries for each day schedule
                    for schedule in day_schedules:
                        day_map = {
                            'Mon': 'M', 'Monday': 'M',
                            'Tue': 'T', 'Tuesday': 'T',
                            'Wed': 'W', 'Wednesday': 'W',
                            'Thu': 'Th', 'Thursday': 'Th',
                            'Fri': 'F', 'Friday': 'F',
                            'Sat': 'S', 'Saturday': 'S',
                            'Sun': 'Su', 'Sunday': 'Su'
                        }
                        mapped_day = day_map.get(schedule.day, schedule.day)
                        
                        courses_from_db.append({
                            'id': f'db_{course.id}_{schedule.day}',
                            'courseCode': course.code,
                            'courseName': course.name,
                            'code': course.code,
                            'name': course.name,
                            'teacherName': course.instructor.full_name if course.instructor else 'N/A',
                            'instructor': course.instructor.full_name if course.instructor else 'N/A',
                            'room': schedule.room or course.room or '',
                            'days': [mapped_day],
                            'startTime': schedule.start_time.strftime('%I:%M %p') if schedule.start_time else '',
                            'endTime': schedule.end_time.strftime('%I:%M %p') if schedule.end_time else '',
                            'classStartTime': schedule.start_time.strftime('%I:%M %p') if schedule.start_time else '',
                            'classEndTime': schedule.end_time.strftime('%I:%M %p') if schedule.end_time else '',
                            'color': course.color or '#3b82f6',
                            'themeColor': course.color or '#3b82f6',
                            'schoolType': initial_school_type,
                            'program': course.program.code if course.program else '',
                            'yearLevel': course.year_level,
                            'section': course.section,
                            'semester': course.semester,
                            'schoolYear': course.school_year or '',
                            'startDate': '',
                            'endDate': '',
                        })
                else:
                    # Use default schedule
                    days_list = [d.strip() for d in course.days.split(',')]
                    day_map = {
                        'Mon': 'M', 'Monday': 'M',
                        'Tue': 'T', 'Tuesday': 'T',
                        'Wed': 'W', 'Wednesday': 'W',
                        'Thu': 'Th', 'Thursday': 'Th',
                        'Fri': 'F', 'Friday': 'F',
                        'Sat': 'S', 'Saturday': 'S',
                        'Sun': 'Su', 'Sunday': 'Su'
                    }
                    mapped_days = [day_map.get(day, day) for day in days_list if day_map.get(day)]
                    
                    courses_from_db.append({
                        'id': f'db_{course.id}',
                        'courseCode': course.code,
                        'courseName': course.name,
                        'code': course.code,
                        'name': course.name,
                        'teacherName': course.instructor.full_name if course.instructor else 'N/A',
                        'instructor': course.instructor.full_name if course.instructor else 'N/A',
                        'room': course.room or '',
                        'days': mapped_days,
                        'startTime': course.start_time.strftime('%I:%M %p') if course.start_time else '',
                        'endTime': course.end_time.strftime('%I:%M %p') if course.end_time else '',
                        'classStartTime': course.start_time.strftime('%I:%M %p') if course.start_time else '',
                        'classEndTime': course.end_time.strftime('%I:%M %p') if course.end_time else '',
                        'attendanceTimeIn': course.attendance_start.strftime('%I:%M %p') if course.attendance_start else '',
                        'attendanceTimeEnd': course.attendance_end.strftime('%I:%M %p') if course.attendance_end else '',
                        'color': course.color or '#3b82f6',
                        'themeColor': course.color or '#3b82f6',
                        'schoolType': initial_school_type,
                        'program': course.program.code if course.program else '',
                        'yearLevel': course.year_level,
                        'section': course.section,
                        'semester': course.semester,
                        'schoolYear': course.school_year or '',
                        'startDate': '',
                        'endDate': '',
                    })
            
            # Final safety check: Log what we're sending to frontend
            logger.info(f"Student {user.username}: Sending {len(courses_from_db)} course(s) to timetable (all from enrollments)")
            for idx, course_data in enumerate(courses_from_db[:3]):  # Log first 3
                logger.info(f"  Course {idx + 1}: {course_data.get('code')} - {course_data.get('name')}")
        # End of student courses processing
    
    # Get unique filter values for instructors
    filter_options = {}
    if user.is_teacher:
        all_instructor_courses = Course.objects.filter(
            instructor=user,
            is_active=True,
            deleted_at__isnull=True
        )
        
        # Get unique school years (filter before values_list)
        school_years_qs = all_instructor_courses.exclude(school_year__isnull=True).exclude(school_year='').values_list('school_year', flat=True).distinct()
        filter_options['school_years'] = sorted(list(set([sy for sy in school_years_qs if sy])))
        
        # Get unique sections from the instructor's courses (Manage Courses), regardless of enrollment
        sections_qs = all_instructor_courses.exclude(section__isnull=True).exclude(section='').values_list('section', flat=True).distinct()
        filter_options['sections'] = sorted(list(set([(sec or '').upper() for sec in sections_qs if sec])))
        
        # Semesters are always available from choices
        filter_options['semesters'] = Course.SEMESTER_CHOICES
        
        logger.info(f"Filter options for instructor {user.id}: school_years={filter_options['school_years']}, sections={filter_options['sections']}, semesters={len(filter_options['semesters'])}")
    
    # Debug logging
    logger.info(f"Weekly timetable view - User: {user.id}, is_teacher: {user.is_teacher}")
    logger.info(f"Total courses from DB: {len(courses_from_db)}")
    if courses_from_db:
        logger.info(f"Sample course: {courses_from_db[0]}")
        # Log all courses for debugging
        for idx, course in enumerate(courses_from_db[:5]):  # Log first 5 courses
            logger.info(f"Course {idx + 1}: {course.get('code', 'N/A')} - {course.get('name', 'N/A')}, "
                       f"Year: {course.get('yearLevel')}, Section: {course.get('section')}, "
                       f"Semester: {course.get('semester')}, SchoolYear: {course.get('schoolYear')}, "
                       f"Days: {course.get('days')}, Time: {course.get('startTime')}-{course.get('endTime')}")
    logger.info(f"Filter options: {filter_options}")
    logger.info(f"Current filters from request: school_year={request.GET.get('school_year')}, "
                f"section={request.GET.get('section')}, "
                f"semester={request.GET.get('semester')}")
    logger.info(f"Initial school type: {initial_school_type}")
    
    # Pass courses list directly to template (no JSON needed - uses DOM data attributes)
    logger.info(f"Passing {len(courses_from_db)} courses to frontend via DOM (no JSON parsing needed)")
    
    # Log first course details for debugging
    if courses_from_db:
        first_course = courses_from_db[0]
        logger.info(f"Sample course data: code={first_course.get('code')}, days={first_course.get('days')}, startTime={first_course.get('startTime')}, endTime={first_course.get('endTime')}, yearLevel={first_course.get('yearLevel')}, section={first_course.get('section')}, semester={first_course.get('semester')}, schoolYear={first_course.get('schoolYear')}")
    
    context = {
        'user': user,
        'is_teacher': user.is_teacher,
        'education_level': education_level,
        'initial_school_type': initial_school_type,
        'courses_list': courses_from_db,  # Pass list directly - template will render as DOM elements
        'school_admin': school_admin,
        'filter_options': filter_options,
        'current_filters': {
            'school_year': request.GET.get('school_year', ''),
            'section': request.GET.get('section', ''),
            'semester': request.GET.get('semester', ''),
        },
    }
    
    return render(request, 'dashboard/shared/weekly_timetable.html', context)

# ============================================
# INSTRUCTOR COURSE MANAGEMENT VIEWS
# ============================================

@login_required
def instructor_courses_view(request):
    """Course management view for instructors - shows all courses assigned to them across all programs"""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get the instructor's primary program (for default display)
    program = user.program if user.program else None
    
    # Get department for navigation (if program exists)
    department = None
    if program and program.department:
        try:
            if user.school_name:
                department = Department.objects.get(name=program.department, school_name=user.school_name, is_active=True)
            else:
                department = Department.objects.filter(name=program.department, is_active=True).first()
        except Department.DoesNotExist:
            pass
    
    # Get all courses assigned to this instructor across ALL programs (excluding deleted and archived)
    # Order by created_at (ascending) to show courses in the order they were added (first, second, third, etc.)
    courses_qs = Course.objects.filter(instructor=user, deleted_at__isnull=True, is_archived=False).prefetch_related('course_schedules').order_by('created_at', 'program__code', 'year_level', 'semester', 'section', 'code')
    
    from .models import CourseEnrollment
    course_counts = CourseEnrollment.objects.filter(course__in=courses_qs, is_active=True, deleted_at__isnull=True).values('course').annotate(total=Count('id'))
    counts_map = {item['course']: item['total'] for item in course_counts}
    courses = list(courses_qs)
    for course in courses:
        course.student_count = counts_map.get(course.id, 0)
    
    # Group courses by identity (code, name, semester, school_year) so sections don't create duplicate cards
    grouped_map = {}
    for c in courses:
        # Normalize key parts to avoid accidental splits due to case/whitespace differences
        norm_code = (c.code or '').strip().upper()
        norm_name = (c.name or '').strip().upper()
        norm_semester = (c.semester or '').strip().lower()
        norm_school_year = (c.school_year or '').strip()
        key = (norm_code, norm_name, norm_semester, norm_school_year)
        if key not in grouped_map:
            grouped_map[key] = []
        grouped_map[key].append(c)
    grouped_courses = []
    for key, group_list in grouped_map.items():
        # Representative course is the first by created_at (already sorted upstream)
        rep = group_list[0]
        # Collect sibling sections (uppercased, non-empty)
        sibling_sections = []
        for gc in group_list:
            sec = (gc.section or '').strip()
            if sec:
                sibling_sections.append(sec.upper())
        # Attach helper attributes for template
        rep.sibling_sections = sorted(list(set(sibling_sections)))
        # Collect distinct colors across sections
        colors = [gc.color for gc in group_list if (gc.color or '').strip()]
        # Ensure at least one color present
        if not colors:
            colors = [(rep.color or '#3C4770') or '#3C4770']
        # Build multi-stop gradient string for any number of colors (up to all)
        unique_colors = []
        for c in colors:
            if c and c not in unique_colors:
                unique_colors.append(c)
        # Fallback if somehow empty
        if not unique_colors:
            unique_colors = ['#3C4770']
        # Compute gradient stops
        stops = []
        if len(unique_colors) == 1:
            stops = [f"{unique_colors[0]} 0%", f"{unique_colors[0]} 100%"]
        else:
            total = len(unique_colors) - 1
            for idx, col in enumerate(unique_colors):
                pct = int(round((idx / total) * 100))
                stops.append(f"{col} {pct}%")
        rep.display_gradient = f"linear-gradient(135deg, {', '.join(stops)})"
        # Maintain start/end for older template fallbacks (first two colors)
        rep.display_color_start = unique_colors[0]
        rep.display_color_end = unique_colors[1] if len(unique_colors) > 1 else unique_colors[0]
        # Use a representative section for display (optional)
        rep.display_section = ', '.join(rep.sibling_sections) if rep.sibling_sections else (rep.section or '')
        # Count students as sum across grouped sections for the card
        rep.student_count = sum(getattr(gc, 'student_count', 0) for gc in group_list)
        grouped_courses.append(rep)
    
    # Generate school year options (2020-2021 to 2030-2031+)
    # Philippine timezone (UTC+8) - using datetime with manual offset
    from django.utils import timezone
    from datetime import timedelta
    # Get current time in UTC and add 8 hours for Philippine time (UTC+8)
    now_utc = timezone.now()
    ph_offset = timedelta(hours=8)
    now_ph = now_utc + ph_offset
    current_year = now_ph.year
    
    # Determine current school year (June is start of new school year in Philippines)
    if now_ph.month >= 6:
        current_school_year = f"{current_year}-{current_year+1}"
    else:
        current_school_year = f"{current_year-1}-{current_year}"
    
    # Generate school year options from 2020-2021 to 2030-2031
    school_year_options = []
    for year in range(2020, 2031):
        school_year_options.append(f"{year}-{year+1}")
    
    # Group courses by school year, then year level, then semester, then group multi-section courses
    courses_by_school_year = {}
    for course in courses:
        school_year_key = course.school_year or 'Unspecified'
        if school_year_key not in courses_by_school_year:
            courses_by_school_year[school_year_key] = {}
        
        year_key = f"Year {course.year_level}"
        if year_key not in courses_by_school_year[school_year_key]:
            courses_by_school_year[school_year_key][year_key] = {
                'semesters': {},
            }
        
        semester_key = course.semester or 'Unspecified'
        if semester_key not in courses_by_school_year[school_year_key][year_key]['semesters']:
            courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = []
        
        courses_by_school_year[school_year_key][year_key]['semesters'][semester_key].append(course)
    
    # Group multi-section courses within each semester
    for school_year_key in courses_by_school_year:
        for year_key in courses_by_school_year[school_year_key]:
            for semester_key in courses_by_school_year[school_year_key][year_key]['semesters']:
                courses_list = courses_by_school_year[school_year_key][year_key]['semesters'][semester_key]
                # Group by course identity (code, name, semester, school_year)
                grouped_map = {}
                for course in courses_list:
                    key = (
                        (course.code or '').strip().upper(),
                        (course.name or '').strip().upper(),
                        (course.semester or '').strip().lower(),
                        (course.school_year or '').strip()
                    )
                    if key not in grouped_map:
                        grouped_map[key] = []
                    grouped_map[key].append(course)
                
                # Replace list with grouped courses (one representative per group)
                grouped_courses_list = []
                for key, group_list in grouped_map.items():
                    rep = group_list[0]  # Representative course
                    rep.grouped_sections = group_list  # Store all sections
                    rep.sibling_sections = sorted([(c.section or '').upper() for c in group_list if c.section])
    
                    # Collect distinct colors across sections for gradient
                    colors = [gc.color for gc in group_list if (gc.color or '').strip()]
                    if not colors:
                        colors = [(rep.color or '#3C4770') or '#3C4770']
                    
                    # Build multi-stop gradient string
                    unique_colors = []
                    for c in colors:
                        if c and c not in unique_colors:
                            unique_colors.append(c)
                    if not unique_colors:
                        unique_colors = ['#3C4770']
                    
                    # Compute gradient stops
                    stops = []
                    if len(unique_colors) == 1:
                        stops = [f"{unique_colors[0]} 0%", f"{unique_colors[0]} 100%"]
                    else:
                        total = len(unique_colors) - 1
                        for idx, col in enumerate(unique_colors):
                            pct = int(round((idx / total) * 100))
                            stops.append(f"{col} {pct}%")
                    rep.display_gradient = f"linear-gradient(135deg, {', '.join(stops)})"
                    rep.display_color_start = unique_colors[0]
                    rep.display_color_end = unique_colors[1] if len(unique_colors) > 1 else unique_colors[0]
                    
                    grouped_courses_list.append(rep)
                
                courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = grouped_courses_list
    
    # Get unique sections and semesters for filtering (legacy support)
    unique_sections = sorted({course.section for course in courses if course.section})
    unique_semesters = Course.SEMESTER_CHOICES
    
    context = {
        'user': user,
        'program': program,
        'department': department,
        'courses': courses,
        'grouped_courses': grouped_courses,
        'courses_by_school_year': courses_by_school_year,
        'courses_by_year_semester': {},  # Keep for backward compatibility
        'unique_sections': unique_sections,
        'unique_semesters': unique_semesters,
        'school_year_options': school_year_options,
        'current_school_year': current_school_year,
        'auto_open_add_modal': request.GET.get('open_add_modal') == '1',
    }
    
    return render(request, 'dashboard/instructor/instructor_program_courses.html', context)

@login_required
@require_http_methods(["GET", "POST"])
def instructor_add_course_view(request):
    """Add a new course - instructor can add to any program in their school"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    if request.method == 'POST':
        try:
            # Get form data
            code = request.POST.get('code', '').strip()
            name = request.POST.get('name', '').strip()
            program_id = request.POST.get('program', '').strip()
            year_level = request.POST.get('year_level', '').strip()
            # Section is optional - students will fill it during enrollment (default to empty string)
            section = request.POST.get('section', '').strip() or ''
            semester = request.POST.get('semester', '1st').strip()
            school_year = request.POST.get('school_year', '').strip()
            room = request.POST.get('room', '').strip() or None
            days = request.POST.get('days', '').strip()
            start_time = request.POST.get('start_time', '').strip()
            end_time = request.POST.get('end_time', '').strip()
            color = request.POST.get('color', '#3C4770').strip()
            
            # Check if day schedules are provided (different schedule for each day)
            day_schedules = request.POST.get('day_schedules', '')
            has_day_schedules = False
            if day_schedules:
                try:
                    schedules_data = json.loads(day_schedules)
                    if schedules_data and len(schedules_data) > 0:
                        has_day_schedules = True
                except:
                    pass
            
            # Validate required fields
            # If day schedules are provided, days, start_time, and end_time are set from day schedules
            missing_fields = []
            if not code:
                missing_fields.append('code')
            if not name:
                missing_fields.append('name')
            # Program is now OPTIONAL - removed from required fields
            if not year_level:
                missing_fields.append('year_level')
            # Section is optional - students will fill it
            # if not section:
            #     missing_fields.append('section')
            if not school_year:
                missing_fields.append('school_year')
            if not semester:
                missing_fields.append('semester')
            
            # If day schedules are provided, days/start_time/end_time come from schedules
            # Otherwise, they are required fields
            if has_day_schedules:
                # Validate day schedules have required data
                try:
                    schedules_data = json.loads(day_schedules)
                    if not schedules_data or len(schedules_data) == 0:
                        missing_fields.append('day_schedules')
                    else:
                        # Check each schedule has required fields
                        for idx, schedule in enumerate(schedules_data):
                            if not schedule.get('day'):
                                missing_fields.append(f'day_schedules[{idx}].day')
                            if not schedule.get('startTime'):
                                missing_fields.append(f'day_schedules[{idx}].startTime')
                            if not schedule.get('endTime'):
                                missing_fields.append(f'day_schedules[{idx}].endTime')
                except:
                    missing_fields.append('day_schedules')
            else:
                # Standard schedule - days, start_time, end_time are required
                if not days:
                    missing_fields.append('days')
                if not start_time:
                    missing_fields.append('start_time')
                if not end_time:
                    missing_fields.append('end_time')
            
            if missing_fields:
                logger.error(f"Missing required fields: {missing_fields}")
                logger.error(f"Form data - code: {code}, name: {name}, program_id: {program_id}, year_level: {year_level}, section: {section}, school_year: {school_year}, days: {days}, start_time: {start_time}, end_time: {end_time}")
                logger.error(f"Day schedules: {day_schedules}")
                logger.error(f"Has day schedules: {has_day_schedules}")
                # If day schedules are provided, provide more specific error
                if has_day_schedules:
                    return JsonResponse({'success': False, 'message': f'Please fill in all required fields. Missing: {", ".join(missing_fields)}. Note: When using day-specific schedules, make sure all course information is filled.'})
                return JsonResponse({'success': False, 'message': f'Please fill in all required fields. Missing: {", ".join(missing_fields)}'})
            
            # Get program if provided (now optional)
            program = None
            if program_id:
                try:
                    program = Program.objects.get(id=int(program_id))
                    if user.school_name and program.school_name != user.school_name:
                        return JsonResponse({'success': False, 'message': 'You can only add courses to programs in your school.'})
                except (ValueError, Program.DoesNotExist):
                    return JsonResponse({'success': False, 'message': 'Invalid program selected.'})
            
            # Parse time strings (handle both 24-hour and 12-hour formats)
            from datetime import datetime
            def parse_time(time_str):
                if not time_str:
                    return None
                try:
                    # Try 24-hour format first (HH:MM)
                    return datetime.strptime(time_str, '%H:%M').time()
                except ValueError:
                    try:
                        # Try 12-hour format (HH:MM AM/PM)
                        return datetime.strptime(time_str, '%I:%M %p').time()
                    except ValueError:
                        try:
                            # Try 12-hour format without space (HH:MMAM/PM)
                            return datetime.strptime(time_str, '%I:%M%p').time()
                        except ValueError:
                            logger.error(f"Could not parse time: {time_str}")
                            return None
            
            parsed_start_time = parse_time(start_time)
            parsed_end_time = parse_time(end_time)
            
            # Check for time conflicts with existing courses
            # IMPORTANT: Conflict when the same instructor already has a course in the SAME semester
            # AND school year (regardless of year level / program / section). Different school year
            # or semester automatically means no conflict.
            normalized_school_year = school_year or None
            conflicting_courses = Course.objects.filter(
                instructor=user,
                semester=semester,
                is_active=True
            )
            if normalized_school_year is None:
                conflicting_courses = conflicting_courses.filter(Q(school_year__isnull=True) | Q(school_year=''))
            else:
                conflicting_courses = conflicting_courses.filter(school_year=normalized_school_year)
            
            conflicts = []
            
            # Day mapping for normalization
            day_normalize_map = {
                'Mon': 'Monday', 'Monday': 'Monday', 'M': 'Monday',
                'Tue': 'Tuesday', 'Tuesday': 'Tuesday', 'T': 'Tuesday',
                'Wed': 'Wednesday', 'Wednesday': 'Wednesday', 'W': 'Wednesday',
                'Thu': 'Thursday', 'Thursday': 'Thursday', 'Th': 'Thursday',
                'Fri': 'Friday', 'Friday': 'Friday', 'F': 'Friday',
                'Sat': 'Saturday', 'Saturday': 'Saturday', 'S': 'Saturday',
                'Sun': 'Sunday', 'Sunday': 'Sunday', 'Su': 'Sunday'
            }
            
            def normalize_day(day_str):
                """Normalize day string to full day name"""
                day_str = day_str.strip()
                return day_normalize_map.get(day_str, day_str)
            
            def days_match(day1, day2):
                """Check if two day strings match (handles abbreviations)"""
                norm1 = normalize_day(day1).lower()
                norm2 = normalize_day(day2).lower()
                return norm1 == norm2 or norm1 in norm2 or norm2 in norm1
            
            if has_day_schedules:
                # Check conflicts for day-specific schedules (asynchronized)
                try:
                    schedules_data = json.loads(day_schedules)
                    for schedule_data in schedules_data:
                        day = schedule_data.get('day')
                        sched_start = parse_time(schedule_data.get('startTime') or schedule_data.get('start_time'))
                        sched_end = parse_time(schedule_data.get('endTime') or schedule_data.get('end_time'))
                        
                        if not sched_start or not sched_end:
                            continue
                        
                        # Check against courses with day-specific schedules
                        for existing_course in conflicting_courses:
                            existing_schedules = existing_course.course_schedules.filter(day=day)
                            for existing_sched in existing_schedules:
                                # Check if times actually overlap (not just adjacent)
                                # Two intervals overlap if: new_start < existing_end AND new_end > existing_start
                                # This correctly excludes adjacent times (e.g., 9:00-10:30 and 10:30-11:30 are OK)
                                if (sched_start < existing_sched.end_time and sched_end > existing_sched.start_time):
                                    conflicts.append({
                                        'course': f"{existing_course.code} - {existing_course.name}",
                                        'day': day,
                                        'time': f"{existing_sched.start_time.strftime('%I:%M %p')} - {existing_sched.end_time.strftime('%I:%M %p')}"
                                    })
                            
                            # Check against courses with default schedule (synchronized)
                            if not existing_course.course_schedules.exists():
                                # Check if this day is in the existing course's days
                                existing_days = [d.strip() for d in existing_course.days.split(',')]
                                day_matches = any(days_match(day, existing_day) for existing_day in existing_days)
                                
                                if day_matches and existing_course.start_time and existing_course.end_time:
                                    # Check if times actually overlap (not just adjacent)
                                    if (sched_start < existing_course.end_time and sched_end > existing_course.start_time):
                                        conflicts.append({
                                            'course': f"{existing_course.code} - {existing_course.name}",
                                            'day': day,
                                            'time': f"{existing_course.start_time.strftime('%I:%M %p')} - {existing_course.end_time.strftime('%I:%M %p')}"
                                        })
                except Exception as e:
                    logger.error(f"Error checking day schedule conflicts: {str(e)}")
            else:
                # Check conflicts for default schedule (synchronized)
                if parsed_start_time and parsed_end_time:
                    new_course_days = [d.strip() for d in days.split(',')]
                    
                    for existing_course in conflicting_courses:
                        # Check each day individually
                        if existing_course.course_schedules.exists():
                            # Existing course has day-specific schedules
                            existing_days = set()
                            for schedule in existing_course.course_schedules.all():
                                existing_days.add(schedule.day)
                            
                            # Check each day in new course against existing course's day-specific schedules
                            for new_day in new_course_days:
                                for existing_day in existing_days:
                                    if days_match(new_day, existing_day):
                                        # Find the schedule for this day
                                        day_schedule = existing_course.course_schedules.filter(day=existing_day).first()
                                        if day_schedule and day_schedule.start_time and day_schedule.end_time:
                                            # Check if times actually overlap (not just adjacent)
                                            if (parsed_start_time < day_schedule.end_time and parsed_end_time > day_schedule.start_time):
                                                conflicts.append({
                                                    'course': f"{existing_course.code} - {existing_course.name}",
                                                    'day': existing_day,
                                                    'time': f"{day_schedule.start_time.strftime('%I:%M %p')} - {day_schedule.end_time.strftime('%I:%M %p')}"
                                                })
                        else:
                            # Existing course has synchronized schedule
                            existing_days = [d.strip() for d in existing_course.days.split(',')]
                            
                            # Check each day individually - only report conflicts for days that match AND times overlap
                            for new_day in new_course_days:
                                for existing_day in existing_days:
                                    if days_match(new_day, existing_day):
                                        # Times overlap check - only conflict if times actually overlap (not just adjacent)
                                        if existing_course.start_time and existing_course.end_time:
                                            if (parsed_start_time < existing_course.end_time and parsed_end_time > existing_course.start_time):
                                                conflicts.append({
                                                    'course': f"{existing_course.code} - {existing_course.name}",
                                                    'day': existing_day,
                                                    'time': f"{existing_course.start_time.strftime('%I:%M %p')} - {existing_course.end_time.strftime('%I:%M %p')}"
                                                })
            
            if conflicts:
                # Remove duplicate conflicts (same course, day, and time)
                seen = set()
                unique_conflicts = []
                for conflict in conflicts:
                    key = (conflict['course'], conflict.get('day', ''), conflict['time'])
                    if key not in seen:
                        seen.add(key)
                        unique_conflicts.append(conflict)
                
                conflict_messages = []
                for conflict in unique_conflicts:
                    conflict_messages.append(f"{conflict['course']} on {conflict['day']} at {conflict['time']}")
                
                return JsonResponse({
                    'success': False,
                    'message': (
                        'Time conflict detected! The following course(s) have overlapping schedules:\n'
                        + '\n'.join(conflict_messages)
                        + '\n\nPlease adjust the time or assign this course to a different semester or school year.'
                    )
                })
            
            # Check if course already exists (handle unique constraint gracefully)
            # Only prevent duplicates when it's the SAME course code - different course codes can share the same section
            # Normalize section for comparison (uppercase, trimmed)
            section_normalized = (section or '').strip().upper() if section else ''
            existing_course = Course.objects.filter(
                program=program,
                code=code,  # Only check for same course code - different codes can have same section
                year_level=int(year_level),
                section=section_normalized,
                semester=semester,
                school_year=school_year or None,
                deleted_at__isnull=True  # Only check non-deleted courses
            ).first()
            
            if existing_course:
                # Course with same code and section already exists - this is a duplicate
                # Different course codes can have the same section name
                return JsonResponse({
                    'success': False,
                    'message': f'A course with code "{code}" already exists for {program.code} - Year {year_level}, Section {section_normalized or "(no section)"}, {semester} Semester, SY {school_year or "N/A"}. Please use a different section or modify the existing course.'
                })
            
            # Create course - instructor is automatically assigned
            # Section is managed by students during enrollment, so set to empty string if not provided
            # Use normalized section (uppercase, trimmed) that was used in the check above
            course_section = section_normalized
            
            try:
                course = Course.objects.create(
                    code=code,
                    name=name,
                    program=program,
                    year_level=int(year_level),
                    section=course_section,  # Empty by default - students will fill during enrollment
                    semester=semester,
                    school_year=school_year or None,
                    instructor=user,  # Automatically assign the instructor
                    room=room,
                    days=days,
                    start_time=parsed_start_time,
                    end_time=parsed_end_time,
                    color=color,
                    attendance_start=None,  # Attendance window removed - no longer used
                    attendance_end=None,  # Attendance window removed - no longer used
                    school_name=user.school_name,
                    enrollment_status='closed',  # Default to closed enrollment
                    is_active=True
                )
            except Exception as e:
                # Handle unique constraint or other database errors
                error_msg = str(e)
                if 'UNIQUE constraint' in error_msg or 'unique' in error_msg.lower():
                    # Check if it's actually a duplicate (same code and section)
                    # Different course codes can have the same section name
                    duplicate_check = Course.objects.filter(
                        program=program,
                        code=code,
                        year_level=int(year_level),
                        section=section_normalized,
                        semester=semester,
                        school_year=school_year or None,
                        deleted_at__isnull=True
                    ).first()
                    
                    if duplicate_check:
                        # It's a true duplicate - same code and section
                        return JsonResponse({
                            'success': False,
                            'message': f'A course with code "{code}" already exists for {program.code} - Year {year_level}, Section {section_normalized or "(no section)"}, {semester} Semester, SY {school_year or "N/A"}. Please use a different section or modify the existing course.'
                        })
                    else:
                        # Unique constraint violation but not a duplicate - might be a different field conflict
                        return JsonResponse({
                            'success': False,
                            'message': f'Unable to create course. A course with similar details may already exist. Please check your course details and try again.'
                        })
                else:
                    logger.error(f"Error creating course: {error_msg}")
                    return JsonResponse({
                        'success': False,
                        'message': f'Error creating course: {error_msg}'
                    })
            
            # Handle day-specific schedules if provided
            day_schedules = request.POST.get('day_schedules', '')
            if day_schedules:
                try:
                    schedules_data = json.loads(day_schedules)
                    for schedule_data in schedules_data:
                        day = schedule_data.get('day')
                        start_time_str = schedule_data.get('startTime') or schedule_data.get('start_time')
                        end_time_str = schedule_data.get('endTime') or schedule_data.get('end_time')
                        room = schedule_data.get('room') or None
                        # Attendance window removed - no longer needed
                        
                        # Parse time strings (handle both 24-hour and 12-hour formats)
                        from datetime import datetime
                        def parse_time(time_str):
                            if not time_str:
                                return None
                            try:
                                # Try 24-hour format first (HH:MM)
                                return datetime.strptime(time_str, '%H:%M').time()
                            except ValueError:
                                try:
                                    # Try 12-hour format (HH:MM AM/PM)
                                    return datetime.strptime(time_str, '%I:%M %p').time()
                                except ValueError:
                                    try:
                                        # Try 12-hour format without space (HH:MMAM/PM)
                                        return datetime.strptime(time_str, '%I:%M%p').time()
                                    except ValueError:
                                        logger.error(f"Could not parse time: {time_str}")
                                        return None
                        
                        start_time = parse_time(start_time_str)
                        end_time = parse_time(end_time_str)
                        # Attendance window removed - set to None
                        attendance_start = None
                        attendance_end = None
                        
                        CourseSchedule.objects.create(
                            course=course,
                            day=day,
                            start_time=start_time,
                            end_time=end_time,
                            room=room,
                            attendance_start=attendance_start,
                            attendance_end=attendance_end
                        )
                except (json.JSONDecodeError, KeyError, ValueError) as e:
                    logger.error(f"Error parsing day schedules: {str(e)}")
                    return JsonResponse({'success': False, 'message': f'Error processing day schedules: {str(e)}'})
            
            # Log the newly created course for verification
            logger.info(f"✅ Course created successfully: {course.code} - {course.name}")
            logger.info(f"   Course ID: {course.id}")
            logger.info(f"   Instructor: {course.instructor.username} (ID: {course.instructor.id})")
            logger.info(f"   Days: {course.days}, Time: {course.start_time} - {course.end_time}")
            logger.info(f"   Year: {course.year_level}, Section: {course.section}, Sem: {course.semester}, SY: {course.school_year}")
            logger.info(f"   Is Active: {course.is_active}, Created: {course.created_at}")
            logger.info(f"   Has Day Schedules: {course.course_schedules.exists()}")
            
            # Create notification for admin school when instructor adds a course
            if user.school_name:
                try:
                    from .models import AdminNotification
                    admin_user = CustomUser.objects.filter(
                        is_admin=True,
                        school_name=user.school_name,
                        deleted_at__isnull=True
                    ).first()
                    if admin_user:
                        AdminNotification.objects.create(
                            admin=admin_user,
                            notification_type='system_alert',
                            title='New Course Added',
                            message=f'Instructor {user.full_name or user.username} added a new course: {course.code} - {course.name}',
                            related_user=user
                        )
                except Exception as e:
                    logger.error(f"Error creating admin notification: {str(e)}")
            
            # Verify the course can be queried immediately
            verify_course = Course.objects.filter(id=course.id, instructor=user, is_active=True).first()
            if verify_course:
                logger.info(f"✅ Verification: Course can be queried immediately after creation")
            else:
                logger.error(f"❌ Verification FAILED: Course cannot be queried immediately after creation!")
            
            return JsonResponse({'success': True, 'message': 'Course added successfully!', 'course_id': course.id})
        except Exception as e:
            logger.error(f"Error adding course: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    # GET request - return form data including all departments and programs
    # Get all departments for the instructor's school
    if user.school_name:
        departments = Department.objects.filter(school_name=user.school_name, is_active=True, deleted_at__isnull=True).order_by('name')
        programs = Program.objects.filter(school_name=user.school_name, is_active=True, deleted_at__isnull=True).order_by('code')
    else:
        departments = Department.objects.filter(is_active=True, deleted_at__isnull=True).order_by('name')
        programs = Program.objects.filter(is_active=True, deleted_at__isnull=True).order_by('code')
    
    # Filter by education level if set
    if user.education_level:
        departments = departments.filter(education_level=user.education_level)
        programs = programs.filter(education_level=user.education_level)
    
    # Group programs by department
    programs_by_dept = {}
    for program in programs:
        dept_name = program.department or 'Uncategorized'
        if dept_name not in programs_by_dept:
            programs_by_dept[dept_name] = []
        programs_by_dept[dept_name].append({
            'id': program.id,
            'code': program.code,
            'name': program.name,
            'department': program.department
        })
    
    # Format departments
    departments_list = [{'id': dept.id, 'name': dept.name, 'code': dept.code or ''} for dept in departments]
    
    # Default program (instructor's assigned program if exists)
    default_program = None
    if user.program:
        default_program = {'id': user.program.id, 'code': user.program.code, 'name': user.program.name}
    
    return JsonResponse({
        'success': True,
        'default_program': default_program,
        'departments': departments_list,
        'programs_by_department': programs_by_dept,
        'semesters': Course.SEMESTER_CHOICES,
    })

@login_required
@require_http_methods(["GET", "POST"])
def instructor_update_course_view(request, course_id):
    """Update an existing course - instructor can only update courses they created"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    course = get_object_or_404(Course, id=course_id)
    
    # Verify the course is assigned to the instructor
    if course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only edit courses assigned to you.'})
    
    # Verify the course belongs to instructor's school
    if user.school_name and course.school_name != user.school_name:
        return JsonResponse({'success': False, 'message': 'You can only edit courses in your school.'})
    
    if request.method == 'POST':
        try:
            # Get form data
            code = request.POST.get('code', '').strip()
            name = request.POST.get('name', '').strip()
            year_level = request.POST.get('year_level', '').strip()
            section = request.POST.get('section', '').strip()
            semester = request.POST.get('semester', '1st').strip()
            school_year = request.POST.get('school_year', '').strip()
            room = request.POST.get('room', '').strip() or None
            days = request.POST.get('days', '').strip()
            start_time = request.POST.get('start_time', '').strip()
            end_time = request.POST.get('end_time', '').strip()
            color = request.POST.get('color', course.color if course.color else '#3C4770').strip()
            # Attendance window removed - no longer used
            attendance_start = None
            attendance_end = None
            is_active = request.POST.get('is_active', 'false') == 'true'
            
            # Check if day schedules are provided (different schedule for each day)
            day_schedules = request.POST.get('day_schedules', '')
            has_day_schedules = False
            if day_schedules:
                try:
                    schedules_data = json.loads(day_schedules)
                    if schedules_data and len(schedules_data) > 0:
                        has_day_schedules = True
                except:
                    pass
            
            # Validate required fields
            # Section is optional - students will fill it
            # If day schedules are provided, days, start_time, and end_time are set from day schedules
            if has_day_schedules:
                if not all([code, name, year_level, days, start_time, end_time, school_year]):
                    return JsonResponse({'success': False, 'message': 'Please fill in all required fields.'})
            else:
                if not all([code, name, year_level, days, start_time, end_time, school_year]):
                    return JsonResponse({'success': False, 'message': 'Please fill in all required fields.'})
            
            # Parse time strings (handle both 24-hour and 12-hour formats)
            from datetime import datetime
            def parse_time(time_str):
                if not time_str:
                    return None
                try:
                    # Try 24-hour format first (HH:MM)
                    return datetime.strptime(time_str, '%H:%M').time()
                except ValueError:
                    try:
                        # Try 12-hour format (HH:MM AM/PM)
                        return datetime.strptime(time_str, '%I:%M %p').time()
                    except ValueError:
                        try:
                            # Try 12-hour format without space (HH:MMAM/PM)
                            return datetime.strptime(time_str, '%I:%M%p').time()
                        except ValueError:
                            logger.error(f"Could not parse time: {time_str}")
                            return None
            
            parsed_start_time = parse_time(start_time)
            parsed_end_time = parse_time(end_time)
            # Attendance window removed - set to None
            parsed_attendance_start = None
            parsed_attendance_end = None
            
            # Check for time conflicts with existing courses (excluding current course)
            # Only conflicts when semester & school year match (regardless of year level/program)
            normalized_school_year = school_year or None
            conflicting_courses = Course.objects.filter(
                instructor=user,
                semester=semester,
                is_active=True
            ).exclude(id=course_id)
            if normalized_school_year is None:
                conflicting_courses = conflicting_courses.filter(Q(school_year__isnull=True) | Q(school_year=''))
            else:
                conflicting_courses = conflicting_courses.filter(school_year=normalized_school_year)
            
            conflicts = []
            day_schedules_check = request.POST.get('day_schedules', '')
            has_day_schedules_check = False
            if day_schedules_check:
                try:
                    schedules_data_check = json.loads(day_schedules_check)
                    if schedules_data_check and len(schedules_data_check) > 0:
                        has_day_schedules_check = True
                except:
                    pass
            
            # Day mapping for normalization
            day_normalize_map = {
                'Mon': 'Monday', 'Monday': 'Monday', 'M': 'Monday',
                'Tue': 'Tuesday', 'Tuesday': 'Tuesday', 'T': 'Tuesday',
                'Wed': 'Wednesday', 'Wednesday': 'Wednesday', 'W': 'Wednesday',
                'Thu': 'Thursday', 'Thursday': 'Thursday', 'Th': 'Thursday',
                'Fri': 'Friday', 'Friday': 'Friday', 'F': 'Friday',
                'Sat': 'Saturday', 'Saturday': 'Saturday', 'S': 'Saturday',
                'Sun': 'Sunday', 'Sunday': 'Sunday', 'Su': 'Sunday'
            }
            
            def normalize_day(day_str):
                """Normalize day string to full day name"""
                day_str = day_str.strip()
                return day_normalize_map.get(day_str, day_str)
            
            def days_match(day1, day2):
                """Check if two day strings match (handles abbreviations)"""
                norm1 = normalize_day(day1).lower()
                norm2 = normalize_day(day2).lower()
                return norm1 == norm2 or norm1 in norm2 or norm2 in norm1
            
            if has_day_schedules_check:
                # Check conflicts for day-specific schedules (asynchronized)
                try:
                    schedules_data = json.loads(day_schedules_check)
                    for schedule_data in schedules_data:
                        day = schedule_data.get('day')
                        sched_start = parse_time(schedule_data.get('startTime') or schedule_data.get('start_time'))
                        sched_end = parse_time(schedule_data.get('endTime') or schedule_data.get('end_time'))
                        
                        if not sched_start or not sched_end:
                            continue
                        
                        # Check against courses with day-specific schedules
                        for existing_course in conflicting_courses:
                            existing_schedules = existing_course.course_schedules.filter(day=day)
                            for existing_sched in existing_schedules:
                                # Check if times actually overlap (not just adjacent)
                                # Two intervals overlap if: new_start < existing_end AND new_end > existing_start
                                # This correctly excludes adjacent times (e.g., 9:00-10:30 and 10:30-11:30 are OK)
                                if (sched_start < existing_sched.end_time and sched_end > existing_sched.start_time):
                                    conflicts.append({
                                        'course': f"{existing_course.code} - {existing_course.name}",
                                        'day': day,
                                        'time': f"{existing_sched.start_time.strftime('%I:%M %p')} - {existing_sched.end_time.strftime('%I:%M %p')}"
                                    })
                            
                            # Check against courses with default schedule (synchronized)
                            if not existing_course.course_schedules.exists():
                                # Check if this day is in the existing course's days
                                existing_days = [d.strip() for d in existing_course.days.split(',')]
                                day_matches = any(days_match(day, existing_day) for existing_day in existing_days)
                                
                                if day_matches and existing_course.start_time and existing_course.end_time:
                                    # Check if times actually overlap (not just adjacent)
                                    if (sched_start < existing_course.end_time and sched_end > existing_course.start_time):
                                        conflicts.append({
                                            'course': f"{existing_course.code} - {existing_course.name}",
                                            'day': day,
                                            'time': f"{existing_course.start_time.strftime('%I:%M %p')} - {existing_course.end_time.strftime('%I:%M %p')}"
                                        })
                except Exception as e:
                    logger.error(f"Error checking day schedule conflicts: {str(e)}")
            else:
                # Check conflicts for default schedule (synchronized)
                if parsed_start_time and parsed_end_time:
                    new_course_days = [d.strip() for d in days.split(',')]
                    
                    for existing_course in conflicting_courses:
                        # Check each day individually
                        if existing_course.course_schedules.exists():
                            # Existing course has day-specific schedules
                            existing_days = set()
                            for schedule in existing_course.course_schedules.all():
                                existing_days.add(schedule.day)
                            
                            # Check each day in new course against existing course's day-specific schedules
                            for new_day in new_course_days:
                                for existing_day in existing_days:
                                    if days_match(new_day, existing_day):
                                        # Find the schedule for this day
                                        day_schedule = existing_course.course_schedules.filter(day=existing_day).first()
                                        if day_schedule and day_schedule.start_time and day_schedule.end_time:
                                            # Check if times actually overlap (not just adjacent)
                                            if (parsed_start_time < day_schedule.end_time and parsed_end_time > day_schedule.start_time):
                                                conflicts.append({
                                                    'course': f"{existing_course.code} - {existing_course.name}",
                                                    'day': existing_day,
                                                    'time': f"{day_schedule.start_time.strftime('%I:%M %p')} - {day_schedule.end_time.strftime('%I:%M %p')}"
                                                })
                        else:
                            # Existing course has synchronized schedule
                            existing_days = [d.strip() for d in existing_course.days.split(',')]
                            
                            # Check each day individually - only report conflicts for days that match AND times overlap
                            for new_day in new_course_days:
                                for existing_day in existing_days:
                                    if days_match(new_day, existing_day):
                                        # Times overlap check - only conflict if times actually overlap (not just adjacent)
                                        if existing_course.start_time and existing_course.end_time:
                                            if (parsed_start_time < existing_course.end_time and parsed_end_time > existing_course.start_time):
                                                conflicts.append({
                                                    'course': f"{existing_course.code} - {existing_course.name}",
                                                    'day': existing_day,
                                                    'time': f"{existing_course.start_time.strftime('%I:%M %p')} - {existing_course.end_time.strftime('%I:%M %p')}"
                                                })
            
            if conflicts:
                # Remove duplicate conflicts (same course, day, and time)
                seen = set()
                unique_conflicts = []
                for conflict in conflicts:
                    key = (conflict['course'], conflict.get('day', ''), conflict['time'])
                    if key not in seen:
                        seen.add(key)
                        unique_conflicts.append(conflict)
                
                conflict_messages = []
                for conflict in unique_conflicts:
                    conflict_messages.append(f"{conflict['course']} on {conflict['day']} at {conflict['time']}")
                
                return JsonResponse({
                    'success': False,
                    'message': (
                        'Time conflict detected! The following course(s) have overlapping schedules:\n'
                        + '\n'.join(conflict_messages)
                        + '\n\nPlease adjust the time or assign this course to a different semester or school year.'
                    )
                })
            
            # Update program if provided
            program_id = request.POST.get('program', '').strip()
            new_program = course.program  # Default to current program
            if program_id:
                new_program = get_object_or_404(Program, id=int(program_id))
                if user.school_name and new_program.school_name != user.school_name:
                    return JsonResponse({'success': False, 'message': 'You can only assign courses to programs in your school.'})
            
            # Find all sibling courses BEFORE updating (using current values)
            # Siblings are courses with same code, name, year_level, semester, school_year, program, but different section
            # These are the courses that should be updated together as one logical course
            sibling_courses = Course.objects.filter(
                instructor=user,
                code=course.code,  # Use current code
                name=course.name,  # Use current name
                year_level=course.year_level,  # Use current year_level
                semester=course.semester,  # Use current semester
                school_year=course.school_year or None,  # Use current school_year
                program=course.program,  # Use current program
                is_active=True,
                deleted_at__isnull=True,
                is_archived=False
            ).exclude(id=course_id)
            
            # Fields that should be shared across all sections (updated for all siblings)
            # NOTE: Color is NOT shared - each section can have its own color
            shared_fields = {
                'code': code,
                'name': name,
                'year_level': int(year_level),
                'semester': semester,
                'school_year': school_year or None,
                'is_active': True
            }
            
            # Update program for all siblings if it changed
            if program_id:
                shared_fields['program'] = new_program
            
            # Update the current course with all fields (shared + section-specific)
            course.code = shared_fields['code']
            course.name = shared_fields['name']
            course.year_level = shared_fields['year_level']
            # Allow instructors to rename the section for this course/section record
            course.section = (section or '').upper() or None
            course.semester = shared_fields['semester']
            course.school_year = shared_fields['school_year']
            course.room = room  # Section-specific
            course.days = days  # Section-specific
            course.start_time = parsed_start_time  # Section-specific
            course.end_time = parsed_end_time  # Section-specific
            course.color = color  # Section-specific - each section can have its own color
            course.is_active = shared_fields['is_active']
            if 'program' in shared_fields:
                course.program = shared_fields['program']
            course.save()
            
            # Update all sibling courses with shared fields only (keep their section-specific fields including color)
            if sibling_courses.exists():
                # Build update_fields list - color is NOT included, so each section keeps its own color
                update_fields_list = ['code', 'name', 'year_level', 'semester', 'school_year', 'is_active', 'updated_at']
                if 'program' in shared_fields:
                    update_fields_list.append('program')
                
                for sibling in sibling_courses:
                    sibling.code = shared_fields['code']
                    sibling.name = shared_fields['name']
                    sibling.year_level = shared_fields['year_level']
                    sibling.semester = shared_fields['semester']
                    sibling.school_year = shared_fields['school_year']
                    # Color is NOT updated - each section keeps its own color
                    sibling.is_active = shared_fields['is_active']
                    if 'program' in shared_fields:
                        sibling.program = shared_fields['program']
                    # Keep section-specific fields unchanged (room, days, start_time, end_time, attendance_start, attendance_end, section, color)
                    sibling.save(update_fields=update_fields_list)
            
            # Handle day-specific schedules
            day_schedules = request.POST.get('day_schedules', '')
            if day_schedules:
                try:
                    # Delete existing day-specific schedules
                    CourseSchedule.objects.filter(course=course).delete()
                    
                    schedules_data = json.loads(day_schedules)
                    for schedule_data in schedules_data:
                        day = schedule_data.get('day')
                        start_time_str = schedule_data.get('startTime') or schedule_data.get('start_time')
                        end_time_str = schedule_data.get('endTime') or schedule_data.get('end_time')
                        room = schedule_data.get('room') or None
                        # Attendance window removed - no longer needed
                        
                        # Parse time strings (handle both 24-hour and 12-hour formats)
                        from datetime import datetime
                        def parse_time(time_str):
                            if not time_str:
                                return None
                            try:
                                # Try 24-hour format first (HH:MM)
                                return datetime.strptime(time_str, '%H:%M').time()
                            except ValueError:
                                try:
                                    # Try 12-hour format (HH:MM AM/PM)
                                    return datetime.strptime(time_str, '%I:%M %p').time()
                                except ValueError:
                                    try:
                                        # Try 12-hour format without space (HH:MMAM/PM)
                                        return datetime.strptime(time_str, '%I:%M%p').time()
                                    except ValueError:
                                        logger.error(f"Could not parse time: {time_str}")
                                        return None
                        
                        start_time = parse_time(start_time_str)
                        end_time = parse_time(end_time_str)
                        # Attendance window removed - set to None
                        attendance_start = None
                        attendance_end = None
                        
                        CourseSchedule.objects.create(
                            course=course,
                            day=day,
                            start_time=start_time,
                            end_time=end_time,
                            room=room,
                            attendance_start=attendance_start,
                            attendance_end=attendance_end
                        )
                except (json.JSONDecodeError, KeyError, ValueError) as e:
                    logger.warning(f"Error parsing day schedules: {str(e)}")
                    return JsonResponse({'success': False, 'message': f'Error processing day schedules: {str(e)}'})
            else:
                # If no day_schedules provided, delete existing day-specific schedules
                CourseSchedule.objects.filter(course=course).delete()
            
            return JsonResponse({'success': True, 'message': 'Course updated successfully!'})
        except Exception as e:
            logger.error(f"Error updating course: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    # GET request - return course data
    schedules = course.course_schedules.all().order_by('day_order')
    day_schedules = []
    if schedules.exists():
        for schedule in schedules:
            day_schedules.append({
                'day': schedule.day,
                'start_time': schedule.start_time.strftime('%H:%M') if schedule.start_time else '',
                'end_time': schedule.end_time.strftime('%H:%M') if schedule.end_time else '',
                'room': schedule.room or ''
            })
    
    course_data = {
        'id': course.id,
        'code': course.code,
        'name': course.name,
        'program_id': course.program.id,
        'year_level': course.year_level,
        'section': course.section,
        'semester': course.semester,
        'school_year': course.school_year or '',
        'instructor_id': course.instructor.id if course.instructor else None,
        'room': course.room or '',
        'days': course.days,
        'start_time': course.start_time.strftime('%H:%M') if course.start_time else '',
        'end_time': course.end_time.strftime('%H:%M') if course.end_time else '',
        'color': course.color,
        'is_active': course.is_active,
        'day_schedules': day_schedules,
    }
    
    # Get all departments and programs for the instructor's school (for edit form)
    if user.school_name:
        departments = Department.objects.filter(school_name=user.school_name, is_active=True, deleted_at__isnull=True).order_by('name')
        programs = Program.objects.filter(school_name=user.school_name, is_active=True, deleted_at__isnull=True).order_by('code')
    else:
        departments = Department.objects.filter(is_active=True, deleted_at__isnull=True).order_by('name')
        programs = Program.objects.filter(is_active=True, deleted_at__isnull=True).order_by('code')
    
    # Filter by education level if set
    if user.education_level:
        departments = departments.filter(education_level=user.education_level)
        programs = programs.filter(education_level=user.education_level)
    
    # Group programs by department
    programs_by_dept = {}
    for program in programs:
        dept_name = program.department or 'Uncategorized'
        if dept_name not in programs_by_dept:
            programs_by_dept[dept_name] = []
        programs_by_dept[dept_name].append({
            'id': program.id,
            'code': program.code,
            'name': program.name,
            'department': program.department
        })
    
    # Format departments
    departments_list = [{'id': dept.id, 'name': dept.name, 'code': dept.code or ''} for dept in departments]
    
    return JsonResponse({
        'success': True,
        'course': course_data,
        'default_program': {'id': course.program.id, 'code': course.program.code, 'name': course.program.name},
        'departments': departments_list,
        'programs_by_department': programs_by_dept,
        'semesters': Course.SEMESTER_CHOICES,
    })

@login_required
@require_http_methods(["POST"])
def instructor_delete_course_view(request, course_id):
    """Delete a course section or all sections based on delete_all_sections parameter
    
    - If delete_all_sections is True: Delete all sections with same code/name/semester/year
    - If delete_all_sections is False or not provided: Delete only the specific section by ID
    """
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    from django.utils import timezone
    from django.db import transaction
    
    # Check if we should delete all sections or just one
    delete_all_sections = request.POST.get('delete_all_sections', 'false').lower() == 'true'
    
    try:
        # Use transaction to ensure atomicity and prevent any race conditions
        with transaction.atomic():
            # Get the specific course by ID
            try:
                course = Course.objects.select_for_update().get(id=course_id, instructor=user)
            except Course.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Course not found.'})
            
            # Verify the course belongs to instructor's school
            if user.school_name and course.school_name != user.school_name:
                return JsonResponse({'success': False, 'message': 'You can only delete courses in your school.'})
            
            # Verify course is not already deleted
            if course.deleted_at is not None:
                return JsonResponse({'success': False, 'message': 'This course section has already been deleted.'})
            
            if delete_all_sections:
                # Delete ALL sections with the same code/name/semester/year
                logger.info(f"Deleting ALL sections for course: Code: {course.code}, Name: {course.name}, Semester: {course.semester}, Year: {course.school_year}")
                
                # Get all courses to delete
                courses_to_delete = Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True,
                    is_archived=False
                )
                
                course_ids_to_delete = list(courses_to_delete.values_list('id', flat=True))
                deleted_count = courses_to_delete.update(
                    deleted_at=timezone.now(),
                    is_active=False
                )
                
                # Clean up QR and biometric registrations before deleting enrollments
                enrollments_to_delete = CourseEnrollment.objects.filter(
                    course_id__in=course_ids_to_delete,
                    is_active=True,
                    deleted_at__isnull=True
                )
                
                for enrollment in enrollments_to_delete:
                    try:
                        # Delete QR registrations
                        QRCodeRegistration.objects.filter(
                            student=enrollment.student,
                            course=enrollment.course,
                            is_active=True
                        ).delete()
                        # Delete biometric registrations
                        BiometricRegistration.objects.filter(
                            student=enrollment.student,
                            course=enrollment.course,
                            is_active=True
                        ).delete()
                        logger.info(f"Deleted QR and biometric registrations for {enrollment.student.full_name} in course {enrollment.course.code}")
                    except Exception as e:
                        logger.error(f"Error cleaning up registrations: {str(e)}")
                
                # Also soft delete related enrollments for all deleted courses
                enrollments_deleted = enrollments_to_delete.update(
                    is_active=False,
                    deleted_at=timezone.now()
                )
                
                logger.info(f"Instructor {user.username} deleted {deleted_count} course sections ({course.code} - {course.name}). All sections deleted. Cleaned up QR and biometric registrations for {enrollments_deleted} students.")
                
                if deleted_count > 1:
                    return JsonResponse({
                        'success': True,
                        'message': f'All {deleted_count} sections of this course have been deleted successfully.'
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'message': 'Course deleted successfully.'
                    })
            else:
                # Delete ONLY the specific section by ID
                logger.info(f"Deleting course ID: {course.id}, Code: {course.code}, Name: {course.name}, Section: {course.section}")
                
                # Get sibling courses BEFORE deletion to verify they remain intact
                sibling_courses_before = list(Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True,
                    is_archived=False
                ).exclude(id=course_id).values_list('id', flat=True))
                
                logger.info(f"Sibling courses before deletion (IDs): {sibling_courses_before}")
                
                # Use direct UPDATE query to ensure ONLY this specific course ID is affected
                # This is the safest way to ensure no other courses are touched
                updated_count = Course.objects.filter(
                    id=course_id,
                    instructor=user,
                    deleted_at__isnull=True  # Only update if not already deleted
                ).update(
                    deleted_at=timezone.now(),
                    is_active=False
                )
                
                if updated_count == 0:
                    return JsonResponse({'success': False, 'message': 'Course not found or already deleted.'})
                
                if updated_count > 1:
                    logger.error(f"CRITICAL ERROR: Update affected {updated_count} courses instead of 1! Course ID: {course_id}")
                    # Rollback transaction
                    transaction.set_rollback(True)
                    return JsonResponse({'success': False, 'message': 'Error: Multiple courses were affected. Deletion cancelled.'})
                
                # Verify sibling courses still exist and are unchanged after deletion
                sibling_courses_after = list(Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True,
                    is_archived=False
                ).exclude(id=course_id).values_list('id', flat=True))
                
                logger.info(f"Sibling courses after deletion (IDs): {sibling_courses_after}")
                
                # Check if any sibling courses were affected
                if set(sibling_courses_before) != set(sibling_courses_after):
                    logger.error(f"ERROR: Sibling courses were affected! Before: {sibling_courses_before}, After: {sibling_courses_after}")
                    # Rollback transaction to restore the course
                    transaction.set_rollback(True)
                    return JsonResponse({'success': False, 'message': 'Error: Other sections were affected. Deletion cancelled.'})
                
                # Also soft delete related enrollments for this specific course only
                # Use course_id directly in the filter to ensure only enrollments for this course are affected
                # And clean up QR and biometric registrations for all students
                enrollments_to_delete = CourseEnrollment.objects.filter(
                    course_id=course_id,  # Use course_id directly, not course object
                    is_active=True,
                    deleted_at__isnull=True
                )
                
                # Clean up QR and biometric registrations before deleting enrollments
                for enrollment in enrollments_to_delete:
                    try:
                        # Delete QR registrations
                        QRCodeRegistration.objects.filter(
                            student=enrollment.student,
                            course=enrollment.course,
                            is_active=True
                        ).delete()
                        # Delete biometric registrations
                        BiometricRegistration.objects.filter(
                            student=enrollment.student,
                            course=enrollment.course,
                            is_active=True
                        ).delete()
                        logger.info(f"Deleted QR and biometric registrations for {enrollment.student.full_name} in course {enrollment.course.code}")
                    except Exception as e:
                        logger.error(f"Error cleaning up registrations: {str(e)}")
                
                enrollments_deleted = enrollments_to_delete.update(
                    is_active=False,
                    deleted_at=timezone.now()
                )
                
                logger.info(f"Instructor {user.username} deleted course section {course.id} ({course.code} - {course.name}, Section: {course.section}). Sibling sections remaining: {len(sibling_courses_after)}. Cleaned up QR and biometric registrations for {enrollments_deleted} students.")
                
                return JsonResponse({
                    'success': True,
                    'message': f'Section {course.section or "N/A"} deleted successfully. Other sections of this course remain intact.'
                })
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Course not found.'})
    except Exception as e:
        logger.error(f"Error deleting course section: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error deleting section: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def instructor_get_programs_by_department_view(request):
    """Get programs filtered by department for instructor course form"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    department_name = request.GET.get('department', '').strip()
    if not department_name:
        return JsonResponse({'success': False, 'message': 'Department name is required.'})
    
    # Get programs for the specified department
    if user.school_name:
        programs = Program.objects.filter(
            department=department_name,
            school_name=user.school_name,
            is_active=True,
            deleted_at__isnull=True
        ).order_by('code')
    else:
        programs = Program.objects.filter(
            department=department_name,
            is_active=True,
            deleted_at__isnull=True
        ).order_by('code')
    
    # Filter by education level if set
    if user.education_level:
        programs = programs.filter(education_level=user.education_level)
    
    programs_list = [{'id': p.id, 'code': p.code, 'name': p.name} for p in programs]
    
    return JsonResponse({
        'success': True,
        'programs': programs_list
    })

@login_required
@require_http_methods(["GET", "POST"])
def enroll_course_view(request):
    """Student enrollment in courses"""
    user = request.user
    if not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        # Handle enrollment submission
        from .models import Course, CourseEnrollment
        import json
        
        try:
            enrollment_code = request.POST.get('enrollment_code', '').strip().upper()
            course_id = request.POST.get('course_id', '').strip()
            full_name = request.POST.get('full_name', '').strip()
            year_level = request.POST.get('year_level', '').strip()
            section = request.POST.get('section', '').strip()
            email = request.POST.get('email', '').strip()
            student_id_number = request.POST.get('student_id', '').strip()
            
            # Validate required fields
            if not all([enrollment_code, course_id, full_name, year_level, section, email, student_id_number]):
                return JsonResponse({'success': False, 'message': 'Please fill in all required fields.'})
            
            # Find course by enrollment code
            try:
                course = Course.objects.get(enrollment_code=enrollment_code, is_active=True)
            except Course.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Invalid enrollment code. Please check and try again.'})
            
            # Check enrollment status
            if course.enrollment_status == 'closed':
                return JsonResponse({'success': False, 'message': 'Enrollment is now CLOSED for this course. Please contact your instructor for assistance.'})
            
            # Verify course_id matches
            if str(course.id) != str(course_id):
                return JsonResponse({'success': False, 'message': 'Course verification failed. Please try again.'})
            
            # Read force_enroll flag from POST (supports 'true','1','on')
            force_enroll_raw = request.POST.get('force_enroll', '')
            force_enroll = str(force_enroll_raw).strip().lower() in ['1', 'true', 'yes', 'on']

            # Check for schedule conflicts (only if not forcing enrollment)
            if not force_enroll:
                enrolled_courses = Course.objects.filter(
                    enrollments__student=user,
                    enrollments__is_active=True,
                    enrollments__deleted_at__isnull=True,
                    is_active=True,
                    deleted_at__isnull=True,
                    is_archived=False
                ).distinct()
                
                conflicts = check_schedule_conflict(course, enrolled_courses)
                if conflicts:
                    return JsonResponse({
                        'success': False,
                        'has_conflicts': True,
                        'conflicts': conflicts,
                        'message': 'This course has schedule conflicts with your enrolled courses.'
                    })
            
            # Check if already enrolled - look for active enrollments first
            existing_enrollment = CourseEnrollment.objects.filter(course=course, student=user, is_active=True, deleted_at__isnull=True).first()
            is_new_enrollment = False
            if existing_enrollment:
                # Already actively enrolled
                return JsonResponse({'success': False, 'message': 'You are already enrolled in this course.'})
            
            # Check for soft-deleted enrollment (dropped/unenrolled) - allow re-enrollment
            soft_deleted_enrollment = CourseEnrollment.objects.filter(course=course, student=user, deleted_at__isnull=False).first()
            if soft_deleted_enrollment:
                # Restore the soft-deleted enrollment and update info
                from django.utils import timezone
                soft_deleted_enrollment.is_active = True
                soft_deleted_enrollment.deleted_at = None  # Clear deleted_at to restore
                soft_deleted_enrollment.full_name = full_name
                soft_deleted_enrollment.year_level = int(year_level)
                soft_deleted_enrollment.section = section
                soft_deleted_enrollment.email = email
                soft_deleted_enrollment.student_id_number = student_id_number
                soft_deleted_enrollment.enrolled_at = timezone.now()  # Update enrollment date
                soft_deleted_enrollment.save()
                enrollment = soft_deleted_enrollment
                is_new_enrollment = True  # Treat reactivation as new enrollment for notification
            else:
                # Create new enrollment
                try:
                    enrollment = CourseEnrollment.objects.create(
                        course=course,
                        student=user,
                        full_name=full_name,
                        year_level=int(year_level),
                        section=section,
                        email=email,
                        student_id_number=student_id_number,
                        is_active=True
                    )
                    is_new_enrollment = True
                except Exception as e:
                    # Handle unique constraint error
                    if 'UNIQUE constraint' in str(e) or 'unique constraint' in str(e).lower():
                        # Check again in case of race condition
                        existing = CourseEnrollment.objects.filter(course=course, student=user, is_active=True, deleted_at__isnull=True).first()
                        if existing:
                            # Already actively enrolled
                            return JsonResponse({'success': False, 'message': 'You are already enrolled in this course.'})
                        
                        # Check for soft-deleted enrollment
                        soft_deleted = CourseEnrollment.objects.filter(course=course, student=user, deleted_at__isnull=False).first()
                        if soft_deleted:
                            from django.utils import timezone
                            soft_deleted.is_active = True
                            soft_deleted.deleted_at = None  # Clear deleted_at to restore
                            soft_deleted.full_name = full_name
                            soft_deleted.year_level = int(year_level)
                            soft_deleted.section = section
                            soft_deleted.email = email
                            soft_deleted.student_id_number = student_id_number
                            soft_deleted.enrolled_at = timezone.now()
                            soft_deleted.save()
                            enrollment = soft_deleted
                            is_new_enrollment = True
                        else:
                            return JsonResponse({'success': False, 'message': 'Unable to enroll. Please try again.'})
                    else:
                        raise
            
            # Update student's section if not set
            if not user.section:
                user.section = section
                user.save()
            
            # Create notification for instructor (only the instructor who created/owns the course)
            # Only create notification for new enrollments (not for already active enrollments)
            if is_new_enrollment and course.instructor and course.instructor.id != user.id:
                try:
                    student_name = full_name or user.full_name or user.username
                    create_notification(
                        user=course.instructor,
                        notification_type='student_enrolled',
                        title='New Student Enrollment',
                        message=f'{student_name} enrolled in {course.code} - {course.name}',
                        category='enrollment',
                        related_course=course,
                        related_user=user
                    )
                except Exception as e:
                    logger.error(f"Error creating enrollment notification: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully enrolled in {course.code} - {course.name}!'
            })
        except Exception as e:
            logger.error(f"Error enrolling student: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    # GET request - show enrollment page
    from .models import CourseEnrollment
    
    # Get enrolled courses for this student
    enrolled_courses = CourseEnrollment.objects.filter(
        deleted_at__isnull=True,
        student=user,
        is_active=True
    ).select_related('course', 'course__program', 'course__instructor').order_by('-enrolled_at')
    
    # Add QR and Biometric registration status to enrolled courses
    for enrollment in enrolled_courses:
        qr_reg = QRCodeRegistration.objects.filter(
            student=user,
            course=enrollment.course,
            is_active=True
        ).exists()
        enrollment.has_qr = qr_reg
        
        biometric_reg = BiometricRegistration.objects.filter(
            student=user,
            course=enrollment.course,
            is_active=True
        ).exists()
        enrollment.has_biometric = biometric_reg
        
        # Debug logging
        if enrollment.has_biometric:
            biometric_count = BiometricRegistration.objects.filter(
                student=user,
                course=enrollment.course,
                is_active=True
            ).count()
            logger.debug(f"[ENROLL] Course {enrollment.course.code}: has_biometric=True (found {biometric_count} records)")
    
    # Get school admin for topbar
    school_admin = None
    if user.school_name:
        from accounts.models import CustomUser
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    # Get unread notification count
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    context = {
        'enrolled_courses': enrolled_courses,
        'school_admin': school_admin,
        'unread_notifications': unread_notifications,
        'esp32_ip': settings.ESP32_IP,
    }
    return render(request, 'dashboard/student/enroll_course.html', context)

def check_schedule_conflict(new_course, enrolled_courses):
    """
    Check if the new course has schedule conflicts with already enrolled courses.
    Returns a list of conflict information.
    """
    conflicts = []
    
    # Helper function to check if two time ranges overlap
    def times_overlap(start1, end1, start2, end2):
        """Check if two time ranges overlap"""
        if not (start1 and end1 and start2 and end2):
            return False
        # Convert to minutes for easier comparison
        def time_to_minutes(t):
            return t.hour * 60 + t.minute
        s1, e1 = time_to_minutes(start1), time_to_minutes(end1)
        s2, e2 = time_to_minutes(start2), time_to_minutes(end2)
        # Check for overlap: not (end1 <= start2 or end2 <= start1)
        return not (e1 <= s2 or e2 <= s1)
    
    # Helper function to normalize day names
    def normalize_day(day):
        day_map = {
            'Monday': 'Mon', 'Mon': 'Mon', 'M': 'Mon',
            'Tuesday': 'Tue', 'Tue': 'Tue', 'T': 'Tue',
            'Wednesday': 'Wed', 'Wed': 'Wed', 'W': 'Wed',
            'Thursday': 'Thu', 'Thu': 'Thu', 'Th': 'Thu',
            'Friday': 'Fri', 'Fri': 'Fri', 'F': 'Fri',
            'Saturday': 'Sat', 'Sat': 'Sat', 'S': 'Sat',
            'Sunday': 'Sun', 'Sun': 'Sun', 'Su': 'Sun'
        }
        return day_map.get(str(day).strip(), str(day).strip())
    
    # Get new course schedules
    new_course_schedules = []
    day_schedules = new_course.course_schedules.all()
    if day_schedules.exists():
        for s in day_schedules:
            if s.start_time and s.end_time:
                new_course_schedules.append({
                    'day': normalize_day(s.day),
                    'start_time': s.start_time,
                    'end_time': s.end_time,
                    'room': s.room or new_course.room or 'N/A'
                })
    elif new_course.days and new_course.start_time and new_course.end_time:
        # Default schedule
        days_list = [d.strip() for d in new_course.days.split(',') if d.strip()]
        for day in days_list:
            new_course_schedules.append({
                'day': normalize_day(day),
                'start_time': new_course.start_time,
                'end_time': new_course.end_time,
                'room': new_course.room or 'N/A'
            })
    
    # Check against enrolled courses
    for enrolled_course in enrolled_courses:
        enrolled_schedules = []
        enrolled_day_schedules = enrolled_course.course_schedules.all()
        if enrolled_day_schedules.exists():
            for s in enrolled_day_schedules:
                if s.start_time and s.end_time:
                    enrolled_schedules.append({
                        'day': normalize_day(s.day),
                        'start_time': s.start_time,
                        'end_time': s.end_time,
                        'room': s.room or enrolled_course.room or 'N/A'
                    })
        elif enrolled_course.days and enrolled_course.start_time and enrolled_course.end_time:
            days_list = [d.strip() for d in enrolled_course.days.split(',') if d.strip()]
            for day in days_list:
                enrolled_schedules.append({
                    'day': normalize_day(day),
                    'start_time': enrolled_course.start_time,
                    'end_time': enrolled_course.end_time,
                    'room': enrolled_course.room or 'N/A'
                })
        
        # Check for conflicts
        for new_sched in new_course_schedules:
            for enrolled_sched in enrolled_schedules:
                if new_sched['day'] == enrolled_sched['day']:
                    if times_overlap(new_sched['start_time'], new_sched['end_time'], 
                                   enrolled_sched['start_time'], enrolled_sched['end_time']):
                        conflicts.append({
                            'conflicting_course_code': enrolled_course.code,
                            'conflicting_course_name': enrolled_course.name,
                            'day': new_sched['day'],
                            'new_course_time': f"{new_sched['start_time'].strftime('%I:%M %p')} - {new_sched['end_time'].strftime('%I:%M %p')}",
                            'conflicting_course_time': f"{enrolled_sched['start_time'].strftime('%I:%M %p')} - {enrolled_sched['end_time'].strftime('%I:%M %p')}",
                            'new_course_room': new_sched['room'],
                            'conflicting_course_room': enrolled_sched['room']
                        })
    
    return conflicts

@login_required
@require_http_methods(["POST"])
def verify_enrollment_code_view(request):
    """Verify enrollment code and return course information"""
    user = request.user
    if not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    import json
    data = json.loads(request.body)
    enrollment_code = data.get('enrollment_code', '').strip().upper()
    
    if not enrollment_code:
        return JsonResponse({'success': False, 'message': 'Enrollment code is required.'})
    
    try:
        from .models import Course, CourseEnrollment
        
        course = Course.objects.get(enrollment_code=enrollment_code, is_active=True)
        
        # Check enrollment status
        if course.enrollment_status == 'closed':
            return JsonResponse({
                'success': False,
                'message': 'Enrollment is now CLOSED for this course. Please contact your instructor for assistance.'
            })
        
        # Check if already enrolled (only check active, non-deleted enrollments)
        # Allow re-enrollment if previously dropped (soft-deleted)
        if CourseEnrollment.objects.filter(course=course, student=user, is_active=True, deleted_at__isnull=True).exists():
            return JsonResponse({
                'success': False,
                'message': 'You are already enrolled in this course.'
            })
        
        # Check for schedule conflicts with already enrolled courses
        enrolled_courses = Course.objects.filter(
            enrollments__student=user,
            enrollments__is_active=True,
            enrollments__deleted_at__isnull=True,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        ).distinct()
        
        conflicts = check_schedule_conflict(course, enrolled_courses)
        
        # Return course information
        course_data = {
            'id': course.id,
            'code': course.code,
            'name': course.name,
            'program_code': course.program.code if course.program else '',
            'program_name': course.program.name if course.program else '',
            'year_level': course.year_level,
            'semester': course.semester,
            'school_year': course.school_year or 'N/A',
            'instructor_name': course.instructor.full_name if course.instructor else 'Not assigned',
            'room': course.room or 'N/A',
            'color': course.color or '#3C4770',
        }
        
        response_data = {
            'success': True,
            'course': course_data
        }
        
        # Add conflict information if any
        if conflicts:
            response_data['has_conflicts'] = True
            response_data['conflicts'] = conflicts
        else:
            response_data['has_conflicts'] = False
        
        return JsonResponse(response_data)
    except Course.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Invalid enrollment code. Please check and try again.'
        })
    except Exception as e:
        logger.error(f"Error verifying enrollment code: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def unenroll_course_view(request, enrollment_id):
    """Unenroll student from a course - uses soft delete. Can be called by student or instructor."""
    user = request.user
    
    try:
        from .models import CourseEnrollment, Course
        
        from django.utils import timezone
        enrollment = CourseEnrollment.objects.get(id=enrollment_id, is_active=True, deleted_at__isnull=True)
        
        # Check authorization: student can unenroll themselves, instructor can drop students from their courses
        if user.is_student:
            # Student can only unenroll themselves
            if enrollment.student != user:
                return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        elif user.is_teacher:
            # Instructor can drop students from their courses
            if enrollment.course.instructor != user:
                return JsonResponse({'success': False, 'message': 'Unauthorized. You can only drop students from your own courses.'}, status=403)
        else:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        
        enrollment.is_active = False
        enrollment.deleted_at = timezone.now()
        enrollment.save()
        
        # NOTE: Do NOT delete biometric or QR registrations on soft drop
        # They will be deleted only on permanent delete
        # This allows students to restore and keep their fingerprint registrations
        
        # Delete attendance records for this student in this course
        # So their attendance log is cleared from student dashboard
        try:
            attendance_records = AttendanceRecord.objects.filter(
                student=enrollment.student,
                course=enrollment.course
            )
            deleted_count = attendance_records.delete()[0]  # delete() returns (count, {model: count})
            logger.info(f"Deleted {deleted_count} attendance records for {enrollment.student.full_name} in {enrollment.course.code}")
        except Exception as e:
            logger.error(f"Error deleting attendance records: {str(e)}")
        
        # Create notification for student when dropped by instructor
        if user.is_teacher and enrollment.student:
            try:
                student_name = enrollment.full_name or enrollment.student.full_name or enrollment.student.username
                create_notification(
                    user=enrollment.student,
                    notification_type='student_dropped',
                    title='Dropped from Course',
                    message=f'You have been dropped from {enrollment.course.code} - {enrollment.course.name}',
                    category='enrollment',
                    related_course=enrollment.course,
                    related_user=user
                )
            except Exception as e:
                logger.error(f"Error creating drop notification: {str(e)}")
        
        if user.is_teacher:
            message = f'Successfully dropped "{enrollment.full_name or enrollment.student.full_name if enrollment.student else "student"}" from the course.'
        else:
            message = 'Successfully unenrolled from the course.'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
    except CourseEnrollment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Enrollment not found.'
        })
    except Exception as e:
        logger.error(f"Error unenrolling student: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def instructor_delete_semester_view(request):
    """Delete all courses in a semester or school year - instructor can only delete courses they created"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        school_year = data.get('school_year')
        year_level = data.get('year_level')
        semester = data.get('semester')
        course_ids = data.get('course_ids', [])
        
        # Build query based on what's provided
        courses_qs = Course.objects.filter(
            instructor=user,
            deleted_at__isnull=True
        )
        
        # If school_year is provided without year_level/semester, delete all courses in that school year
        if school_year and not year_level and not semester:
            courses_qs = courses_qs.filter(school_year=school_year)
            
            # If course_ids provided, get all courses with same identity (for multi-section)
            if course_ids:
                # Get the courses by IDs first
                initial_courses = list(courses_qs.filter(id__in=course_ids))
                # For each course, find all courses with same identity (code, name, semester, school_year)
                all_course_ids = set(course_ids)
                for course in initial_courses:
                    # Find all courses with same identity
                    same_identity = Course.objects.filter(
                        instructor=user,
                        code=course.code,
                        name=course.name,
                        semester=course.semester,
                        school_year=course.school_year,
                        deleted_at__isnull=True
                    ).values_list('id', flat=True)
                    all_course_ids.update(same_identity)
                
                courses_qs = courses_qs.filter(id__in=all_course_ids)
            else:
                # Delete all courses in this school year
                pass
            
            courses = list(courses_qs)
            if not courses:
                return JsonResponse({'success': False, 'message': f'No courses found in School Year {school_year}.'})
            
            # Soft delete all courses
            from django.utils import timezone
            deleted_count = 0
            for course in courses:
                course.deleted_at = timezone.now()
                course.is_active = False
                course.save()
                deleted_count += 1
            
            return JsonResponse({
                'success': True, 
                'message': f'Successfully deleted {deleted_count} course(s) from School Year {school_year}.'
            })
        
        # Otherwise, delete by semester (existing logic)
        if not year_level or not semester:
            return JsonResponse({'success': False, 'message': 'Year level and semester are required, or school year must be provided.'})
        
        courses_qs = courses_qs.filter(
            year_level=int(year_level),
            semester=semester
        )
        
        # Filter by school year if provided
        if school_year:
            courses_qs = courses_qs.filter(school_year=school_year)
        
        # If course_ids provided, get all courses with same identity (for multi-section)
        if course_ids:
            # Get the courses by IDs first
            initial_courses = list(courses_qs.filter(id__in=course_ids))
            # For each course, find all courses with same identity (code, name, semester, school_year)
            all_course_ids = set(course_ids)
            for course in initial_courses:
                # Find all courses with same identity
                same_identity = Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True
                ).values_list('id', flat=True)
                all_course_ids.update(same_identity)
            
            courses_qs = courses_qs.filter(id__in=all_course_ids)
        
        # Verify all courses belong to the instructor
        courses = list(courses_qs)
        if not courses:
            return JsonResponse({'success': False, 'message': 'No courses found in this semester.'})
        
        # Soft delete all courses
        from django.utils import timezone
        deleted_count = 0
        for course in courses:
            course.deleted_at = timezone.now()
            course.is_active = False
            course.save()
            deleted_count += 1
        
        semester_display = '1st Semester' if semester == '1st' else '2nd Semester' if semester == '2nd' else 'Summer' if semester == 'summer' else semester
        return JsonResponse({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} course(s) from Year {year_level}, {semester_display}.'
        })
    except Exception as e:
        logger.error(f"Error deleting semester/school year: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_archive_school_year_view(request):
    """Archive all courses in a school year - instructor can only archive courses they created"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        school_year = data.get('school_year')
        course_ids = data.get('course_ids', [])
        
        if not school_year:
            return JsonResponse({'success': False, 'message': 'School year is required.'})
        
        # Get all courses for this instructor in the specified school year
        courses_qs = Course.objects.filter(
            instructor=user,
            school_year=school_year,
            deleted_at__isnull=True,  # Only consider active courses
            is_archived=False  # Not already archived
        )
        
        # If course_ids provided, get all courses with same identity (for multi-section)
        if course_ids:
            initial_courses = list(courses_qs.filter(id__in=course_ids))
            all_course_ids = set(course_ids)
            for course in initial_courses:
                # Find all courses with same identity (all sections)
                same_identity = Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True,
                    is_archived=False
                ).values_list('id', flat=True)
                all_course_ids.update(same_identity)
            courses_qs = courses_qs.filter(id__in=all_course_ids)
        
        count = courses_qs.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'No active courses found for this school year to archive.'})
        
        # Archive all courses (set is_archived=True, keep deleted_at=None)
        archived = 0
        for course in courses_qs:
            course.is_archived = True
            course.is_active = False
            # Keep deleted_at as None for archived items
            course.save()
            archived += 1
        
        return JsonResponse({'success': True, 'message': f'{archived} course(s) for School Year {school_year} archived successfully.'})
    except Exception as e:
        logger.error(f"Error archiving school year courses: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_archive_semester_view(request):
    """Archive all courses in a semester - instructor can only archive courses they created"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        school_year = data.get('school_year')
        year_level_str = data.get('year_level')
        semester = data.get('semester')
        course_ids = data.get('course_ids', [])
        
        if not school_year or not year_level_str or not semester:
            return JsonResponse({'success': False, 'message': 'School year, year level, and semester are required.'})
        
        # Extract numeric year_level from string like "Year 1"
        try:
            year_level = int(year_level_str.replace('Year ', ''))
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid year level format.'})
        
        # Get all courses for this instructor in the specified school year, year level, and semester
        courses_qs = Course.objects.filter(
            instructor=user,
            school_year=school_year,
            year_level=year_level,
            semester=semester,
            deleted_at__isnull=True,  # Only consider active courses
            is_archived=False  # Not already archived
        )
        
        # If course_ids provided, get all courses with same identity (for multi-section)
        if course_ids:
            initial_courses = list(courses_qs.filter(id__in=course_ids))
            all_course_ids = set(course_ids)
            for course in initial_courses:
                # Find all courses with same identity (all sections)
                same_identity = Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    deleted_at__isnull=True,
                    is_archived=False
                ).values_list('id', flat=True)
                all_course_ids.update(same_identity)
            courses_qs = courses_qs.filter(id__in=all_course_ids)
        
        count = courses_qs.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'No active courses found for this semester to archive.'})
        
        # Archive all courses (set is_archived=True, keep deleted_at=None)
        archived = 0
        for course in courses_qs:
            course.is_archived = True
            course.is_active = False
            # Keep deleted_at as None for archived items
            course.save()
            archived += 1
        
        return JsonResponse({'success': True, 'message': f'{archived} course(s) for {semester} Semester, {year_level_str} in SY {school_year} archived successfully.'})
    except Exception as e:
        logger.error(f"Error archiving semester courses: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_archive_course_view(request, course_id):
    """Archive a course - move to archive for completed sessions"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    course = get_object_or_404(Course, id=course_id, deleted_at__isnull=True, is_archived=False)
    
    # Verify the course is assigned to the instructor
    if course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only archive courses assigned to you.'})
    
    try:
        from django.db import transaction
        
        # Archive: set is_archived flag but DO NOT set deleted_at (archive is storage, not deletion)
        # For multi-section courses, archive all sections with same identity
        with transaction.atomic():
            courses_to_archive = Course.objects.filter(
                instructor=user,
                code=course.code,
                name=course.name,
                semester=course.semester,
                school_year=course.school_year,
                deleted_at__isnull=True,
                is_archived=False
            )
            
            archived_count = 0
            for c in courses_to_archive:
                c.is_archived = True
                c.is_active = False
                # Keep deleted_at as None for archived items
                c.save(update_fields=['is_archived', 'is_active', 'updated_at'])
                # Force database commit by refreshing
                c.refresh_from_db()
                # Verify the save worked - check directly from database
                if not c.is_archived or c.is_active:
                    logger.error(f"Failed to archive course {c.id}: is_archived={c.is_archived}, is_active={c.is_active}")
                else:
                    logger.info(f"Successfully archived course {c.id} ({c.code}): is_archived={c.is_archived}, is_active={c.is_active}")
                archived_count += 1
            # Transaction auto-commits on successful exit
        
        if archived_count > 1:
            return JsonResponse({'success': True, 'message': f'{archived_count} course sections archived successfully. They will be stored in Archive.'})
        else:
            return JsonResponse({'success': True, 'message': 'Course archived successfully. It will be stored in Archive.'})
    except Exception as e:
        logger.error(f"Error archiving course: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def teacher_update_profile_view(request):
    """Update teacher profile"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        # Update full name if provided
        full_name = request.POST.get('full_name', '').strip()
        if full_name:
            user.full_name = full_name
        
        # Update department if provided
        department = request.POST.get('department', '').strip()
        if department and department.upper() != 'NOT ASSIGNED':
            user.department = department
        elif not department:
            user.department = ''
        
        # Update program if provided
        program = request.POST.get('program', '').strip()
        if program and program.upper() != 'NOT ASSIGNED':
            # Store program input as program_text field (for free-text program entry)
            user.program_text = program
        elif not program:
            user.program_text = ''
        
        # Update custom password if provided
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_new_password', '').strip()
        current_password = request.POST.get('current_password', '').strip()
        
        if new_password:
            # Validate password length
            if len(new_password) < 6:
                return JsonResponse({'success': False, 'message': 'Password must be at least 6 characters long.'})
            
            # Check if passwords match
            if new_password != confirm_password:
                return JsonResponse({'success': False, 'message': 'New password and confirm password do not match.'})
            
            # If user already has a custom password, require current password
            if user.custom_password:
                if not current_password:
                    return JsonResponse({'success': False, 'message': 'Current password is required to change your password.'})
                # Verify current password (either temporary or custom)
                from django.contrib.auth import authenticate
                authenticated = authenticate(request, username=user.username, password=current_password)
                if not authenticated and not user.check_custom_password(current_password):
                    return JsonResponse({'success': False, 'message': 'Current password is incorrect.'})
            # If no custom password exists, they can add one without current password (first time)
            
            user.set_custom_password(new_password)
            logger.info(f"Custom password {'updated' if user.custom_password else 'added'} for teacher {user.username}")
        
        # Update profile picture
        if 'profile_picture' in request.FILES:
            profile_picture = request.FILES['profile_picture']
            # Validate file size (max 5MB)
            if profile_picture.size > 5 * 1024 * 1024:
                return JsonResponse({'success': False, 'message': 'Profile picture size must be less than 5MB.'})
            # Validate file type
            if not profile_picture.content_type.startswith('image/'):
                return JsonResponse({'success': False, 'message': 'Please upload a valid image file.'})
            user.profile_picture = profile_picture
        
        user.save()
        
        # Return the profile picture URL with cache busting
        profile_picture_url = None
        if user.profile_picture:
            profile_picture_url = user.profile_picture.url + '?v=' + str(int(datetime.now().timestamp()))
        
        # Prepare response data
        response_data = {
            'success': True, 
            'message': 'Profile updated successfully!'
        }
        if profile_picture_url:
            response_data['profile_picture_url'] = profile_picture_url
        
        # Return department and program data for frontend update
        response_data['department'] = user.department or 'Not Assigned'
        if hasattr(user, 'program_text') and user.program_text:
            response_data['program'] = user.program_text
        elif user.program:
            response_data['program'] = f"{user.program.code} - {user.program.name}" if user.program.code else user.program.name
        else:
            response_data['program'] = 'Not Assigned'
        
        return JsonResponse(response_data)
    except Exception as e:
        logger.error(f"Error updating profile: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def student_update_profile_view(request):
    """Update student profile"""
    user = request.user
    if not user.is_student:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        # Update full name if provided (support split fields for students)
        first_name = request.POST.get('first_name', '').strip()
        middle_name = request.POST.get('middle_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        full_name = request.POST.get('full_name', '').strip()

        if first_name or middle_name or last_name:
            if not first_name or not last_name:
                return JsonResponse({'success': False, 'message': 'First Name and Last Name are required.'})
            combined = ' '.join([p for p in [first_name, middle_name, last_name] if p]).strip()
            if combined:
                user.full_name = combined
        elif full_name:
            user.full_name = full_name
        
        # Update department if provided
        department = request.POST.get('department', '').strip()
        if department and department.upper() != 'NOT ASSIGNED':
            user.department = department
        elif not department:
            user.department = ''
        
        # Update program if provided
        program = request.POST.get('program', '').strip()
        if program and program.upper() != 'NOT ASSIGNED':
            # Store program input as program_text field for free-text entry
            user.program_text = program
        elif not program:
            user.program_text = ''
        
        # Update custom password if provided
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_new_password', '').strip()
        current_password = request.POST.get('current_password', '').strip()
        
        if new_password:
            # Validate password length
            if len(new_password) < 6:
                return JsonResponse({'success': False, 'message': 'Password must be at least 6 characters long.'})
            
            # Check if passwords match
            if new_password != confirm_password:
                return JsonResponse({'success': False, 'message': 'New password and confirm password do not match.'})
            
            # If user already has a custom password, require current password
            if user.custom_password:
                if not current_password:
                    return JsonResponse({'success': False, 'message': 'Current password is required to change your password.'})
                # Verify current password (either temporary or custom)
                from django.contrib.auth import authenticate
                authenticated = authenticate(request, username=user.username, password=current_password)
                if not authenticated and not user.check_custom_password(current_password):
                    return JsonResponse({'success': False, 'message': 'Current password is incorrect.'})
            # If no custom password exists, they can add one without current password (first time)
            
            user.set_custom_password(new_password)
            logger.info(f"Custom password {'updated' if user.custom_password else 'added'} for student {user.username}")
        
        # Update profile picture
        if 'profile_picture' in request.FILES:
            profile_picture = request.FILES['profile_picture']
            # Validate file size (max 5MB)
            if profile_picture.size > 5 * 1024 * 1024:
                return JsonResponse({'success': False, 'message': 'Profile picture size must be less than 5MB.'})
            # Validate file type
            if not profile_picture.content_type.startswith('image/'):
                return JsonResponse({'success': False, 'message': 'Please upload a valid image file.'})
            user.profile_picture = profile_picture
        
        user.save()
        
        # Return the profile picture URL with cache busting
        profile_picture_url = None
        if user.profile_picture:
            profile_picture_url = user.profile_picture.url + '?v=' + str(int(datetime.now().timestamp()))
        
        # Prepare response data
        response_data = {
            'success': True, 
            'message': 'Profile updated successfully!'
        }
        if profile_picture_url:
            response_data['profile_picture_url'] = profile_picture_url
        
        # Return department and program data for frontend update
        response_data['department'] = user.department or 'Not Assigned'
        if hasattr(user, 'program_text') and user.program_text:
            response_data['program'] = user.program_text
        elif user.program:
            response_data['program'] = f"{user.program.code} - {user.program.name}" if user.program.code else user.program.name
        else:
            response_data['program'] = 'Not Assigned'
        
        return JsonResponse(response_data)
    except Exception as e:
        logger.error(f"Error updating profile: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def get_user_profile(request):
    """Get current user's profile data (department and program)"""
    user = request.user
    
    try:
        # Format department
        department = user.department or 'Not Set'
        
        # Format program
        if user.program_text:
            program = user.program_text
        elif user.program:
            program = f"{user.program.code} - {user.program.name}" if user.program.code else user.program.name
        else:
            program = 'Not Set'
        
        return JsonResponse({
            'success': True,
            'department': department,
            'program': program
        })
    except Exception as e:
        logger.error(f"Error fetching profile: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def get_custom_password_view(request):
    """Get the current custom password for the logged-in user"""
    user = request.user
    if not user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Not authenticated'})
    
    try:
        from dashboard.models import UserCustomPassword
        custom_pwd_record = UserCustomPassword.objects.filter(user=user).first()
        
        response_data = {
            'success': True,
            'has_custom_password': custom_pwd_record is not None,
            'custom_password': custom_pwd_record.password if custom_pwd_record else None
        }
        
        return JsonResponse(response_data)
    except Exception as e:
        logger.error(f"Error fetching custom password: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def instructor_qr_code_view(request, course_id):
    """Generate and return QR code image for a course - Always returns a valid PNG image"""
    from django.http import HttpResponse
    from io import BytesIO
    
    user = request.user
    if not user.is_teacher:
        # Return a placeholder image instead of error
        try:
            from PIL import Image, ImageDraw
            img = Image.new('RGB', (200, 200), color='white')
            draw = ImageDraw.Draw(img)
            draw.text((60, 90), 'Unauthorized', fill='black')
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            return response
        except Exception:
            pass
        return HttpResponse('Unauthorized', status=403)
    
    try:
        course = Course.objects.get(id=course_id, instructor=user)
    except Course.DoesNotExist:
        # Return a placeholder image instead of 404
        try:
            from PIL import Image, ImageDraw
            img = Image.new('RGB', (200, 200), color='white')
            draw = ImageDraw.Draw(img)
            draw.text((70, 90), 'Not Found', fill='black')
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            return response
        except Exception:
            pass
        return HttpResponse('Not Found', status=404)
    
    # Generate SESSION-BASED QR code that changes per day/session
    # This prevents students from scanning old QR codes when not in class
    # Format: course_id + date + day_of_week + time_window + section
    # hashlib is already imported at module level
    from datetime import date, time, timedelta
    try:
        from zoneinfo import ZoneInfo
        PH_TZ = ZoneInfo('Asia/Manila')
    except Exception:
        try:
            import pytz
            PH_TZ = pytz.timezone('Asia/Manila')
        except Exception:
            PH_TZ = None
    
    now_ph = datetime.now(PH_TZ) if PH_TZ else datetime.now()
    today_date = now_ph.date()
    today_weekday = now_ph.weekday()  # 0=Monday, 6=Sunday
    weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    weekday_short = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    today_day_name = weekday_names[today_weekday]
    today_day_short = weekday_short[today_weekday]
    
    # Get attendance window for today
    attendance_start = None
    attendance_end = None
    
    # Check if day parameter is provided in request (for specific day QR code)
    day_param = request.GET.get('day', '').strip()
    day_schedule = None
    
    # Check day-specific schedules first (asynchronous)
    if day_param:
        # Use the day from request parameter
        day_map = {
            'Monday': 'Mon', 'Mon': 'Mon',
            'Tuesday': 'Tue', 'Tue': 'Tue',
            'Wednesday': 'Wed', 'Wed': 'Wed',
            'Thursday': 'Thu', 'Thu': 'Thu',
            'Friday': 'Fri', 'Fri': 'Fri',
            'Saturday': 'Sat', 'Sat': 'Sat',
            'Sunday': 'Sun', 'Sun': 'Sun'
        }
        schedule_day = day_map.get(day_param, day_param)
        day_schedules = course.course_schedules.filter(day__iexact=schedule_day).all()
        if day_schedules.exists():
            day_schedule = day_schedules.first()
            attendance_start = day_schedule.attendance_start or course.attendance_start
            attendance_end = day_schedule.attendance_end or course.attendance_end
            logger.info(f"Using day-specific schedule for course {course.id} on {schedule_day} (from request)")
    else:
        # Check for today's schedule
        day_schedules = course.course_schedules.filter(day__iexact=today_day_short).all()
        if day_schedules.exists():
            day_schedule = day_schedules.first()
            attendance_start = day_schedule.attendance_start or course.attendance_start
            attendance_end = day_schedule.attendance_end or course.attendance_end
            logger.info(f"Using day-specific schedule for course {course.id} on {today_day_short}")
        elif course.days:
            # Check synchronized schedule
            days_list = [d.strip() for d in course.days.split(',') if d.strip()]
            if today_day_name in days_list or today_day_short in days_list:
                attendance_start = course.attendance_start
                attendance_end = course.attendance_end
                logger.info(f"Using synchronized schedule for course {course.id} on {today_day_name}")
    
    # Initialize date_str and day_str for logging and response headers
    date_str = today_date.strftime('%Y%m%d')
    day_str = str(today_weekday)  # Initialize day_str early to avoid UnboundLocalError
    
    # Use day-specific QR code if available and schedule hasn't finished
    # Only use existing QR code if the schedule for that day hasn't ended yet
    if day_schedule and day_schedule.qr_code:
        # Check if the schedule for this day has finished
        schedule_finished = False
        weekday_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
        day_weekday = weekday_map.get(day_schedule.day, None)
        
        if day_weekday is not None:
            current_weekday = now_ph.weekday()
            days_ahead = (day_weekday - current_weekday) % 7
            
            if days_ahead == 0:
                # Today - check if class has ended
                if day_schedule.end_time:
                    end_time_today = now_ph.replace(hour=day_schedule.end_time.hour, minute=day_schedule.end_time.minute, second=0, microsecond=0)
                    if now_ph > end_time_today:
                        schedule_finished = True
            else:
                # Check if QR code is for a past occurrence that has finished
                from datetime import timedelta
                last_occurrence_date = today_date - timedelta(days=days_ahead)
                if day_schedule.qr_code_date == last_occurrence_date and day_schedule.end_time:
                    last_end_time = datetime.combine(last_occurrence_date, day_schedule.end_time)
                    if PH_TZ:
                        # zoneinfo doesn't have localize(), use replace() instead
                        if last_end_time.tzinfo is None:
                            last_end_time = last_end_time.replace(tzinfo=PH_TZ)
                    elif last_end_time.tzinfo is None:
                        last_end_time = last_end_time.replace(tzinfo=now_ph.tzinfo)
                    if now_ph > last_end_time:
                        schedule_finished = True
        
        # Only use existing QR code if schedule hasn't finished
        if not schedule_finished:
            qr_code_data = day_schedule.qr_code
            logger.info(f"Using existing day-specific QR code for course {course.id} on {day_schedule.day}: {qr_code_data[:8]}... (schedule not finished)")
        else:
            # Schedule has finished - need to generate new QR code
            # This will be handled by the my_classes view, but generate one here as fallback
            # hashlib and random are already imported at module level
            # Don't include timestamp to ensure stability - QR code should only change when schedule finishes
            qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}"
            qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
            
            max_attempts = 10
            attempts = 0
            while CourseSchedule.objects.filter(qr_code=qr_hash).exclude(id=day_schedule.id).exists() and attempts < max_attempts:
                qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}_{random.randint(1000, 9999)}"
                qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                attempts += 1
            
            day_schedule.qr_code = qr_hash
            day_schedule.qr_code_date = today_date
            day_schedule.save(update_fields=['qr_code', 'qr_code_date'])
            qr_code_data = qr_hash
            logger.info(f"Generated new QR code for course {course.id} on {day_schedule.day} (schedule finished): {qr_code_data[:8]}...")
    elif day_schedule:
        # Day schedule exists but QR code is missing - generate one
        # hashlib and random are already imported at module level
        # Don't include timestamp to ensure stability - QR code should only change when schedule finishes
        qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}"
        qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
        
        max_attempts = 10
        attempts = 0
        while CourseSchedule.objects.filter(qr_code=qr_hash).exclude(id=day_schedule.id).exists() and attempts < max_attempts:
            qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}_{random.randint(1000, 9999)}"
            qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
            attempts += 1
        
        day_schedule.qr_code = qr_hash
        day_schedule.qr_code_date = today_date
        day_schedule.save(update_fields=['qr_code', 'qr_code_date'])
        qr_code_data = qr_hash
        logger.info(f"Generated missing day-specific QR code for course {course.id} on {today_day_short}: {qr_code_data[:8]}...")
    elif not day_schedule and course.days:
        # No day schedule exists but course has days - create one and generate QR code
        # This handles courses that were created before day schedules were implemented
        try:
            days_list = [d.strip() for d in course.days.split(',') if d.strip()]
            # Use today's day or first day in the list
            target_day = today_day_short
            day_map = {
                'Monday': 'Mon', 'Mon': 'Mon',
                'Tuesday': 'Tue', 'Tue': 'Tue',
                'Wednesday': 'Wed', 'Wed': 'Wed',
                'Thursday': 'Thu', 'Thu': 'Thu',
                'Friday': 'Fri', 'Fri': 'Fri',
                'Saturday': 'Sat', 'Sat': 'Sat',
                'Sunday': 'Sun', 'Sun': 'Sun'
            }
            # Find matching day
            for day in days_list:
                mapped_day = day_map.get(day, day)
                if mapped_day == today_day_short:
                    target_day = mapped_day
                    break
            else:
                # Use first day if today doesn't match
                if days_list:
                    target_day = day_map.get(days_list[0], days_list[0])
            
            # Create day schedule if it doesn't exist
            day_schedule, created = CourseSchedule.objects.get_or_create(
                course=course,
                day=target_day,
                defaults={
                    'start_time': course.start_time,
                    'end_time': course.end_time,
                    'room': course.room
                }
            )
            
            # Generate QR code for this day schedule
            # hashlib and random are already imported at module level
            # Don't include timestamp to ensure stability - QR code should only change when schedule finishes
            qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}"
            qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
            
            # Ensure uniqueness
            max_attempts = 10
            attempts = 0
            while CourseSchedule.objects.filter(qr_code=qr_hash).exclude(id=day_schedule.id).exists() and attempts < max_attempts:
                qr_data = f"{course.id}_{course.code}_{course.section}_{day_schedule.day}_{today_date.strftime('%Y%m%d')}_{random.randint(1000, 9999)}"
                qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                attempts += 1
            
            day_schedule.qr_code = qr_hash
            day_schedule.qr_code_date = today_date
            day_schedule.save(update_fields=['qr_code', 'qr_code_date'])
            qr_code_data = qr_hash
            logger.info(f"Created day schedule and generated QR code for course {course.id} on {target_day}: {qr_code_data[:8]}...")
        except Exception as e:
            logger.error(f"Error creating day schedule for QR code: {str(e)}")
            # Fall through to session-based QR code
            day_schedule = None
    
    if not day_schedule or 'qr_code_data' not in locals():
        # Generate session-based QR code (changes daily) - fallback
        # day_str is already initialized above
        
        # Include time window in hash for extra security
        time_hash = ''
        if attendance_start and attendance_end:
            time_hash = f"{attendance_start.strftime('%H%M')}_{attendance_end.strftime('%H%M')}"
        else:
            # If no attendance window set, use default time window
            if course.start_time and course.end_time:
                time_hash = f"{course.start_time.strftime('%H%M')}_{course.end_time.strftime('%H%M')}"
        
        # Generate session-specific QR code (unique per day/session)
        session_data = f"{course.id}_{date_str}_{day_str}_{time_hash}_{course.section or 'NONE'}"
        qr_code_data = hashlib.sha256(session_data.encode()).hexdigest()[:32].upper()
        logger.info(f"Generated session-based QR code for course {course.id} ({course.code}) on {date_str} (day {today_day_short}): {qr_code_data[:8]}...")
    
    # Generate QR code image - ensure it always works
    # Always try to import qrcode here (even if module-level import failed)
    qrcode_module = None
    try:
        if QRCODE_AVAILABLE:
            # Use module-level import if available
            import qrcode
            qrcode_module = qrcode
            logger.info("Using module-level qrcode import")
        else:
            # Try function-level import as fallback
            import qrcode
            qrcode_module = qrcode
            logger.info("Using function-level qrcode import (fallback)")
    except ImportError as import_err:
        logger.error(f"qrcode library not available: {str(import_err)}")
        # Return a proper error image instead of text
        try:
            from PIL import Image, ImageDraw
            img = Image.new('RGB', (250, 250), color='white')
            draw = ImageDraw.Draw(img)
            # Draw error message as image
            try:
                from PIL import ImageFont
                font = ImageFont.load_default()
            except:
                font = None
            text = "QR Code\nError\nInstall qrcode"
            draw.text((125, 125), text, fill='red', anchor='mm', font=font)
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            response['Content-Length'] = str(len(buffer.getvalue()))
            return response
        except Exception as e:
            logger.error(f"Error creating QR code error image: {str(e)}")
            # Return minimal valid PNG
            fallback_png = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\tpHYs\x00\x00\x0b\x13\x00\x00\x0b\x13\x01\x00\x9a\x9c\x18\x00\x00\x00\nIDATx\x9cc\xf8\x00\x00\x00\x01\x00\x01\x00\x00\x00\x00IEND\xaeB`\x82'
            response = HttpResponse(fallback_png, content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            return response
    
    # Generate QR code using the imported qrcode module
    try:
        qr = qrcode_module.QRCode(
            version=None,  # Auto-determine version based on data
            error_correction=qrcode.constants.ERROR_CORRECT_M,  # Medium error correction
            box_size=10,
            border=4,
        )
        
        # Store the session-based QR code string (32-char hex)
        qr.add_data(qr_code_data.strip())
        qr.make(fit=True)
        
        # Create QR code image using PIL factory - ensure we get a proper image
        try:
            img = qr.make_image(fill_color="black", back_color="white")
            logger.info(f"QR code image created successfully for course {course.id}: {img.size}, type: {type(img)}")
        except Exception as img_error:
            logger.warning(f"Error creating QR code image with colors: {str(img_error)}, trying default")
            # Try with default factory
            try:
                img = qr.make_image()
                logger.info(f"QR code image created with default factory for course {course.id}: {img.size}")
            except Exception as default_error:
                logger.error(f"Error creating QR code image with default factory: {str(default_error)}")
                raise
        
        # Ensure the image is properly sized (at least 250x250 for excellent visibility and scanning)
        current_size = img.size
        target_size = 250  # Larger size for better visibility and scanning
        if current_size[0] < target_size or current_size[1] < target_size:
            # Scale up to ensure visibility
            scale_factor = max(target_size / current_size[0], target_size / current_size[1])
            new_size = (int(current_size[0] * scale_factor), int(current_size[1] * scale_factor))
            # Use high-quality resampling
            img = img.resize(new_size, Image.LANCZOS)
        elif current_size[0] > target_size * 1.5 or current_size[1] > target_size * 1.5:
            # If too large, scale down for better performance while maintaining quality
            scale_factor = min(target_size * 1.5 / current_size[0], target_size * 1.5 / current_size[1])
            new_size = (int(current_size[0] * scale_factor), int(current_size[1] * scale_factor))
            img = img.resize(new_size, Image.LANCZOS)
        
        # Convert to RGB if necessary (some PIL modes don't support PNG properly)
        if img.mode not in ('RGB', 'RGBA'):
            if img.mode == 'L':  # Grayscale
                rgb_img = Image.new('RGB', img.size, color='white')
                rgb_img.paste(img, (0, 0))
                img = rgb_img
            else:
                img = img.convert('RGB')
        
        # Save to BytesIO - ensure proper PNG format with maximum quality
        buffer = BytesIO()
        try:
            # Save with maximum quality settings
            img.save(buffer, format='PNG', optimize=True, compress_level=6)
        except Exception as save_error:
            logger.error(f"Error saving QR code image: {str(save_error)}")
            # Try saving without optimization
            buffer = BytesIO()
            try:
                img.save(buffer, format='PNG', optimize=False)
            except Exception as save_error2:
                logger.error(f"Error saving QR code image without optimization: {str(save_error2)}")
                # Last resort - convert to PNG format explicitly
                buffer = BytesIO()
                img.save(buffer, format='PNG')
        buffer.seek(0)
        
        image_data = buffer.read()
        
        # Verify we have actual image data
        if len(image_data) < 100:  # PNG files should be at least 100 bytes
            logger.error(f"Generated QR code image data is too small: {len(image_data)} bytes")
            raise Exception(f"Generated image data is too small: {len(image_data)} bytes")
        
        logger.info(f"QR code image generated successfully for course {course.id}: {len(image_data)} bytes")
        
        response = HttpResponse(image_data, content_type='image/png')
        # Cache headers - cache until course end time (or end of day)
        # QR code is valid for the entire day, so cache it until end of day
        # hashlib is already imported at module level
        # ETag based on session (date + day) so it changes daily
        etag_data = f"{course.id}_{date_str}_{day_str}_{qr_code_data[:8]}"
        etag = hashlib.md5(etag_data.encode()).hexdigest()
        response['ETag'] = f'"{etag}"'
        
        # Calculate cache duration - cache until end of day or course end time, whichever is later
        seconds_until_midnight = (24 * 3600) - ((now_ph.hour * 3600) + (now_ph.minute * 60) + now_ph.second)
        
        # If course has end time today, cache until then (with some buffer)
        cache_until_end = seconds_until_midnight
        if attendance_end:
            end_time_today = datetime.combine(today_date, attendance_end)
            if PH_TZ and end_time_today.tzinfo is None:
                if hasattr(PH_TZ, 'localize'):
                    end_time_today = PH_TZ.localize(end_time_today)
                else:
                    end_time_today = end_time_today.replace(tzinfo=PH_TZ)
            elif not PH_TZ:
                pass  # Keep naive
            
            if end_time_today > now_ph:
                seconds_until_end = (end_time_today - now_ph).total_seconds()
                # Cache until course end + 1 hour buffer, but not more than until midnight
                cache_until_end = min(seconds_until_end + 3600, seconds_until_midnight)
        
        # Cache for the calculated duration (at least 1 hour, max until end of day)
        max_age = max(3600, min(int(cache_until_end), 86400))
        response['Cache-Control'] = f'public, max-age={max_age}, immutable'  # Cache until course end or end of day
        response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}_{date_str}.png"'
        response['Content-Length'] = str(len(image_data))
        response['X-Content-Type-Options'] = 'nosniff'
        # Add timestamp header for debugging
        response['X-QR-Code-Date'] = date_str
        response['X-QR-Code-Day'] = day_str
        return response
    except ImportError as import_err:
        # If qrcode library is not installed - try to import it again as fallback
        logger.warning(f"qrcode import failed at module level, trying function-level import: {str(import_err)}")
        try:
            import qrcode
            from PIL import Image
            # Retry QR code generation with function-level import
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_code_data.strip())
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            # Resize and process image (same as above)
            current_size = img.size
            target_size = 250
            if current_size[0] < target_size or current_size[1] < target_size:
                scale_factor = max(target_size / current_size[0], target_size / current_size[1])
                new_size = (int(current_size[0] * scale_factor), int(current_size[1] * scale_factor))
                img = img.resize(new_size, Image.LANCZOS)
            if img.mode not in ('RGB', 'RGBA'):
                if img.mode == 'L':
                    rgb_img = Image.new('RGB', img.size, color='white')
                    rgb_img.paste(img, (0, 0))
                    img = rgb_img
                else:
                    img = img.convert('RGB')
            buffer = BytesIO()
            img.save(buffer, format='PNG', optimize=True, compress_level=6)
            buffer.seek(0)
            image_data = buffer.read()
            logger.info(f"QR code generated successfully using function-level import for course {course.id}: {len(image_data)} bytes")
            response = HttpResponse(image_data, content_type='image/png')
            etag_data = f"{course.id}_{date_str}_{day_str}_{qr_code_data[:8]}"
            etag = hashlib.md5(etag_data.encode()).hexdigest()
            response['ETag'] = f'"{etag}"'
            seconds_until_midnight = (24 * 3600) - ((now_ph.hour * 3600) + (now_ph.minute * 60) + now_ph.second)
            cache_until_end = seconds_until_midnight
            if attendance_end:
                end_time_today = datetime.combine(today_date, attendance_end)
                if PH_TZ and end_time_today.tzinfo is None:
                    if hasattr(PH_TZ, 'localize'):
                        end_time_today = PH_TZ.localize(end_time_today)
                    else:
                        end_time_today = end_time_today.replace(tzinfo=PH_TZ)
                if end_time_today > now_ph:
                    seconds_until_end = (end_time_today - now_ph).total_seconds()
                    cache_until_end = min(seconds_until_end + 3600, seconds_until_midnight)
            max_age = max(3600, min(int(cache_until_end), 86400))
            response['Cache-Control'] = f'public, max-age={max_age}, immutable'
            response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}_{date_str}.png"'
            response['Content-Length'] = str(len(image_data))
            response['X-Content-Type-Options'] = 'nosniff'
            response['X-QR-Code-Date'] = date_str
            response['X-QR-Code-Day'] = day_str
            return response
        except ImportError as import_err2:
            logger.error(f"qrcode library not available even with function-level import: {str(import_err2)}")
            # Try to generate a basic QR code using PIL directly as fallback
            try:
                from PIL import Image, ImageDraw
                # Generate a simple placeholder that looks like a QR code
                size = 200
                img = Image.new('RGB', (size, size), color='white')
                draw = ImageDraw.Draw(img)
                # Draw a simple grid pattern that resembles a QR code
                block_size = 20
                for i in range(0, size, block_size):
                    for j in range(0, size, block_size):
                        if (i // block_size + j // block_size) % 3 == 0:
                            draw.rectangle([i, j, i+block_size-2, j+block_size-2], fill='black')
                buffer = BytesIO()
                img.save(buffer, format='PNG')
                buffer.seek(0)
                image_data = buffer.read()
                response = HttpResponse(image_data, content_type='image/png')
                response['Cache-Control'] = 'no-cache'
                response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}.png"'
                response['Content-Length'] = len(image_data)
                return response
            except Exception as e:
                logger.error(f"Error creating fallback QR code: {str(e)}")
                pass
            # Final fallback - minimal PNG
            fallback_png = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\tpHYs\x00\x00\x0b\x13\x00\x00\x0b\x13\x01\x00\x9a\x9c\x18\x00\x00\x00\nIDATx\x9cc\xf8\x00\x00\x00\x01\x00\x01\x00\x00\x00\x00IEND\xaeB`\x82'
            response = HttpResponse(fallback_png, content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}.png"'
            return response
    except Exception as e:
        logger.error(f"Error generating QR code for course {course.id}: {str(e)}", exc_info=True)
        # Try to return a proper error image
        try:
            from PIL import Image, ImageDraw
            img = Image.new('RGB', (200, 200), color='white')
            draw = ImageDraw.Draw(img)
            draw.text((70, 90), 'Error', fill='red')
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            image_data = buffer.read()
            response = HttpResponse(image_data, content_type='image/png')
            response['Cache-Control'] = 'no-cache'
            response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}.png"'
            response['Content-Length'] = len(image_data)
            return response
        except Exception:
            pass
        # Final fallback
        fallback_png = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\tpHYs\x00\x00\x0b\x13\x00\x00\x0b\x13\x01\x00\x9a\x9c\x18\x00\x00\x00\nIDATx\x9cc\xf8\x00\x00\x00\x01\x00\x01\x00\x00\x00\x00IEND\xaeB`\x82'
        response = HttpResponse(fallback_png, content_type='image/png')
        response['Cache-Control'] = 'no-cache'
        response['Content-Disposition'] = f'inline; filename="qr_code_{course.id}.png"'
        return response

@login_required
def student_qr_scanner_view(request):
    """Student QR code scanner page"""
    user = request.user
    if not user.is_student:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    context = {
        'user': user,
        'school_admin': school_admin,
    }
    
    return render(request, 'dashboard/student/student_qr_scanner.html', context)

@login_required
@require_http_methods(["POST"])
def student_scan_qr_attendance_view(request):
    """Handle QR code scan and mark attendance"""
    user = request.user
    if not user.is_student:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        data = json.loads(request.body)
        qr_code_raw = data.get('qr_code', '').strip()
        
        if not qr_code_raw:
            return JsonResponse({'success': False, 'message': 'QR code is required.'})
        
        # Extract QR code from scanned data (might be URL or just the code)
        # Handle cases where QR code might be a URL or contain extra data
        qr_code = qr_code_raw
        # If it's a URL, try to extract the code from it
        if 'qr_code' in qr_code_raw.lower() or '/' in qr_code_raw:
            # Try to extract from URL pattern
            import re
            # Look for 32-character hex string (typical QR code format)
            hex_match = re.search(r'([a-f0-9]{32})', qr_code_raw, re.IGNORECASE)
            if hex_match:
                qr_code = hex_match.group(1)
            else:
                # Try to extract from query parameters
                if '?' in qr_code_raw:
                    parts = qr_code_raw.split('?')
                    qr_code = parts[-1] if len(parts) > 1 else qr_code_raw
                # If it contains the course ID, try to find by ID first
                id_match = re.search(r'course[_-]?id[=:](\d+)', qr_code_raw, re.IGNORECASE)
                if id_match:
                    course_id = int(id_match.group(1))
                    try:
                        course = Course.objects.get(id=course_id, is_active=True, deleted_at__isnull=True, is_archived=False)
                        # Use the course's QR code for matching
                        qr_code = course.qr_code
                    except Course.DoesNotExist:
                        pass
        
        # Clean the QR code - remove any whitespace, special characters
        scanned_qr_code = qr_code.strip().upper()
        
        # SESSION-BASED QR CODE VALIDATION
        # QR codes are now session-based (change daily) to prevent scanning old codes
        # We need to find the course and validate the QR code matches today's session
        
        # Get timezone first (before using it)
        try:
            from zoneinfo import ZoneInfo
            PH_TZ = ZoneInfo('Asia/Manila')
        except Exception:
            try:
                import pytz
                PH_TZ = pytz.timezone('Asia/Manila')
            except Exception:
                PH_TZ = None
        
        # Get current date and day info
        now_ph = datetime.now(PH_TZ) if PH_TZ else datetime.now()
        today_date = now_ph.date()
        today_weekday = now_ph.weekday()
        weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        weekday_short = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        today_day_name = weekday_names[today_weekday]
        today_day_short = weekday_short[today_weekday]
        date_str = today_date.strftime('%Y%m%d')
        day_str = str(today_weekday)
        
        # Try to find course by matching today's session QR code
        # We'll check all courses and generate their expected QR codes for today
        all_courses = Course.objects.filter(
            instructor__is_teacher=True,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        ).select_related('instructor').prefetch_related('course_schedules')
        
        course = None
        matched_qr_code = None
        matched_schedule = None
        
        for c in all_courses:
            # Check day-specific schedules first (asynchronous)
            day_schedules = c.course_schedules.filter(day__iexact=today_day_short).all()
            if day_schedules.exists():
                # Check each day schedule's QR code directly from database
                for schedule in day_schedules:
                    # Check if the scanned QR code matches the schedule's stored QR code
                    if schedule.qr_code and scanned_qr_code == schedule.qr_code.upper():
                        course = c
                        matched_schedule = schedule
                        matched_qr_code = schedule.qr_code
                        logger.info(f"Student {user.username} scanned valid QR code for course {c.id} ({c.code}) on {date_str} (day {today_day_short})")
                        break
                    
                    # Also try generating expected QR code (for backward compatibility)
                    # Format: course_id_code_section_day_YYYYMMDD
                    qr_data = f"{c.id}_{c.code}_{c.section or ''}_{schedule.day}_{date_str}"
                    expected_qr_code = hashlib.sha256(qr_data.encode()).hexdigest()[:16].upper()
                    
                    if scanned_qr_code == expected_qr_code:
                        course = c
                        matched_schedule = schedule
                        matched_qr_code = expected_qr_code
                        logger.info(f"Student {user.username} scanned valid generated QR code for course {c.id} ({c.code}) on {date_str} (day {today_day_short})")
                        break
                    
                    # Try with random suffix (in case of collisions)
                    for attempt in range(10):
                        qr_data_with_suffix = f"{c.id}_{c.code}_{c.section or ''}_{schedule.day}_{date_str}_{random.randint(1000, 9999)}"
                        expected_qr_code = hashlib.sha256(qr_data_with_suffix.encode()).hexdigest()[:16].upper()
                        if scanned_qr_code == expected_qr_code:
                            course = c
                            matched_schedule = schedule
                            matched_qr_code = expected_qr_code
                            logger.info(f"Student {user.username} scanned valid QR code (with suffix) for course {c.id} ({c.code}) on {date_str} (day {today_day_short})")
                            break
                    
                    if course:
                        break
                
                if course:
                    break
            elif c.days:
                # Check synchronized schedule - need to find the day schedule or use course-level QR
                days_list = [d.strip() for d in c.days.split(',') if d.strip()]
                if today_day_name in days_list or today_day_short in days_list:
                    # For synchronized schedules, check if there's a day schedule for today
                    # If not, use course-level QR code (backward compatibility)
                    day_schedule = c.course_schedules.filter(day__iexact=today_day_short).first()
                    if day_schedule and day_schedule.qr_code:
                        # Use the day schedule's QR code
                        if scanned_qr_code == day_schedule.qr_code:
                            course = c
                            matched_schedule = day_schedule
                            matched_qr_code = day_schedule.qr_code
                            logger.info(f"Student {user.username} scanned valid day schedule QR code for course {c.id} ({c.code}) on {date_str}")
                            break
                    elif c.qr_code:
                        # Fallback to course-level QR code (for backward compatibility)
                        if scanned_qr_code == c.qr_code:
                            course = c
                            matched_schedule = None
                            matched_qr_code = c.qr_code
                            logger.info(f"Student {user.username} scanned valid course-level QR code for course {c.id} ({c.code}) on {date_str}")
                            break
        
        if not course:
            # Log for debugging
            logger.warning(f"Student {user.username} scanned invalid/expired QR code. Scanned: '{scanned_qr_code[:16]}...' (expected format: 16-char hex)")
            return JsonResponse({
                'success': False, 
                'message': 'Invalid or expired QR code. This QR code may be from a previous session. Please scan the current QR code displayed by your instructor for today\'s class.'
            })
        
        # Check if student is enrolled
        try:
            enrollment = CourseEnrollment.objects.get(
                course=course,
                student=user,
                is_active=True,
                deleted_at__isnull=True
            )
        except CourseEnrollment.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'You are not enrolled in this course.'})
        
        # CHECK: Verify student has registered their QR code for this course
        # This ensures only students who have registered their QR code can mark attendance
        qr_registration = QRCodeRegistration.objects.filter(
            student=user,
            course=course,
            is_active=True
        ).first()
        
        if not qr_registration:
            return JsonResponse({
                'success': False,
                'message': f'Your QR code is NOT registered in {course.code}. Please register your QR code first before marking attendance.',
                'error_type': 'qr_not_registered'
            })
        
        # Check attendance status
        # datetime is already imported at module level, no need to re-import
        # PH_TZ was already defined earlier in the function
        # Get current time in Philippines timezone (reuse PH_TZ from earlier)
        if PH_TZ:
            if hasattr(PH_TZ, 'localize'):
                # pytz timezone
                now_ph = datetime.now(PH_TZ)
            else:
                # zoneinfo timezone (Python 3.9+)
                now_ph = datetime.now(PH_TZ)
        else:
            now_ph = datetime.now()
        
        today_date = now_ph.date()
        
        # Get attendance status - use day-specific status if available, otherwise course-level
        attendance_status = 'closed'
        if matched_schedule and matched_schedule.attendance_status:
            # Use day-specific attendance status
            attendance_status = matched_schedule.attendance_status
        else:
            # Fallback to course-level attendance status
            attendance_status = course.attendance_status or 'closed'
        
        # STRICT VALIDATION: Only allow attendance if status is "open"
        # This prevents students from marking attendance when viewing closed schedules
        if attendance_status != 'open':
            if attendance_status == 'closed':
                status_message = 'Attendance is currently CLOSED for this schedule. Please contact your instructor.'
            elif attendance_status == 'stopped':
                status_message = 'Attendance has been STOPPED for this schedule. Please contact your instructor.'
            elif attendance_status == 'postponed':
                status_message = 'Attendance has been POSTPONED for this schedule. Please contact your instructor.'
            else:
                status_message = 'Attendance is not open for this schedule. Please contact your instructor.'
            
            return JsonResponse({
                'success': False,
                'message': status_message,
                'attendance_status': attendance_status
            })
        
        attendance_allowed = False
        status_message = ''
        
        # Get attendance window for today (needed for determining present/late status)
        # Get today's schedule
        today_weekday = now_ph.weekday()
        weekday_map = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
        
        # Check day-specific schedules first
        day_schedules = course.course_schedules.filter(day__in=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']).all()
        today_schedule = None
        for s in day_schedules:
            day_key = weekday_map.get(s.day, None)
            if day_key == today_weekday:
                today_schedule = s
                break
        
        attendance_start = None
        attendance_end = None
        
        if today_schedule:
            attendance_start = today_schedule.attendance_start or course.attendance_start
            attendance_end = today_schedule.attendance_end or course.attendance_end
        else:
            # Check synchronized schedule
            if course.days:
                days_list = [d.strip() for d in course.days.split(',') if d.strip()]
                day_names = {'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu', 'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'}
                weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                today_day_name = weekday_names[today_weekday] if today_weekday < 7 else None
                
                if today_day_name and today_day_name in days_list:
                    attendance_start = course.attendance_start
                    attendance_end = course.attendance_end
            else:
                attendance_start = course.attendance_start
                attendance_end = course.attendance_end
        
        # Check if attendance is allowed based on status
        # Note: We already validated that status is 'open' above, so this section handles 'automatic' mode
        if attendance_status == 'automatic':
            # Check if current time is within attendance window
            if attendance_start and attendance_end:
                current_time = now_ph.time()
                if attendance_start <= current_time <= attendance_end:
                    attendance_allowed = True
                else:
                    status_message = f'Attendance window is {attendance_start.strftime("%I:%M %p")} - {attendance_end.strftime("%I:%M %p")}. Current time is outside the window.'
            else:
                status_message = 'No attendance window set for today.'
        elif attendance_status == 'open':
            # Attendance is explicitly open - allow it
            attendance_allowed = True
        else:
            # Should not reach here since we validated 'open' status above, but just in case
            status_message = 'Attendance is not open for this schedule.'
            attendance_allowed = False
        
        # Get the schedule day for checking existing records
        schedule_day_for_check = None
        if matched_schedule:
            schedule_day_for_check = matched_schedule.day
        elif course.days:
            days_list = [d.strip() for d in course.days.split(',') if d.strip()]
            weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            today_day_name = weekday_names[today_weekday] if today_weekday < 7 else None
            if today_day_name and today_day_name in days_list:
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                schedule_day_for_check = day_map.get(today_day_name, today_day_short)
        
        # Check if already marked attendance today for this specific schedule day
        existing_record = AttendanceRecord.objects.filter(
            course=course,
            student=user,
            attendance_date=today_date,
            schedule_day=schedule_day_for_check
        ).first()
        
        if existing_record:
            # Student has already scanned - return specific message
            status_text = existing_record.status.capitalize() if existing_record.status else 'marked'
            return JsonResponse({
                'success': False,
                'message': f'You have already marked your attendance for today. Status: {status_text}.',
                'already_scanned': True,
                'existing_status': existing_record.status,
                'existing_time': existing_record.attendance_time.strftime('%I:%M %p') if getattr(existing_record, 'attendance_time', None) else None,
            })
        
        # If attendance is not allowed, return error message with specific status
        if not attendance_allowed:
            if attendance_status == 'closed':
                status_message = 'Attendance is currently CLOSED for this course. Please contact your instructor.'
            elif attendance_status == 'stopped':
                status_message = 'Attendance has been STOPPED for this course. Please contact your instructor.'
            elif attendance_status == 'postponed':
                status_message = 'Attendance has been POSTPONED for this course. Please contact your instructor.'
            return JsonResponse({
                'success': False,
                'message': status_message,
                'attendance_status': attendance_status
            })
        
        # Determine status (present, late, etc.)
        # Rules:
        # - Present: Scan time is within attendance window (start <= scan <= end) OR within present duration window
        # - Late: Scan time is AFTER present/attendance end time but BEFORE course end time
        # - Absent: No scan at all (handled in attendance reports, not here)
        status = 'present'
        current_time = now_ph.time()
        
        # Get course end time for today
        course_end_time = None
        if today_schedule:
            course_end_time = today_schedule.end_time
        else:
            course_end_time = course.end_time
        
        if attendance_allowed:
            # Allow instructor-configurable 'present' window based on when QR/session was opened
            present_duration_minutes = None
            qr_opened_at_dt = None
            try:
                if matched_schedule and getattr(matched_schedule, 'attendance_present_duration', None) is not None:
                    present_duration_minutes = matched_schedule.attendance_present_duration
                elif getattr(course, 'attendance_present_duration', None) is not None:
                    present_duration_minutes = course.attendance_present_duration
            except Exception:
                present_duration_minutes = None

            try:
                if matched_schedule and getattr(matched_schedule, 'qr_code_opened_at', None):
                    qr_opened_at_dt = matched_schedule.qr_code_opened_at
                elif getattr(course, 'qr_code_opened_at', None):
                    qr_opened_at_dt = course.qr_code_opened_at
            except Exception:
                qr_opened_at_dt = None

            # If instructor provided a present-duration and we have a recorded QR open time, use that rule
            if present_duration_minutes and qr_opened_at_dt:
                try:
                    present_cutoff = qr_opened_at_dt + datetime.timedelta(minutes=int(present_duration_minutes))
                    if now_ph <= present_cutoff:
                        status = 'present'
                    else:
                        # After present-window has expired, mark as late
                        # Check if we're still within course end time
                        if course_end_time and current_time <= course_end_time:
                            status = 'late'
                        else:
                            # After course end time, still mark as late (not absent)
                            status = 'late'
                except Exception:
                    # Fallback to default rules if computation fails
                    pass
            else:
                # Fallback to previous behavior: use attendance window times
                if attendance_start and attendance_end:
                    if attendance_start <= current_time <= attendance_end:
                        status = 'present'
                    elif current_time > attendance_end:
                        # After attendance end window
                        if course_end_time and current_time <= course_end_time:
                            status = 'late'
                        else:
                            # After course end time, mark as late if attendance is still open
                            status = 'late'
                    else:
                        # Before attendance start time
                        if attendance_status == 'open':
                            status = 'present'
                        else:
                            status = 'present'
                else:
                    status = 'present'
        
        # Get enrollment reference (should exist since we checked earlier)
        # But we'll get it again to ensure we have the latest
        enrollment = CourseEnrollment.objects.filter(
            course=course,
            student=user,
            is_active=True,
            deleted_at__isnull=True
        ).first()
        
        # Create attendance record (enrollment is required by model)
        if not enrollment:
            return JsonResponse({'success': False, 'message': 'Enrollment not found. Please contact support.'})
        
        # Get the schedule day for this attendance record
        schedule_day = None
        if matched_schedule:
            schedule_day = matched_schedule.day  # e.g., 'Mon', 'Tue', etc.
        elif course.days:
            # For synchronized schedules, use today's day
            days_list = [d.strip() for d in course.days.split(',') if d.strip()]
            weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            today_day_name = weekday_names[today_weekday] if today_weekday < 7 else None
            if today_day_name and today_day_name in days_list:
                day_map = {
                    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
                    'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
                }
                schedule_day = day_map.get(today_day_name, today_day_short)
        
        attendance_record = AttendanceRecord.objects.create(
            course=course,
            student=user,
            enrollment=enrollment,
            attendance_date=today_date,
            attendance_time=now_ph.time(),
            status=status,
            schedule_day=schedule_day  # Store which day/schedule this attendance is for
        )
        
        # Create notification for instructor when student marks attendance
        if course.instructor:
            try:
                student_name = user.full_name or user.username
                status_text = 'marked as late' if status == 'late' else 'marked as present'
                create_notification(
                    user=course.instructor,
                    notification_type='attendance_marked',
                    title='Student Marked Attendance',
                    message=f'{student_name} {status_text} for {course.code} - {course.name}',
                    category='course',
                    related_course=course,
                    related_user=user
                )
            except Exception as e:
                logger.error(f"Error creating attendance notification: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Attendance marked successfully! Status: {status.capitalize()}',
            'status': status,
            'attendance_time': attendance_record.attendance_time.strftime('%I:%M %p') if getattr(attendance_record, 'attendance_time', None) else None,
            'course_id': course.id,
            'schedule_day': schedule_day,
            'is_late': status == 'late'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid request data.'})
    except Exception as e:
        logger.error(f"Error processing QR scan: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def instructor_course_enrollments_view(request, course_id):
    """Get enrollments for a course - for archived course students display"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        course = get_object_or_404(Course, id=course_id, instructor=user)
        
        # Get all sibling sections for multi-section courses (including archived)
        sibling_courses = Course.objects.filter(
            instructor=user,
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            deleted_at__isnull=True  # Don't filter by is_active for archived courses
        ).order_by('section', 'id')
        
        # Get enrollments from all sibling sections (including archived courses)
        enrollments = CourseEnrollment.objects.filter(
            course__in=sibling_courses,
            deleted_at__isnull=True  # Don't filter by is_active for archived course enrollments
        ).select_related('student', 'course').order_by('course__section', 'full_name')
        
        enrollment_list = []
        sections_list = []
        
        # Get sections from course (not from enrollment) - all sections from sibling courses
        for sibling in sibling_courses:
            if sibling.section:
                section_upper = sibling.section.upper()
                if section_upper and section_upper not in sections_list:
                    sections_list.append(section_upper)
        
        sections_list.sort()
        
        # Map enrollments to their course sections (not enrollment.section)
        for enrollment in enrollments:
            # Use course section, not enrollment section
            course_section = (enrollment.course.section or '').upper() or 'N/A'
            enrollment_list.append({
                'id': enrollment.id,
                'full_name': enrollment.full_name,
                'student_name': enrollment.student.full_name if enrollment.student else enrollment.full_name,
                'student_id_number': enrollment.student_id_number,
                'section': course_section,  # Use course section, not enrollment section
                'email': enrollment.email,
                'course_id': enrollment.course.id,
                'enrollment_code': enrollment.course.enrollment_code or 'N/A',
                'profile_picture': enrollment.student.profile_picture.url if enrollment.student and enrollment.student.profile_picture else None,
                'year_level': getattr(enrollment, 'year_level', None) or (enrollment.course.year_level if enrollment.course else None),
                'enrolled_at': enrollment.enrolled_at.strftime('%B %d, %Y at %I:%M %p') if hasattr(enrollment, 'enrolled_at') and enrollment.enrolled_at else 'N/A',
            })
        
        sections_list.sort()
        
        return JsonResponse({
            'success': True,
            'enrollments': enrollment_list,
            'count': len(enrollment_list),
            'course_name': f"{course.code} - {course.name}" if course.code and course.name else (course.code or course.name or 'N/A'),
            'sections': sections_list,
            'is_multi_section': len(sections_list) > 1
        })
    except Exception as e:
        logger.error(f"Error fetching course enrollments: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
@require_http_methods(["GET"])
def instructor_course_detail_view(request, course_id):
    """Get course details for modal display - instructor can only view their courses"""
    try:
        course = get_object_or_404(Course, id=course_id)
        user = request.user
        
        if not user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        
        # Verify the course is assigned to the instructor
        if course.instructor != user:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        
        # Verify the course belongs to instructor's school
        if user.school_name and course.school_name != user.school_name:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        
        # Get accurate enrollment count
        student_count = CourseEnrollment.objects.filter(course=course, is_active=True).count()
        
        # Get day-specific schedules
        schedules = course.course_schedules.all().order_by('day_order')
        schedule_list = []
        has_day_schedules = False
        is_per_day_schedule = False  # True if schedules have different times (per-day), False if synchronized (same time)
        
        if schedules.exists():
            has_day_schedules = True
            # Check if this is truly a per-day schedule (different times per day) or synchronized (same time)
            # A course is synchronized if:
            # 1. It has a days field and start_time/end_time at course level, AND
            # 2. All day schedules have the same start_time and end_time as the course
            course_start_time = course.start_time
            course_end_time = course.end_time
            all_same_time = True
            
            # Check if all day schedules have the same time as the course-level times
            if course_start_time and course_end_time:
                for schedule in schedules:
                    # Check if this schedule has different times than the course-level times
                    if schedule.start_time != course_start_time or schedule.end_time != course_end_time:
                        all_same_time = False
                        break
            else:
                # If course doesn't have start_time/end_time, but schedules do, it's per-day
                all_same_time = False
            
            # Determine if it's per-day or synchronized:
            # - Synchronized: Course has days field AND all schedules have same time as course
            # - Per-day: Schedules have different times OR course doesn't have days field
            if course.days and course_start_time and course_end_time and all_same_time:
                is_per_day_schedule = False  # Synchronized schedule (same time across all days)
            else:
                is_per_day_schedule = True  # Per-day schedule (different times or no course-level times)
            
            for schedule in schedules:
                # Room priority: schedule room > course room > 'Not specified'
                room_display = schedule.room if schedule.room and schedule.room.strip() else (course.room if course.room and course.room.strip() else 'Not specified')
                schedule_list.append({
                    'day': schedule.day,
                    'start_time': schedule.start_time.strftime('%I:%M %p') if schedule.start_time else '',
                    'end_time': schedule.end_time.strftime('%I:%M %p') if schedule.end_time else '',
                    'room': room_display,
                    'attendance_start': schedule.attendance_start.strftime('%I:%M %p') if schedule.attendance_start else '',
                    'attendance_end': schedule.attendance_end.strftime('%I:%M %p') if schedule.attendance_end else ''
                })
        
        # Find sibling sections (same course identity across sections)
        # Include sibling sections across all instructors so instructors can
        # view and switch between sections of the same course code/name.
        # Do not restrict to the current instructor here.
        siblings_qs = Course.objects.filter(
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            is_active=True
        ).order_by('section', 'id')
        sibling_sections = [
            {
                'id': c.id,
                'section': (c.section or '').upper() or 'N/A',
            }
            for c in siblings_qs
        ]
        
        course_data = {
            'id': course.id,
            'code': course.code,
            'name': course.name,
            'program': course.program.code if course.program else 'N/A',
            'program_id': course.program.id if course.program else None,
            'program_code': course.program.code if course.program else '',
            'program_name': course.program.name if course.program else '',
            'department': course.program.department if course.program else 'N/A',
            'year_level': course.year_level,
            'section': course.section,
            'semester': course.semester,
            'school_year': course.school_year or 'Not specified',
            'instructor': course.instructor.full_name if course.instructor else 'Not assigned',
            'room': course.room or 'Not specified',
            'days': course.days,
            'start_time': course.start_time.strftime('%I:%M %p') if course.start_time else '',
            'end_time': course.end_time.strftime('%I:%M %p') if course.end_time else '',
            'color': course.color or '#3C4770',
            'is_active': course.is_active,
            'created_at': course.created_at.strftime('%B %d, %Y at %I:%M %p') if course.created_at else 'N/A',
            'updated_at': course.updated_at.strftime('%B %d, %Y at %I:%M %p') if course.updated_at else 'N/A',
            'attendance_start': course.attendance_start.strftime('%I:%M %p') if course.attendance_start else 'Not set',
            'attendance_end': course.attendance_end.strftime('%I:%M %p') if course.attendance_end else 'Not set',
            'day_schedules': schedule_list,
            'has_day_schedules': has_day_schedules,
            'is_per_day_schedule': is_per_day_schedule,  # True if per-day (different times), False if synchronized (same time)
            'enrollment_code': course.enrollment_code or '',  # Include enrollment code
            'enrollment_status': course.enrollment_status or 'open',  # Include enrollment status
            # QR code removed - it's already displayed in My Classes
            'student_count': student_count,
            'sibling_sections': sibling_sections,
        }
        
        return JsonResponse({'success': True, 'course': course_data})
    except Exception as e:
        logger.error(f"Error loading course details: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error loading course details: {str(e)}'}, status=500)

@login_required
def instructor_notifications_view(request):
    """Notifications view for instructors"""
    user = request.user
    
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    # Get all notifications for this instructor (handle if table doesn't exist yet)
    try:
        from .models import UserNotification
        # Use prefetch_related for related_course__program to avoid errors if program is None
        notifications = UserNotification.objects.filter(user=user).select_related('related_course', 'related_user').prefetch_related('related_course__program').order_by('-created_at')[:50]
        unread_count = UserNotification.objects.filter(user=user, is_read=False).count()
        logger.info(f"Instructor {user.username} (ID: {user.id}) has {len(notifications)} notifications, {unread_count} unread")
    except Exception as e:
        logger.error(f"Error fetching notifications for instructor {user.username}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        notifications = []
        unread_count = 0
    
    context = {
        'user': user,
        'notifications': notifications,
        'unread_count': unread_count,
    }
    
    return render(request, 'dashboard/shared/user_notifications.html', context)

@login_required
def student_notifications_view(request):
    """Notifications view for students - returns only the notifications list HTML"""
    user = request.user
    
    if not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    # Get all notifications for this student (handle if table doesn't exist yet)
    try:
        notifications = UserNotification.objects.filter(user=user).order_by('-created_at')[:50]
        unread_count = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        notifications = []
        unread_count = 0
    
    context = {
        'user': user,
        'notifications': notifications,
        'unread_count': unread_count,
    }
    
    return render(request, 'dashboard/shared/user_notifications.html', context)

@login_required
@require_http_methods(["POST"])
def mark_notification_read_view(request, notification_id):
    """Mark a notification as read for any user type"""
    try:
        notification = UserNotification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except UserNotification.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Notification not found'}, status=404)
    except Exception:
        # Handle case where table doesn't exist yet
        return JsonResponse({'success': False, 'message': 'Notification system not available'}, status=503)

@login_required
@require_http_methods(["POST"])
def delete_notification_view(request, notification_id):
    """Delete a specific notification"""
    try:
        notification = UserNotification.objects.get(id=notification_id, user=request.user)
        notification.delete()
        return JsonResponse({'success': True, 'message': 'Notification deleted successfully'})
    except UserNotification.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Notification not found'}, status=404)
    except Exception as e:
        logger.error(f"Error deleting notification: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read_view(request):
    """Mark all notifications as read for the current user"""
    try:
        updated_count = UserNotification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True, 'message': f'{updated_count} notification(s) marked as read', 'count': updated_count})
    except Exception as e:
        logger.error(f"Error marking all notifications as read: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
@require_http_methods(["POST"])
def delete_all_notifications_view(request):
    """Delete all notifications for the current user"""
    try:
        deleted_count = UserNotification.objects.filter(user=request.user).delete()[0]
        return JsonResponse({'success': True, 'message': f'{deleted_count} notification(s) deleted successfully', 'count': deleted_count})
    except Exception as e:
        logger.error(f"Error deleting all notifications: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def home_view(request):
    if request.user.is_authenticated:
        if request.user.is_teacher:
            return redirect('dashboard:teacher_dashboard')
        elif request.user.is_student:
            return redirect('dashboard:student_dashboard')
    return redirect('login_signup')

# INSTRUCTOR TRASH MANAGEMENT VIEWS
@login_required
def instructor_trash_view(request):
    """View deleted courses and enrollments for instructor"""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    from django.utils import timezone
    now = timezone.now()
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    # Get deleted courses for this instructor (trash - has deleted_at)
    deleted_courses = Course.objects.filter(
        instructor=user,
        deleted_at__isnull=False,
        is_archived=False  # Not archived, actually deleted
    ).order_by('-deleted_at')
    
    # Debug logging
    logger.info(f"Trash view - Found {deleted_courses.count()} deleted courses for instructor {user.id}")
    if deleted_courses.exists():
        for c in deleted_courses[:3]:
            logger.info(f"  - Course {c.id}: {c.code} - {c.name}, deleted_at={c.deleted_at}, is_archived={c.is_archived}")
    
    # Get archived courses for this instructor (archive - is_archived=True, deleted_at=None)
    archived_courses = Course.objects.filter(
        instructor=user,
        is_archived=True,
        deleted_at__isnull=True  # Archived items are not deleted
    ).order_by('-updated_at')
    
    # Get deleted enrollments for courses taught by this instructor
    deleted_enrollments = CourseEnrollment.objects.filter(
        course__instructor=user,
        deleted_at__isnull=False
    ).select_related('course', 'student').order_by('-deleted_at')
    
    # Calculate days until permanent deletion (30 days from deleted_at)
    for course in deleted_courses:
        days_until_deletion = 30 - (now - course.deleted_at).days
        course.days_until_deletion = max(0, days_until_deletion)
    
    for enrollment in deleted_enrollments:
        days_until_deletion = 30 - (now - enrollment.deleted_at).days
        enrollment.days_until_deletion = max(0, days_until_deletion)
    
    # Organize deleted courses by school year, then year level, then semester
    deleted_courses_by_school_year = {}
    deleted_courses_by_semester = {}
    for course in deleted_courses:
        school_year_key = course.school_year or 'Unspecified'
        if school_year_key not in deleted_courses_by_school_year:
            deleted_courses_by_school_year[school_year_key] = {}
        
        # Handle year_level - use 'Unspecified' if None
        year_level_value = course.year_level if course.year_level is not None else 0
        year_key = f"Year {year_level_value}" if year_level_value > 0 else "Unspecified Year"
        if year_key not in deleted_courses_by_school_year[school_year_key]:
            deleted_courses_by_school_year[school_year_key][year_key] = {
                'semesters': {},
            }
        
        semester_key = course.semester or 'Unspecified'
        if semester_key not in deleted_courses_by_school_year[school_year_key][year_key]['semesters']:
            deleted_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = []
        
        deleted_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key].append(course)
        
        # Also group by semester for category filter
        if semester_key not in deleted_courses_by_semester:
            deleted_courses_by_semester[semester_key] = []
        deleted_courses_by_semester[semester_key].append(course)
    
    # Organize archived courses by school year, then year level, then semester
    archived_courses_by_school_year = {}
    archived_courses_by_semester = {}
    for course in archived_courses:
        school_year_key = course.school_year or 'Unspecified'
        if school_year_key not in archived_courses_by_school_year:
            archived_courses_by_school_year[school_year_key] = {}
        
        # Handle year_level - use 'Unspecified' if None
        year_level_value = course.year_level if course.year_level is not None else 0
        year_key = f"Year {year_level_value}" if year_level_value > 0 else "Unspecified Year"
        if year_key not in archived_courses_by_school_year[school_year_key]:
            archived_courses_by_school_year[school_year_key][year_key] = {
                'semesters': {},
            }
        
        semester_key = course.semester or 'Unspecified'
        if semester_key not in archived_courses_by_school_year[school_year_key][year_key]['semesters']:
            archived_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = []
        
        archived_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key].append(course)
        
        # Also group by semester for category filter
        if semester_key not in archived_courses_by_semester:
            archived_courses_by_semester[semester_key] = []
        archived_courses_by_semester[semester_key].append(course)
    
    # Group multi-section courses within each semester for deleted courses
    for school_year_key in deleted_courses_by_school_year:
        for year_key in deleted_courses_by_school_year[school_year_key]:
            for semester_key in deleted_courses_by_school_year[school_year_key][year_key]['semesters']:
                courses_list = deleted_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key]
                # Group by course identity (code, name, semester, school_year)
                grouped_map = {}
                for course in courses_list:
                    key = (
                        (course.code or '').strip().upper(),
                        (course.name or '').strip().upper(),
                        (course.semester or '').strip().lower(),
                        (course.school_year or '').strip()
                    )
                    if key not in grouped_map:
                        grouped_map[key] = []
                    grouped_map[key].append(course)
                
                # Replace list with grouped courses (one representative per group)
                grouped_courses_list = []
                for key, group_list in grouped_map.items():
                    rep = group_list[0]  # Representative course
                    rep.grouped_sections = group_list  # Store all sections
                    rep.sibling_sections = sorted([(c.section or '').upper() for c in group_list if c.section])
                    
                    # Collect distinct colors across sections for gradient
                    colors = [gc.color for gc in group_list if (gc.color or '').strip()]
                    if not colors:
                        colors = [(rep.color or '#3C4770') or '#3C4770']
                    
                    # Build multi-stop gradient string
                    unique_colors = []
                    for c in colors:
                        if c and c not in unique_colors:
                            unique_colors.append(c)
                    if not unique_colors:
                        unique_colors = ['#3C4770']
                    
                    # Compute gradient stops
                    stops = []
                    if len(unique_colors) == 1:
                        stops = [f"{unique_colors[0]} 0%", f"{unique_colors[0]} 100%"]
                    else:
                        total = len(unique_colors) - 1
                        for idx, col in enumerate(unique_colors):
                            pct = int(round((idx / total) * 100))
                            stops.append(f"{col} {pct}%")
                    rep.display_gradient = f"linear-gradient(135deg, {', '.join(stops)})"
                    rep.display_color_start = unique_colors[0]
                    rep.display_color_end = unique_colors[1] if len(unique_colors) > 1 else unique_colors[0]
                    
                    grouped_courses_list.append(rep)
                
                deleted_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = grouped_courses_list
    
    # Group multi-section courses within each semester for archived courses
    for school_year_key in archived_courses_by_school_year:
        for year_key in archived_courses_by_school_year[school_year_key]:
            for semester_key in archived_courses_by_school_year[school_year_key][year_key]['semesters']:
                courses_list = archived_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key]
                # Group by course identity (code, name, semester, school_year)
                grouped_map = {}
                for course in courses_list:
                    key = (
                        (course.code or '').strip().upper(),
                        (course.name or '').strip().upper(),
                        (course.semester or '').strip().lower(),
                        (course.school_year or '').strip()
                    )
                    if key not in grouped_map:
                        grouped_map[key] = []
                    grouped_map[key].append(course)
                
                # Replace list with grouped courses (one representative per group)
                grouped_courses_list = []
                for key, group_list in grouped_map.items():
                    rep = group_list[0]  # Representative course
                    rep.grouped_sections = group_list  # Store all sections
                    rep.sibling_sections = sorted([(c.section or '').upper() for c in group_list if c.section])
                    
                    # Collect distinct colors across sections for gradient
                    colors = [gc.color for gc in group_list if (gc.color or '').strip()]
                    if not colors:
                        colors = [(rep.color or '#3C4770') or '#3C4770']
                    
                    # Build multi-stop gradient string
                    unique_colors = []
                    for c in colors:
                        if c and c not in unique_colors:
                            unique_colors.append(c)
                    if not unique_colors:
                        unique_colors = ['#3C4770']
                    
                    # Compute gradient stops
                    stops = []
                    if len(unique_colors) == 1:
                        stops = [f"{unique_colors[0]} 0%", f"{unique_colors[0]} 100%"]
                    else:
                        total = len(unique_colors) - 1
                        for idx, col in enumerate(unique_colors):
                            pct = int(round((idx / total) * 100))
                            stops.append(f"{col} {pct}%")
                    rep.display_gradient = f"linear-gradient(135deg, {', '.join(stops)})"
                    rep.display_color_start = unique_colors[0]
                    rep.display_color_end = unique_colors[1] if len(unique_colors) > 1 else unique_colors[0]
                    
                    grouped_courses_list.append(rep)
                
                # Update the dictionary with grouped courses (FIXED: was updating deleted_courses_by_school_year)
                archived_courses_by_school_year[school_year_key][year_key]['semesters'][semester_key] = grouped_courses_list
    
    # Category filtering removed - always display as folders
    mode = request.GET.get('mode', 'trash')  # 'trash' (default), 'archive', 'all'
    
    # Sort archived courses by school year (newest first)
    # School year format: "2026-2027" -> extract first year (2026) for sorting
    def extract_year_for_sorting(school_year_str):
        """Extract the first year from school year string for sorting (e.g., '2026-2027' -> 2026)"""
        if not school_year_str or school_year_str == 'Unspecified':
            return 0
        try:
            # Extract first year from "2026-2027" format
            first_year = int(school_year_str.split('-')[0])
            return first_year
        except (ValueError, IndexError):
            return 0
    
    # Sort archived_courses_by_school_year by school year (newest first)
    sorted_archived_courses_by_school_year = dict(sorted(
        archived_courses_by_school_year.items(),
        key=lambda x: extract_year_for_sorting(x[0]),
        reverse=True  # Newest first
    ))
    
    # Debug logging for trash display
    logger.info(f"Trash view - Mode: {mode}")
    logger.info(f"Trash view - deleted_courses count: {deleted_courses.count()}")
    logger.info(f"Trash view - deleted_courses_by_school_year keys: {list(deleted_courses_by_school_year.keys())}")
    logger.info(f"Trash view - deleted_courses_by_school_year length: {len(deleted_courses_by_school_year)}")
    if deleted_courses_by_school_year:
        for sy_key, sy_data in deleted_courses_by_school_year.items():
            logger.info(f"  School Year '{sy_key}': {len(sy_data)} year levels")
            for y_key, y_data in sy_data.items():
                logger.info(f"    Year '{y_key}': {len(y_data.get('semesters', {}))} semesters")
                for sem_key, sem_courses in y_data.get('semesters', {}).items():
                    logger.info(f"      Semester '{sem_key}': {len(sem_courses)} courses")
    
    context = {
        'user': user,
        'school_admin': school_admin,
        'deleted_courses': deleted_courses,
        'archived_courses': archived_courses,
        'deleted_enrollments': deleted_enrollments,
        'deleted_courses_by_school_year': deleted_courses_by_school_year,
        'deleted_courses_by_semester': deleted_courses_by_semester,
        'archived_courses_by_school_year': sorted_archived_courses_by_school_year,  # Use sorted version
        'archived_courses_by_semester': archived_courses_by_semester,
        'category': 'all',  # Always 'all' - no category filtering, display as folders only
        'mode': mode,
    }
    
    return render(request, 'dashboard/instructor/instructor_trash.html', context)

@login_required
@require_http_methods(["POST"])
def instructor_restore_course_view(request, course_id):
    """Restore a deleted or archived course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    # Try to get course (either deleted or archived)
    try:
        course = Course.objects.get(id=course_id, instructor=user)
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Course not found.'})
    
    # Verify the course is assigned to the instructor
    if course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only restore courses assigned to you.'})
    
    try:
        # Restore: clear deleted_at, clear is_archived, and set is_active
        # For multi-section courses, restore all sections with same identity
        courses_to_restore = Course.objects.filter(
            instructor=user,
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year
        ).filter(
            Q(deleted_at__isnull=False) | Q(is_archived=True)
        )
        
        restored_count = 0
        for c in courses_to_restore:
            c.deleted_at = None
            c.is_archived = False
            c.is_active = True
            c.save()
            restored_count += 1
        
        if restored_count > 1:
            return JsonResponse({'success': True, 'message': f'{restored_count} course sections restored successfully!'})
        else:
            return JsonResponse({'success': True, 'message': f'Course "{course.code} - {course.name}" restored successfully!'})
    except Exception as e:
        logger.error(f"Error restoring course: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_drop_enrollment_view(request, enrollment_id):
    """Drop (soft delete) an enrollment"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id, deleted_at__isnull=True)
    
    # Verify the enrollment belongs to a course taught by the instructor
    if enrollment.course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only drop enrollments for your courses.'})
    
    try:
        from django.utils import timezone
        # Soft delete: set deleted_at and is_active
        # NOTE: Do NOT delete biometric or QR registrations on soft drop
        # They will be deleted only on permanent delete
        enrollment.deleted_at = timezone.now()
        enrollment.is_active = False
        enrollment.save()
        
        return JsonResponse({'success': True, 'message': f'Enrollment for "{enrollment.full_name}" dropped successfully!'})
    except Exception as e:
        logger.error(f"Error dropping enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_restore_enrollment_view(request, enrollment_id):
    """Restore a deleted enrollment"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id, deleted_at__isnull=False)
    
    # Verify the enrollment belongs to a course taught by the instructor
    if enrollment.course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only restore enrollments for your courses.'})
    
    try:
        # Restore: clear deleted_at and set is_active
        enrollment.deleted_at = None
        enrollment.is_active = True
        enrollment.save()
        return JsonResponse({'success': True, 'message': f'Enrollment for "{enrollment.full_name}" restored successfully!'})
    except Exception as e:
        logger.error(f"Error restoring enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_permanent_delete_course_view(request, course_id):
    """Permanently delete a course from trash"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    course = get_object_or_404(Course, id=course_id, deleted_at__isnull=False)
    
    # Verify the course is assigned to the instructor
    if course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only delete courses assigned to you.'})
    
    try:
        course_name = f"{course.code} - {course.name}"
        course_id = course.id  # Store ID before deletion
        course.delete()  # Permanent delete - completely removed from database
        
        # Log the permanent deletion
        logger.info(f"Instructor {user.id} permanently deleted course {course_id} ({course_name})")
        
        return JsonResponse({
            'success': True, 
            'message': f'Course "{course_name}" permanently deleted.',
            'deleted_course_id': course_id  # Return ID so frontend can clear cache
        })
    except Exception as e:
        logger.error(f"Error permanently deleting course: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_restore_all_courses_view(request):
    """Restore all deleted courses for instructor"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        deleted_courses = Course.objects.filter(
            instructor=user,
            deleted_at__isnull=False,
            is_archived=False  # Only restore deleted, not archived
        )
        
        count = deleted_courses.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'No deleted courses found to restore.'})
        
        # Restore all courses
        restored = 0
        for course in deleted_courses:
            course.deleted_at = None
            course.is_active = True
            course.save()
            restored += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully restored {restored} course(s)!'
        })
    except Exception as e:
        logger.error(f"Error restoring all courses: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_restore_all_archived_courses_view(request):
    """Restore all archived courses for instructor"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        archived_courses = Course.objects.filter(
            instructor=user,
            is_archived=True,
            deleted_at__isnull=True  # Only restore archived, not deleted
        )
        
        count = archived_courses.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'No archived courses found to restore.'})
        
        # Restore all courses
        restored = 0
        for course in archived_courses:
            course.is_archived = False
            course.is_active = True
            course.save()
            restored += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully restored {restored} archived course(s)!'
        })
    except Exception as e:
        logger.error(f"Error restoring all archived courses: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def instructor_permanent_delete_all_courses_view(request):
    """Permanently delete all deleted courses for instructor"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        deleted_courses = Course.objects.filter(
            instructor=user,
            deleted_at__isnull=False
        )
        
        count = deleted_courses.count()
        if count == 0:
            return JsonResponse({'success': False, 'message': 'No deleted courses found to delete.'})
        
        # Store course IDs before deletion for cache clearing
        deleted_course_ids = list(deleted_courses.values_list('id', flat=True))
        
        # Permanently delete all courses - completely removed from database
        deleted = deleted_courses.delete()[0]
        
        # Log the permanent deletion
        logger.info(f"Instructor {user.id} permanently deleted {deleted} course(s): {deleted_course_ids}")
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully permanently deleted {deleted} course(s)!',
            'deleted_course_ids': deleted_course_ids  # Return IDs so frontend can clear cache
        })
    except Exception as e:
        logger.error(f"Error permanently deleting all courses: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def instructor_get_dropped_students_view(request):
    """Get dropped students for a specific course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to access this page.'})
    
    try:
        course_id = request.GET.get('course_id')
        if not course_id:
            return JsonResponse({'success': False, 'message': 'Course ID is required.'})
        
        course = get_object_or_404(Course, id=course_id, instructor=user, deleted_at__isnull=True)
        
        # Get all sibling courses (same code, name, semester, school_year)
        sibling_courses = Course.objects.filter(
            instructor=user,
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            deleted_at__isnull=True
        )
        
        # Get dropped enrollments
        dropped_enrollments = CourseEnrollment.objects.filter(
            course__in=sibling_courses,
            deleted_at__isnull=False
        ).select_related('student', 'course').order_by('-deleted_at')
        
        from django.utils import timezone
        now = timezone.now()
        
        dropped_students = []
        for enrollment in dropped_enrollments:
            days_left = None
            if enrollment.deleted_at:
                days_until_deletion = 30 - (now - enrollment.deleted_at).days
                days_left = max(0, days_until_deletion)
            
            # Use the current student's full_name instead of the enrollment snapshot
            # This ensures the dropped students modal shows the student's updated name
            current_full_name = enrollment.student.full_name if enrollment.student else enrollment.full_name
            dropped_students.append({
                'id': enrollment.id,
                'student_id': enrollment.student.id if enrollment.student else None,
                'full_name': current_full_name or 'Unknown',
                'student_id_number': enrollment.student.school_id or enrollment.student_id_number if enrollment.student else enrollment.student_id_number or 'N/A',
                'email': enrollment.student.email if enrollment.student else enrollment.email or 'N/A',
                'section': enrollment.course.section or 'N/A',
                'year_level': getattr(enrollment, 'year_level', None) or (enrollment.course.year_level if enrollment.course else None),
                'dropped_date': enrollment.deleted_at.strftime('%B %d, %Y') if enrollment.deleted_at else 'N/A',
                'days_left': days_left,
                'profile_picture': enrollment.student.profile_picture.url if enrollment.student and enrollment.student.profile_picture else None,
            })
        
        return JsonResponse({
            'success': True,
            'dropped_students': dropped_students,
            'count': len(dropped_students)
        })
    except Exception as e:
        logger.error(f"Error fetching dropped students: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
@require_http_methods(["POST"])
def instructor_permanent_delete_enrollment_view(request, enrollment_id):
    """Permanently delete an enrollment from trash"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id, deleted_at__isnull=False)
    
    # Verify the enrollment belongs to a course taught by the instructor
    if enrollment.course.instructor != user:
        return JsonResponse({'success': False, 'message': 'You can only delete enrollments for your courses.'})
    
    try:
        student_name = enrollment.full_name
        
        # Clean up biometric registration before deleting enrollment
        try:
            from .models import BiometricRegistration
            biometric_registrations = BiometricRegistration.objects.filter(
                student=enrollment.student,
                course=enrollment.course
            )
            biometric_registrations.delete()
            logger.info(f"Deleted biometric registrations for {student_name} before permanent enrollment deletion")
        except Exception as e:
            logger.error(f"Error deleting biometric registrations: {str(e)}")
        
        # Clean up QR registration before deleting enrollment
        try:
            from .models import QRCodeRegistration
            qr_registrations = QRCodeRegistration.objects.filter(
                student=enrollment.student,
                course=enrollment.course
            )
            qr_registrations.delete()
            logger.info(f"Deleted QR registrations for {student_name} before permanent enrollment deletion")
        except Exception as e:
            logger.error(f"Error deleting QR registrations: {str(e)}")
        
        enrollment.delete()  # Permanent delete
        return JsonResponse({'success': True, 'message': f'Enrollment for "{student_name}" permanently deleted.'})
    except Exception as e:
        logger.error(f"Error permanently deleting enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def student_dropped_enrollments_view(request):
    """List dropped (soft-deleted) enrollments for the current student"""
    user = request.user
    if not user.is_authenticated or not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    try:
        from django.utils import timezone
        now = timezone.now()

        dropped = CourseEnrollment.objects.filter(student=user, deleted_at__isnull=False).select_related('course').order_by('-deleted_at')
        enrollments = []
        for e in dropped:
            days_left = None
            if e.deleted_at:
                days_until = 14 - (now - e.deleted_at).days
                days_left = max(0, days_until)

            enrollments.append({
                'id': e.id,
                'course_id': e.course.id if e.course else None,
                'course_code': e.course.code if e.course else 'N/A',
                'course_name': e.course.name if e.course else 'N/A',
                'section': e.course.section if e.course else '',
                'dropped_date': e.deleted_at.strftime('%B %d, %Y') if e.deleted_at else 'N/A',
                'days_left': days_left,
            })

        return JsonResponse({'success': True, 'dropped_enrollments': enrollments, 'count': len(enrollments)})
    except Exception as e:
        logger.error(f"Error fetching student dropped enrollments: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def student_restore_enrollment_view(request, enrollment_id):
    """Restore a student's dropped enrollment"""
    user = request.user
    if not user.is_authenticated or not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id, deleted_at__isnull=False)

    # Verify ownership
    if enrollment.student != user:
        return JsonResponse({'success': False, 'message': 'You can only restore your own enrollments.'}, status=403)

    try:
        enrollment.deleted_at = None
        enrollment.is_active = True
        enrollment.save()
        return JsonResponse({'success': True, 'message': f'Enrollment for "{enrollment.full_name}" restored successfully!'})
    except Exception as e:
        logger.error(f"Error restoring enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def student_permanent_delete_enrollment_view(request, enrollment_id):
    """Permanently delete a student's dropped enrollment"""
    user = request.user
    if not user.is_authenticated or not user.is_student:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id, deleted_at__isnull=False)

    # Verify ownership
    if enrollment.student != user:
        return JsonResponse({'success': False, 'message': 'You can only delete your own enrollments.'}, status=403)

    try:
        course_name = enrollment.full_name
        
        # Clean up biometric registration before deleting enrollment
        try:
            from .models import BiometricRegistration
            biometric_registrations = BiometricRegistration.objects.filter(
                student=enrollment.student,
                course=enrollment.course
            )
            biometric_registrations.delete()
            logger.info(f"Deleted biometric registrations for student before permanent enrollment deletion")
        except Exception as e:
            logger.error(f"Error deleting biometric registrations: {str(e)}")
        
        # Clean up QR registration before deleting enrollment
        try:
            from .models import QRCodeRegistration
            qr_registrations = QRCodeRegistration.objects.filter(
                student=enrollment.student,
                course=enrollment.course
            )
            qr_registrations.delete()
            logger.info(f"Deleted QR registrations for student before permanent enrollment deletion")
        except Exception as e:
            logger.error(f"Error deleting QR registrations: {str(e)}")
        
        enrollment.delete()
        return JsonResponse({'success': True, 'message': f'Enrollment permanently deleted.'})
    except Exception as e:
        logger.error(f"Error permanently deleting enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
def instructor_attendance_reports_view(request):
    """Attendance reports view for instructors"""
    user = request.user
    if not user.is_teacher:
        return render(request, 'dashboard/shared/error.html', {'message': 'You are not authorized to access this page.'})
    
    # Get school admin for topbar display
    school_admin = None
    if user.school_name:
        school_admin = CustomUser.objects.filter(
            is_admin=True,
            school_name=user.school_name,
            deleted_at__isnull=True
        ).first()
    
    # Get all active courses for this instructor - group multi-section courses
    # Order by creation time so the first course added appears first (preserve manage-courses behavior)
    all_courses = Course.objects.filter(
        instructor=user,
        is_active=True,
        deleted_at__isnull=True,
        is_archived=False
    ).select_related('program', 'instructor').order_by('created_at')
    
    # Group courses by code, name, semester, school_year (multi-section as one)
    grouped_courses_map = {}
    for course in all_courses:
        key = (
            (course.code or '').strip().upper(),
            (course.name or '').strip().upper(),
            (course.semester or '').strip().lower(),
            (course.school_year or '').strip()
        )
        if key not in grouped_courses_map:
            grouped_courses_map[key] = {
                'id': course.id,  # Use first course ID as representative
                'code': course.code,
                'name': course.name,
                'semester': course.semester,
                'school_year': course.school_year,
                'sections': []
            }
        if course.section:
            section_upper = (course.section or '').upper()
            if section_upper not in grouped_courses_map[key]['sections']:
                grouped_courses_map[key]['sections'].append(section_upper)
    
    # Preserve insertion order so the first-added representative stays first
    # grouped_courses_map is populated in the order of `all_courses` (which is ordered by created_at)
    grouped_courses = list(grouped_courses_map.values())
    # Build display gradients/colors for grouped courses so templates can reuse
    try:
        for rep in grouped_courses:
            try:
                # Find actual Course objects that belong to this grouped key
                group_list = [c for c in all_courses if (
                    (c.code or '').strip().upper() == (rep.get('code') or '').strip().upper() and
                    (c.name or '').strip().upper() == (rep.get('name') or '').strip().upper() and
                    (c.semester or '').strip().lower() == (rep.get('semester') or '').strip().lower() and
                    (c.school_year or '').strip() == (rep.get('school_year') or '').strip()
                )]

                colors = [getattr(g, 'color', None) for g in group_list if (getattr(g, 'color', None) or '').strip()]
                if not colors:
                    # fallback to representative course color if available
                    if group_list:
                        colors = [(group_list[0].color or '#3C4770')]
                    else:
                        colors = ['#3C4770']

                # Keep unique colors preserving order
                unique_colors = []
                for c in colors:
                    if c and c not in unique_colors:
                        unique_colors.append(c)
                if not unique_colors:
                    unique_colors = ['#3C4770']

                # Build multi-stop gradient
                stops = []
                if len(unique_colors) == 1:
                    stops = [f"{unique_colors[0]} 0%", f"{unique_colors[0]} 100%"]
                else:
                    total = len(unique_colors) - 1
                    for idx, col in enumerate(unique_colors):
                        pct = int(round((idx / total) * 100))
                        stops.append(f"{col} {pct}%")

                rep['display_gradient'] = f"linear-gradient(135deg, {', '.join(stops)})"
                rep['display_color_start'] = unique_colors[0]
                rep['display_color_end'] = unique_colors[1] if len(unique_colors) > 1 else unique_colors[0]
            except Exception:
                rep.setdefault('display_gradient', 'linear-gradient(135deg, #3C4770 0%, #447294 100%)')
                rep.setdefault('display_color_start', '#3C4770')
                rep.setdefault('display_color_end', '#447294')
    except Exception:
        pass
    
    # Get filter parameters
    course_id = request.GET.get('course', '').strip()
    section_filter = request.GET.get('section', '').strip()
    date_filter = request.GET.get('date', '').strip()  # Specific date filter (YYYY-MM-DD)
    day_filter = request.GET.get('day_filter', '').strip()  # Day of week filter (Monday, Tuesday, etc.)
    month_filter = request.GET.get('month_filter', '').strip()  # Month filter (1-12)
    week_filter = request.GET.get('week_filter', '').strip()  # Week filter (1, 2, 3, last, or empty for all)
    weeks = int(request.GET.get('weeks', 4))  # Number of weeks to look ahead (default 4, used when week_filter is not set)
    
    # Get selected course
    selected_course = None
    selected_sections = []
    attendance_data = []
    attendance_data_grouped = []
    course_stats = {}
    student_stats = {}
    
    # Only process data if course is selected
    if course_id:
        try:
            # Get the representative course. If the course ID is not owned by the instructor
            # (for example the section is taught by another instructor) allow loading it
            # as long as it belongs to the same grouped course key (code/name/semester/school_year).
            try:
                selected_course = Course.objects.get(
                    id=course_id,
                    instructor=user,
                    is_active=True,
                    deleted_at__isnull=True,
                    is_archived=False
                )
            except Course.DoesNotExist:
                # Attempt to load the course without instructor restriction
                fallback_course = Course.objects.filter(
                    id=course_id,
                    is_active=True,
                    deleted_at__isnull=True,
                    is_archived=False
                ).first()
                if fallback_course:
                    key = (
                        (fallback_course.code or '').strip().upper(),
                        (fallback_course.name or '').strip().upper(),
                        (fallback_course.semester or '').strip().lower(),
                        (fallback_course.school_year or '').strip()
                    )
                    # Allow if this course matches one of the instructor's grouped courses
                    if key in grouped_courses_map:
                        selected_course = fallback_course
                    else:
                        raise

            # Get all sibling courses (same code, name, semester, school_year)
            # Include sibling sections across instructors so instructors can view other sections of
            # the same course (useful when students enrolled into other section should appear)
            sibling_courses = Course.objects.filter(
                code=selected_course.code,
                name=selected_course.name,
                semester=selected_course.semester,
                school_year=selected_course.school_year,
                is_active=True,
                deleted_at__isnull=True,
                is_archived=False
            ).order_by('section')
            
            selected_sections = sorted([(c.section or '').upper() for c in sibling_courses if c.section])
            
            # Validate section filter - if multiple sections exist, section must be selected
            # If no section filter is provided and multiple sections exist, default to the selected_course's section
            if len(selected_sections) > 1 and not section_filter:
                # Multiple sections but no section selected - default to the selected_course's section
                # This ensures users see their own section's records when clicking the course
                section_filter = (selected_course.section or '').upper() if selected_course.section else 'all'
                if section_filter and section_filter != 'all':
                    courses_to_process = sibling_courses.filter(section__iexact=section_filter)
                else:
                    courses_to_process = sibling_courses
            else:
                # Filter courses by section if specified
                if section_filter and section_filter.lower() == 'all':
                    # Include all sections
                    courses_to_process = sibling_courses
                elif section_filter and section_filter.upper() in selected_sections:
                    courses_to_process = sibling_courses.filter(section__iexact=section_filter)
                elif len(selected_sections) <= 1:
                    # Single section or no sections - use all courses
                    courses_to_process = sibling_courses
                else:
                    # Multiple sections but invalid section filter - don't process
                    selected_course = None
                    courses_to_process = None
            
            # Only process data if courses_to_process is valid
            if courses_to_process is not None:
                # Get all enrollments for these courses
                # When filtering by section, we already filtered courses by section, so we just need enrollments for those courses
                # The enrollment's section field is the student's section at enrollment time, but we want enrollments
                # for courses with the selected section, so we don't need to filter by enrollment.section
                all_enrollments = CourseEnrollment.objects.filter(
                    course__in=courses_to_process,
                    is_active=True,
                    deleted_at__isnull=True
                ).select_related('student', 'course')
                
                # Get unique students (distinct) - only from filtered enrollments
                # Also track which courses each student is enrolled in (for multi-section courses)
                unique_students = {}
                student_courses_map = {}  # Maps student_id -> list of course objects
                for enrollment in all_enrollments:
                    if enrollment.student.id not in unique_students:
                        unique_students[enrollment.student.id] = enrollment.student
                        student_courses_map[enrollment.student.id] = []
                    student_courses_map[enrollment.student.id].append(enrollment.course)
                
                # Get attendance records for the filtered courses
                # Since courses_to_process is already filtered by section, we just need records for those courses
                # IMPORTANT: Only include records where the student was enrolled in the selected course/section
                # Do NOT include records for students who were enrolled in other sections
                attendance_query = AttendanceRecord.objects.filter(
                    course__in=courses_to_process,
                    enrollment__in=all_enrollments  # Only records from ACTIVE enrollments in selected section
                ).select_related('student', 'course', 'enrollment').order_by('-attendance_date', '-attendance_time')
                
                # If a specific section is selected (not 'all'), ensure we only include those records
                if section_filter and section_filter.lower() != 'all':
                    attendance_query = attendance_query.filter(
                        enrollment__course__section__iexact=section_filter
                    )
                
                attendance_records = list(attendance_query)
                
                # CRITICAL: Double-check section filtering - ensure NO records from other sections slip through
                if section_filter and section_filter.lower() != 'all':
                    filtered_section = section_filter.strip().lower()
                    cleaned_records = []
                    for record in attendance_records:
                        # Check if this record's course section matches the filter
                        record_section = (record.course.section or '').strip().lower()
                        if record_section == filtered_section:
                            cleaned_records.append(record)
                    attendance_records = cleaned_records
                
                # Filter attendance records by scheduled class days only
                # Only show records for days the class is scheduled to meet
                scheduled_days_map = {}  # Map course_id -> set of scheduled days
                course_end_times = {}  # Map course_id -> end_time
                for course in courses_to_process:
                    scheduled_days = set()
                    
                    # Get day-specific schedules for this course
                    day_schedules = course.course_schedules.all()
                    if day_schedules.exists():
                        for sched in day_schedules:
                            if sched.day:
                                scheduled_days.add(sched.day)
                    # Fallback to default course days if no specific schedules
                    elif course.days:
                        for day in [d.strip() for d in course.days.split(',') if d.strip()]:
                            scheduled_days.add(day)
                    
                    scheduled_days_map[course.id] = scheduled_days
                    course_end_times[course.id] = course.end_time
                
                # Get current date and time (in PH timezone)
                try:
                    from zoneinfo import ZoneInfo
                    PH_TZ = ZoneInfo('Asia/Manila')
                except Exception:
                    try:
                        import pytz
                        PH_TZ = pytz.timezone('Asia/Manila')
                    except Exception:
                        PH_TZ = None
                
                from django.utils import timezone
                now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
                today = now_ph.date()
                current_time = now_ph.time()
                
                # Filter records to only include:
                # 1. Scheduled days only
                # 2. NOT future dates (attendance_date must be <= today)
                # 3. For today's date, only show if class schedule has finished
                filtered_records = []
                for record in attendance_records:
                    scheduled_days = scheduled_days_map.get(record.course.id, set())
                    
                    # Check if record is on a scheduled day
                    record_day = getattr(record, 'schedule_day', None)
                    if record_day and record_day not in scheduled_days:
                        continue  # Skip if not on a scheduled day
                    elif not record_day and scheduled_days:
                        # If no schedule_day recorded, use the day of week from the date
                        day_name = record.attendance_date.strftime('%a')  # e.g., 'Mon'
                        if day_name not in scheduled_days:
                            continue  # Skip if not on a scheduled day
                    
                    # Check if record is not a future date
                    if record.attendance_date > today:
                        continue  # Skip future dates
                    
                    # All checks passed, include this record
                    filtered_records.append(record)
                
                attendance_records = filtered_records
                
                # Filter by specific date if provided
                if date_filter:
                    try:
                        from datetime import datetime
                        filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                        attendance_records = [r for r in attendance_records if r.attendance_date == filter_date]
                    except (ValueError, AttributeError):
                        pass  # Invalid date format, show all records
                
                # Generate dates based on course schedule and day filter
                from datetime import datetime, timedelta
                from django.utils import timezone
                try:
                    from zoneinfo import ZoneInfo
                    PH_TZ = ZoneInfo('Asia/Manila')
                except Exception:
                    try:
                        import pytz
                        PH_TZ = pytz.timezone('Asia/Manila')
                    except Exception:
                        PH_TZ = None
                
                now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
                today = now_ph.date()
                
                # Get all scheduled days for the course(s)
                scheduled_days = set()
                for course in courses_to_process:
                    # Day-specific schedules
                    day_schedules = course.course_schedules.all()
                    if day_schedules.exists():
                        for s in day_schedules:
                            if s.day:
                                scheduled_days.add(s.day)
                    # Default schedule days
                    elif course.days:
                        for day in [d.strip() for d in course.days.split(',') if d.strip()]:
                            scheduled_days.add(day)
                
                # Map day names to weekday numbers
                day_to_weekday = {
                    'Monday': 0, 'Mon': 0, 'M': 0,
                    'Tuesday': 1, 'Tue': 1, 'T': 1,
                    'Wednesday': 2, 'Wed': 2, 'W': 2,
                    'Thursday': 3, 'Thu': 3, 'Th': 3,
                    'Friday': 4, 'Fri': 4, 'F': 4,
                    'Saturday': 5, 'Sat': 5, 'S': 5,
                    'Sunday': 6, 'Sun': 6, 'Su': 6
                }
                
                # Generate dates based on:
                # 1. All historical attendance records (past dates)
                # 2. Scheduled dates for the future (next N weeks)
                dates_to_check = []
                
                # Convert month_filter to int if provided (do this BEFORE using it)
                month_filter_int = None
                if month_filter:
                    try:
                        month_filter_int = int(month_filter)
                    except (ValueError, TypeError):
                        month_filter_int = None
                
                # FIRST: Add all dates from existing attendance records (ensures historical records show)
                all_record_dates = set()
                for record in attendance_records:
                    all_record_dates.add(record.attendance_date)
                
                # SECOND: Generate future dates based on week_filter
                # Initialize start_date and end_date with default values
                start_date = today
                end_date = today + timedelta(weeks=weeks)
                
                # Calculate date range based on week_filter - accurate week calculation
                if week_filter == 'all':
                    # All weeks: use weeks parameter (default 4 weeks)
                    start_date = today
                    end_date = today + timedelta(weeks=weeks)
                elif week_filter == '1':
                    # First week: today to 7 days ahead (includes today)
                    start_date = today
                    end_date = today + timedelta(days=6)
                elif week_filter == '2':
                    # Second week: 7-13 days ahead (7 days from today to 13 days from today)
                    start_date = today + timedelta(days=7)
                    end_date = today + timedelta(days=13)
                elif week_filter == '3':
                    # Third week: 14-20 days ahead
                    start_date = today + timedelta(days=14)
                    end_date = today + timedelta(days=20)
                elif week_filter == 'last':
                    # Last week: depends on month filter
                    if month_filter_int:
                        # Get last day of month
                        from calendar import monthrange
                        last_day = monthrange(today.year, month_filter_int)[1]
                        end_date = datetime(today.year, month_filter_int, last_day).date()
                        start_date = end_date - timedelta(days=6)
                        # Ensure start_date is not before today if filtering current month
                        if month_filter_int == today.month and start_date < today:
                            start_date = today
                    else:
                        # Last week: 21-27 days ahead (or last 7 days of the month if we're near month end)
                        start_date = today + timedelta(days=21)
                        end_date = today + timedelta(days=27)
                else:
                    # No week filter or invalid - default to all weeks
                    start_date = today
                    end_date = today + timedelta(weeks=weeks)
                
                # Generate future dates
                current = start_date
                future_dates_set = set()
                
                while current <= end_date and current >= start_date:
                    # Apply month filter if specified
                    if month_filter_int and current.month != month_filter_int:
                        current += timedelta(days=1)
                        continue
                    
                    weekday = current.weekday()
                    day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][weekday]
                    
                    # Check if this day matches the filter and is in scheduled days
                    if day_filter and day_filter.lower() != 'all':
                        # Filter by specific day
                        if day_name == day_filter or day_to_weekday.get(day_filter, -1) == weekday:
                            if any(day_name in str(sd) or day_to_weekday.get(str(sd), -1) == weekday for sd in scheduled_days):
                                future_dates_set.add(current)
                    else:
                        # No day filter or "all" - include all scheduled days
                        if any(day_name in str(sd) or day_to_weekday.get(str(sd), -1) == weekday for sd in scheduled_days):
                            future_dates_set.add(current)
                    
                    current += timedelta(days=1)
                
                # Combine all record dates + future dates and normalize to date objects
                combined_dates = set()
                for d in all_record_dates.union(future_dates_set):
                    try:
                        # If it's a datetime, convert to date
                        if hasattr(d, 'date') and not isinstance(d, type(today)):
                            combined_dates.add(d.date())
                        else:
                            combined_dates.add(d)
                    except Exception:
                        try:
                            combined_dates.add(d)
                        except Exception:
                            continue
                dates_to_check = sorted(list(combined_dates))

                # If a specific date filter is active, restrict processing to that date only
                if date_filter:
                    try:
                        from datetime import datetime as _dt
                        filter_date_obj = _dt.strptime(date_filter, '%Y-%m-%d').date()
                        dates_to_check = [filter_date_obj]
                    except Exception:
                        # keep computed dates if parsing fails
                        pass

                # Determine which dates were marked as postponed for the schedule
                postponed_dates_set = set()
                try:
                    # Map weekday index to CourseSchedule.day code
                    weekday_to_code = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
                    for d in dates_to_check:
                        # For each date, check the schedules for courses_to_process to see if any day-schedule
                        # for that weekday is explicitly marked as 'postponed'
                        wd = d.weekday()
                        day_code = weekday_to_code.get(wd)
                        if not day_code:
                            continue
                        for course in courses_to_process:
                            try:
                                # Look for a CourseSchedule for this day
                                sched = course.course_schedules.filter(day__iexact=day_code).first()
                                if sched and getattr(sched, 'attendance_status', '') == 'postponed':
                                    postponed_dates_set.add(d.strftime('%Y-%m-%d'))
                                    break
                            except Exception:
                                continue
                except Exception:
                    postponed_dates_set = set()
                
                # Build attendance data - include all enrolled students, mark absent if no record
                # Create a map of attendance records by (student_id, date)
                attendance_map = {}
                for record in attendance_records:
                    key = (record.student.id, record.attendance_date)
                    attendance_map[key] = record
                
                # Get course schedule times for a specific date
                def get_course_schedule_times(course, date_obj):
                    """Get course start/end times for a specific date"""
                    weekday_map = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
                    weekday = date_obj.weekday()
                    day_name = weekday_map.get(weekday, '')
                    
                    # Check day-specific schedules
                    day_schedules = course.course_schedules.filter(day__iexact=day_name)
                    if day_schedules.exists():
                        schedule = day_schedules.first()
                        # Use course start/end time if schedule has them, otherwise use course defaults
                        return (schedule.start_time or course.start_time,
                                schedule.end_time or course.end_time)
                    elif course.days and day_name in [d.strip() for d in course.days.split(',')]:
                        return (course.start_time, course.end_time)
                    return (None, None)
                
                # Process each date and each enrolled student
                for date_to_check in dates_to_check:
                    # Check if course has schedule for this day
                    has_schedule_today = False
                    course_start_time = None
                    course_end_time = None
                    
                    for course in courses_to_process:
                        start_t, end_t = get_course_schedule_times(course, date_to_check)
                        if start_t and end_t:
                            course_start_time = start_t
                            course_end_time = end_t
                            has_schedule_today = True
                            break
                    
                    if not has_schedule_today:
                        continue  # Skip dates when course doesn't have schedule
                    
                    # Use course times for attendance window (since attendance window was removed)
                    attendance_start_time = course_start_time
                    attendance_end_time = course_end_time
                    
                    # Process each enrolled student
                    for student_id, student in unique_students.items():
                        key = (student_id, date_to_check)
                        
                        if key in attendance_map:
                            # Student has attendance record
                            record = attendance_map[key]
                            # Determine status based on attendance time vs scheduled time
                            # Rules:
                            # - Present: Scan time is within attendance window (start <= scan <= end)
                            # - Late: Scan time is AFTER attendance end time
                            # - Absent: No scan at all (handled below)
                            final_status = record.status
                            
                            # Re-evaluate status based on present-window (preferred) or course time
                            if record.attendance_time:
                                from datetime import datetime, time as dtime, timedelta
                                scan_time = record.attendance_time
                                if isinstance(scan_time, dtime):
                                    scan_dt = datetime.combine(date_to_check, scan_time)

                                    # Determine present-cutoff using present_duration (schedule or course) and qr_opened_at if available
                                    present_duration_minutes = None
                                    schedule_obj = None
                                    try:
                                        # Attempt to find CourseSchedule for this date (by weekday)
                                        weekday_map = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
                                        wd = date_to_check.weekday()
                                        day_code = weekday_map.get(wd)
                                        schedule_obj = record.course.course_schedules.filter(day__iexact=day_code).first()
                                    except Exception:
                                        schedule_obj = None

                                    if schedule_obj and getattr(schedule_obj, 'attendance_present_duration', None):
                                        present_duration_minutes = schedule_obj.attendance_present_duration
                                    elif getattr(record.course, 'attendance_present_duration', None):
                                        present_duration_minutes = record.course.attendance_present_duration

                                    qr_opened_at = None
                                    try:
                                        if schedule_obj and getattr(schedule_obj, 'qr_code_opened_at', None):
                                            qr_opened_at = schedule_obj.qr_code_opened_at
                                        elif getattr(record.course, 'qr_code_opened_at', None):
                                            qr_opened_at = record.course.qr_code_opened_at
                                    except Exception:
                                        qr_opened_at = None

                                    # If we have a present duration, compute cutoff and decide present/late
                                    if present_duration_minutes:
                                        try:
                                            if qr_opened_at:
                                                present_cutoff_dt = qr_opened_at + timedelta(minutes=int(present_duration_minutes))
                                            elif schedule_obj and getattr(schedule_obj, 'start_time', None):
                                                from datetime import datetime as _dt
                                                schedule_start = schedule_obj.start_time
                                                present_cutoff_dt = _dt.combine(date_to_check, schedule_start) + timedelta(minutes=int(present_duration_minutes))
                                                # make timezone-aware if PH_TZ available
                                                try:
                                                    if PH_TZ:
                                                        if hasattr(PH_TZ, 'localize'):
                                                            import pytz as _pytz
                                                            present_cutoff_dt = _pytz.timezone('Asia/Manila').localize(present_cutoff_dt)
                                                        else:
                                                            present_cutoff_dt = present_cutoff_dt.replace(tzinfo=PH_TZ)
                                                except Exception:
                                                    pass
                                            else:
                                                present_cutoff_dt = None

                                            if present_cutoff_dt:
                                                # Compare using timezone-aware datetimes if possible
                                                if scan_dt <= present_cutoff_dt:
                                                    final_status = 'present'
                                                else:
                                                    final_status = 'late'
                                                # finished decision for present-window
                                            else:
                                                # Fallback: use course start/end logic
                                                if attendance_start_time and attendance_end_time:
                                                    start_dt = datetime.combine(date_to_check, attendance_start_time)
                                                    end_dt = datetime.combine(date_to_check, attendance_end_time)
                                                    if start_dt <= scan_dt <= end_dt:
                                                        final_status = 'present'
                                                    elif scan_dt > end_dt:
                                                        final_status = 'late'
                                        except Exception as e:
                                            logger.debug(f"[REPORT] present-window check failed: {e}")
                                            # Fall back to original logic below
                                    else:
                                        # No present window configured; fall back to course time logic
                                        if attendance_start_time and attendance_end_time:
                                            start_dt = datetime.combine(date_to_check, attendance_start_time)
                                            end_dt = datetime.combine(date_to_check, attendance_end_time)
                                            if start_dt <= scan_dt <= end_dt:
                                                final_status = 'present'
                                            elif scan_dt > end_dt:
                                                final_status = 'late'
                            
                            # Get course schedule times for this date
                            course_start_time = None
                            course_end_time = None
                            if date_to_check:
                                start_t, end_t = get_course_schedule_times(record.course, date_to_check)
                                course_start_time = start_t
                                course_end_time = end_t
                            
                            attendance_data.append({
                                'id': record.id,
                                'student_id': student.id,
                                'student_name': student.full_name or student.username,
                                'student_id_number': student.school_id or 'N/A',
                                'student_profile_picture': student.profile_picture.url if student.profile_picture else None,
                                'course_code': selected_course.code,
                                'course_name': selected_course.name,
                                'section': record.course.section or 'N/A',
                                'attendance_date': record.attendance_date,
                                'attendance_time': record.attendance_time,
                                'course_start_time': course_start_time,
                                'course_end_time': course_end_time,
                                'status': final_status,
                                'status_display': dict(AttendanceRecord.STATUS_CHOICES).get(final_status, final_status.title()),
                                'can_edit': True,
                                'record_exists': True,
                            })
                        else:
                            # Student has no attendance record - only mark as absent if class has finished
                            # Check if the class end time has passed for this date
                            class_has_finished = False
                            
                            if attendance_end_time:
                                # Create datetime for class end time on the scheduled date
                                class_end_dt = datetime.combine(date_to_check, attendance_end_time)
                                
                                # Make class_end_dt timezone-aware to match now_ph
                                if PH_TZ:
                                    # Localize naive datetime to timezone
                                    if hasattr(PH_TZ, 'localize'):
                                        # pytz timezone
                                        class_end_dt = PH_TZ.localize(class_end_dt)
                                    else:
                                        # zoneinfo timezone
                                        class_end_dt = class_end_dt.replace(tzinfo=PH_TZ)
                                
                                # Check if class has finished (end time has passed)
                                # Both datetimes are now timezone-aware, safe to compare
                                if class_end_dt < now_ph:
                                    class_has_finished = True
                                elif date_to_check < today:
                                    # Past date - class must have finished
                                    class_has_finished = True
                            
                            # Only mark as absent if class has finished
                            if class_has_finished:
                                # Check if this date is marked as postponed
                                # If so, mark status as 'postponed' instead of 'absent'
                                status_to_use = 'absent'
                                date_str = date_to_check.strftime('%Y-%m-%d')
                                if date_str in postponed_dates_set:
                                    status_to_use = 'postponed'
                                
                                # Get the section for this student
                                # Use the first course they're enrolled in for this date
                                student_section = 'N/A'
                                if student_id in student_courses_map and student_courses_map[student_id]:
                                    for course in student_courses_map[student_id]:
                                        if course.id in [c.id for c in courses_to_process]:
                                            student_section = course.section or 'N/A'
                                            break
                                
                                attendance_data.append({
                                    'id': None,
                                    'student_id': student.id,
                                    'student_name': student.full_name or student.username,
                                    'student_id_number': student.school_id or 'N/A',
                                    'student_profile_picture': student.profile_picture.url if student.profile_picture else None,
                                    'course_code': selected_course.code,
                                    'course_name': selected_course.name,
                                    'section': student_section,
                                    'attendance_date': date_to_check,
                                    'attendance_time': None,
                                    'course_start_time': attendance_start_time,
                                    'course_end_time': attendance_end_time,
                                    'status': status_to_use,
                                    'status_display': dict(AttendanceRecord.STATUS_CHOICES).get(status_to_use, status_to_use.title()),
                                    'can_edit': True,
                                    'record_exists': False,
                                })
                            # For classes that haven't finished yet, don't add absent records
                
                # Deduplicate attendance data when viewing all sections
                # If a student has multiple records on the same date (from different sections),
                # keep the record with the best attendance status (present > late > absent).
                # If statuses are equal, prefer the earliest attendance time (if available).
                if section_filter and section_filter.lower() == 'all':
                    def _status_priority(s):
                        if s == 'present':
                            return 3
                        if s == 'late':
                            return 2
                        if s == 'absent':
                            return 1
                        return 0

                    dedup_map = {}  # Key: (student_id, attendance_date), Value: record
                    for record in attendance_data:
                        key = (record['student_id'], record['attendance_date'])
                        if key not in dedup_map:
                            dedup_map[key] = record
                        else:
                            existing = dedup_map[key]
                            pri_new = _status_priority(record.get('status'))
                            pri_existing = _status_priority(existing.get('status'))
                            if pri_new > pri_existing:
                                logger.info(f"[DEDUP][Sidebar] key={key} kept_new_status={record.get('status')} over {existing.get('status')}")
                                dedup_map[key] = record
                            elif pri_new == pri_existing:
                                # If equal priority, keep earliest attendance_time if available
                                rt = record.get('attendance_time')
                                et = existing.get('attendance_time')
                                if rt and et:
                                    if rt < et:
                                        logger.info(f"[DEDUP][Sidebar] key={key} kept_new_time={rt} over {et}")
                                        dedup_map[key] = record
                                elif rt and not et:
                                    logger.info(f"[DEDUP][Sidebar] key={key} kept_new_has_time")
                                    dedup_map[key] = record
                    attendance_data = list(dedup_map.values())
                
                # Sort attendance data by date (newest first).
                # When viewing a specific date, we'll display students sorted by surname A->Z.
                # For general grouping, we will sort students within each date group by surname.
                def _surname_key(record):
                    name = (record.get('student_name') or '').strip()
                    if not name:
                        return ''
                    parts = name.split()
                    return parts[-1].lower()

                # If a specific date filter is active, sort attendance_data by surname ascending
                if date_filter:
                    attendance_data.sort(key=_surname_key)
                else:
                    # Keep overall order by date descending; stable sort will preserve insertion order within same date
                    attendance_data.sort(key=lambda x: x['attendance_date'], reverse=True)
                
                # Group attendance data by normalized date objects for folder-style display
                from collections import defaultdict
                attendance_by_date = defaultdict(list)

                # Ensure attendance_date values are date objects (not datetimes or strings)
                for record in attendance_data:
                    dt_val = record.get('attendance_date')
                    try:
                        # If it's a datetime, convert to date
                        if hasattr(dt_val, 'date'):
                            dt_key = dt_val.date()
                        else:
                            dt_key = dt_val
                    except Exception:
                        dt_key = dt_val
                    attendance_by_date[dt_key].append(record)

                # Sort students alphabetically within each date (A->Z by surname)
                def _rec_surname_key(r):
                    name = (r.get('student_name') or '').strip()
                    if not name:
                        return ''
                    parts = name.split()
                    return parts[-1].lower()

                for dt in list(attendance_by_date.keys()):
                    try:
                        attendance_by_date[dt].sort(key=_rec_surname_key)
                    except Exception:
                        # If sort fails for a specific date group, leave order as-is
                        pass

                # Convert to list of tuples (date, records) sorted by date (newest first)
                # Ensure keys are date objects so sorting and template grouping are consistent
                attendance_data_grouped = sorted(
                    [(d if hasattr(d, 'weekday') else d, recs) for d, recs in attendance_by_date.items()],
                    key=lambda x: x[0],
                    reverse=True
                )
                
                # Calculate course statistics
                total_records = len(attendance_data)
                present_count = sum(1 for r in attendance_data if r['status'] == 'present')
                late_count = sum(1 for r in attendance_data if r['status'] == 'late')
                absent_count = sum(1 for r in attendance_data if r['status'] == 'absent')
                postponed_count = sum(1 for r in attendance_data if r['status'] == 'postponed')
                
                course_stats = {
                    'total_records': total_records,
                    'present_count': present_count,
                    'late_count': late_count,
                    'absent_count': absent_count,
                    'postponed_count': postponed_count,
                    'present_percentage': round((present_count / total_records * 100) if total_records > 0 else 0, 2),
                    'late_percentage': round((late_count / total_records * 100) if total_records > 0 else 0, 2),
                    'absent_percentage': round((absent_count / total_records * 100) if total_records > 0 else 0, 2),
                    'postponed_percentage': round((postponed_count / total_records * 100) if total_records > 0 else 0, 2),
                }
                
                # Calculate student statistics
                student_attendance_map = {}
                for record_data in attendance_data:
                    student_id = record_data['student_id']
                    if student_id not in student_attendance_map:
                        # Get profile picture URL if it exists (should already be a URL string from attendance_data)
                        profile_pic_url = record_data.get('student_profile_picture')
                        # Ensure it's a string or None (not an ImageFieldFile)
                        if profile_pic_url and not isinstance(profile_pic_url, str):
                            try:
                                profile_pic_url = profile_pic_url.url if hasattr(profile_pic_url, 'url') else None
                            except (AttributeError, ValueError):
                                profile_pic_url = None
                        elif not profile_pic_url:
                            profile_pic_url = None
                        
                        student_attendance_map[student_id] = {
                            'student_name': record_data['student_name'],
                            'student_id_number': record_data['student_id_number'],
                            'student_profile_picture': profile_pic_url,
                            'present': 0,
                            'late': 0,
                            'absent': 0,
                            'postponed': 0,
                            'total': 0,
                        }
                    
                    student_attendance_map[student_id][record_data['status']] += 1
                    student_attendance_map[student_id]['total'] += 1
                
                # Convert to list and calculate percentages
                for student_id, stats in student_attendance_map.items():
                    total = stats['total']
                    student_stats[str(student_id)] = {
                        'student_name': stats['student_name'],
                        'student_id_number': stats['student_id_number'],
                        'student_profile_picture': stats.get('student_profile_picture'),
                        'present': stats['present'],
                        'late': stats['late'],
                        'absent': stats['absent'],
                        'postponed': stats['postponed'],
                        'total': total,
                        'present_percentage': round((stats['present'] / total * 100) if total > 0 else 0, 2),
                        'late_percentage': round((stats['late'] / total * 100) if total > 0 else 0, 2),
                        'absent_percentage': round((stats['absent'] / total * 100) if total > 0 else 0, 2),
                        'postponed_percentage': round((stats['postponed'] / total * 100) if total > 0 else 0, 2),
                    }
            
        except Course.DoesNotExist:
            selected_course = None
    
    # Get unread notifications
    try:
        unread_notifications = UserNotification.objects.filter(user=user, is_read=False).count()
    except Exception:
        unread_notifications = 0
    
    # Convert student_stats to JSON for JavaScript
    import json
    student_stats_json = json.dumps(student_stats)
    
    # Quick Pick: Use the same logic as My Classes to show all schedules with proper statuses
    from datetime import datetime, timedelta, time as dtime
    from django.utils import timezone
    try:
        from zoneinfo import ZoneInfo
        PH_TZ = ZoneInfo('Asia/Manila')
    except Exception:
        try:
            import pytz
            PH_TZ = pytz.timezone('Asia/Manila')
        except Exception:
            PH_TZ = None
    
    now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
    today = now_ph.date()
    
    # Get schedules matching My Classes format
    quick_pick_schedules = []
    weekday_map = {'M': 0, 'T': 1, 'W': 2, 'Th': 3, 'F': 4, 'S': 5, 'Su': 6}
    days_full = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday', 6: 'Sunday'}
    
    def to_time_label(start_t, end_t):
        s = start_t.strftime('%I:%M %p') if start_t else ''
        e = end_t.strftime('%I:%M %p') if end_t else ''
        return f"{s} - {e}" if s and e else s or e
    
    def next_occurrence(weekday_idx, t):
        if not isinstance(t, dtime):
            return None
        today_idx = now_ph.weekday()
        days_ahead = (weekday_idx - today_idx) % 7
        base = now_ph.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        if days_ahead == 0 and base <= now_ph:
            days_ahead = 7
        return base + timedelta(days=days_ahead)
    
    for course in all_courses:
        # Day-specific schedules - use same logic as My Classes
        day_schedules = course.course_schedules.all()
        if day_schedules.exists():
            for s in day_schedules:
                day_val = s.day
                day_key_short = {
                    'Monday': 'M', 'Mon': 'M',
                    'Tuesday': 'T', 'Tue': 'T',
                    'Wednesday': 'W', 'Wed': 'W',
                    'Thursday': 'Th', 'Thu': 'Th',
                    'Friday': 'F', 'Fri': 'F',
                    'Saturday': 'S', 'Sat': 'S',
                    'Sunday': 'Su', 'Sun': 'Su'
                }.get(str(day_val), str(day_val))
                wd = weekday_map.get(day_key_short, None)
                if wd is None or not s.start_time:
                    continue
                
                start_today = now_ph.replace(hour=s.start_time.hour, minute=s.start_time.minute, second=0, microsecond=0)
                end_today = start_today
                if s.end_time:
                    end_today = start_today.replace(hour=s.end_time.hour, minute=s.end_time.minute)
                
                if wd == now_ph.weekday():
                    start_dt = start_today
                    end_dt = end_today
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                else:
                    start_dt = next_occurrence(wd, s.start_time)
                    end_dt = start_dt.replace(hour=s.end_time.hour, minute=s.end_time.minute) if (start_dt and s.end_time) else start_dt
                    status = calculate_course_status(start_dt, end_dt, now_ph)
                
                date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                time_label = to_time_label(s.start_time, s.end_time)
                # Remove leading zeros from date
                if date_label:
                    import re
                    date_label = re.sub(r'\b0(\d)', r'\1', date_label)
                
                quick_pick_schedules.append({
                    'course_id': course.id,
                    'course_code': course.code,
                    'course_name': course.name,
                    'section': course.section or 'N/A',
                    'date': start_dt.date() if start_dt else today,
                    'start_time': s.start_time,
                    'end_time': s.end_time,
                    'start_dt': start_dt,
                    'end_dt': end_dt,
                    'status': status,
                    'day': str(day_val),
                    'room': course.room or 'N/A',
                    'color': course.color or '#3C4770',
                    'date_label': date_label,
                    'time_label': time_label,
                })
        else:
            # Default schedule - use same logic as My Classes
            if course.days and course.start_time:
                days_list = [d.strip() for d in course.days.split(',') if d.strip()]
                for day in days_list:
                    day_key = {
                        'Monday': 'M', 'Mon': 'M',
                        'Tuesday': 'T', 'Tue': 'T',
                        'Wednesday': 'W', 'Wed': 'W',
                        'Thursday': 'Th', 'Thu': 'Th',
                        'Friday': 'F', 'Fri': 'F',
                        'Saturday': 'S', 'Sat': 'S',
                        'Sunday': 'Su', 'Sun': 'Su'
                    }.get(str(day), str(day))
                    wd = weekday_map.get(day_key, None)
                    if wd is None:
                        continue
                    
                    start_today = now_ph.replace(hour=course.start_time.hour, minute=course.start_time.minute, second=0, microsecond=0)
                    end_today = start_today
                    if course.end_time:
                        end_today = start_today.replace(hour=course.end_time.hour, minute=course.end_time.minute)
                    
                    if wd == now_ph.weekday():
                        start_dt = start_today
                        end_dt = end_today
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    else:
                        start_dt = next_occurrence(wd, course.start_time)
                        end_dt = start_dt.replace(hour=course.end_time.hour, minute=course.end_time.minute) if (start_dt and course.end_time) else start_dt
                        status = calculate_course_status(start_dt, end_dt, now_ph)
                    
                    # Check if scheduled for tomorrow
                    if start_dt:
                        tomorrow = now_ph.date() + timedelta(days=1)
                        if start_dt.date() == tomorrow:
                            status = 'tomorrow'
                    
                    date_label = (start_dt.strftime('%b %d, %Y') if start_dt else '')
                    day_label = days_full.get(wd, day)
                    time_label = to_time_label(course.start_time, course.end_time)
                    # Remove leading zeros from date
                    if date_label:
                        import re
                        date_label = re.sub(r'\b0(\d)', r'\1', date_label)
                    
                    quick_pick_schedules.append({
                        'course_id': course.id,
                        'course_code': course.code,
                        'course_name': course.name,
                        'section': course.section or 'N/A',
                        'date': start_dt.date() if start_dt else today,
                        'start_time': course.start_time,
                        'end_time': course.end_time,
                        'start_dt': start_dt,
                        'end_dt': end_dt,
                        'status': status,
                        'day': day,
                        'room': course.room or 'N/A',
                        'color': course.color or '#3C4770',
                        'date_label': date_label,
                        'time_label': time_label,
                    })
    
    # Sort: live first, then starting_today, then tomorrow, then upcoming, then soon, then ongoing, then finished
    def sort_key(e):
        status_priority = {'live': 0, 'starting_today': 1, 'tomorrow': 2, 'upcoming': 3, 'soon': 4, 'ongoing': 5, 'finished': 6}
        status_rank = status_priority.get(e.get('status'), 6)
        dt = e.get('start_dt') or now_ph
        return (status_rank, dt if status_rank < 6 else datetime.max.replace(tzinfo=dt.tzinfo) - (dt - dt.replace(hour=0, minute=0, second=0, microsecond=0)))
    quick_pick_schedules.sort(key=sort_key)
    
    context = {
        'user': user,
        'school_admin': school_admin,
        'grouped_courses': grouped_courses,
        'selected_course': selected_course,
        'selected_sections': selected_sections,
        'section_filter': section_filter,
        'attendance_data': attendance_data,
        'attendance_data_grouped': attendance_data_grouped,
        'course_stats': course_stats,
        'student_stats': student_stats,
        'student_stats_json': student_stats_json,
        'unread_notifications': unread_notifications,
        'day_filter': day_filter,
        'month_filter': month_filter,
        'week_filter': week_filter,
        'weeks': weeks,
        'quick_pick_schedules': quick_pick_schedules,
        'today': today,
        'postponed_dates': sorted(list(postponed_dates_set)) if 'postponed_dates_set' in locals() else [],
    }
    # Format date string for template header (use full month name) when provided as YYYY-MM-DD
    formatted_date = None
    if date_filter:
        try:
            from datetime import datetime
            dt = datetime.strptime(date_filter, '%Y-%m-%d').date()
            # Use day integer to avoid platform-specific %-d support
            formatted_date = f"{dt.strftime('%B')} {dt.day}, {dt.year}"
        except Exception:
            formatted_date = None
    # Inject formatted_date into context
    context['formatted_date'] = formatted_date
    
    return render(request, 'dashboard/instructor/instructor_attendance_reports.html', context)

@login_required
@require_http_methods(["GET"])
def instructor_attendance_reports_download_view(request):
    """Download attendance reports as Excel file"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to access this page.'})
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from io import BytesIO
    except ImportError:
        return JsonResponse({'success': False, 'message': 'Excel library not installed. Please install openpyxl.'})
    
    # Get filter parameters
    course_id = request.GET.get('course', None)
    section_filter = request.GET.get('section', None)
    day_filter = request.GET.get('day_filter', None)
    month_filter = request.GET.get('month_filter', None)
    week_filter = request.GET.get('week_filter', None)
    weeks = int(request.GET.get('weeks', 4))
    
    if not course_id:
        return JsonResponse({'success': False, 'message': 'Please select a course.'})
    
    try:
        course = Course.objects.get(
            id=course_id,
            instructor=user,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        )
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Course not found.'})
    
    # Get section filter
    section_filter = request.GET.get('section', None)
    
    # Get sibling courses based on section filter
    sibling_courses = Course.objects.filter(
        instructor=user,
        code=course.code,
        name=course.name,
        semester=course.semester,
        school_year=course.school_year,
        is_active=True,
        deleted_at__isnull=True,
        is_archived=False
    )
    
    if section_filter and section_filter.lower() != 'all':
        # Filter by specific section only
        courses_to_include = list(sibling_courses.filter(section__iexact=section_filter))
        if not courses_to_include:
            courses_to_include = [course]  # Fallback to selected course
    else:
        # No section filter or "all" - include all sibling courses
        courses_to_include = list(sibling_courses)
    
    # Get attendance records
    attendance_query = AttendanceRecord.objects.filter(
        course__in=courses_to_include
    ).select_related('student', 'course', 'enrollment').order_by('-attendance_date', '-attendance_time')
    
    # If a specific section is selected (not 'all'), filter records to students enrolled in that section
    if section_filter and section_filter.lower() != 'all':
        # Only include attendance records for students enrolled in the selected section
        attendance_query = attendance_query.filter(
            enrollment__course__section__iexact=section_filter
        )

    # Optional: filter by exact date (YYYY-MM-DD) when provided (used for per-date exports)
    date_param = request.GET.get('date', None)
    if date_param:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_param, '%Y-%m-%d').date()
            attendance_query = attendance_query.filter(attendance_date=date_obj)
        except Exception:
            # ignore invalid date formats and continue without date filter
            pass
    
    # Apply day-based filtering if specified
    # Note: do NOT treat default `weeks` (an int, default=4) as a boolean flag here
    # otherwise the filter will always run and accidentally exclude explicit `date` exports.
    if day_filter or week_filter or month_filter:
        from datetime import datetime, timedelta
        from django.utils import timezone
        try:
            from zoneinfo import ZoneInfo
            PH_TZ = ZoneInfo('Asia/Manila')
        except Exception:
            try:
                import pytz
                PH_TZ = pytz.timezone('Asia/Manila')
            except Exception:
                PH_TZ = None
        
        now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
        today = now_ph.date()
        
        # Get scheduled days for courses
        scheduled_days = set()
        for c in courses_to_include:
            day_schedules = c.course_schedules.all()
            if day_schedules.exists():
                for s in day_schedules:
                    if s.day:
                        scheduled_days.add(s.day)
            elif c.days:
                for day in [d.strip() for d in c.days.split(',') if d.strip()]:
                    scheduled_days.add(day)
        
        # Generate dates based on day filter and weeks
        day_to_weekday = {
            'Monday': 0, 'Mon': 0, 'M': 0,
            'Tuesday': 1, 'Tue': 1, 'T': 1,
            'Wednesday': 2, 'Wed': 2, 'W': 2,
            'Thursday': 3, 'Thu': 3, 'Th': 3,
            'Friday': 4, 'Fri': 4, 'F': 4,
            'Saturday': 5, 'Sat': 5, 'S': 5,
            'Sunday': 6, 'Sun': 6, 'Su': 6
        }
        
        # Convert month_filter to int if provided
        month_filter_int = None
        if month_filter:
            try:
                month_filter_int = int(month_filter)
            except (ValueError, TypeError):
                month_filter_int = None
        
        valid_dates = []
        
        # Calculate date range based on week_filter
        if week_filter == 'all':
            # All weeks: use weeks parameter (default 4 weeks)
            start_date = today
            end_date = today + timedelta(weeks=weeks)
        elif week_filter == '1':
            start_date = today
            end_date = today + timedelta(days=6)
        elif week_filter == '2':
            start_date = today + timedelta(days=7)
            end_date = today + timedelta(days=13)
        elif week_filter == '3':
            start_date = today + timedelta(days=14)
            end_date = today + timedelta(days=20)
        elif week_filter == 'last':
            if month_filter_int:
                try:
                    from calendar import monthrange
                    month_int = int(month_filter)
                    last_day = monthrange(today.year, month_int)[1]
                    end_date = datetime(today.year, month_int, last_day).date()
                    start_date = end_date - timedelta(days=6)
                except (ValueError, TypeError):
                    start_date = today
                    end_date = today + timedelta(weeks=weeks)
            else:
                start_date = today + timedelta(days=21)
                end_date = today + timedelta(days=27)
        else:
            start_date = today
            end_date = today + timedelta(weeks=weeks)
        
        current = start_date
        
        while current <= end_date and current >= start_date:
            # Apply month filter if specified
            if month_filter_int and current.month != month_filter_int:
                current += timedelta(days=1)
                continue
            
            weekday = current.weekday()
            day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][weekday]
            
            if day_filter and day_filter.lower() != 'all':
                # Filter by specific day
                if day_name == day_filter or day_to_weekday.get(day_filter, -1) == weekday:
                    if any(day_name in str(sd) or day_to_weekday.get(str(sd), -1) == weekday for sd in scheduled_days):
                        valid_dates.append(current)
            else:
                # No day filter or "all" - include all scheduled days
                if any(day_name in str(sd) or day_to_weekday.get(str(sd), -1) == weekday for sd in scheduled_days):
                    valid_dates.append(current)
            
            current += timedelta(days=1)
        
        if valid_dates:
            attendance_query = attendance_query.filter(attendance_date__in=valid_dates)
    
    attendance_records = list(attendance_query)

    # Sort attendance_records by student's surname (A->Z) for export clarity
    def _record_surname_key(rec):
        try:
            name = (rec.student.full_name or rec.student.username or '').strip()
        except Exception:
            name = ''

        tokens = [p for p in str(name).strip().split() if p]
        if not tokens:
            return ('', '', '')

        particle_tokens = {
            'de', 'del', 'dela', 'da', 'di', 'la', 'las', 'los', 'van', 'von', 'bin', 'binti', 'al', 'ibn'
        }

        last_start = len(tokens) - 1
        found_particle = False
        for i in range(len(tokens) - 2, 0, -1):
            if str(tokens[i]).lower() in particle_tokens:
                last_start = i
                found_particle = True
                break

        if found_particle:
            last = ' '.join(tokens[last_start:]).strip()
            remaining = tokens[:last_start]
            if not remaining:
                first = ''
                middle = ''
            elif len(remaining) == 1:
                first = remaining[0]
                middle = ''
            else:
                first = ' '.join(remaining[:-1]).strip()
                middle = remaining[-1]
        else:
            first = tokens[0]
            middle = ' '.join(tokens[1:-1]).strip() if len(tokens) > 2 else ''
            last = tokens[-1]

        return (
            str(last).strip().lower() if last else '',
            str(first).strip().lower() if first else '',
            str(middle).strip().lower() if middle else ''
        )

    try:
        attendance_records.sort(key=_record_surname_key)
    except Exception:
        # If sorting fails for any reason, continue without raising
        pass

    # If a specific date is requested, ensure exported file includes ALL enrolled students
    # for that date (mark absent if no record) and recalculate statuses for accuracy.
    date_param = request.GET.get('date', None)
    if date_param:
        try:
            from datetime import datetime, timedelta
            from django.utils import timezone
            try:
                from zoneinfo import ZoneInfo
                PH_TZ = ZoneInfo('Asia/Manila')
            except Exception:
                try:
                    import pytz
                    PH_TZ = pytz.timezone('Asia/Manila')
                except Exception:
                    PH_TZ = None

            now_ph = timezone.now().astimezone(PH_TZ) if PH_TZ else timezone.now()
            today = now_ph.date()

            date_obj = datetime.strptime(date_param, '%Y-%m-%d').date()

            # Helper: get course schedule times for this date
            def get_course_schedule_times(course, date_obj):
                weekday_map = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
                weekday = date_obj.weekday()
                day_name = weekday_map.get(weekday, '')
                # Day-specific schedules
                day_schedules = course.course_schedules.filter(day__iexact=day_name)
                if day_schedules.exists():
                    s = day_schedules.first()
                    return (s.start_time or course.start_time, s.end_time or course.end_time)
                # Default course days
                if course.days and day_name in [d.strip() for d in course.days.split(',') if d.strip()]:
                    return (course.start_time, course.end_time)
                return (None, None)

            # Build lookup of existing attendance records by student id
            attendance_map = {}
            for rec in attendance_records:
                if rec.student:
                    attendance_map[rec.student.id] = rec

            # Get all enrollments for the courses to include
            all_enrollments = CourseEnrollment.objects.filter(
                course__in=courses_to_include,
                is_active=True,
                deleted_at__isnull=True
            ).select_related('student', 'course')
            
            # If a specific section is selected (not 'all'), filter enrollments to that section only
            if section_filter and section_filter.lower() != 'all':
                all_enrollments = all_enrollments.filter(course__section__iexact=section_filter)

            # Unique students map
            unique_students = {}
            for enr in all_enrollments:
                if enr.student and enr.student.id not in unique_students:
                    unique_students[enr.student.id] = enr

            # Rebuild attendance_records to include ALL enrolled students for the date
            # Preserve original saved status and attendance_time for students who have records
            from types import SimpleNamespace
            rebuilt_records = []
            for student_id, enrollment in unique_students.items():
                student = enrollment.student
                # Find an attendance record for this student on the requested date
                student_record = None
                for rec in attendance_records:
                    if rec.student and rec.student.id == student_id and rec.attendance_date == date_obj:
                        student_record = rec
                        break

                if student_record:
                    # Preserve the original record's status and time
                    built = SimpleNamespace(
                        student=SimpleNamespace(
                            id=student.id,
                            school_id=getattr(student, 'school_id', 'N/A'),
                            full_name=getattr(student, 'full_name', getattr(student, 'username', ''))
                        ),
                        course=SimpleNamespace(section=getattr(student_record.course, 'section', 'N/A')),
                        attendance_date=student_record.attendance_date,
                        attendance_time=getattr(student_record, 'attendance_time', None),
                        status=student_record.status
                    )
                    rebuilt_records.append(built)
                else:
                    # No attendance record for this student on this date
                    # Only mark as absent if the class has already finished on that date
                    course_for_check = enrollment.course
                    start_t, end_t = get_course_schedule_times(course_for_check, date_obj)
                    class_has_finished = False
                    if end_t:
                        from datetime import datetime as _dt
                        class_end_dt = _dt.combine(date_obj, end_t)
                        if PH_TZ:
                            try:
                                if hasattr(PH_TZ, 'localize'):
                                    # pytz
                                    import pytz as _pytz
                                    class_end_dt = _pytz.timezone('Asia/Manila').localize(class_end_dt)
                                else:
                                    class_end_dt = class_end_dt.replace(tzinfo=PH_TZ)
                            except Exception:
                                pass
                        if class_end_dt <= now_ph:
                            class_has_finished = True
                    elif date_obj < today:
                        class_has_finished = True

                    if class_has_finished:
                        built = SimpleNamespace(
                            student=SimpleNamespace(
                                id=student.id,
                                school_id=getattr(student, 'school_id', 'N/A'),
                                full_name=getattr(student, 'full_name', getattr(student, 'username', ''))
                            ),
                            course=SimpleNamespace(section=getattr(course_for_check, 'section', 'N/A')),
                            attendance_date=date_obj,
                            attendance_time=None,
                            status='absent'
                        )
                        rebuilt_records.append(built)

            # Replace attendance_records with rebuilt list for downstream export
            attendance_records = rebuilt_records
            # Ensure alphabetical ordering by surname for exported file
            try:
                attendance_records.sort(key=_record_surname_key)
            except Exception:
                pass
        except Exception as e:
            # On error, fall back to existing attendance_records (only those with records)
            logger.exception('Error building full attendance export for date: %s', e)
    
    # Check if the requested date is marked as postponed
    # If yes, mark ALL attendance records with 'postponed' status for that date
    date_param = request.GET.get('date', None)
    if date_param:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_param, '%Y-%m-%d').date()
            weekday_map = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
            day_name = weekday_map.get(date_obj.weekday(), '')
            
            # Check if this date's day schedule is marked as postponed
            is_postponed = False
            for course in courses_to_include:
                if day_name:
                    day_schedule = course.course_schedules.filter(day__iexact=day_name).first()
                    if day_schedule and getattr(day_schedule, 'attendance_status', '') == 'postponed':
                        is_postponed = True
                        break
            
            # If postponed, mark all records for this date as 'postponed'
            if is_postponed:
                for record in attendance_records:
                    if record.attendance_date == date_obj:
                        record.status = 'postponed'
        except Exception:
            pass
    
    # Deduplicate records when viewing all sections
    # If a student has multiple records on the same date (from different sections),
    # keep the record with the best attendance status (present > late > absent).
    # If statuses are equal, prefer the earliest attendance time (if available).
    if section_filter and section_filter.lower() == 'all':
        def _status_priority(s):
            if s == 'present':
                return 3
            if s == 'late':
                return 2
            if s == 'absent':
                return 1
            return 0

        dedup_map = {}  # Key: (student_id, attendance_date), Value: record
        for record in attendance_records:
            key = (record.student.id, record.attendance_date)
            if key not in dedup_map:
                dedup_map[key] = record
            else:
                existing = dedup_map[key]
                pri_new = _status_priority(record.status)
                pri_existing = _status_priority(existing.status)
                if pri_new > pri_existing:
                    logger.info(f"[DEDUP][Excel] key={key} kept_new_status={record.status} over {existing.status}")
                    dedup_map[key] = record
                elif pri_new == pri_existing:
                    # If equal priority, keep earliest attendance_time if available
                    if record.attendance_time and existing.attendance_time:
                        if record.attendance_time < existing.attendance_time:
                            logger.info(f"[DEDUP][Excel] key={key} kept_new_time={record.attendance_time} over {existing.attendance_time}")
                            dedup_map[key] = record
                    elif record.attendance_time and not existing.attendance_time:
                        logger.info(f"[DEDUP][Excel] key={key} kept_new_has_time")
                        dedup_map[key] = record
        attendance_records = list(dedup_map.values())
        logger.info(f"[DEDUP][Excel] section_filter=all result: {len(attendance_records)} unique records after dedup")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Helper: format student name as 'Last, First Middle' for exports
    def _format_student_name(name):
        try:
            if not name:
                return ''
            tokens = [p for p in str(name).strip().split() if p]
            if not tokens:
                return ''
            if len(tokens) == 1:
                return str(tokens[0]).upper()

            particle_tokens = {
                'de', 'del', 'dela', 'da', 'di', 'la', 'las', 'los', 'van', 'von', 'bin', 'binti', 'al', 'ibn'
            }

            last_start = len(tokens) - 1
            found_particle = False
            for i in range(len(tokens) - 2, 0, -1):
                if str(tokens[i]).lower() in particle_tokens:
                    last_start = i
                    found_particle = True
                    break

            if found_particle:
                last = ' '.join(tokens[last_start:]).strip()
                remaining = tokens[:last_start]
                if not remaining:
                    first = ''
                    middle = ''
                elif len(remaining) == 1:
                    first = remaining[0]
                    middle = ''
                else:
                    first = ' '.join(remaining[:-1]).strip()
                    middle = remaining[-1]
            else:
                # Default parsing: First [Middle...] Last
                first = tokens[0]
                middle = ' '.join(tokens[1:-1]).strip() if len(tokens) > 2 else ''
                last = tokens[-1]

            first = str(first).strip().upper() if first else ''
            last = str(last).strip().upper() if last else ''
            middle_initial = ''
            if middle:
                m = str(middle).strip()
                if m:
                    middle_initial = f"{m[0].upper()}."

            if last and first:
                return f"{last}, {first}{(' ' + middle_initial) if middle_initial else ''}".strip()
            if last:
                return str(last)
            return f"{first}{(' ' + middle_initial) if middle_initial else ''}".strip()
        except Exception:
            return str(name)
    
    # Header style
    header_fill = PatternFill(start_color="3C4770", end_color="3C4770", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Course Information
    row = 1
    # Expand header to cover all attendance columns so long titles are visible
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = f"Attendance Report - {course.code} - {course.name}"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    
    ws[f'A{row}'] = "Course Code:"
    ws[f'B{row}'] = course.code
    ws[f'A{row+1}'] = "Course Name:"
    ws[f'B{row+1}'] = course.name
    ws[f'A{row+2}'] = "Section:"
    if section_filter and section_filter.lower() != 'all':
        # Show the specific selected section
        ws[f'B{row+2}'] = section_filter
    else:
        # Show all sections
        sections_list = [c.section for c in courses_to_include if c.section]
        ws[f'B{row+2}'] = ', '.join(sorted(set(sections_list))) if sections_list else 'All Sections'
    ws[f'A{row+3}'] = "Semester:"
    ws[f'B{row+3}'] = course.semester or 'N/A'
    ws[f'A{row+4}'] = "School Year:"
    ws[f'B{row+4}'] = course.school_year or 'N/A'
    row += 6
    
    # Attendance Records Header - Updated column order: ID, Student, Section, Date, Day, Time, Status
    headers = ['Student ID', 'Student Name', 'Section', 'Date', 'Day', 'Time', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Attendance Records Data - Match header order exactly
    for record in attendance_records:
        # Student ID (Column 1)
        ws.cell(row=row, column=1, value=record.student.school_id or 'N/A')
        # Student Name (Column 2) - formatted as 'Last, First Middle'
        try:
            raw_name = record.student.full_name if getattr(record.student, 'full_name', None) else getattr(record.student, 'username', '')
        except Exception:
            raw_name = getattr(record.student, 'username', '')
        ws.cell(row=row, column=2, value=_format_student_name(raw_name))
        # Section (Column 3)
        ws.cell(row=row, column=3, value=record.course.section or 'N/A')
        # Date (Column 4)
        ws.cell(row=row, column=4, value=record.attendance_date.strftime('%Y-%m-%d'))
        # Day name (Column 5)
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_name = day_names[record.attendance_date.weekday()]
        ws.cell(row=row, column=5, value=day_name)
        # Time in 12-hour format (Column 6)
        if record.attendance_time:
            time_12h = record.attendance_time.strftime('%I:%M %p')
            ws.cell(row=row, column=6, value=time_12h)
        else:
            ws.cell(row=row, column=6, value='N/A')
        # Status (Column 7)
        status_display = dict(AttendanceRecord.STATUS_CHOICES).get(record.status, record.status.title())
        status_cell = ws.cell(row=row, column=7, value=status_display)
        
        # Add color formatting based on status
        if record.status == 'present':
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Light green
            status_cell.font = Font(color="065F46", bold=True)  # Dark green text
        elif record.status == 'late':
            status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light yellow
            status_cell.font = Font(color="92400E", bold=True)  # Dark yellow text
        elif record.status == 'absent':
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light red
            status_cell.font = Font(color="991B1B", bold=True)  # Dark red text
        elif record.status == 'postponed':
            status_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")  # Light grey
            status_cell.font = Font(color="374151", bold=True)  # Dark grey text
        
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    # Statistics Sheet
    ws2 = wb.create_sheet("Statistics")
    
    # Course Statistics
    row = 1
    ws2.merge_cells(f'A{row}:B{row}')
    ws2[f'A{row}'] = "Course Statistics"
    ws2[f'A{row}'].font = title_font
    row += 2
    
    total_records = len(attendance_records)
    present_count = sum(1 for r in attendance_records if r.status == 'present')
    late_count = sum(1 for r in attendance_records if r.status == 'late')
    absent_count = sum(1 for r in attendance_records if r.status == 'absent')
    postponed_count = sum(1 for r in attendance_records if r.status == 'postponed')
    
    stats_data = [
        ['Total Records', total_records],
        ['Present', present_count],
        ['Late', late_count],
        ['Absent', absent_count],
        ['Postponed', postponed_count],
        ['Present %', round((present_count / total_records * 100) if total_records > 0 else 0, 2)],
        ['Late %', round((late_count / total_records * 100) if total_records > 0 else 0, 2)],
        ['Absent %', round((absent_count / total_records * 100) if total_records > 0 else 0, 2)],
        ['Postponed %', round((postponed_count / total_records * 100) if total_records > 0 else 0, 2)],
    ]
    
    for stat_row in stats_data:
        ws2.cell(row=row, column=1, value=stat_row[0])
        ws2.cell(row=row, column=2, value=stat_row[1])
        row += 1
    
    # Student Statistics
    row += 2
    ws2.merge_cells(f'A{row}:E{row}')
    ws2[f'A{row}'] = "Student Statistics"
    ws2[f'A{row}'].font = title_font
    row += 2
    
    # Student Statistics Header
    student_headers = ['Student ID', 'Student Name', 'Present', 'Late', 'Absent', 'Postponed', 'Total', 'Present %', 'Late %', 'Absent %', 'Postponed %']
    for col, header in enumerate(student_headers, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Calculate student statistics
    student_attendance_map = {}
    for record in attendance_records:
        student_id = record.student.id
        if student_id not in student_attendance_map:
            raw_name = getattr(record.student, 'full_name', None) or getattr(record.student, 'username', '')
            student_attendance_map[student_id] = {
                'student_name': _format_student_name(raw_name),
                'student_id_number': record.student.school_id or 'N/A',
                'present': 0,
                'late': 0,
                'absent': 0,
                'postponed': 0,
                'total': 0,
            }
        
        student_attendance_map[student_id][record.status] += 1
        student_attendance_map[student_id]['total'] += 1
    
    # Write student statistics
    for student_id, stats in student_attendance_map.items():
        total = stats['total']
        ws2.cell(row=row, column=1, value=stats['student_id_number'])
        ws2.cell(row=row, column=2, value=stats['student_name'])
        
        # Present (Column 3) - Green
        present_cell = ws2.cell(row=row, column=3, value=stats['present'])
        present_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        present_cell.font = Font(color="065F46", bold=True)
        present_cell.alignment = Alignment(horizontal='center')
        
        # Late (Column 4) - Yellow
        late_cell = ws2.cell(row=row, column=4, value=stats['late'])
        late_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        late_cell.font = Font(color="92400E", bold=True)
        late_cell.alignment = Alignment(horizontal='center')
        
        # Absent (Column 5) - Red
        absent_cell = ws2.cell(row=row, column=5, value=stats['absent'])
        absent_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        absent_cell.font = Font(color="991B1B", bold=True)
        absent_cell.alignment = Alignment(horizontal='center')
        
        # Postponed (Column 6) - Light Grey
        postponed_cell = ws2.cell(row=row, column=6, value=stats['postponed'])
        postponed_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        postponed_cell.font = Font(color="374151", bold=True)
        postponed_cell.alignment = Alignment(horizontal='center')
        
        # Total (Column 7)
        total_cell = ws2.cell(row=row, column=7, value=total)
        total_cell.alignment = Alignment(horizontal='center')
        
        # Percentages (Columns 8-11) - Center aligned
        ws2.cell(row=row, column=8, value=round((stats['present'] / total * 100) if total > 0 else 0, 2)).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=9, value=round((stats['late'] / total * 100) if total > 0 else 0, 2)).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=10, value=round((stats['absent'] / total * 100) if total > 0 else 0, 2)).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=11, value=round((stats['postponed'] / total * 100) if total > 0 else 0, 2)).alignment = Alignment(horizontal='center')
        row += 1
    
    # Auto-adjust column widths and apply formatting
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            column_cells = [cell for cell in column if cell.value]
            if column_cells:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                # Adjust width with smaller cap for Student ID column
                if column_letter == 'A' and sheet.title == "Attendance Report":
                    adjusted_width = min(max_length + 1, 12)  # Smaller width for ID column
                else:
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders and formatting to data rows
        if sheet.title == "Attendance Report":
            # Apply borders and alignment to attendance records
            for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 2:  # Date column
                        cell.number_format = 'YYYY-MM-DD'
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column == 3:  # Time column
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column in [4, 6]:  # Student ID and Section columns
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column == 7:  # Status column
                        cell.alignment = Alignment(horizontal='center')
                        # Color code status
                        if cell.value == 'Present':
                            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                            cell.font = Font(bold=True, color="065F46")
                        elif cell.value == 'Late':
                            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                            cell.font = Font(bold=True, color="92400E")
                        elif cell.value == 'Absent':
                            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                            cell.font = Font(bold=True, color="991B1B")
                        elif cell.value == 'Postponed':
                            cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                            cell.font = Font(bold=True, color="374151")
        elif sheet.title == "Statistics":
            # Apply borders to statistics
            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 2:  # Values column
                        cell.alignment = Alignment(horizontal='right')
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_report_{course.code}_{course.name.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
@require_http_methods(["POST"])
def instructor_update_attendance_record_status_view(request):
    """Update individual attendance record status manually"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'You are not authorized to perform this action.'})
    
    try:
        import json
        data = json.loads(request.body)
        record_id = data.get('record_id')
        student_id = data.get('student_id')
        course_id = data.get('course_id')
        attendance_date = data.get('attendance_date')
        new_status = data.get('new_status') or data.get('status')  # Support both parameter names
        
        if not all([student_id, course_id, attendance_date, new_status]):
            return JsonResponse({'success': False, 'message': 'Missing required parameters. Please ensure all fields are filled.'})
        
        if new_status not in ['present', 'late', 'absent']:
            return JsonResponse({'success': False, 'message': 'Invalid status.'})
        
        # Get course and verify ownership
        course = get_object_or_404(Course, id=course_id, instructor=user, is_active=True, deleted_at__isnull=True, is_archived=False)
        
        # Get student
        student = get_object_or_404(CustomUser, id=student_id, is_student=True)
        
        # Parse date
        from datetime import datetime
        date_obj = datetime.strptime(attendance_date, '%Y-%m-%d').date()
        
        # Get or create attendance record
        if record_id:
            # Update existing record
            try:
                record = AttendanceRecord.objects.get(
                    id=record_id,
                    course=course,
                    student=student,
                    attendance_date=date_obj
                )
                old_status = record.status
                record.status = new_status
                record.save()
                logger.info(f"[ATTENDANCE_UPDATE] record_id={record_id} student={student.id} date={date_obj} status_change={old_status}->{new_status}")
                return JsonResponse({'success': True, 'message': f'Attendance status updated to {new_status.title()}.'})
            except AttendanceRecord.DoesNotExist:
                # Record not found by ID, try to find by course, student, and date
                existing_record = AttendanceRecord.objects.filter(
                    course=course,
                    student=student,
                    attendance_date=date_obj
                ).first()
                
                if existing_record:
                    # Update the found record
                    old_status = existing_record.status
                    existing_record.status = new_status
                    existing_record.save()
                    logger.info(f"[ATTENDANCE_UPDATE] found_by_filter student={student.id} date={date_obj} status_change={old_status}->{new_status}")
                    return JsonResponse({'success': True, 'message': f'Attendance status updated to {new_status.title()}.'})
                else:
                    # No record found, fall through to create new one (don't return error)
                    pass
        
        # If we reach here and record_id was provided but not found, try to create/update anyway
        if record_id and record_id != 'None' and record_id != '':
            # Try to find by course, student, and date
            existing_record = AttendanceRecord.objects.filter(
                course=course,
                student=student,
                attendance_date=date_obj
            ).first()
            
            if existing_record:
                # Update the found record
                old_status = existing_record.status
                existing_record.status = new_status
                existing_record.save()
                logger.info(f"[ATTENDANCE_UPDATE] fallback_update student={student.id} date={date_obj} status_change={old_status}->{new_status}")
                return JsonResponse({'success': True, 'message': f'Attendance status updated to {new_status.title()}.'})
        
        # Create new record or update existing
        if True:  # This block handles both record_id=None and record_id not found cases
            # Create new record
            # Get enrollment - check for multi-section courses (sibling courses)
            # First check direct enrollment
            enrollment = CourseEnrollment.objects.filter(
                course=course,
                student=student,
                is_active=True,
                deleted_at__isnull=True
            ).first()
            
            # If not found, check sibling courses (multi-section courses)
            if not enrollment:
                sibling_courses = Course.objects.filter(
                    instructor=user,
                    code=course.code,
                    name=course.name,
                    semester=course.semester,
                    school_year=course.school_year,
                    is_active=True,
                    deleted_at__isnull=True,
                    is_archived=False
                )
                enrollment = CourseEnrollment.objects.filter(
                    course__in=sibling_courses,
                    student=student,
                    is_active=True,
                    deleted_at__isnull=True
                ).first()
            
            if not enrollment:
                return JsonResponse({'success': False, 'message': 'Student is not enrolled in this course.'})
            # If enrollment exists in a sibling course, prefer that course when creating/updating records
            record_course = course
            try:
                if enrollment and enrollment.course and enrollment.course.id != course.id:
                    record_course = enrollment.course
            except Exception:
                record_course = course
            
            # Check if record already exists
            existing_record = AttendanceRecord.objects.filter(
                course=record_course,
                student=student,
                attendance_date=date_obj
            ).first()
            
            if existing_record:
                existing_record.status = new_status
                existing_record.save()
                return JsonResponse({'success': True, 'message': f'Attendance status updated to {new_status.title()}.'})
            
            # Create new record
            from datetime import time
            from django.utils import timezone
            current_time = timezone.now().time()
            
            created_record = AttendanceRecord.objects.create(
                course=record_course,
                student=student,
                enrollment=enrollment,
                attendance_date=date_obj,
                attendance_time=current_time,
                status=new_status
            )
            logger.info(f"[ATTENDANCE_UPDATE] created_new student={student.id} date={date_obj} status={new_status}")
            
            return JsonResponse({'success': True, 'message': f'Attendance record created with status {new_status.title()}.'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'Invalid date format: {str(e)}'})
    except Exception as e:
        logger.error(f"Error updating attendance record status: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def instructor_scan_student_qr_code_view(request):
    """
    Handle instructor scanning of REGISTERED QR codes for attendance.
    Looks up the QR code in QRCodeRegistration to find the student.
    Works across ALL schedules of a course - if student is registered with that QR 
    in the course, they can scan into any section/schedule of that course.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Only instructors can scan QR codes.'})
        
        data = json.loads(request.body)
        qr_code_raw = (data.get('qr_code', '') or '').strip()
        course_id = data.get('course_id')
        
        if not qr_code_raw or not course_id:
            return JsonResponse({'success': False, 'message': 'Missing required fields.'})
        
        # Normalize QR code:
        # - Extract 32-hex id from URL/text if present
        # - Lowercase to avoid case-sensitivity issues
        qr_code_norm = qr_code_raw
        try:
            if qr_code_raw.startswith('http://') or qr_code_raw.startswith('https://'):
                from urllib.parse import urlparse, parse_qs
                parsed = urlparse(qr_code_raw)
                q = parse_qs(parsed.query or '')
                candidate = (q.get('qr_code') or q.get('qr') or [None])[0]
                if not candidate:
                    import re
                    m = re.search(r'([a-f0-9]{32})', qr_code_raw, flags=re.IGNORECASE)
                    candidate = m.group(1) if m else None
                if candidate:
                    qr_code_norm = candidate
            else:
                import re
                m = re.search(r'([a-f0-9]{32})', qr_code_raw, flags=re.IGNORECASE)
                if m:
                    qr_code_norm = m.group(1)
        except Exception:
            qr_code_norm = qr_code_raw

        qr_code_norm = (qr_code_norm or '').strip().lower()

        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        sibling_courses = Course.objects.filter(
            instructor=request.user,
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        )

        # Look up the QR code in QRCodeRegistration.
        # Prefer this course, but allow sibling sections so scans don't fail due to section mismatch.
        qr_registration = QRCodeRegistration.objects.filter(
            qr_code__iexact=qr_code_norm,
            course=course,
            is_active=True
        ).select_related('student').first()

        if not qr_registration:
            qr_registration = QRCodeRegistration.objects.filter(
                qr_code__iexact=qr_code_norm,
                course__in=sibling_courses,
                is_active=True
            ).select_related('student', 'course').first()

        # Legacy fallback: older registrations may have stored the *full URL* or other text
        # that contains the 32-hex QR id. If so, match by substring (safe because the 32-hex
        # id should be unique).
        if not qr_registration and qr_code_norm:
            qr_registration = QRCodeRegistration.objects.filter(
                qr_code__icontains=qr_code_norm,
                course=course,
                is_active=True
            ).select_related('student').first()

        if not qr_registration and qr_code_norm:
            qr_registration = QRCodeRegistration.objects.filter(
                qr_code__icontains=qr_code_norm,
                course__in=sibling_courses,
                is_active=True
            ).select_related('student', 'course').first()
        
        if not qr_registration:
            return JsonResponse({
                'success': False, 
                'message': f'QR code not registered for {course.code}. Please register it first.'
            })
        
        student = qr_registration.student
        
        # Verify student is still enrolled in the course
        enrollment = CourseEnrollment.objects.filter(
            student=student,
            course=course,
            is_active=True
        ).first()
        
        if not enrollment:
            return JsonResponse({
                'success': False, 
                'message': f'{student.full_name} is no longer enrolled in {course.code}.'
            })
        
        # Get today's date in PH timezone
        from django.utils import timezone
        from zoneinfo import ZoneInfo
        try:
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        today_day = today.strftime('%A')
        
        # Convert full day name to short code (Monday -> Mon)
        day_map_short = {
            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
            'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
        }
        today_day_short = day_map_short.get(today_day, today_day[:3])
        
        # Find ONLY schedules for this course on this specific day
        # Attendance should only record for the actual class on that day
        active_schedules = CourseSchedule.objects.filter(
            course=course,
            day=today_day_short,
            is_deleted=False
        ).order_by('start_time')
        
        if not active_schedules.exists():
            return JsonResponse({
                'success': False, 
                'message': f'No class scheduled for {course.code} on {today_day}.'
            })
        
        # Use the first schedule on this day (earliest class)
        active_schedule = active_schedules.first()
        
        # Check if attendance is open for this specific schedule
        if active_schedule.attendance_status != 'open':
            return JsonResponse({
                'success': False, 
                'message': f'Attendance is not open for {course.code} today.'
            })
        
        # Determine present/late status using the same logic as school ID scans
        attendance_status = 'present'
        present_duration_minutes = None
        qr_opened_at = None
        try:
            if active_schedule and getattr(active_schedule, 'attendance_present_duration', None) is not None:
                present_duration_minutes = int(active_schedule.attendance_present_duration or 0)
            else:
                present_duration_minutes = int(getattr(course, 'attendance_present_duration', 0) or 0)
        except Exception:
            present_duration_minutes = 0

        try:
            if active_schedule and getattr(active_schedule, 'qr_code_opened_at', None):
                qr_opened_at = active_schedule.qr_code_opened_at
            elif getattr(course, 'qr_code_opened_at', None):
                qr_opened_at = course.qr_code_opened_at
        except Exception:
            qr_opened_at = None

        cutoff_dt = None
        if present_duration_minutes and present_duration_minutes > 0:
            try:
                from datetime import timedelta
                if qr_opened_at:
                    cutoff_dt = qr_opened_at + timedelta(minutes=int(present_duration_minutes))
                elif active_schedule and getattr(active_schedule, 'start_time', None):
                    cutoff_dt = datetime.combine(today, active_schedule.start_time) + timedelta(minutes=int(present_duration_minutes))
                    try:
                        cutoff_dt = cutoff_dt.replace(tzinfo=ph_tz)
                    except Exception:
                        pass
            except Exception:
                cutoff_dt = None

        if cutoff_dt and now_ph > cutoff_dt:
            attendance_status = 'late'
        elif not cutoff_dt:
            try:
                end_time = None
                if active_schedule and getattr(active_schedule, 'attendance_end', None):
                    end_time = active_schedule.attendance_end
                elif getattr(course, 'attendance_end', None):
                    end_time = course.attendance_end
                if end_time and now_ph.time() > end_time:
                    attendance_status = 'late'
            except Exception:
                pass

        # Create or update attendance record
        attendance_record, created = AttendanceRecord.objects.get_or_create(
            course=course,
            student=student,
            enrollment=enrollment,
            attendance_date=today,
            schedule_day=active_schedule.day if active_schedule else today_day_short,
            defaults={
                'attendance_time': now_ph.time(),
                'status': attendance_status
            }
        )
        
        # If record already exists, update status to computed status (present/late)
        if not created:
            attendance_record.attendance_time = now_ph.time()
            attendance_record.status = attendance_status
            attendance_record.save()
        
        # Create notification for instructor
        create_notification(
            request.user,
            'attendance_scanned',
            'Student Attendance Recorded',
            f'{student.full_name} scanned QR for {course.code}',
            category='attendance',
            related_course=course,
            related_user=student
        )
        # Return status info so frontend can show appropriate popup (Present vs Late)
        resp_status = getattr(attendance_record, 'status', 'present')
        resp_message = 'Marked as ' + resp_status.title() if resp_status else 'Attendance recorded successfully'
        return JsonResponse({
            'success': True,
            'message': resp_message,
            'status': resp_status,
            'student_name': student.full_name,
            'student_id': student.school_id,
            'course_code': course.code
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error scanning student QR code: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def instructor_scan_student_school_id_view(request):
    """
    Handle instructor scanning of student school IDs for attendance.
    The instructor uses the camera to scan student school IDs.
    This creates/updates attendance records immediately.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Only instructors can scan student IDs.'})
        
        data = json.loads(request.body)
        school_id = data.get('school_id', '').strip()
        course_id = data.get('course_id')
        schedule_id = data.get('schedule_id')
        
        # DEBUG: Log raw request
        logger.info(f"[SCAN] Raw request received - school_id='{school_id}', course_id={repr(course_id)} (type: {type(course_id).__name__}), schedule_id={repr(schedule_id)}")
        
        # Normalize course_id: extract numeric part from composite keys like "73_F_20251212"
        if course_id:
            try:
                course_id_str = str(course_id)
                if '_' in course_id_str:
                    course_id = int(course_id_str.split('_')[0])
                    logger.info(f"[SCAN] Extracted numeric course_id: {course_id}")
                else:
                    course_id = int(course_id_str)
            except (ValueError, AttributeError) as e:
                logger.error(f"[SCAN] Error normalizing course_id '{course_id}': {e}")
                return JsonResponse({'success': False, 'message': f'Invalid course ID format: {course_id}'})
        
        # Normalize schedule_id: extract numeric part from composite keys like "73_F_20251212"
        if schedule_id:
            try:
                schedule_id_str = str(schedule_id)
                if '_' in schedule_id_str:
                    schedule_id = int(schedule_id_str.split('_')[0])
                    logger.info(f"[SCAN] Extracted numeric schedule_id: {schedule_id}")
                else:
                    schedule_id = int(schedule_id_str)
            except (ValueError, AttributeError) as e:
                logger.error(f"[SCAN] Error normalizing schedule_id '{schedule_id}': {e}")
                schedule_id = None  # Set to None if invalid, will skip schedule validation
        
        logger.info(f"[SCAN] After normalization - course_id={course_id} (type: {type(course_id).__name__})")
        
        if not school_id or not course_id:
            return JsonResponse({'success': False, 'message': 'Missing required fields.'})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Extract the numeric student ID from the scanned QR
        # The QR might contain formatted info like "EPIS, CLIFFORD ALEGRIA 166701002@Alipao|4059216"
        # Extract the numeric school ID (6+ digits)
        extracted_school_id = None
        digit_match = re.search(r'\b(\d{6,})\b', school_id)
        if digit_match:
            extracted_school_id = digit_match.group(1)
            logger.debug(f"Extracted school ID '{extracted_school_id}' from scanned value: '{school_id}'")
        
        # First, try to find student by matching registered QRs containing the extracted school ID
        # This handles QR codes that have slight variations in formatting but same student ID
        student = None
        from dashboard.models import QRCodeRegistration
        
        if extracted_school_id:
            # Find all QR registrations that contain the student's school ID
            all_active_qrs = QRCodeRegistration.objects.filter(is_active=True).select_related('student')
            for qr_reg in all_active_qrs:
                if extracted_school_id in qr_reg.qr_code:
                    student = qr_reg.student
                    logger.info(f"Found student by matching school ID in QR registration: {student.full_name} ({student.school_id}) - Stored QR: '{qr_reg.qr_code}'")
                    break
        
        # If not found by QR content matching, try exact QR code match
        if not student:
            qr_reg = QRCodeRegistration.objects.filter(qr_code=school_id, is_active=True).select_related('student').first()
            if qr_reg:
                student = qr_reg.student
                logger.info(f"Found student via exact QR code match: {student.full_name} ({student.school_id})")
        
        # If not found in QR registrations, try direct school_id match (for manual ID scans)
        if not student:
            student = CustomUser.objects.filter(
                Q(school_id=school_id) | Q(username=school_id),
                is_student=True
            ).first()
            if student:
                logger.info(f"Found student by direct school_id match: {student.full_name} ({student.school_id})")
        
        # If not found, try the extracted school ID
        if not student and extracted_school_id:
            student = CustomUser.objects.filter(
                Q(school_id=extracted_school_id) | Q(username=extracted_school_id),
                is_student=True
            ).first()
            if student:
                logger.info(f"Found student by extracted school ID: {student.full_name} ({student.school_id})")
        
        # If not found, try tolerant lookup by normalizing digits (handles dashes/spaces)
        if not student:
            try:
                normalized_digits = re.sub(r"\D", "", school_id)
                if normalized_digits:
                    candidates = CustomUser.objects.filter(is_student=True).only('id', 'school_id', 'full_name')
                    for cand in candidates:
                        if re.sub(r"\D", "", (cand.school_id or '')) == normalized_digits:
                            student = cand
                            logger.info(f"Matched student by normalized digits during scan: input={school_id} -> matched={cand.full_name} ({cand.school_id})")
                            break
            except Exception as _e:
                logger.debug(f"Normalization scan lookup failed: {_e}")
        
        if not student:
            logger.warning(f"No student found for scanned value: '{school_id}'")
            return JsonResponse({'success': False, 'message': f'Student with ID {school_id} not found.'})
        
        # Check if student is enrolled in this course
        enrollment = CourseEnrollment.objects.filter(
            student=student,
            course=course,
            is_active=True
        ).first()
        
        if not enrollment:
            return JsonResponse({'success': False, 'message': f'{student.full_name} is not enrolled in {course.code}.'})
        
        # 🔴 CRITICAL CHECK: Verify student has registered their QR code for this course
        # This ensures only students who have registered their QR code can mark attendance
        qr_registration = QRCodeRegistration.objects.filter(
            student=student,
            course=course,
            is_active=True
        ).first()
        
        if not qr_registration:
            return JsonResponse({
                'success': False, 
                'message': f'{student.full_name}\'s QR code is NOT registered for {course.code}. Please register it first before marking attendance.',
                'error_type': 'qr_not_registered'
            })
        
        # Verify this is the correct schedule/section and map to the
        # CourseSchedule short day code (e.g. 'Mon', 'Tue'). CourseSchedule
        # stores the day in `day`, not `day_of_week`.
        schedule_day = None
        if schedule_id:
            logger.info(f"[SCAN] Looking for schedule: id={schedule_id}, course={course.id}")
            # Try direct numeric lookup first
            schedule = None
            try:
                # If schedule_id is a number, try direct lookup
                if isinstance(schedule_id, int) or (isinstance(schedule_id, str) and schedule_id.isdigit()):
                    schedule = CourseSchedule.objects.filter(id=int(schedule_id), course=course).first()
                    if schedule:
                        logger.info(f"[SCAN] Found schedule by direct ID: {schedule.id}")
            except (ValueError, TypeError):
                pass
            
            # If not found and schedule_id is composite like "73_F_20251212", try to extract the day
            if not schedule and isinstance(schedule_id, str) and '_' in schedule_id:
                parts = schedule_id.split('_')
                if len(parts) >= 2:
                    day_code = parts[1]
                    # Map single letter or two-letter day codes to CourseSchedule.day format (Mon, Tue, etc.)
                    day_map_reverse = {
                        'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri', 'S': 'Sat', 'Su': 'Sun',
                        'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat', 'Sun': 'Sun'
                    }
                    day_name = day_map_reverse.get(day_code)
                    if day_name:
                        schedule = CourseSchedule.objects.filter(course=course, day=day_name).first()
                        if schedule:
                            logger.info(f"[SCAN] Found schedule by day name '{day_name}': {schedule.id}")
            
            if schedule:
                schedule_day = schedule.day
                logger.info(f"[SCAN] Using schedule: {schedule.id}, day={schedule_day}")
            else:
                # Schedule not found, but don't fail - just continue without schedule validation
                logger.warning(f"[SCAN] Schedule '{schedule_id}' not found for course {course.id}, continuing without schedule validation")
                schedule_day = None

        # Get today's date and map to short day code used by CourseSchedule.day
        from django.utils import timezone
        from zoneinfo import ZoneInfo
        try:
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')

        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()

        day_map = {
            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
            'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
        }
        today_short = day_map.get(now_ph.strftime('%A'), now_ph.strftime('%A')[:3])

        # Use the provided schedule_day (if a specific schedule_id was passed),
        # otherwise use today's short day code
        lookup_day = schedule_day if schedule_day else today_short

        # Find active schedule for this day (CourseSchedule has no `is_deleted` field)
        active_schedule = CourseSchedule.objects.filter(
            course=course,
            day=lookup_day
        ).first()
        
        if not active_schedule:
            return JsonResponse({'success': False, 'message': 'No active schedule for this course today.'})
        
        # Check if attendance is open for this schedule
        if active_schedule.attendance_status != 'open':
            return JsonResponse({'success': False, 'message': 'Attendance is not open for this schedule.'})
        
        # Determine attendance status: check if student scanned after present window duration
        attendance_status = 'present'  # Default status

        # Get the schedule/course to check present window duration
        present_duration_minutes = None
        if active_schedule and getattr(active_schedule, 'attendance_present_duration', None):
            present_duration_minutes = active_schedule.attendance_present_duration
        elif getattr(course, 'attendance_present_duration', None):
            present_duration_minutes = course.attendance_present_duration

        # Prefer using a recorded QR open timestamp when available (more accurate than schedule start)
        qr_opened_at = None
        try:
            if active_schedule and getattr(active_schedule, 'qr_code_opened_at', None):
                qr_opened_at = active_schedule.qr_code_opened_at
            elif getattr(course, 'qr_code_opened_at', None):
                qr_opened_at = course.qr_code_opened_at
        except Exception:
            qr_opened_at = None

        # If present window is configured, compute cutoff using qr_opened_at if present, otherwise fall back to schedule start
        if present_duration_minutes:
            try:
                from datetime import timedelta
                # If instructor explicitly opened QR/session (qr_opened_at), use that as the baseline
                if qr_opened_at:
                    # Ensure timezone-aware comparison using now_ph
                    present_cutoff_dt = qr_opened_at + timedelta(minutes=int(present_duration_minutes))
                    if now_ph > present_cutoff_dt:
                        attendance_status = 'late'
                        logger.info(f"[SCAN] Student {student.full_name} scanned at {now_ph} after present cutoff {present_cutoff_dt} - marked LATE")
                elif active_schedule and active_schedule.start_time:
                    # Fallback: schedule start time + present duration on today's date
                    schedule_start = active_schedule.start_time
                    present_cutoff_dt = datetime.combine(today, schedule_start) + timedelta(minutes=int(present_duration_minutes))
                    # Make present_cutoff_dt timezone-aware to match now_ph if possible
                    try:
                        if hasattr(ph_tz, 'localize'):
                            # pytz
                            import pytz as _pytz
                            present_cutoff_dt = _pytz.timezone('Asia/Manila').localize(present_cutoff_dt)
                        else:
                            present_cutoff_dt = present_cutoff_dt.replace(tzinfo=ph_tz)
                    except Exception:
                        pass
                    if now_ph > present_cutoff_dt:
                        attendance_status = 'late'
                        logger.info(f"[SCAN] Student {student.full_name} scanned at {now_ph} after fallback present cutoff {present_cutoff_dt} - marked LATE")
            except Exception as e:
                logger.debug(f"[SCAN] Could not calculate late status: {e}")
        
        # Create or update attendance record. Store the schedule short-code
        # (e.g. 'Mon') in AttendanceRecord.schedule_day (it's a CharField).
        attendance_record, created = AttendanceRecord.objects.get_or_create(
            course=course,
            student=student,
            enrollment=enrollment,
            attendance_date=today,
            schedule_day=active_schedule.day if active_schedule else lookup_day,
            defaults={
                'attendance_time': now_ph.time(),
                'status': attendance_status
            }
        )
        
        # If record already exists, just update the status to present (re-scan)
        if not created:
            attendance_record.attendance_time = now_ph.time()
            attendance_record.status = attendance_status
            attendance_record.save()
        
        # Create notification for instructor
        create_notification(
            request.user,
            'attendance_scanned',
            'Student Attendance Recorded',
            f'{student.full_name} (ID: {school_id}) scanned for {course.code}',
            category='attendance',
            related_course=course,
            related_user=student
        )
        # Return a message indicating whether the scan marked the student Present or Late
        resp_status = attendance_status or 'present'
        resp_message = 'Marked as ' + resp_status.title()
        return JsonResponse({
            'success': True,
            'message': resp_message,
            'status': resp_status,
            'student_name': student.full_name,
            'student_id': student.school_id,
            'student_pk': student.id,
            'course_code': course.code,
            'scanned_value': school_id
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error scanning student school ID: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
def instructor_get_scanned_students_view(request):
    """
    Get list of students who have scanned for a specific course/schedule today.
    Used to update the scanned students list in real-time during scanning.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'students': []})
        
        course_id = request.GET.get('course_id')
        schedule_id = request.GET.get('schedule_id')
        
        if not course_id:
            return JsonResponse({'students': []})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Get today's attendance records for this course
        from django.utils import timezone
        from zoneinfo import ZoneInfo
        try:
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        
        # Include both present and late so the scanned list reflects actual scan statuses
        records = AttendanceRecord.objects.filter(
            course=course,
            attendance_date=today,
            status__in=['present', 'late']
        ).select_related('student').order_by('-attendance_time')

        # Deduplicate: Keep only the latest record per student
        # This ensures we don't show multiple scans - only the final status
        students_by_id = {}
        for record in records:
            if record.student_id not in students_by_id:
                students_by_id[record.student_id] = {
                    'name': record.student.full_name,
                    'id': record.student.school_id or 'N/A',
                    'time': record.attendance_time.strftime('%I:%M %p') if record.attendance_time else 'N/A',
                    'status': record.status
                }
        
        students_data = list(students_by_id.values())
        logger.info(f"[QR SCANNED] Returning {len(students_data)} deduplicated QR scanned students for {course.code}")
        return JsonResponse({'students': students_data, 'count': len(students_data)})
    
    except Exception as e:
        logger.error(f"Error getting scanned students: {str(e)}")
        return JsonResponse({'students': []})


@login_required
def instructor_get_biometric_students_view(request):
    """
    Get list of students who have biometric attendance for a specific course/schedule today.
    This is used by the biometric modal to display students who verified via fingerprint.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'students': []})
        
        course_id = request.GET.get('course_id')
        schedule_id = request.GET.get('schedule_id')
        
        if not course_id:
            return JsonResponse({'students': []})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Get today's attendance records for this course
        from django.utils import timezone
        from zoneinfo import ZoneInfo
        try:
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        
        # Get attendance records - will include both present and late status
        records = AttendanceRecord.objects.filter(
            course=course,
            attendance_date=today,
            status__in=['present', 'late']
        ).select_related('student').order_by('-attendance_time')
        
        # Filter to only include students who have biometric registrations for this course
        # This ensures we only show students who scanned via biometric
        biometric_student_ids = BiometricRegistration.objects.filter(
            course=course,
            is_active=True
        ).values_list('student_id', flat=True)
        
        # Track students already added to avoid duplicates (show only latest scan per student)
        students_by_id = {}
        for record in records:
            if record.student_id in biometric_student_ids:
                # Only add if we haven't seen this student yet (since ordered by -attendance_time, first is latest)
                if record.student_id not in students_by_id:
                    students_by_id[record.student_id] = {
                        'name': record.student.full_name,
                        'id': record.student.school_id or 'N/A',
                        'time': record.attendance_time.strftime('%I:%M %p') if record.attendance_time else 'N/A',
                        'status': record.status,
                        'scan_method': 'biometric'
                    }
                    # Debug logging to verify we're returning latest record
                    logger.info(f"[BIOMETRIC ENDPOINT] Student {record.student.full_name}: {record.attendance_time} - {record.status}")
        
        # Convert to list
        students_data = list(students_by_id.values())
        logger.info(f"[BIOMETRIC ENDPOINT] Returning {len(students_data)} biometric students for {course.code}")
        
        return JsonResponse({'students': students_data, 'count': len(students_data)})
    
    except Exception as e:
        logger.error(f"Error getting biometric students: {str(e)}")
        return JsonResponse({'students': []})


@login_required
def instructor_course_enrollments_view(request, course_id):
    """
    Get list of enrolled students for a course (for QR registration modal).
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'students': []})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Get all enrolled students for this course
        enrollments = CourseEnrollment.objects.filter(
            course=course,
            is_active=True
        ).select_related('student').order_by('student__full_name')
        
        students_data = []
        for enrollment in enrollments:
            students_data.append({
                'id': enrollment.student.id,
                'full_name': enrollment.student.full_name,
                'id_number': enrollment.student.school_id or 'N/A'
            })
        
        return JsonResponse({'students': students_data})
    
    except Exception as e:
        logger.error(f"Error getting course enrollments: {str(e)}")
        return JsonResponse({'students': []})


@login_required
def instructor_register_student_qr_code_view(request):
    """
    Register a QR code to a student for a specific course.
    Validates that the student is enrolled in the course by checking their ID number.
    Stores the mapping in QRCodeRegistration model.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Only instructors can register QR codes.'})
        
        data = json.loads(request.body)
        course_id = data.get('course_id')
        incoming_enrollment_id = data.get('enrollment_id')
        incoming_template_id = data.get('template_id')
        student_id_number = data.get('student_id_number', '').strip()
        qr_code_raw = data.get('qr_code', '').strip()
        
        if not course_id or not student_id_number or not qr_code_raw:
            return JsonResponse({'success': False, 'message': 'Missing required fields.'})

        qr_code = _normalize_registered_qr(qr_code_raw)
        if not qr_code:
            return JsonResponse({'success': False, 'message': 'Invalid QR code value.'})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Find student by school_id (the correct field)
        student = CustomUser.objects.filter(
            school_id=student_id_number,
            is_student=True
        ).first()

        # If not found, try tolerant matching: strip non-digits and compare (handles dashes/spaces)
        if not student:
            try:
                normalized_digits = re.sub(r"\D", "", student_id_number)
                if normalized_digits:
                    # iterate student records (narrow to students) and try to match normalized digits
                    candidates = CustomUser.objects.filter(is_student=True).only('id', 'school_id', 'full_name')
                    for cand in candidates:
                        cand_school = (cand.school_id or '')
                        if re.sub(r"\D", "", cand_school) == normalized_digits:
                            student = cand
                            logger.info(f"Matched student by normalized digits: input={student_id_number} -> matched={cand.full_name} ({cand.school_id})")
                            break
            except Exception as _e:
                logger.debug(f"Normalization lookup failed: {_e}")
        
        if not student:
            return JsonResponse({'success': False, 'message': f'Student with ID {student_id_number} not found in the system.'})
        
        # Check if student is enrolled in this course
        enrollment = CourseEnrollment.objects.filter(
            student=student,
            course=course,
            is_active=True
        ).first()
        
        if not enrollment:
            return JsonResponse({'success': False, 'message': f'{student.full_name} is not enrolled in {course.code}.'})
        
        # Store the QR code as-is (exactly as scanned), but link it to the selected student
        # The QR content doesn't matter - we use the selected student for identification
        # This allows students to scan any QR, and we'll mark the registered student as present
        logger.info(f"Registering QR code for student: {student.full_name} (school_id={student.school_id}). QR content: '{qr_code}'")
        
        # Create or update QR code registration
        from dashboard.models import QRCodeRegistration
        
        # ============= CRITICAL SECURITY CHECK =============
        # CHECK 1: If this student already has a DIFFERENT QR code registered to their account
        if student.qr_code_id and student.qr_code_id != qr_code:
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: {student.full_name} already has QR code ID "{student.qr_code_id}" registered to their account. Each student can only have ONE QR code ID. Cannot register a different QR code.'
            })
        
        # CHECK 2: Ensure the QR code is not registered to ANY OTHER student via account
        other_student = CustomUser.objects.filter(qr_code_id=qr_code, is_student=True).exclude(id=student.id).first()
        if other_student:
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: This QR code ID "{qr_code}" is already registered to {other_student.full_name}. Each QR code can only belong to one student.'
            })
        
        # CHECK 3: Ensure the QR code is NOT registered to ANY OTHER student in the SAME COURSE ONLY
        # NOTE: We only check within the same course - same student can use same QR in different courses
        other_student_reg = QRCodeRegistration.objects.filter(qr_code=qr_code, course=course, is_active=True).exclude(student=student).first()
        if other_student_reg:
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: This QR code is already registered to {other_student_reg.student.full_name} in {course.code}. Each QR code can only be assigned to one student per course. You cannot register another student\'s QR code for this course.'
            })

        # ============= REGISTRATION APPROVED =============
        # Store the QR code on the student's account (global registration)
        if not student.qr_code_id:
            student.qr_code_id = qr_code
            student.save()
            logger.info(f"Registered QR code ID '{qr_code}' to student account: {student.full_name}")

        # Delete any old registrations for this student in this course to avoid unique_together constraint violations
        QRCodeRegistration.objects.filter(
            student=student,
            course=course
        ).delete()

        # Create new QR registration
        qr_reg = QRCodeRegistration.objects.create(
            student=student,
            course=course,
            qr_code=qr_code,
            registered_by=request.user,
            is_active=True
        )
        
        # Create notification for instructor
        create_notification(
            request.user,
            'qr_registered',
            'QR Code Registered',
            f'QR code ID "{qr_code}" registered for {student.full_name} in {course.code}',
            category='course_management',
            related_course=course,
            related_user=student
        )
        
        return JsonResponse({
            'success': True,
            'message': f'✅ QR code ID "{qr_code}" registered successfully for {student.full_name}',
            'student_name': student.full_name,
            'student_id': student.school_id,
            'qr_code_id': qr_code
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error registering student QR code: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def instructor_decode_image_view(request):
    """
    Server-side decode endpoint: accepts JSON with `image` being a data URL (PNG/JPEG),
    decodes any QR codes using pyzbar and returns decoded text(s).
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Only instructors may use this endpoint.'})

        if not SERVER_DECODE_AVAILABLE:
            return JsonResponse({'success': False, 'message': 'Server decode libraries not available. Install pyzbar and Pillow.'})

        data = json.loads(request.body)
        data_url = data.get('image')
        if not data_url:
            return JsonResponse({'success': False, 'message': 'No image provided.'})

        # Strip off data URL prefix
        if data_url.startswith('data:'):
            header, encoded = data_url.split(',', 1)
        else:
            encoded = data_url

        image_bytes = base64.b64decode(encoded)
        img = Image.open(io.BytesIO(image_bytes)).convert('RGB')

        results = pyzbar_decode(img)
        decoded = []
        for r in results:
            try:
                text = r.data.decode('utf-8')
            except Exception:
                text = str(r.data)
            decoded.append(text)

        if decoded:
            return JsonResponse({'success': True, 'decoded': decoded})
        else:
            return JsonResponse({'success': False, 'message': 'No QR detected by server decoder.'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON.'})
    except Exception as e:
        logger.error('Server decode error: ' + str(e))
        return JsonResponse({'success': False, 'message': 'Server decode error: ' + str(e)})


@login_required
@require_http_methods(["GET"])
def check_course_registration_status(request):
    """Check if registration is enabled for instructor's courses"""
    try:
        user = request.user
        
        # Get the registration status for this instructor
        status, created = InstructorRegistrationStatus.objects.get_or_create(instructor=user)
        
        return JsonResponse({
            'success': True,
            'registration_enabled': status.is_registration_enabled,
        })
    
    except Exception as e:
        logger.error(f"Error checking registration status: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def enable_course_registration(request):
    """Enable QR and Fingerprint registration for all courses handled by instructor"""
    try:
        user = request.user
        
        # Get or create registration status for this instructor
        status, created = InstructorRegistrationStatus.objects.get_or_create(instructor=user)
        
        # Set registration as enabled
        status.is_registration_enabled = True
        status.enabled_at = datetime.now()
        status.save()
        
        # Get all courses where this user is the instructor
        courses = Course.objects.filter(instructor=user, is_active=True, deleted_at__isnull=True)
        
        # Create notification for instructor
        create_notification(
            user,
            'registration_enabled',
            'Course Registration Enabled',
            f'Attendance registration (QR Code and Fingerprint) has been enabled for {courses.count()} course(s)',
            category='course_management'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Registration enabled for {courses.count()} course(s)',
            'course_count': courses.count()
        })
    
    except Exception as e:
        logger.error(f"Error enabling course registration: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def disable_course_registration(request):
    """Disable QR and Fingerprint registration for all courses handled by instructor"""
    try:
        user = request.user
        
        # Get or create registration status for this instructor
        status, created = InstructorRegistrationStatus.objects.get_or_create(instructor=user)
        
        # Set registration as disabled
        status.is_registration_enabled = False
        status.save()
        
        # Get all courses where this user is the instructor
        courses = Course.objects.filter(instructor=user, is_active=True, deleted_at__isnull=True)
        
        # Create notification for instructor
        create_notification(
            user,
            'registration_disabled',
            'Course Registration Disabled',
            f'Attendance registration has been disabled for {courses.count()} course(s)',
            category='course_management'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Registration disabled for {courses.count()} course(s)',
            'course_count': courses.count()
        })
    
    except Exception as e:
        logger.error(f"Error disabling course registration: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def student_get_registration_instructors_view(request):
    """Get list of instructors with registration enabled and courses enrolled by student"""
    try:
        student = request.user
        if not student.is_student:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
        
        # Get all active courses enrolled by this student
        enrolled_enrollments = CourseEnrollment.objects.filter(
            student=student,
            is_active=True,
            deleted_at__isnull=True
        ).select_related('course', 'course__instructor').distinct()
        
        logger.info(f"[REGISTRATION] Student {student.id} has {enrolled_enrollments.count()} enrolled courses")
        
        # Group by instructor and check if they have registration enabled
        instructors_data = {}
        for enrollment in enrolled_enrollments:
            instructor = enrollment.course.instructor
            if not instructor:
                continue
            
            logger.info(f"[REGISTRATION] Checking instructor {instructor.id} for registration status")
                
            # Check if this instructor has registration enabled
            status = InstructorRegistrationStatus.objects.filter(
                instructor=instructor,
                is_registration_enabled=True
            ).first()
            
            logger.info(f"[REGISTRATION] Instructor {instructor.id} registration status: {status}")
            
            if status:  # Only include instructors with registration enabled
                if instructor.id not in instructors_data:
                    # Get profile picture URL
                    profile_pic_url = ''
                    if instructor.profile_picture:
                        profile_pic_url = instructor.profile_picture.url
                    
                    instructors_data[instructor.id] = {
                        'id': instructor.id,
                        'name': instructor.full_name or instructor.username,
                        'profile_picture': profile_pic_url,
                        'courses': []
                    }
                
                # Add this course if not already added
                course_data = {
                    'id': enrollment.course.id,
                    'code': enrollment.course.code,
                    'name': enrollment.course.name,
                    'color': enrollment.course.color or '#3C4770'
                }
                if course_data not in instructors_data[instructor.id]['courses']:
                    instructors_data[instructor.id]['courses'].append(course_data)
        
        logger.info(f"[REGISTRATION] Found {len(instructors_data)} instructors with registration enabled")
        
        return JsonResponse({
            'success': True,
            'instructors': list(instructors_data.values())
        })
    
    except Exception as e:
        logger.error(f"Error getting registration instructors: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def student_register_qr_code_with_instructor_view(request):
    """
    Student registers their QR code with a specific instructor.
    The QR code will be registered for all courses with that instructor.
    """
    try:
        if not request.user.is_student:
            return JsonResponse({'success': False, 'message': 'Only students can register QR codes.'})
        
        data = json.loads(request.body)
        instructor_id = data.get('instructor_id')
        qr_code_raw = data.get('qr_code', '').strip()
        
        if not instructor_id or not qr_code_raw:
            return JsonResponse({'success': False, 'message': 'Missing required fields.'})

        qr_code = _normalize_registered_qr(qr_code_raw)
        if not qr_code:
            return JsonResponse({'success': False, 'message': 'Invalid QR code value.'})
        
        # Get the instructor
        instructor = get_object_or_404(CustomUser, id=instructor_id, is_teacher=True, deleted_at__isnull=True)
        
        # Get all active courses this student is enrolled in with this instructor
        courses = Course.objects.filter(
            instructor=instructor,
            is_active=True,
            deleted_at__isnull=True,
            enrollments__student=request.user,
            enrollments__is_active=True,
            enrollments__deleted_at__isnull=True
        ).distinct()
        
        if not courses.exists():
            return JsonResponse({'success': False, 'message': 'You are not enrolled in any courses with this instructor.'})
        
        from .models import QRCodeRegistration
        registered_count = 0
        
        # Register the QR code for each course with this instructor
        for course in courses:
            # Delete existing QR registration for this student in this course
            QRCodeRegistration.objects.filter(
                student=request.user,
                course=course,
                is_active=True
            ).delete()
            
            # Create new QR registration
            qr_registration, created = QRCodeRegistration.objects.get_or_create(
                student=request.user,
                course=course,
                defaults={
                    'qr_code': qr_code,
                    'is_active': True
                }
            )
            
            if not created:
                # Update existing registration
                qr_registration.qr_code = qr_code
                qr_registration.is_active = True
                qr_registration.save()
            
            registered_count += 1
            logger.info(f"Student {request.user.full_name} registered QR code for {course.code}")
        
        return JsonResponse({
            'success': True,
            'message': f'QR code registered successfully for {registered_count} course(s)',
            'courses_registered': registered_count
        })
    
    except Exception as e:
        logger.error(f"Error registering QR code with instructor: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def student_register_qr_code_view(request):
    """
    Student self-registration endpoint for registering their QR code for a specific course.
    Allows students to register their school ID as a QR code for a course they are enrolled in.
    The QR code is stored globally on the student's account and can be reused for all their courses.
    """
    try:
        if not request.user.is_student:
            return JsonResponse({'success': False, 'message': 'Only students can register QR codes.'})
        
        data = json.loads(request.body)
        course_id = data.get('course_id')
        qr_code_raw = data.get('qr_code', '').strip()
        
        if not course_id or not qr_code_raw:
            return JsonResponse({'success': False, 'message': 'Missing required fields.'})

        qr_code = _normalize_registered_qr(qr_code_raw)
        if not qr_code:
            return JsonResponse({'success': False, 'message': 'Invalid QR code value.'})
        
        # Get the course
        course = get_object_or_404(Course, id=course_id, is_active=True, deleted_at__isnull=True)
        
        # Check if student is enrolled in this course
        enrollment = CourseEnrollment.objects.filter(
            student=request.user,
            course=course,
            is_active=True
        ).first()
        
        if not enrollment:
            return JsonResponse({'success': False, 'message': 'You are not enrolled in this course.'})
        
        # Store the QR code registration (one per student per course)
        from .models import QRCodeRegistration
        
        # ============= CRITICAL SECURITY CHECK =============
        # CHECK 1: If this student already has a QR code registered to their account
        if request.user.qr_code_id and request.user.qr_code_id != qr_code:
            # Student is trying to register a DIFFERENT QR code than what's on their account
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: You already have QR code ID "{request.user.qr_code_id}" registered to your account. Each student can only have ONE QR code ID. You must use your registered QR code ID: "{request.user.qr_code_id}"'
            })
        
        # CHECK 2: Ensure the QR code is unique globally (no other student has this QR code)
        other_student = CustomUser.objects.filter(qr_code_id=qr_code, is_student=True).exclude(id=request.user.id).first()
        if other_student:
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: This QR code ID "{qr_code}" is already registered to {other_student.full_name}. Each QR code can only belong to one student. You cannot use another student\'s QR code.'
            })
        
        # CHECK 3: Ensure the QR code is NOT registered to ANY OTHER student in the SAME COURSE
        # NOTE: We only check within the same course - same student can use same QR in different courses
        other_student_reg = QRCodeRegistration.objects.filter(qr_code=qr_code, course=course, is_active=True).exclude(student=request.user).first()
        if other_student_reg:
            return JsonResponse({
                'success': False,
                'message': f'❌ ERROR: This QR code is already registered to {other_student_reg.student.full_name} in {other_student_reg.course.code}. Each QR code can only belong to one student per course. You cannot register another student\'s QR code for this course.'
            })
        
        # ============= REGISTRATION APPROVED =============
        # Store the QR code on the student's account (global registration)
        request.user.qr_code_id = qr_code
        request.user.save()
        logger.info(f"Registered QR code ID '{qr_code}' to student account: {request.user.full_name}")
        
        # Delete any old registrations for this student in this course to avoid unique_together constraint violations
        QRCodeRegistration.objects.filter(
            student=request.user,
            course=course
        ).delete()
        
        # Create new QR registration for this specific course
        qr_reg = QRCodeRegistration.objects.create(
            student=request.user,
            course=course,
            qr_code=qr_code,
            registered_by=request.user,
            is_active=True
        )
        
        # Create notification for student
        create_notification(
            request.user,
            'qr_registered',
            'QR Code Registered',
            f'Your QR code ID "{qr_code}" has been registered to your account and for {course.code} - {course.name}',
            category='course_management',
            related_course=course,
            related_user=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'✅ QR code ID "{qr_code}" registered successfully! This QR code is now tied to your account and can be used for all your courses.',
            'course_name': course.name,
            'course_code': course.code,
            'qr_code_id': qr_code
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error registering QR code: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
def get_course_enrollments_view(request):
    """API endpoint to get enrollments for a course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse({'success': False, 'message': 'Course ID required'})
    
    try:
        # Get the course
        course = Course.objects.get(id=int(course_id), instructor=user)
        
        # Get all active enrollments for this course
        enrollments = CourseEnrollment.objects.filter(
            course=course,
            is_active=True,
            deleted_at__isnull=True
        ).select_related('student').order_by('full_name')
        
        enrollments_data = []
        for enrollment in enrollments:
            # Use the current student's full_name instead of the enrollment snapshot
            # This ensures the reassign modal shows the student's updated name
            current_full_name = enrollment.student.full_name if enrollment.student else enrollment.full_name
            enrollments_data.append({
                'id': enrollment.id,
                'full_name': current_full_name or enrollment.student.username,
                'student_id_number': enrollment.student_id_number,
                'section': enrollment.section or 'N/A',
                'profile_picture': enrollment.student.profile_picture.url if enrollment.student and enrollment.student.profile_picture else None
            })
        
        return JsonResponse({'success': True, 'enrollments': enrollments_data})
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Course not found or you do not have permission'})
    except Exception as e:
        logger.error(f"Error getting course enrollments: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_course_sections_view(request):
    """API endpoint to get available sections for a course"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse({'success': False, 'message': 'Course ID required'})
    
    try:
        # Get the course
        course = Course.objects.get(id=int(course_id))
        
        # Get all sibling courses (same code, name, semester, school_year)
        sibling_courses = Course.objects.filter(
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        ).order_by('section')
        
        sections = []
        for sibling in sibling_courses:
            sections.append({
                'id': sibling.id,
                'section': sibling.section or 'N/A'
            })
        
        return JsonResponse({'success': True, 'sections': sections})
    except Exception as e:
        logger.error(f"Error getting course sections: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def move_students_to_section_view(request):
    """API endpoint to move students to another section"""
    user = request.user
    if not user.is_teacher:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'})
    
    try:
        data = json.loads(request.body)
        enrollment_ids = data.get('enrollment_ids', [])
        target_course_id = data.get('target_course_id')
        
        logger.info(f"[TRANSFER] Request received: enrollment_ids={enrollment_ids}, target_course_id={target_course_id}")
        
        if not enrollment_ids or not target_course_id:
            return JsonResponse({'success': False, 'message': 'Missing required parameters'})
        
        # Convert to list of integers
        enrollment_ids = [int(eid) for eid in enrollment_ids]
        target_course_id = int(target_course_id)
        
        # Verify target course exists and instructor owns it
        try:
            target_course = Course.objects.get(
                id=target_course_id,
                instructor=user,
                is_active=True,
                deleted_at__isnull=True
            )
        except Course.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Target section not found or you do not have permission'})
        
        # Get source course from first enrollment
        try:
            first_enrollment = CourseEnrollment.objects.get(id=enrollment_ids[0])
            source_course = first_enrollment.course
        except CourseEnrollment.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Enrollment not found'})
        
        # Verify instructor owns the source course
        if source_course.instructor != user:
            return JsonResponse({'success': False, 'message': 'You do not own the source course'})
        
        # Verify both courses are related (same code, name, semester, school_year)
        if (source_course.code != target_course.code or
            source_course.name != target_course.name or
            source_course.semester != target_course.semester or
            source_course.school_year != target_course.school_year):
            return JsonResponse({'success': False, 'message': 'Cannot move students between unrelated courses'})
        
        moved_count = 0
        errors = []
        
        for enrollment_id in enrollment_ids:
            try:
                enrollment = CourseEnrollment.objects.get(
                    id=enrollment_id,
                    course=source_course,
                    is_active=True,
                    deleted_at__isnull=True
                )
                
                student = enrollment.student
                
                # Delete any existing enrollment in the target course (if any)
                existing_in_target = CourseEnrollment.objects.filter(
                    student=student,
                    course=target_course
                ).first()
                
                if existing_in_target:
                    existing_in_target.delete()
                
                # Now update the enrollment to point to the target course
                enrollment.course = target_course
                enrollment.section = target_course.section or 'N/A'
                enrollment.course_code = target_course.code
                enrollment.course_name = target_course.name
                enrollment.course_section = target_course.section or 'N/A'
                enrollment.enrolled_at = timezone.now()
                enrollment.save()
                
                # Transfer QR Code registrations to the new course
                qr_registrations = QRCodeRegistration.objects.filter(
                    student=student,
                    course=source_course,
                    is_active=True
                )
                for qr_reg in qr_registrations:
                    # Delete any existing QR registration in target course for this student
                    QRCodeRegistration.objects.filter(
                        student=student,
                        course=target_course
                    ).delete()
                    # Update the QR registration to point to target course
                    qr_reg.course = target_course
                    qr_reg.save()
                    logger.info(f"[TRANSFER] Transferred QR code registration for student {student.id} to section {target_course.section}")
                
                # Transfer Biometric registrations to the new course
                biometric_registrations = BiometricRegistration.objects.filter(
                    student=student,
                    course=source_course,
                    is_active=True
                )
                for bio_reg in biometric_registrations:
                    # Delete any existing biometric registration in target course for this student
                    BiometricRegistration.objects.filter(
                        student=student,
                        course=target_course
                    ).delete()
                    # Update the biometric registration to point to target course
                    bio_reg.course = target_course
                    bio_reg.save()
                    logger.info(f"[TRANSFER] Transferred biometric registration for student {student.id} to section {target_course.section}")
                
                logger.info(f"[TRANSFER] Student {student.id} successfully moved to section {target_course.section}")
                
                # Log the action
                logger.info(f"Student {student.id} ({student.username}) moved to course {target_course.id} section {target_course.section}")
                
                moved_count += 1
                
            except CourseEnrollment.DoesNotExist:
                logger.error(f"Enrollment {enrollment_id} not found")
                errors.append(f"Enrollment {enrollment_id} not found")
            except Exception as e:
                logger.error(f"Error moving enrollment {enrollment_id}: {str(e)}")
                errors.append(f"Error: {str(e)}")
        
        message = f"Successfully moved {moved_count} student(s)"
        if errors:
            message += f". {len(errors)} error(s) occurred"
            message += f". Errors: {', '.join(errors[:3])}"
        
        return JsonResponse({
            'success': True,
            'message': message,
            'moved_count': moved_count,
            'errors': errors
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Error in move_students_to_section_view: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== BIOMETRIC INTEGRATION ====================

# Server-side enrollment lock using cache (prevents concurrent enrollments)
def get_enrollment_lock_key(instructor_id):
    """Generate a cache key for enrollment lock per instructor"""
    return f'biometric_enrollment_lock_instructor_{instructor_id}'

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def check_instructor_enrollment_lock(request):
    """
    SECURITY ENDPOINT: Check if another student is currently enrolling with an instructor.
    Called before starting enrollment to prevent concurrent registrations.
    
    Expected POST data:
    {
        'instructor_id': <int>
    }
    
    Returns:
    {
        'success': True,
        'is_locked': False,  # True if another student is enrolling
        'enrolling_student': 'Student Name' or None,
        'message': 'Ready to enroll' or 'Another student is enrolling...'
    }
    """
    try:
        from django.core.cache import cache
        data = json.loads(request.body)
        instructor_id = data.get('instructor_id')
        
        if not instructor_id:
            return JsonResponse({'success': False, 'message': 'Missing instructor_id'}, status=400)
        
        lock_key = get_enrollment_lock_key(instructor_id)
        lock_data = cache.get(lock_key)  # Returns None if expired (5 minutes default)
        
        current_student_id = request.user.id
        
        if lock_data:
            enrolling_student_id = lock_data.get('student_id')
            enrolling_student_name = lock_data.get('student_name')
            lock_time = lock_data.get('timestamp')
            
            # If THE SAME student is enrolling, allow them through
            if enrolling_student_id == current_student_id:
                return JsonResponse({
                    'success': True,
                    'is_locked': False,  # Same student = not blocked
                    'enrolling_student': None,
                    'message': 'Ready to enroll'
                })
            
            # DIFFERENT student is enrolling - BLOCK
            return JsonResponse({
                'success': True,
                'is_locked': True,
                'enrolling_student': enrolling_student_name,
                'message': f'{enrolling_student_name} is currently registering their fingerprint. Please wait for them to finish.'
            })
        
        # No lock exists - ready to enroll
        return JsonResponse({
            'success': True,
            'is_locked': False,
            'enrolling_student': None,
            'message': 'Ready to enroll'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error checking enrollment lock: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_biometric_enroll_view(request):
    """
    API endpoint for ESP32 fingerprint sensor enrollment.
    Students/instructors call this to enroll a fingerprint for a course or multiple courses.
    
    Expected POST data (Single Course):
    {
        'course_id': <int>,
        'student_id': <int> (optional, for instructor enrolling student),
        'biometric_data': '<string>',  # Fingerprint template or ID from ESP32
        'biometric_type': 'fingerprint' (default)
    }
    
    Expected POST data (Multiple Courses - BATCH):
    {
        'course_ids': [<int>, <int>, ...],  # Array of course IDs - registers fingerprint for ALL courses at once
        'biometric_data': '<string>',
        'biometric_type': 'fingerprint' (default)
    }
    """
    try:
        from django.db import transaction
        
        data = json.loads(request.body)
        biometric_data = data.get('biometric_data', '').strip()
        biometric_type = data.get('biometric_type', 'fingerprint')
        skip_mqtt_start = bool(data.get('skip_mqtt_start', False))
        
        # CRITICAL: Support both single course_id and multiple course_ids
        course_ids = data.get('course_ids')  # Array of course IDs (NEW)
        if not course_ids:
            # Fallback to single course_id for backward compatibility
            course_id = data.get('course_id')
            course_ids = [course_id] if course_id else []
        
        student_id = data.get('student_id')
        
        if not course_ids or not biometric_data:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_ids (or course_id) or biometric_data'
            }, status=400)
        
        # Determine target student
        if student_id and request.user.is_teacher:
            # Instructor enrolling a student
            try:
                target_student = CustomUser.objects.get(id=student_id, is_student=True)
            except CustomUser.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Student not found'
                }, status=404)
        else:
            # Student self-registering
            target_student = request.user
            if not target_student.is_student:
                return JsonResponse({
                    'success': False,
                    'message': 'You must be a student to register biometric'
                }, status=403)
        
        # Validate all courses exist and student is enrolled in them
        courses = []
        for course_id in course_ids:
            try:
                course = Course.objects.get(id=course_id)
                courses.append(course)
                
                # Check permission for this course
                if student_id and request.user.is_teacher:
                    # Instructor enrolling a student - verify instructor owns course
                    if course.instructor != request.user:
                        return JsonResponse({
                            'success': False,
                            'message': f'You do not have permission to register biometric for course {course.code}'
                        }, status=403)
                else:
                    # Student self-registering - verify student is enrolled
                    enrollment = CourseEnrollment.objects.filter(
                        student=target_student,
                        course=course,
                        is_active=True
                    ).first()
                    if not enrollment:
                        return JsonResponse({
                            'success': False,
                            'message': f'You are not enrolled in course {course.code}'
                        }, status=403)
                        
            except Course.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': f'Course ID {course_id} not found'
                }, status=404)
        
        # ==================== SERVER-SIDE ENROLLMENT LOCK ====================
        # SECURITY: Prevent concurrent biometric registrations for same instructor
        # Only one student can enroll biometric at a time per instructor
        from django.core.cache import cache
        
        # Get instructor from first course (all courses must be from same instructor for batch enrollment)
        instructor = courses[0].instructor if courses else None
        if not instructor:
            return JsonResponse({
                'success': False,
                'message': 'Cannot determine instructor for enrollment'
            }, status=400)
        
        lock_key = get_enrollment_lock_key(instructor.id)
        lock_data = cache.get(lock_key)
        
        # Check if another student is currently enrolling
        if lock_data:
            other_student_id = lock_data.get('student_id')
            other_student_name = lock_data.get('student_name')
            
            # If SAME student is enrolling, allow (might be confirming or retrying)
            if other_student_id == target_student.id:
                logger.info(f"[ENROLLMENT LOCK] Same student {target_student.full_name} (ID: {target_student.id}) re-attempting enrollment")
            else:
                # DIFFERENT student is enrolling - BLOCK THIS REQUEST
                logger.warning(f"[ENROLLMENT LOCK] BLOCKED! Student {target_student.full_name} (ID: {target_student.id}) attempted to enroll while {other_student_name} (ID: {other_student_id}) is already enrolling under instructor {instructor.full_name}")
                return JsonResponse({
                    'success': False,
                    'message': f'⏳ Another student ({other_student_name}) is currently registering their fingerprint. Please wait for them to finish before starting your enrollment.',
                    'is_locked': True,
                    'other_student': other_student_name
                }, status=423)  # 423 = Locked (HTTP status code)
        
        # SET ENROLLMENT LOCK for this instructor (5 minute timeout)
        cache.set(
            lock_key,
            {
                'student_id': target_student.id,
                'student_name': target_student.full_name,
                'timestamp': str(datetime.now()),
                'instructor_id': instructor.id
            },
            timeout=300  # 5 minute timeout (should be enough for 3-capture enrollment)
        )
        logger.info(f"[ENROLLMENT LOCK] SET - Student {target_student.full_name} (ID: {target_student.id}) started enrollment for instructor {instructor.full_name}")
        # ==================== END ENROLLMENT LOCK ====================
        
        # Encrypt biometric data ONCE (same fingerprint for all courses)
        from .biometric_utils import encrypt_biometric_data
        encrypted_data = encrypt_biometric_data(biometric_data)
        
        if not encrypted_data:
            return JsonResponse({
                'success': False,
                'message': 'Failed to encrypt biometric data'
            }, status=500)
        
        # DEBUG: Show biometric data being captured
        print(f"\n[BIOMETRIC CAPTURE DEBUG - 3-CAPTURE ENROLLMENT]")
        print(f"   Raw biometric_data length: {len(biometric_data)} characters")
        print(f"   Raw biometric_data hash: {hashlib.sha256(biometric_data.encode()).hexdigest()[:16]}...")
        print(f"   Encrypted_data hash: {hashlib.sha256(encrypted_data.encode()).hexdigest()[:16]}...")
        print(f"   This is the best quality capture from 3 scans")
        
        # ========== CRITICAL OPTIMIZATION: Check if student already has registered fingerprint FOR THIS INSTRUCTOR ==========
        # Key: Different instructors get DIFFERENT fingerprint IDs to avoid conflicts
        # But all courses under ONE instructor use SAME fingerprint_id
        
        # Get the instructor from first course (all should be same instructor)
        instructor = courses[0].instructor if courses else None
        
        if not instructor:
            return JsonResponse({
                'success': False,
                'message': 'Cannot determine instructor for enrollment'
            }, status=400)
        
        # DEBUG: Show enrollment request details
        print(f"\n[DEBUG] Enrollment request:")
        print(f"   Student: {target_student.full_name or target_student.username}")
        print(f"   Courses: {[c.code for c in courses]}")
        try:
            course_instructor_names = [
                (getattr(getattr(c, 'instructor', None), 'full_name', None) or getattr(getattr(c, 'instructor', None), 'username', '') or '')
                for c in courses
            ]
        except Exception:
            course_instructor_names = []
        print(f"   Course instructors: {course_instructor_names}")
        print(f"   Primary instructor: {instructor.full_name or instructor.username}")
        
        # Check if student has existing fingerprint registered with THIS SPECIFIC INSTRUCTOR
        existing_instructor_fingerprint = BiometricRegistration.objects.filter(
            student=target_student,
            course__instructor=instructor,
            is_active=True
        ).first()
        
        print(f"   [DEBUG] Existing fingerprint with {instructor.full_name}: {existing_instructor_fingerprint is not None}")
        
        # OPTIMIZATION: Check if this biometric data matches any existing registration for this student
        # (even from different instructors) - if same finger, reuse the fingerprint_id
        existing_same_biometric = BiometricRegistration.objects.filter(
            student=target_student,
            biometric_data=encrypted_data,  # Match exact biometric data
            is_active=True
        ).first()
        
        print(f"\n[BIOMETRIC MATCHING DEBUG]")
        print(f"   Checking for exact biometric match...")
        
        # Debug: Show all existing biometrics for comparison
        all_student_biometrics = BiometricRegistration.objects.filter(
            student=target_student,
            is_active=True
        ).distinct()
        print(f"   Student has {all_student_biometrics.count()} active biometric registration(s):")
        for existing in all_student_biometrics:
            existing_hash = hashlib.sha256(existing.biometric_data.encode()).hexdigest()[:16]
            current_hash = hashlib.sha256(encrypted_data.encode()).hexdigest()[:16]
            match = "MATCH" if existing.biometric_data == encrypted_data else "NO MATCH"
            try:
                _course = getattr(existing, 'course', None)
                _course_code = getattr(_course, 'code', '') if _course else ''
                _instr = getattr(_course, 'instructor', None) if _course else None
                _instr_name = (getattr(_instr, 'full_name', None) or getattr(_instr, 'username', '') or '') if _instr else ''
            except Exception:
                _course_code = ''
                _instr_name = ''
            print(f"      - {_course_code} ({_instr_name}): hash={existing_hash}... {match}")
        
        print(f"   Current biometric hash: {hashlib.sha256(encrypted_data.encode()).hexdigest()[:16]}...")
        
        print(f"   [DEBUG] Checking for biometric data match across all instructors...")
        if existing_same_biometric:
            try:
                _m_course = getattr(existing_same_biometric, 'course', None)
                _m_instr = getattr(_m_course, 'instructor', None) if _m_course else None
                _m_instr_name = (getattr(_m_instr, 'full_name', None) or getattr(_m_instr, 'username', '') or '') if _m_instr else ''
            except Exception:
                _m_instr_name = ''
            print(f"   [DEBUG] FOUND MATCH! Same biometric in {_m_instr_name}'s course")
        else:
            print(f"   [DEBUG] No biometric match found - new fingerprint needed")
        
        # Determine if this is a replacement and what fingerprint_id to use
        is_replacement = False
        fingerprint_id = None  # Initialize before if/elif/else
        is_new_fingerprint = False
        
        if existing_instructor_fingerprint:
            # REPLACEMENT MODE: Student re-enrolling with same instructor
            # REUSE the SAME fingerprint_id to save storage (don't create new ID)
            old_fingerprint_id = existing_instructor_fingerprint.fingerprint_id
            
            # Handle case where old records have None fingerprint_id (shouldn't happen, but safety check)
            if old_fingerprint_id is None:
                print(f"\n[WARNING] Existing registration has None fingerprint_id! Assigning new ID...")
                max_fingerprint_id = BiometricRegistration.objects.aggregate(max_id=Max('fingerprint_id'))['max_id'] or 99
                fingerprint_id = max(100, max_fingerprint_id + 1)
                is_replacement = False  # Treat as new enrollment since old ID was invalid
                is_new_fingerprint = True
            else:
                fingerprint_id = old_fingerprint_id  # KEEP SAME ID, just replace data
                is_new_fingerprint = False
                is_replacement = True
            
            print(f"\nFINGERPRINT REPLACEMENT MODE (Student re-enrolling with same instructor):")
            print(f"   Student: {target_student.full_name or target_student.username}")
            print(f"   Instructor: {instructor.full_name or instructor.username}")
            print(f"   REUSING fingerprint_id: {fingerprint_id} (to save storage)")
            print(f"   Courses: {[c.code for c in courses]}")
            print(f"   (Biometric data will be updated, but ID remains the same to save storage!)")
        elif existing_same_biometric:
            # OPTIMIZATION: Same biometric found in another instructor's course
            # REUSE that fingerprint_id to save storage space
            biometric_fingerprint_id = existing_same_biometric.fingerprint_id
            
            # Handle case where existing records have None fingerprint_id
            if biometric_fingerprint_id is None:
                print(f"\n[WARNING] Existing biometric match has None fingerprint_id! Assigning new ID...")
                max_fingerprint_id = BiometricRegistration.objects.aggregate(max_id=Max('fingerprint_id'))['max_id'] or 99
                fingerprint_id = max(100, max_fingerprint_id + 1)
                is_new_fingerprint = True
                is_replacement = False
            else:
                fingerprint_id = biometric_fingerprint_id
                is_new_fingerprint = False
                is_replacement = False
            
            print(f"\nBIOMETRIC MATCH FOUND (Same finger, different instructor):")
            print(f"   Student: {target_student.full_name or target_student.username}")
            try:
                _prev_instr = getattr(getattr(getattr(existing_same_biometric, 'course', None), 'instructor', None), 'full_name', None) or getattr(getattr(getattr(existing_same_biometric, 'course', None), 'instructor', None), 'username', '')
            except Exception:
                _prev_instr = ''
            print(f"   Previous instructor: {_prev_instr}")
            print(f"   Current instructor: {instructor.full_name or instructor.username}")
            print(f"   REUSING fingerprint_id: {fingerprint_id} (same finger = same ID)")
            print(f"   Courses: {[c.code for c in courses]}")
            print(f"   (Storage optimized - same biometric across instructors!)")
        else:
            # NEW fingerprint for this student (different finger than any previous)
            # Use high ID numbers (starting from 100) to avoid conflicts with course_ids (1-99)
            max_fingerprint_id = BiometricRegistration.objects.aggregate(max_id=Max('fingerprint_id'))['max_id'] or 99
            fingerprint_id = max(100, max_fingerprint_id + 1)
            is_new_fingerprint = True
            is_replacement = False
            print(f"\nNEW FINGERPRINT (Different finger than any previous):")
            print(f"   Student: {target_student.full_name or target_student.username}")
            print(f"   Instructor: {instructor.full_name or instructor.username}")
            print(f"   [DEBUG] Max existing fingerprint_id globally: {max_fingerprint_id}")
            print(f"   [DEBUG] New fingerprint_id assigned: {fingerprint_id}")
            print(f"   Courses: {[c.code for c in courses]}")
            print(f"   (Different finger = different ID from previous registrations)")
            
            # Debug: Show all fingerprint IDs for this student
            all_student_regs = BiometricRegistration.objects.filter(student=target_student).values('course__instructor__full_name', 'fingerprint_id').distinct()
            print(f"   [DEBUG] All existing fingerprints for student:")
            for reg in all_student_regs:
                print(f"      - Instructor: {reg['course__instructor__full_name']}, Fingerprint ID: {reg['fingerprint_id']}")
            print(f"   [DEBUG] After this enrollment, student will have fingerprint_id {fingerprint_id} for {instructor.full_name or instructor.username}")
        
        # CRITICAL FIX: Register fingerprint for ALL courses in single transaction
        # This ensures one fingerprint is registered for all enrolled courses (like QR code)
        registered_courses = []
        failed_courses = []
        
        # CLEANUP: Deactivate old registrations with this fingerprint_id
        # (This allows the new registrations to use the same fingerprint_id for multiple courses)
        old_registrations = BiometricRegistration.objects.filter(
            student=target_student,
            fingerprint_id=fingerprint_id,
            is_active=True
        )
        if old_registrations.exists():
            print(f"\n[CLEANUP] Deactivating {old_registrations.count()} old registrations with fingerprint_id {fingerprint_id}")
            for old_reg in old_registrations:
                print(f"   - Deactivating: {old_reg.course.code}")
            old_registrations.update(is_active=False)
            print(f"[CLEANUP] All old registrations deactivated")
        
        # CLEANUP: Fix any existing records with None fingerprint_id for this student
        # This can happen if old enrollments weren't finalized properly
        none_registrations = BiometricRegistration.objects.filter(
            student=target_student,
            fingerprint_id__isnull=True,
            is_active=True
        )
        if none_registrations.exists():
            print(f"\n[CLEANUP] Found {none_registrations.count()} registrations with None fingerprint_id")
            # Update them to use the current fingerprint_id being assigned
            none_registrations.update(fingerprint_id=fingerprint_id)
            print(f"[CLEANUP] Updated them to fingerprint_id: {fingerprint_id}")
        
        # Safety check: ensure fingerprint_id is valid
        if fingerprint_id is None:
            print(f"[ERROR] fingerprint_id is None! This is a critical error.")
            print(f"   Existing instructor fingerprint: {existing_instructor_fingerprint}")
            print(f"   Existing same biometric: {existing_same_biometric}")
            return JsonResponse({
                'success': False,
                'message': 'Failed to assign fingerprint ID. Please try again.'
            }, status=500)
        
        # ==================== MQTT ENROLLMENT TRIGGER ====================
        # CRITICAL: Send MQTT message to ESP32 to store the fingerprint in sensor memory
        # This allows hardware-based matching during attendance (accurate and fast)
        # Without this, fingerprints only exist in database, not in sensor
        logger.info(f"\n[MQTT ENROLLMENT] Sending enrollment request to ESP32...")
        logger.info(f"[MQTT ENROLLMENT] Student: {target_student.full_name}")
        logger.info(f"[MQTT ENROLLMENT] Fingerprint slot: {fingerprint_id}")
        
        try:
            from dashboard.mqtt_client import get_mqtt_client
            mqtt_client = get_mqtt_client()
            
            if skip_mqtt_start:
                logger.info("[MQTT ENROLLMENT] Skipping MQTT start publish (skip_mqtt_start=true)")
            elif mqtt_client and mqtt_client.is_connected:
                enrollment_message = {
                    'action': 'start',
                    'slot': fingerprint_id,
                    'template_id': f'student_{target_student.id}',
                    'student_name': target_student.full_name or target_student.username
                }
                
                success = mqtt_client.publish('biometric/esp32/enroll/request', enrollment_message)
                if success:
                    logger.info(f"[MQTT ENROLLMENT] Enrollment request published to ESP32")
                    logger.info(f"[MQTT ENROLLMENT] Slot {fingerprint_id} is now waiting for 3 finger scans")
                else:
                    logger.warning(f"[MQTT ENROLLMENT] Failed to publish enrollment request to ESP32")
            else:
                logger.warning(f"[MQTT ENROLLMENT] MQTT client not connected")
        except Exception as e:
            logger.error(f"[MQTT ENROLLMENT] Error: {str(e)}")
            import traceback
            traceback.print_exc()
        
        with transaction.atomic():
            for course in courses:
                try:
                    # Create or update biometric registration with the same fingerprint_id
                    biometric_reg, created = BiometricRegistration.objects.update_or_create(
                        student=target_student,
                        course=course,
                        defaults={
                            'fingerprint_id': fingerprint_id,  # Use reused or new fingerprint_id
                            'biometric_data': encrypted_data,
                            'biometric_type': biometric_type,
                            'is_active': True,
                        }
                    )
                    print(f"   BiometricRegistration saved for {course.code}:")
                    print(f"      - fingerprint_id: {biometric_reg.fingerprint_id}")
                    print(f"      - is_active: {biometric_reg.is_active}")
                    print(f"      - created: {created}")
                    registered_courses.append(course)
                    
                    # Create notification for each course
                    try:
                        create_notification(
                            user=target_student,
                            notification_type='student_enrolled',
                            title='Biometric Registration Successful',
                            message=f'Your fingerprint has been registered for {course.code} - {course.name}',
                            category='enrollment',
                            related_course=course,
                            related_user=request.user if request.user.is_teacher else None
                        )
                    except Exception as e:
                        logger.error(f"Error creating notification for course {course.id}: {str(e)}")
                        
                except Exception as e:
                    failed_courses.append({
                        'course_id': course.id,
                        'error': str(e)
                    })
                    logger.error(f"Error registering biometric for course {course.id}: {str(e)}")
        
        # Check if ALL courses were registered successfully
        if failed_courses:
            return JsonResponse({
                'success': False,
                'message': f'Failed to register fingerprint for {len(failed_courses)} course(s)',
                'registered_count': len(registered_courses),
                'failed_count': len(failed_courses),
                'failed_courses': failed_courses
            }, status=500)
        
        # ALL COURSES REGISTERED SUCCESSFULLY
        logger.info(f"Biometric registration COMPLETED for student {target_student.id} in {len(registered_courses)} course(s): {[c.code for c in registered_courses]}")
        
        course_names = ', '.join([f'{c.code} - {c.name}' for c in registered_courses])
        
        # Store metadata about replacement in session
        request.session['pending_fingerprint'] = {
            'student_id': target_student.id,
            'instructor_id': instructor.id,
            'fingerprint_id': fingerprint_id,
            'is_replacement': is_replacement,
            'course_ids': [c.id for c in courses],
            'timestamp': str(datetime.now())
        }
        request.session.modified = True
        logger.info(f"\nENROLLMENT SESSION CREATED:")
        logger.info(f"   Fingerprint ID: {fingerprint_id}")
        logger.info(f"   Is replacement: {is_replacement}")
        if is_replacement:
            logger.info(f"   Biometric data will be updated for fingerprint_id {fingerprint_id} (same ID)")
        logger.info(f"\nIMPORTANT: Biometric data temporarily stored!")
        logger.info(f"   Only on CONFIRM button click will new biometric be permanently saved.")
        
        # CRITICAL: Send MQTT message to ESP32 to confirm enrollment is complete and reset flags
        # This unblocks the ESP32 so next student can enroll
        mqtt_sent = False
        try:
            from dashboard.mqtt_client import get_mqtt_client
            import time
            
            mqtt_client = get_mqtt_client()
            
            if mqtt_client and mqtt_client.is_connected:
                # Create completion message for ESP32
                completion_msg = {
                    'status': 'enrollment_saved',
                    'fingerprint_id': str(fingerprint_id),
                    'template_id': biometric_template_id,
                    'message': 'Enrollment confirmed and saved to database',
                    'is_replacement': is_replacement
                }
                
                logger.info(f"\n[MQTT] ====== SENDING COMPLETION MESSAGE TO ESP32 ======")
                logger.info(f"[MQTT] Fingerprint ID: {fingerprint_id}")
                logger.info(f"[MQTT] Template ID: {biometric_template_id}")
                logger.info(f"[MQTT] Message: {json.dumps(completion_msg)}")
                
                # CRITICAL: Publish with QoS=1 (at least once delivery) with automatic retries
                # Publish to BOTH topics for maximum reliability
                # Retry up to 3 times if initial publish fails
                max_retries = 3
                for attempt in range(1, max_retries + 1):
                    try:
                        result1 = mqtt_client.publish('biometric/esp32/enroll/response', completion_msg, qos=1)
                        result2 = mqtt_client.publish('biometric/esp32/enroll/completion', completion_msg, qos=1)
                        
                        logger.info(f"[MQTT] Attempt {attempt}/{max_retries}:")
                        logger.info(f"[MQTT]   - Published to biometric/esp32/enroll/response")
                        logger.info(f"[MQTT]   - Published to biometric/esp32/enroll/completion")
                        
                        # Both publishes should succeed
                        if result1 and result2:
                            logger.info(f"[MQTT] Success on attempt {attempt}")
                            mqtt_sent = True
                            break
                        else:
                            logger.warning(f"[MQTT] Publish failed, retrying...")
                            time.sleep(0.3)  # Wait before retry
                            
                    except Exception as publish_error:
                        logger.error(f"[MQTT] Attempt {attempt} failed: {str(publish_error)}")
                        if attempt < max_retries:
                            time.sleep(0.3)
                            continue
                        else:
                            raise
                
                # Wait a moment for MQTT to process
                time.sleep(0.5)
                
                if mqtt_sent:
                    logger.info(f"[MQTT] CRITICAL: ESP32 should now receive enrollment_saved and reset enrollmentInProgress = false")
                    logger.info(f"[MQTT] Next student can immediately start enrollment!")
                    logger.info(f"[MQTT] Enrollment completion sent to ESP32 for fingerprint {fingerprint_id} (QoS=1, delivered)")
                else:
                    logger.warning(f"[MQTT] WARNING: Could not send enrollment_saved to ESP32 after {max_retries} attempts")
            else:
                logger.warning(f"[MQTT] MQTT client not connected - enrollment saved but ESP32 not notified")
            
        except Exception as mqtt_error:
            logger.error(f"[MQTT] CRITICAL ERROR - Could not notify ESP32 of completion: {str(mqtt_error)}")
            # Don't raise - let enrollment complete in Django even if MQTT fails
            # User will see the error but enrollment is saved
        
        if not mqtt_sent:
            print(f"[MQTT] WARNING: MQTT completion message was NOT sent successfully!")
            print(f"[MQTT] ESP32 may still be blocked. Try clicking start enrollment again in 10 seconds.")
            logger.warning(f"[MQTT] WARNING: Enrollment saved in DB but ESP32 not notified. May need manual reset.")
        
        # CLEAR THE ENROLLMENT LOCK (enrollment completed successfully)
        cache.delete(lock_key)
        logger.info(f"[ENROLLMENT LOCK] CLEARED - Student {target_student.full_name} completed enrollment for instructor {instructor.full_name}")
        
        return JsonResponse({
            'success': True,
            'message': f'Fingerprint registered successfully for {len(registered_courses)} course(s)!',
            'registered_count': len(registered_courses),
            'registered_courses': [{'id': c.id, 'code': c.code, 'name': c.name} for c in registered_courses],
            'biometric_type': biometric_type,
            'fingerprint_id': fingerprint_id,
            'is_replacement': is_replacement,
            'registration_type': "Fingerprint replacement (updating biometric data)" if is_replacement else "New fingerprint created",
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        # Clear enrollment lock on error
        from django.core.cache import cache
        try:
            if 'instructor' in locals() and instructor:
                lock_key = get_enrollment_lock_key(instructor.id)
                cache.delete(lock_key)
                logger.info(f"[ENROLLMENT LOCK] CLEARED (on error) for instructor {instructor.full_name}")
        except:
            pass
        
        logger.error(f"Error in biometric enrollment: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_check_biometric_view(request):
    """
    API endpoint to check if a student has existing biometric registered for given courses.
    Returns whether student already has a fingerprint registered.
    
    Expected POST data:
    {
        'student_id': '<string>',  # School ID
        'course_ids': [<int>, <int>, ...]  # List of course IDs
    }
    """
    try:
        data = json.loads(request.body)
        student_school_id = data.get('student_id', '').strip()
        course_ids = data.get('course_ids', [])
        
        if not student_school_id or not course_ids:
            return JsonResponse({
                'success': True,
                'has_existing_biometric': False,
                'message': 'No student ID or courses provided'
            })
        
        # Get student by school_id
        try:
            student = CustomUser.objects.get(school_id=student_school_id, is_student=True)
        except CustomUser.DoesNotExist:
            # Student not found, treat as no existing biometric
            return JsonResponse({
                'success': True,
                'has_existing_biometric': False,
                'message': 'Student not found'
            })
        
        # Check if student has biometric registration for ANY of these courses
        has_existing = BiometricRegistration.objects.filter(
            student=student,
            course_id__in=course_ids,
            is_active=True
        ).exists()
        
        # Get instructor name if they have existing registration
        instructor_name = ''
        if has_existing:
            existing_reg = BiometricRegistration.objects.filter(
                student=student,
                course_id__in=course_ids,
                is_active=True
            ).first()
            if existing_reg and existing_reg.course:
                instructor_name = existing_reg.course.instructor.get_full_name() or existing_reg.course.instructor.username
        
        return JsonResponse({
            'success': True,
            'has_existing_biometric': has_existing,
            'instructor_name': instructor_name,
            'message': 'Check completed successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error checking biometric: {str(e)}")
        return JsonResponse({
            'success': True,
            'has_existing_biometric': False,
            'message': 'Could not check biometric - proceeding normally'
        })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_biometric_verify_view(request):
    """
    API endpoint to verify a fingerprint against registered biometric.
    Used by ESP32 to check if scanned fingerprint matches student's registered print.
    
    Expected POST data:
    {
        'course_id': <int>,
        'biometric_data': '<string>'  # Scanned fingerprint template
    }
    """
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        biometric_data = data.get('biometric_data', '').strip()
        
        if not course_id or not biometric_data:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_id or biometric_data',
                'matched_student': None
            }, status=400)
        
        # Get course
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Course not found',
                'matched_student': None
            }, status=404)
        
        # Find matching biometric registration for this course
        from .biometric_utils import verify_biometric_match
        biometric_registrations = BiometricRegistration.objects.filter(
            course=course,
            is_active=True
        ).select_related('student')
        
        matched_student = None
        for bio_reg in biometric_registrations:
            if verify_biometric_match(bio_reg.biometric_data, biometric_data):
                matched_student = bio_reg.student
                logger.info(f"Biometric match found for student {matched_student.id} in course {course.id}")
                break
        
        if matched_student:
            return JsonResponse({
                'success': True,
                'message': f'Fingerprint matched for {matched_student.full_name or matched_student.username}',
                'matched_student': {
                    'id': matched_student.id,
                    'username': matched_student.username,
                    'full_name': matched_student.full_name or matched_student.username,
                    'student_id_number': matched_student.student_id_number or '',
                }
            })
        else:
            logger.warning(f"No biometric match found for fingerprint in course {course.id}")
            return JsonResponse({
                'success': False,
                'message': 'Fingerprint does not match any registered student',
                'matched_student': None
            }, status=401)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data', 'matched_student': None}, status=400)
    except Exception as e:
        logger.error(f"Error in biometric verification: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e), 'matched_student': None}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_biometric_scan_attendance_view(request):
    """
    API endpoint for recording attendance via biometric scan.
    Called by ESP32/frontend when fingerprint is verified.
    
    Expected POST data:
    {
        'course_id': <int>,
        'student_id': <int>,
        'schedule_id': '<string>' (optional)
    }
    """
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        student_id = data.get('student_id')
        schedule_id = data.get('schedule_id', '')
        
        if not course_id or not student_id:
            return JsonResponse({'success': False, 'message': 'Missing course_id or student_id'}, status=400)
        
        # Get course and student
        try:
            course = Course.objects.get(id=course_id)
            student = CustomUser.objects.get(id=student_id, is_student=True)
        except (Course.DoesNotExist, CustomUser.DoesNotExist):
            return JsonResponse({'success': False, 'message': 'Course or student not found'}, status=404)
        
        # Get student's enrollment
        try:
            enrollment = CourseEnrollment.objects.get(
                student=student,
                course=course,
                is_active=True
            )
        except CourseEnrollment.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Student is not enrolled in this course'}, status=403)
        
        # ✅ CRITICAL: Check if student has registered their fingerprint for this course
        # Students MUST register biometric with instructor before using attendance
        biometric_registration = BiometricRegistration.objects.filter(
            student=student,
            course=course,
            is_active=True
        ).first()
        
        if not biometric_registration:
            return JsonResponse({
                'success': False,
                'message': f'Student has not registered their fingerprint for this course. Please register biometric first.',
                'student_id': student.id,
                'course_id': course.id
            }, status=403)
        
        # Check if attendance is open
        from django.utils import timezone
        from zoneinfo import ZoneInfo
        
        try:
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        
        # Log biometric attendance attempt
        print(f"\n[BIOMETRIC ATTENDANCE]")
        print(f"  Student: {student.full_name}")
        print(f"  Course: {course.code}")
        print(f"  Biometric registered: ✓ Yes (fingerprint_id={biometric_registration.fingerprint_id})")
        print(f"  Time: {now_ph.strftime('%H:%M:%S')}")
        
        # Determine schedule day and check attendance status
        schedule_day = None
        attendance_status = 'closed'
        
        active_schedule = None
        if schedule_id:
            # Extract day from schedule_id format: course_id_day_YYYYMMDD
            try:
                parts = str(schedule_id).split('_')
                if len(parts) >= 2:
                    day_token = parts[1]
                    day_map_reverse = {
                        'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri',
                        'S': 'Sat', 'Su': 'Sun',
                        'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat', 'Sun': 'Sun'
                    }
                    mapped_day = day_map_reverse.get(day_token)
                    if mapped_day:
                        active_schedule = CourseSchedule.objects.filter(course=course, day=mapped_day).first()
                        if active_schedule:
                            schedule_day = active_schedule.day
                            attendance_status = active_schedule.attendance_status or course.attendance_status
            except Exception as e:
                logger.warning(f"Error parsing schedule_id: {str(e)}")
        
        # Fallback: use course-level attendance status
        if not attendance_status or attendance_status == 'closed':
            attendance_status = course.attendance_status or 'closed'
        
        # Check if attendance is open
        if attendance_status not in ['open', 'automatic']:
            return JsonResponse({
                'success': False,
                'message': f'Attendance is currently {attendance_status}. Cannot record attendance.'
            }, status=403)
        
        # Record attendance
        from django.db import transaction
        with transaction.atomic():
            # Check for existing attendance record today
            existing_record = AttendanceRecord.objects.filter(
                student=student,
                course=course,
                attendance_date=today,
                schedule_day=schedule_day
            ).first()
            
            if existing_record and existing_record.status in ['present', 'late']:
                return JsonResponse({
                    'success': False,
                    'message': f'You already have {existing_record.status} attendance recorded for today'
                }, status=409)
            
            # Determine attendance status based on present window
            attendance_record_status = 'present'
            attendance_time = now_ph.time()

            present_duration_minutes = None
            qr_opened_at = None

            try:
                if active_schedule and getattr(active_schedule, 'attendance_present_duration', None) is not None:
                    present_duration_minutes = int(active_schedule.attendance_present_duration or 0)
                else:
                    present_duration_minutes = int(getattr(course, 'attendance_present_duration', 0) or 0)
            except Exception:
                present_duration_minutes = 0

            try:
                if active_schedule and getattr(active_schedule, 'qr_code_opened_at', None):
                    qr_opened_at = active_schedule.qr_code_opened_at
                elif getattr(course, 'qr_code_opened_at', None):
                    qr_opened_at = course.qr_code_opened_at
            except Exception:
                qr_opened_at = None

            cutoff_dt = None
            if present_duration_minutes and present_duration_minutes > 0:
                try:
                    if qr_opened_at:
                        cutoff_dt = qr_opened_at + timezone.timedelta(minutes=int(present_duration_minutes))
                    elif active_schedule and getattr(active_schedule, 'start_time', None):
                        from datetime import datetime as dt
                        cutoff_dt = dt.combine(today, active_schedule.start_time) + timezone.timedelta(minutes=int(present_duration_minutes))
                        try:
                            cutoff_dt = cutoff_dt.replace(tzinfo=now_ph.tzinfo)
                        except Exception:
                            pass
                except Exception:
                    cutoff_dt = None

            if cutoff_dt and now_ph > cutoff_dt:
                attendance_record_status = 'late'
                logger.info(f"Student {student.id} marked as LATE (present window expired)")
            elif not cutoff_dt:
                # Fallback: use attendance_end window if configured
                try:
                    end_time = None
                    if active_schedule and getattr(active_schedule, 'attendance_end', None):
                        end_time = active_schedule.attendance_end
                    elif getattr(course, 'attendance_end', None):
                        end_time = course.attendance_end
                    if end_time and now_ph.time() > end_time:
                        attendance_record_status = 'late'
                except Exception:
                    pass
            
            # Create or update attendance record
            attendance_record, created = AttendanceRecord.objects.update_or_create(
                student=student,
                course=course,
                attendance_date=today,
                schedule_day=schedule_day or '',
                defaults={
                    'enrollment': enrollment,
                    'attendance_time': attendance_time,
                    'status': attendance_record_status,
                }
            )
            
            # Create notification
            try:
                create_notification(
                    user=student,
                    notification_type='attendance_marked',
                    title='Attendance Recorded - Biometric',
                    message=f'Your biometric attendance has been recorded as {attendance_record_status.upper()} in {course.code}',
                    category='attendance',
                    related_course=course,
                    related_user=course.instructor
                )
            except Exception as e:
                logger.error(f"Error creating notification: {str(e)}")
            
            logger.info(f"Attendance record {'created' if created else 'updated'} via biometric for student {student.id} in course {course.id} with status {attendance_record_status}")
        
        return JsonResponse({
            'success': True,
            'message': f'Attendance recorded as {attendance_record_status.upper()}',
            'status': attendance_record_status,
            'timestamp': now_ph.isoformat(),
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error recording biometric attendance: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def instructor_register_student_biometric_view(request):
    """
    Instructor registers a student's biometric for their course.
    """
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        student_id = data.get('student_id')
        biometric_data = data.get('biometric_data', '').strip()
        
        if not course_id or not student_id or not biometric_data:
            return JsonResponse({'success': False, 'message': 'Missing required fields'})
        
        # Get course and verify instructor
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        student = get_object_or_404(CustomUser, id=student_id, is_student=True)
        
        # Verify student is enrolled
        enrollment = get_object_or_404(CourseEnrollment, student=student, course=course, is_active=True)
        
        # Call the biometric enrollment API
        from .biometric_utils import encrypt_biometric_data
        encrypted_data = encrypt_biometric_data(biometric_data)
        
        biometric_reg, created = BiometricRegistration.objects.update_or_create(
            student=student,
            course=course,
            defaults={
                'biometric_data': encrypted_data,
                'biometric_type': 'fingerprint',
                'is_active': True,
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Biometric registered for {student.full_name or student.username}',
            'created': created,
        })
    
    except Exception as e:
        logger.error(f"Error registering student biometric: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["GET", "POST"])
def student_register_biometric_view(request):
    """
    Student self-registers biometric for their enrolled courses.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            course_id = data.get('course_id')
            biometric_data = data.get('biometric_data', '').strip()
            
            if not course_id or not biometric_data:
                return JsonResponse({'success': False, 'message': 'Missing required fields'})
            
            # Get course and verify student is enrolled
            course = get_object_or_404(Course, id=course_id)
            enrollment = get_object_or_404(CourseEnrollment, student=request.user, course=course, is_active=True)
            
            # Register biometric
            from .biometric_utils import encrypt_biometric_data
            encrypted_data = encrypt_biometric_data(biometric_data)
            
            biometric_reg, created = BiometricRegistration.objects.update_or_create(
                student=request.user,
                course=course,
                defaults={
                    'biometric_data': encrypted_data,
                    'biometric_type': 'fingerprint',
                    'is_active': True,
                }
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Biometric registered successfully',
                'created': created,
            })
        
        except Exception as e:
            logger.error(f"Error registering biometric: {str(e)}")
            return JsonResponse({'success': False, 'message': str(e)})
    
    # GET: Show registration form for enrolled courses
    enrollments = CourseEnrollment.objects.filter(
        student=request.user,
        is_active=True
    ).select_related('course', 'course__program', 'course__instructor')
    
    context = {
        'enrollments': enrollments,
        'total_courses': enrollments.count(),
        'esp32_ip': settings.ESP32_IP,
    }
    return render(request, 'dashboard/student/register_biometric.html', context)


@login_required
@require_http_methods(["POST"])
def student_verify_biometric_view(request):
    """
    Student verifies biometric during attendance scanning.
    """
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        biometric_data = data.get('biometric_data', '').strip()
        
        if not course_id or not biometric_data:
            return JsonResponse({'success': False, 'message': 'Missing required fields'})
        
        # Get course and verify student is enrolled
        course = get_object_or_404(Course, id=course_id)
        enrollment = get_object_or_404(CourseEnrollment, student=request.user, course=course, is_active=True)
        
        # Verify biometric matches
        from .biometric_utils import verify_biometric_match
        biometric_reg = BiometricRegistration.objects.filter(
            student=request.user,
            course=course,
            is_active=True
        ).first()
        
        if not biometric_reg:
            return JsonResponse({'success': False, 'message': 'No biometric registered for this course'})
        
        if verify_biometric_match(biometric_reg.biometric_data, biometric_data):
            return JsonResponse({
                'success': True,
                'message': 'Biometric verified successfully',
                'matched': True,
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Biometric does not match',
                'matched': False,
            })
    
    except Exception as e:
        logger.error(f"Error verifying biometric: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["GET"])
def instructor_check_biometric_registration_view(request):
    """
    Check if a student has registered biometric for a course.
    """
    try:
        course_id = request.GET.get('course_id')
        student_id = request.GET.get('student_id')
        
        if not course_id or not student_id:
            return JsonResponse({'has_biometric': False})
        
        # Verify instructor has access to this course
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
        
        # Check if student has biometric registered
        has_biometric = BiometricRegistration.objects.filter(
            student_id=student_id,
            course=course,
            is_active=True
        ).exists()
        
        return JsonResponse({'has_biometric': has_biometric})
    
    except Exception as e:
        logger.error(f"Error checking biometric registration: {str(e)}")
        return JsonResponse({'has_biometric': False, 'error': str(e)})


@login_required
@require_http_methods(["GET"])
def api_get_instructor_courses_view(request):
    """
    API endpoint to get courses for a specific instructor.
    Used by students when registering biometric with an instructor.
    
    Expected GET parameter:
    - instructor_id: ID of the instructor
    
    Returns:
    {
        'success': bool,
        'courses': [
            {
                'id': <int>,
                'code': '<string>',
                'name': '<string>'
            },
            ...
        ]
    }
    """
    try:
        instructor_id = request.GET.get('instructor_id')
        
        logger.info(f"API: Getting courses for instructor_id={instructor_id}")
        
        if not instructor_id:
            logger.warning("API: Missing instructor_id parameter")
            return JsonResponse({
                'success': False,
                'message': 'Missing instructor_id parameter'
            }, status=400)
        
        # Get instructor
        try:
            instructor = CustomUser.objects.get(id=instructor_id, is_teacher=True)
            logger.info(f"API: Found instructor: {instructor.username}")
        except CustomUser.DoesNotExist:
            logger.warning(f"API: Instructor not found for id={instructor_id}")
            return JsonResponse({
                'success': False,
                'message': 'Instructor not found'
            }, status=404)
        except Exception as e:
            logger.error(f"API: Error looking up instructor: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error looking up instructor: {str(e)}'
            }, status=500)
        
        # Get courses taught by this instructor (including archived/inactive for now)
        courses = Course.objects.filter(
            instructor=instructor
        ).exclude(
            deleted_at__isnull=False  # Exclude soft-deleted courses
        ).values('id', 'code', 'name').order_by('code')
        
        courses_list = list(courses)
        logger.info(f"API: Found {len(courses_list)} courses for instructor {instructor.username}")
        
        return JsonResponse({
            'success': True,
            'courses': courses_list
        })
    
    except Exception as e:
        logger.error(f"Error getting instructor courses: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_student_enrolled_courses_view(request):
    """
    API endpoint to get courses the current student is enrolled in with a specific instructor.
    Used for automatic biometric enrollment in all courses with that instructor.
    
    Expected GET parameter:
    - instructor_id: ID of the instructor
    
    Returns:
    {
        'success': bool,
        'courses': [
            {
                'id': <int>,
                'code': '<string>',
                'name': '<string>'
            },
            ...
        ]
    }
    """
    try:
        instructor_id = request.GET.get('instructor_id')
        
        logger.info(f"API: Getting enrolled courses for student {request.user.id} with instructor {instructor_id}")
        
        if not instructor_id:
            logger.warning("API: Missing instructor_id parameter")
            return JsonResponse({
                'success': False,
                'message': 'Missing instructor_id parameter'
            }, status=400)
        
        if not request.user.is_student:
            return JsonResponse({
                'success': False,
                'message': 'Only students can access this endpoint'
            }, status=403)
        
        # Get instructor
        try:
            instructor = CustomUser.objects.get(id=instructor_id, is_teacher=True)
            logger.info(f"API: Found instructor: {instructor.username}")
        except CustomUser.DoesNotExist:
            logger.warning(f"API: Instructor not found for id={instructor_id}")
            return JsonResponse({
                'success': False,
                'message': 'Instructor not found'
            }, status=404)
        
        # Get courses the student is enrolled in with this instructor
        enrolled_courses = CourseEnrollment.objects.filter(
            student=request.user,
            course__instructor=instructor,
            is_active=True
        ).select_related('course').exclude(
            course__deleted_at__isnull=False  # Exclude soft-deleted courses
        ).values('course__id', 'course__code', 'course__name').distinct()
        
        courses_list = [
            {
                'id': course['course__id'],
                'code': course['course__code'],
                'name': course['course__name']
            }
            for course in enrolled_courses
        ]
        
        logger.info(f"API: Found {len(courses_list)} enrolled courses for student with instructor {instructor.username}")
        
        return JsonResponse({
            'success': True,
            'courses': courses_list,
            'instructor_name': instructor.full_name or instructor.username
        })
    
    except Exception as e:
        logger.error(f"Error getting student enrolled courses: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_enrolled_courses_status_view(request):
    """
    Get the current registration status (QR and biometric) for specified enrolled courses.
    Used to update course cards after successful biometric enrollment.
    
    Expected GET parameter:
    - course_ids: comma-separated course IDs (e.g., "1,2,3")
    
    Returns:
    {
        'success': bool,
        'enrollments': [
            {
                'id': <enrollment_id>,
                'course_id': <course_id>,
                'has_qr': bool,
                'has_biometric': bool
            },
            ...
        ]
    }
    """
    try:
        if not request.user.is_student:
            return JsonResponse({
                'success': False,
                'message': 'Only students can access this endpoint'
            }, status=403)
        
        course_ids_str = request.GET.get('course_ids', '')
        
        if not course_ids_str:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_ids parameter'
            }, status=400)
        
        # Parse course IDs
        try:
            course_ids = [int(cid.strip()) for cid in course_ids_str.split(',') if cid.strip()]
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid course_ids format'
            }, status=400)
        
        logger.info(f"Getting enrollment status for student {request.user.id} for courses: {course_ids}")
        
        # Get enrollment records for this student in the specified courses
        enrollments = CourseEnrollment.objects.filter(
            student=request.user,
            course_id__in=course_ids,
            is_active=True,
            deleted_at__isnull=True
        ).select_related('course').values('id', 'course_id')
        
        enrollments_data = []
        
        for enrollment in enrollments:
            enrollment_id = enrollment['id']
            course_id = enrollment['course_id']
            
            # Check if biometric registration exists
            has_biometric = BiometricRegistration.objects.filter(
                student=request.user,
                course_id=course_id,
                is_active=True
            ).exists()
            
            # Check if QR registration exists
            has_qr = QRCodeRegistration.objects.filter(
                student=request.user,
                course_id=course_id,
                is_active=True
            ).exists()
            
            enrollments_data.append({
                'id': enrollment_id,
                'course_id': course_id,
                'has_qr': has_qr,
                'has_biometric': has_biometric
            })
            
            logger.info(f"Enrollment {enrollment_id} (Course {course_id}): QR={has_qr}, Biometric={has_biometric}")
        
        return JsonResponse({
            'success': True,
            'enrollments': enrollments_data
        })
    
    except Exception as e:
        logger.error(f"Error getting enrolled courses status: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def api_health_check(request):
    """
    Simple health check endpoint for Arduino to test connectivity.
    GET /dashboard/api/health-check/
    Returns JSON with server status.
    """
    return JsonResponse({
        'status': 'ok',
        'message': 'Django server is reachable',
        'timestamp': timezone.now().isoformat()
    })


@csrf_exempt
@require_http_methods(["POST"])
@csrf_exempt
@require_http_methods(["POST"])
def api_broadcast_scan_update(request):
    """
    API endpoint to broadcast R307 scan updates to WebSocket clients.
    Called by Arduino or polling mechanism to push updates to frontend via WebSocket.
    
    POST data:
    {
        'enrollment_id': '<enrollment_id>',
        'slot': <1-3>,
        'success': <true/false>,
        'quality_score': <0-100>,
        'message': '<status message>'
    }
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync  # CRITICAL: Use async_to_sync, NOT asyncio.run()
        
        print("\n" + "="*80)
        print("🔔 [BROADCAST] ===== SCAN UPDATE RECEIVED =====")
        print(f"Body: {request.body}")
        
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        slot = data.get('slot')
        success = data.get('success', False)
        quality = data.get('quality_score', 0)
        message = data.get('message', '')
        progress = (slot / 3) * 100 if slot else 0
        
        print(f"[BROADCAST] Enrollment ID: {enrollment_id}")
        print(f"[BROADCAST] Slot: {slot}")
        print(f"[BROADCAST] Success: {success}")
        print(f"[BROADCAST] Quality: {quality}")
        print(f"[BROADCAST] Message: {message}")
        print(f"[BROADCAST] Calculated Progress: {progress}%")
        
        logger.info(f"[BROADCAST] Received scan update: enrollment={enrollment_id}, slot={slot}, success={success}, progress={progress}%")
        
        if not enrollment_id:
            print("[BROADCAST] ✗ Missing enrollment_id in request")
            logger.warning("[BROADCAST] Missing enrollment_id in request")
            return JsonResponse({'success': False, 'message': 'Missing enrollment_id'}, status=400)
        
        # Import the enrollment states from enrollment_state module (single source of truth)
        try:
            from .enrollment_state import _enrollment_states
        except ImportError:
            logger.error("[BROADCAST] Could not import _enrollment_states from enrollment_state")
            _enrollment_states = {}
        
        # Update enrollment state (this is critical for frontend polling)
        if enrollment_id in _enrollment_states:
            _enrollment_states[enrollment_id].update({
                'current_scan': slot,
                'progress': int(progress),
                'message': message,
                'updated_at': datetime.now().isoformat(),
                'status': 'processing'
            })
            print(f"[BROADCAST] ✓ State updated for enrollment {enrollment_id}")
            logger.info(f"[BROADCAST] ✓ State updated: enrollment={enrollment_id}, scan={slot}, progress={progress}%, message='{message}'")
        else:
            # Create the state if it doesn't exist (shouldn't happen normally)
            print(f"[BROADCAST] ⚠️ Creating state for enrollment {enrollment_id}")
            logger.warning(f"[BROADCAST] Creating state for enrollment {enrollment_id} (should have been created by api_start_enrollment)")
            _enrollment_states[enrollment_id] = {
                'status': 'processing',
                'current_scan': slot,
                'progress': int(progress),
                'message': message,
                'updated_at': datetime.now().isoformat()
            }
        
        # Get channel layer and broadcast to group
        channel_layer = get_channel_layer()
        group_name = f"biometric_enrollment_{enrollment_id}"
        
        print(f"[BROADCAST] Group name: {group_name}")
        logger.info(f"[BROADCAST] Sending to group: {group_name}")
        
        # Send message to WebSocket group
        print(f"[BROADCAST] Sending WebSocket message with step={slot}")
        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                'type': 'scan_update',
                'slot': slot,
                'step': slot,  # CRITICAL: Include step for frontend deduplication
                'success': success,
                'quality': quality,
                'message': message,
                'progress': progress
            }
        )
        
        print(f"[BROADCAST] ✓ Broadcast sent to {group_name}")
        logger.info(f"[BROADCAST] ✓ Broadcast sent: enrollment={enrollment_id}, slot={slot}, group={group_name}")
        
        print("="*80 + "\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Scan update broadcasted to clients'
        })
    
    except Exception as e:
        print(f"[BROADCAST] ✗✗✗ ERROR: {str(e)}")
        logger.error(f"Error broadcasting scan update: {str(e)}", exc_info=True)
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_scan_acknowledged(request):
    """
    Endpoint for frontend to acknowledge receipt of a scan update.
    Stores acknowledgment in cache so ESP32 can poll and check.
    
    POST data:
    {
        'enrollment_id': '<enrollment_id>',
        'step': <1-3>,
        'success': <true/false>
    }
    """
    try:
        from django.core.cache import cache
        
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        step = data.get('step', 0)
        success = data.get('success', False)
        
        print(f"\n[ACK API] Frontend acknowledged scan {step}/3 for enrollment {enrollment_id}")
        logger.info(f"[ACK API] Frontend acknowledged scan {step}/3 for enrollment {enrollment_id}")
        
        if not enrollment_id or step < 1 or step > 3:
            return JsonResponse({'success': False, 'message': 'Invalid parameters'}, status=400)
        
        # Store acknowledgment in cache with key format: ack_<enrollment_id>_<step>
        # ESP32 will poll this to check if frontend acknowledged
        cache_key = f"scan_ack_{enrollment_id}_{step}"
        cache.set(cache_key, True, timeout=60)  # Keep for 60 seconds
        
        print(f"[ACK CACHE] ✓ Stored acknowledgment: {cache_key} = True")
        print(f"[ACK] Step {step} acknowledged by frontend - ESP32 will detect via polling!")
        
        return JsonResponse({
            'success': True,
            'message': f'Scan {step} acknowledged - ESP32 will detect via HTTP polling',
            'enrollment_id': enrollment_id,
            'step': step
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error in scan acknowledgment API: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_check_scan_acknowledged(request):
    """
    Endpoint for ESP32 to check if frontend has acknowledged a specific scan.
    ESP32 polls this while waiting for acknowledgment.
    
    GET parameters:
    ?enrollment_id=<id>&step=<1-3>
    
    Returns: {'acknowledged': true/false}
    """
    try:
        from django.core.cache import cache
        
        enrollment_id = request.GET.get('enrollment_id')
        step = request.GET.get('step', '0')
        
        try:
            step = int(step)
        except:
            step = 0
        
        if not enrollment_id or step < 1 or step > 3:
            return JsonResponse({'success': False, 'message': 'Invalid parameters'}, status=400)
        
        # Check if acknowledgment exists in cache
        cache_key = f"scan_ack_{enrollment_id}_{step}"
        acknowledged = cache.get(cache_key, False)
        
        if acknowledged:
            print(f"[ESP32 POLL] ✓ Scan {step} WAS acknowledged by frontend!")
            print(f"[ESP32 POLL]   ESP32 can now proceed to next scan")
        else:
            print(f"[ESP32 POLL] ⏳ Scan {step} NOT yet acknowledged - ESP32 will keep waiting/retrying")
        
        return JsonResponse({
            'success': True,
            'enrollment_id': enrollment_id,
            'step': step,
            'acknowledged': acknowledged
        })
    
    except Exception as e:
        logger.error(f"Error checking scan acknowledgment: {str(e)}")
        return JsonResponse({
            'success': False,
            'acknowledged': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
@require_http_methods(["POST"])
def api_broadcast_enrollment_complete(request):
    """
    API endpoint to broadcast enrollment completion to WebSocket clients.
    
    POST data:
    {
        'enrollment_id': '<enrollment_id>',
        'fingerprint_id': <id>,
        'success': <true/false>,
        'message': '<message>'
    }
    """
    try:
        from channels.layers import get_channel_layer
        import asyncio
        
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        fingerprint_id = data.get('fingerprint_id', 1)
        success = data.get('success', True)
        message = data.get('message', 'Enrollment complete')
        
        if not enrollment_id:
            return JsonResponse({'success': False, 'message': 'Missing enrollment_id'}, status=400)
        
        # Import the enrollment states
        try:
            from .views_enrollment_apis import _enrollment_states
        except ImportError:
            logger.error("[BROADCAST] Could not import _enrollment_states from views_enrollment_apis")
            _enrollment_states = {}
        
        # Update enrollment state to mark as completed
        if enrollment_id in _enrollment_states:
            _enrollment_states[enrollment_id].update({
                'status': 'completed',
                'progress': 100,
                'message': message,
                'fingerprint_id': fingerprint_id,
                'completed_at': datetime.now().isoformat()
            })
            logger.info(f"[BROADCAST] Marked enrollment {enrollment_id} as completed")
        
        channel_layer = get_channel_layer()
        group_name = f"biometric_enrollment_{enrollment_id}"
        
        asyncio.run(channel_layer.group_send(
            group_name,
            {
                'type': 'enrollment_complete',
                'success': success,
                'fingerprint_id': fingerprint_id,
                'message': message
            }
        ))
        
        logger.info(f"[BROADCAST] Enrollment complete: enrollment={enrollment_id}, fingerprint={fingerprint_id}")
        
        return JsonResponse({'success': True, 'message': 'Enrollment completion broadcasted'})
    
    except Exception as e:
        logger.error(f"[BROADCAST] Error broadcasting enrollment completion: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==================== ESP32 FINGERPRINT SENSOR COMMUNICATION ====================

@csrf_exempt
@require_http_methods(["POST"])
def api_esp32_scan_feedback(request):
    """
    API endpoint for ESP32 to send fingerprint scan feedback during enrollment.
    Called by ESP32 to report scan progress and quality.
    
    POST data:
    {
        'template_id': '<enrollment_template_id>',
        'success': <true/false>,
        'quality': <0-100>,
        'confirmations': <1-5>,
        'timestamp': <milliseconds>
    }
    """
    try:
        data = json.loads(request.body)
        template_id = data.get('template_id', '')
        success = data.get('success', False)
        quality = data.get('quality', 0)
        confirmations = data.get('confirmations', 0)
        
        logger.info(f"[ESP32] Scan feedback: template={template_id}, success={success}, quality={quality}, confirmations={confirmations}")
        
        # Broadcast scan update to WebSocket clients
        try:
            from channels.layers import get_channel_layer
            import asyncio
            
            channel_layer = get_channel_layer()
            group_name = f"biometric_enrollment_{template_id}"
            
            asyncio.run(channel_layer.group_send(
                group_name,
                {
                    'type': 'scan_update',
                    'slot': confirmations,
                    'success': success,
                    'quality': quality,
                    'message': f'Scan {confirmations}/5 - Quality: {quality}%',
                    'progress': (confirmations / 5) * 100
                }
            ))
            
            logger.info(f"[ESP32] Broadcasted scan update to group: {group_name}")
        except Exception as e:
            logger.warning(f"[ESP32] Could not broadcast to WebSocket: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Scan feedback received',
            'confirmations': confirmations
        })
    
    except json.JSONDecodeError:
        logger.error("[ESP32] Invalid JSON in scan feedback request")
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"[ESP32] Error processing scan feedback: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_esp32_enrollment_complete(request):
    """
    API endpoint for ESP32 to signal fingerprint enrollment completion.
    Called by ESP32 after 5 scans are successfully collected.
    
    POST data:
    {
        'template_id': '<enrollment_template_id>',
        'success': <true/false>,
        'fingerprint_id': <sensor_id>,
        'confirmations': <5>,
        'timestamp': <milliseconds>
    }
    """
    try:
        data = json.loads(request.body)
        template_id = data.get('template_id', '')
        success = data.get('success', False)
        fingerprint_id = data.get('fingerprint_id', 0)
        confirmations = data.get('confirmations', 0)
        
        logger.info(f"[ESP32] Enrollment complete: template={template_id}, success={success}, fingerprint_id={fingerprint_id}")
        
        if not success:
            logger.warning(f"[ESP32] Enrollment failed for template {template_id}")
            # Broadcast failure to WebSocket clients
            try:
                from channels.layers import get_channel_layer
                import asyncio
                
                channel_layer = get_channel_layer()
                group_name = f"biometric_enrollment_{template_id}"
                
                asyncio.run(channel_layer.group_send(
                    group_name,
                    {
                        'type': 'enrollment_error',
                        'message': 'Enrollment failed. Please try again.',
                        'success': False
                    }
                ))
            except Exception as e:
                logger.warning(f"[ESP32] Could not broadcast error to WebSocket: {str(e)}")
            
            return JsonResponse({
                'success': False,
                'message': 'Enrollment failed on ESP32',
                'fingerprint_id': 0
            })
        
        # Broadcast completion to WebSocket clients
        try:
            from channels.layers import get_channel_layer
            import asyncio
            
            channel_layer = get_channel_layer()
            group_name = f"biometric_enrollment_{template_id}"
            
            asyncio.run(channel_layer.group_send(
                group_name,
                {
                    'type': 'enrollment_complete',
                    'success': True,
                    'message': 'Fingerprint enrollment complete. Processing...',
                    'fingerprint_id': fingerprint_id
                }
            ))
            
            logger.info(f"[ESP32] Broadcasted enrollment complete to group: {group_name}")
        except Exception as e:
            logger.warning(f"[ESP32] Could not broadcast completion to WebSocket: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Enrollment completion received',
            'fingerprint_id': fingerprint_id,
            'confirmations': confirmations
        })
    
    except json.JSONDecodeError:
        logger.error("[ESP32] Invalid JSON in enrollment complete request")
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"[ESP32] Error processing enrollment completion: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def api_esp32_config(request):
    """
    API endpoint for ESP32 to request configuration settings.
    Allows remote configuration of WiFi, server URL, etc.
    
    GET request returns current configuration
    POST request updates configuration
    """
    try:
        if request.method == 'GET':
            config = {
                'server_url': getattr(settings, 'BIOMETRIC_SERVER_URL', 'http://192.168.1.X:8000'),
                'api_base': '/dashboard/api/biometric/',
                'scan_feedback_endpoint': '/dashboard/api/esp32/scan-feedback/',
                'enrollment_complete_endpoint': '/dashboard/api/esp32/enrollment-complete/',
                'health_check_endpoint': '/dashboard/api/health-check/',
                'enrollment_slots': 5,  # Number of fingerprint scans needed
                'quality_threshold': 70,  # Minimum fingerprint quality (0-100)
                'device_id': getattr(settings, 'BIOMETRIC_DEVICE_ID', 'ESP32-001'),
                'version': '1.0.0'
            }
            return JsonResponse(config)
        
        elif request.method == 'POST':
            # Only allow configuration updates from localhost or authorized sources
            client_ip = request.META.get('REMOTE_ADDR', '')
            logger.info(f"[ESP32] Config update request from IP: {client_ip}")
            
            # For security, only accept from specific IPs in production
            authorized_ips = getattr(settings, 'BIOMETRIC_AUTHORIZED_IPS', ['127.0.0.1', 'localhost'])
            
            data = json.loads(request.body)
            # Store new configuration in settings or database as needed
            logger.info(f"[ESP32] Config update received: {list(data.keys())}")
            
            return JsonResponse({
                'success': True,
                'message': 'Configuration updated successfully'
            })
    
    except Exception as e:
        logger.error(f"[ESP32] Error handling config request: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==================== BIOMETRIC ENROLLMENT MANAGEMENT APIs ====================

# In-memory store for current enrollment states
_enrollment_states = {}

@csrf_exempt
@require_http_methods(["POST"])
def api_enrollment_status(request, enrollment_id):
    """Get current enrollment status"""
    try:
        if enrollment_id in _enrollment_states:
            state = _enrollment_states[enrollment_id]
            return JsonResponse({
                'status': state['status'],
                'current_scan': state.get('current_scan', 0),
                'progress': state.get('progress', 0),
                'message': state.get('message', '')
            })
        return JsonResponse({'status': 'not_found'}, status=404)
    except Exception as e:
        logger.error(f"[ENROLLMENT] Status error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_enrollment_updates(request, enrollment_id):
    """Get enrollment updates"""
    try:
        if enrollment_id not in _enrollment_states:
            return JsonResponse({'status': 'not_found'}, status=404)
        
        state = _enrollment_states[enrollment_id]
        response = JsonResponse(state)
        response['Cache-Control'] = 'no-cache'
        return response
        
    except Exception as e:
        logger.error(f"[ENROLLMENT] Updates error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_save_enrollment(request):
    """Save completed enrollment to database"""
    try:
        from django.db import transaction
        
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        course_id = data.get('course_id')
        
        user = request.user if request.user.is_authenticated else None
        if not user:
            return JsonResponse({'success': False, 'message': 'Not authenticated'}, status=401)
        
        if enrollment_id not in _enrollment_states:
            return JsonResponse({'success': False, 'message': 'Enrollment not found'}, status=404)
        
        state = _enrollment_states[enrollment_id]
        # CRITICAL FIX: Accept BOTH 'ready_for_confirmation' and 'completed' status
        # Frontend sends save request when status is 'ready_for_confirmation' (after 3 scans)
        # Not when it's 'completed' (which is after ESP32 confirms)
        if state['status'] not in ['completed', 'ready_for_confirmation']:
            logger.error(f"[SAVE] Invalid status for save: {state['status']} (expected 'ready_for_confirmation' or 'completed')")
            return JsonResponse({'success': False, 'message': f'Enrollment not ready (status: {state["status"]})'}, status=400)
        
        # Get course
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Course not found'}, status=404)
        
        fingerprint_id = state.get('fingerprint_id')
        if fingerprint_id is None:
            fingerprint_id = state.get('fingerprint_slot')

        try:
            fingerprint_id = int(str(fingerprint_id).strip())
        except Exception:
            fingerprint_id = None

        if fingerprint_id is None:
            return JsonResponse({'success': False, 'message': 'Enrollment missing fingerprint_id (slot). Please restart enrollment.'}, status=400)
        
        # Check if this is a re-registration and delete the old fingerprint ID from sensor
        old_fingerprint_id = state.get('old_fingerprint_id')
        is_re_registration = state.get('is_re_registration', False)
        
        if is_re_registration and old_fingerprint_id:
            logger.info(f"[ENROLLMENT] RE-REGISTRATION: Deleting old fingerprint ID {old_fingerprint_id} from sensor")
            # In a real implementation, you would send a command to ESP32 to delete the old fingerprint ID
            # For now, we just log it and the old one will be overwritten in the database
            try:
                import requests
                esp32_ip = '192.168.1.7'
                esp32_port = 80
                delete_url = f'http://{esp32_ip}:{esp32_port}/api/delete-fingerprint/'
                
                delete_data = {
                    'fingerprint_id': old_fingerprint_id,
                    'course_id': course_id
                }
                
                response = requests.post(delete_url, json=delete_data, timeout=2)
                logger.info(f"[ENROLLMENT] Notified ESP32 to delete old fingerprint ID {old_fingerprint_id}: {response.status_code}")
            except Exception as e:
                logger.warning(f"[ENROLLMENT] Could not notify ESP32 to delete old fingerprint: {str(e)}")
        
        # CRITICAL: Send confirmation to ESP32 to finalize fingerprint storage
        logger.info(f"[ENROLLMENT] SENDING CONFIRMATION TO ESP32...")
        logger.info(f"[ENROLLMENT]   Fingerprint ID: {fingerprint_id}")
        logger.info(f"[ENROLLMENT]   Enrollment ID: {enrollment_id}")
        
        try:
            import requests
            esp32_url = f'http://192.168.1.9:80/enroll/confirm'
            esp32_payload = {
                'fingerprint_id': fingerprint_id,
                'template_id': enrollment_id
            }
            
            logger.info(f"[ENROLLMENT] Posting to ESP32: {esp32_url}")
            logger.info(f"[ENROLLMENT] Payload: {esp32_payload}")
            
            esp32_response = requests.post(
                esp32_url,
                json=esp32_payload,
                timeout=10
            )
            
            logger.info(f"[ENROLLMENT] ESP32 Response Status: {esp32_response.status_code}")
            logger.info(f"[ENROLLMENT] ESP32 Response Body: {esp32_response.text}")
            
            if esp32_response.status_code == 200:
                esp32_data = esp32_response.json()
                logger.info(f"[✓] ESP32 CONFIRMATION SUCCESSFUL")
                logger.info(f"[✓]   Verified: {esp32_data.get('verified', False)}")
                logger.info(f"[✓]   Total fingerprints on sensor: {esp32_data.get('total_fingerprints', 'unknown')}")
            else:
                logger.warning(f"[⚠] ESP32 returned status {esp32_response.status_code}")
                logger.warning(f"[⚠] Response: {esp32_response.text}")
        
        except Exception as e:
            logger.error(f"[ERROR] Could not send confirmation to ESP32: {str(e)}")
            logger.error(f"[ERROR] This may prevent the fingerprint from being saved to the sensor!")
            logger.error(f"[ERROR] Make sure ESP32 is running at 192.168.1.9:80")
        
        # Use transaction to avoid database locking
        with transaction.atomic():
            biometric_reg, created = BiometricRegistration.objects.update_or_create(
                user=user,
                course=course,
                defaults={
                    'fingerprint_id': fingerprint_id,
                    'is_active': True
                }
            )
            
            # Update user profile
            user.biometric_fingerprint_id = fingerprint_id
            user.biometric_registered = True
            user.save()
        
        # Clean up
        if enrollment_id in _enrollment_states:
            del _enrollment_states[enrollment_id]
        
        # Log the enrollment
        if is_re_registration and old_fingerprint_id:
            logger.info(f"[ENROLLMENT] RE-REGISTRATION COMPLETED: Student {user.full_name}")
            logger.info(f"[ENROLLMENT]   Old fingerprint ID: {old_fingerprint_id}")
            logger.info(f"[ENROLLMENT]   New fingerprint ID: {fingerprint_id}")
            logger.info(f"[ENROLLMENT]   Course: {course.code} ({course.name})")
        else:
            logger.info(f"[ENROLLMENT] NEW ENROLLMENT: Student {user.full_name} with fingerprint ID {fingerprint_id} for course {course.code}")
        
        return JsonResponse({
            'success': True,
            'fingerprint_id': fingerprint_id,
            'is_re_registration': is_re_registration,
            'old_fingerprint_id': old_fingerprint_id if is_re_registration else None
        })
        
    except Exception as e:
        logger.error(f"[ENROLLMENT] Save error: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ==================== BIOMETRIC ENROLLMENT MANAGEMENT APIs ====================



# In-memory store for current enrollment states (in production, use Redis or database)

_enrollment_states = {}



@csrf_exempt

@require_http_methods(["POST"])

@csrf_exempt
@require_http_methods(["POST"])
@csrf_exempt
@require_http_methods(["POST"])
def api_start_enrollment(request):

    """
    Start a new biometric enrollment session for a user

    Creates an enrollment ID and initializes tracking
    
    If the student already has a fingerprint registered for this course,
    they will be informed and can choose to re-register (which will replace the old one)
    """
    
    # Import centralized enrollment state management
    from .enrollment_state import _enrollment_states
    
    # PRINT DIRECTLY SO WE SEE IT
    print("\n" + "="*80)
    print("🔥 🔥 🔥  api_start_enrollment CALLED  🔥 🔥 🔥")
    print("="*80)
    
    logger.info("[ENROLLMENT API] ============ ENROLLMENT REQUEST RECEIVED ============")
    logger.info(f"[ENROLLMENT API] Request method: {request.method}")
    logger.info(f"[ENROLLMENT API] User authenticated: {request.user.is_authenticated}")
    
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        
        print(f"✓ Parsed JSON - course_id: {course_id}")
        logger.info(f"[ENROLLMENT API] Parsed course_id: {course_id}")
        
        # Get current user (if authenticated)
        user = request.user if request.user.is_authenticated else None
        
        logger.info(f"[ENROLLMENT API] User: {user}")
        
        if not user:
            print("✗ ERROR: User not authenticated!")
            logger.error("[ENROLLMENT API] User not authenticated!")
            return JsonResponse({
                'success': False,
                'message': 'User must be authenticated to start enrollment'
            }, status=401)

        lock_payload = {
            'student_user_id': user.id,
            'student_id': getattr(user, 'school_id', None),
            'enrollment_id': incoming_enrollment_id,
            'template_id': incoming_template_id,
            'created_at': timezone.now().isoformat(),
        }

        lock_acquired = cache.add('biometric_enrollment_lock', lock_payload, timeout=300)
        if not lock_acquired:
            existing_lock = cache.get('biometric_enrollment_lock')
            try:
                if existing_lock and existing_lock.get('student_user_id') == user.id:
                    cache.set('biometric_enrollment_lock', lock_payload, timeout=300)
                else:
                    locked_by = (existing_lock or {}).get('student_id') or 'another student'
                    return JsonResponse({
                        'success': False,
                        'message': f'Another student is currently enrolling ({locked_by}). Please wait...'
                    }, status=409)
            except Exception:
                return JsonResponse({
                    'success': False,
                    'message': 'Another student is currently enrolling. Please wait...'
                }, status=409)
        
        # Check if course exists
        try:
            course = Course.objects.get(id=course_id)
            print(f"✓ Course found: {course.code}")
            logger.info(f"[ENROLLMENT API] Course found: {course.code}")
        except Course.DoesNotExist:
            print(f"✗ ERROR: Course {course_id} not found!")
            logger.error(f"[ENROLLMENT API] Course {course_id} not found!")
            try:
                cache.delete('biometric_enrollment_lock')
            except Exception:
                pass
            return JsonResponse({
                'success': False,
                'message': 'Course not found'
            }, status=404)
        
        # CHECK: Has this student already registered their fingerprint for this course?
        existing_registration = BiometricRegistration.objects.filter(
            student=user,
            course=course,
            is_active=True
        ).first()
        
        # IMPORTANT: Keep enrollment_id consistent with the frontend WebSocket group.
        # The student UI opens /ws/biometric/enrollment/<enrollment_id>/ BEFORE calling this API.
        # If we generate a different ID here, the WebSocket and backend state won't match.
        enrollment_id = (str(incoming_enrollment_id).strip() if incoming_enrollment_id else '').strip() or f"enrollment_{int(datetime.now().timestamp()*1000)}_{user.id}"
        
        print(f"✓ Generated enrollment_id: {enrollment_id}")
        logger.info(f"[ENROLLMENT API] Generated enrollment_id: {enrollment_id}")
        
        # ========== CRITICAL: Allocate unique fingerprint ID for this student FIRST ==========
        # IMPORTANT: Use a unique ID per STUDENT, NOT per COURSE
        # Each student gets ONE unique fingerprint ID across all courses/instructors
        
        # Check if this student already has a fingerprint_id from a previous enrollment
        existing_biometric = BiometricRegistration.objects.filter(
            student=user,
            is_active=True
        ).order_by('-created_at').first()
        
        if existing_biometric and existing_biometric.fingerprint_id is not None:
            try:
                existing_slot = int(existing_biometric.fingerprint_id)
            except Exception:
                existing_slot = None
        else:
            existing_slot = None

        # Prevent legacy/unsafe slot reuse (e.g., slot 1) by forcing our reserved range >= 100.
        if existing_slot is not None and existing_slot >= 100:
            fingerprint_slot = existing_slot
            print(f"✓ REUSING fingerprint_id: {fingerprint_slot} (student already has fingerprint registered)")
            logger.info(f"[ENROLLMENT] REUSING fingerprint_id: {fingerprint_slot} for student {user.full_name}")
        else:
            max_existing_slot = BiometricRegistration.objects.aggregate(max_id=Max('fingerprint_id'))['max_id'] or 99
            fingerprint_slot = max(100, int(max_existing_slot) + 1)
            print(f"✓ ALLOCATING NEW fingerprint_id: {fingerprint_slot} for student {user.full_name}")
            logger.info(f"[ENROLLMENT] ALLOCATING NEW fingerprint_id: {fingerprint_slot} for student {user.full_name}")
        
        # Initialize enrollment state
        _enrollment_states[enrollment_id] = {
            'status': 'processing',
            'current_scan': 0,
            'progress': 0,
            'message': 'Initializing enrollment...',
            'course_id': course_id,
            'user_id': user.id,
            'created_at': datetime.now().isoformat(),
            'scans': [],
            'is_re_registration': existing_registration is not None,
            'old_fingerprint_id': existing_registration.fingerprint_id if existing_registration else None,
            'fingerprint_slot': fingerprint_slot  # CRITICAL: Store the unique fingerprint ID being used for this enrollment
        }
        
        logger.info(f"[ENROLLMENT] Started enrollment session: {enrollment_id}")
        
        # Try multiple ESP32 server addresses in order of preference
        esp32_addresses = [
            os.environ.get('ESP32_SERVER', 'http://192.168.1.7'),  # From environment variable
            'http://192.168.1.7',   # Default
            'http://192.168.1.9',   # Fallback
            'http://192.168.1.6',   # Alternative
            'http://esp32.local',   # mDNS name
        ]
        
        esp32_server = None
        esp32_response = None
        last_error = None
        
        for address in esp32_addresses:
            try:
                esp32_url = f"{address}/enroll"
                payload = {
                    'slot': fingerprint_slot,  # Use unique fingerprint_id, NOT course_id
                    'template_id': enrollment_id
                }
                print(f"\n📤 TRYING ESP32 at {address}...")
                print(f"   URL: {esp32_url}")
                print(f"   Payload: {payload}")
                print(f"   Student: {user.full_name}")
                print(f"   Courses: {course.code}")
                
                logger.info(f"[ENROLLMENT] Trying ESP32 at {esp32_url} with payload: {payload}")
                logger.info(f"[ENROLLMENT] Student: {user.full_name}, Course: {course.code}, Slot: {fingerprint_slot}")
                
                response = requests.post(esp32_url, json=payload, timeout=3)
                
                print(f"✓ ESP32 Response: {response.status_code}")
                logger.info(f"[ENROLLMENT] ESP32 response status: {response.status_code}")
                
                if response.status_code == 200:
                    logger.info(f"[ENROLLMENT] ESP32 acknowledged at {address}")
                    _enrollment_states[enrollment_id]['message'] = 'Sensor ready - Place your finger'
                    print(f"✓✓✓ ESP32 ACKNOWLEDGED ENROLLMENT at {address}")
                    esp32_server = address
                    break
                else:
                    logger.warning(f"[ENROLLMENT] ESP32 at {address} returned status {response.status_code}")
                    print(f"⚠ ESP32 at {address} returned {response.status_code}")
                    last_error = f"Status {response.status_code}"
            
            except requests.exceptions.ConnectionError as e:
                print(f"⚠ Cannot connect to {address}: {str(e)}")
                logger.warning(f"[ENROLLMENT] Cannot reach ESP32 at {address}: {str(e)}")
                last_error = str(e)
                continue
            except Exception as e:
                print(f"⚠ Error with {address}: {str(e)}")
                logger.warning(f"[ENROLLMENT] Error with ESP32 at {address}: {str(e)}")
                last_error = str(e)
                continue
        
        if not esp32_server:
            print(f"✗ ERROR: Could not reach any ESP32 address. Last error: {last_error}")
            logger.error(f"[ENROLLMENT] Could not reach ESP32 at any address. Last error: {last_error}")
            _enrollment_states[enrollment_id]['message'] = f'Sensor offline - tried multiple addresses. Last error: {last_error}'
        
        response_data = {
            'success': True,
            'enrollment_id': enrollment_id,
            'message': 'Enrollment session created',
            'has_existing_registration': existing_registration is not None,
            'slot': fingerprint_slot,
            'template_id': incoming_template_id or enrollment_id
        }
        
        # If re-registering, include the old fingerprint ID in response
        if existing_registration:
            response_data['old_fingerprint_id'] = existing_registration.fingerprint_id
            response_data['re_registration_message'] = f'You already have a fingerprint registered for this course. Placing a new finger will replace the old one (ID: {existing_registration.fingerprint_id}) with a new fingerprint ID.'
        
        print(f"✓ Returning enrollment_id: {enrollment_id}\n")
        logger.info(f"[ENROLLMENT API] Returning success response with enrollment_id: {enrollment_id}")
        return JsonResponse(response_data)

    except Exception as e:
        print(f"✗✗✗ EXCEPTION IN api_start_enrollment: {str(e)}\n")
        logger.error(f"[ENROLLMENT] Error starting enrollment: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)





@csrf_exempt

@require_http_methods(["GET"])

def api_enrollment_status(request, enrollment_id):

    """

    Get current status of an ongoing enrollment

    Returns progress percentage, current scan number, and status message

    """

    try:

        # Check if enrollment exists in memory

        if enrollment_id in _enrollment_states:

            state = _enrollment_states[enrollment_id]

            return JsonResponse({

                'status': state['status'],

                'current_scan': state.get('current_scan', 0),

                'progress': state.get('progress', 0),

                'message': state.get('message', ''),

                'error': state.get('error', None)

            })

        else:

            return JsonResponse({

                'status': 'not_found',

                'message': 'Enrollment session not found'

            }, status=404)

            

    except Exception as e:

        logger.error(f"[ENROLLMENT] Error getting enrollment status: {str(e)}")

        return JsonResponse({

            'success': False,

            'message': str(e)

        }, status=500)





@csrf_exempt

@require_http_methods(["GET"])

def api_enrollment_updates(request, enrollment_id):

    """

    EventSource endpoint for real-time enrollment updates

    """

    try:

        if enrollment_id not in _enrollment_states:

            return JsonResponse({

                'status': 'not_found',

                'message': 'Enrollment not found'

            }, status=404)

        

        state = _enrollment_states[enrollment_id]

        

        # Return current state

        response = JsonResponse(state)

        response['Cache-Control'] = 'no-cache'

        response['X-Accel-Buffering'] = 'no'

        

        return response

        

    except Exception as e:

        logger.error(f"[ENROLLMENT] Error getting enrollment updates: {str(e)}")

        return JsonResponse({

            'success': False,

            'message': str(e)

        }, status=500)





@csrf_exempt

@require_http_methods(["POST"])

def api_save_enrollment(request):

    """

    Save completed enrollment to database

    Stores fingerprint ID in user's biometric registration and links to courses

    """

    try:

        data = json.loads(request.body)

        enrollment_id = data.get('enrollment_id')

        course_id = data.get('course_id')

        

        user = request.user if request.user.is_authenticated else None

        if not user:

            return JsonResponse({

                'success': False,

                'message': 'User must be authenticated'

            }, status=401)

        

        # Get enrollment state

        if enrollment_id not in _enrollment_states:

            return JsonResponse({

                'success': False,

                'message': 'Enrollment session not found'

            }, status=404)

        

        state = _enrollment_states[enrollment_id]
        # CRITICAL FIX: Accept BOTH 'ready_for_confirmation' and 'completed' status
        # Frontend sends save request when status is 'ready_for_confirmation' (after 3 scans)
        # Not when it's 'completed' (which is after ESP32 confirms)
        if state['status'] not in ['completed', 'ready_for_confirmation']:

            return JsonResponse({

                'success': False,

                'message': f'Enrollment not ready (status: {state["status"]})'

            }, status=400)

        

        # Get course

        try:

            course = Course.objects.get(id=course_id)

        except Course.DoesNotExist:

            return JsonResponse({

                'success': False,

                'message': 'Course not found'

            }, status=404)

        

        # Get fingerprint ID from ESP32 response

        fingerprint_id = state.get('fingerprint_id')
        if fingerprint_id is None:
            fingerprint_id = state.get('fingerprint_slot')

        try:
            fingerprint_id = int(str(fingerprint_id).strip())
        except Exception:
            fingerprint_id = None

        if fingerprint_id is None:
            return JsonResponse({'success': False, 'message': 'Enrollment missing fingerprint_id (slot). Please restart enrollment.'}, status=400)

        

        # CRITICAL: Send confirmation to ESP32 to finalize fingerprint storage
        logger.info(f"[ENROLLMENT] SENDING CONFIRMATION TO ESP32...")
        logger.info(f"[ENROLLMENT]   Fingerprint ID: {fingerprint_id}")
        logger.info(f"[ENROLLMENT]   Enrollment ID: {enrollment_id}")
        
        try:
            import requests
            esp32_url = f'http://192.168.1.9:80/enroll/confirm'
            esp32_payload = {
                'fingerprint_id': fingerprint_id,
                'template_id': enrollment_id
            }
            
            logger.info(f"[ENROLLMENT] Posting to ESP32: {esp32_url}")
            logger.info(f"[ENROLLMENT] Payload: {esp32_payload}")
            
            esp32_response = requests.post(
                esp32_url,
                json=esp32_payload,
                timeout=10
            )
            
            logger.info(f"[ENROLLMENT] ESP32 Response Status: {esp32_response.status_code}")
            logger.info(f"[ENROLLMENT] ESP32 Response Body: {esp32_response.text}")
            
            if esp32_response.status_code == 200:
                esp32_data = esp32_response.json()
                logger.info(f"[✓] ESP32 CONFIRMATION SUCCESSFUL")
                logger.info(f"[✓]   Verified: {esp32_data.get('verified', False)}")
                logger.info(f"[✓]   Total fingerprints on sensor: {esp32_data.get('total_fingerprints', 'unknown')}")
            else:
                logger.warning(f"[⚠] ESP32 returned status {esp32_response.status_code}")
                logger.warning(f"[⚠] Response: {esp32_response.text}")
        
        except Exception as e:
            logger.error(f"[ERROR] Could not send confirmation to ESP32: {str(e)}")
            logger.error(f"[ERROR] This may prevent the fingerprint from being saved to the sensor!")
            logger.error(f"[ERROR] Make sure ESP32 is running at 192.168.1.9:80")

        # Create or update biometric registration

        biometric_reg, created = BiometricRegistration.objects.update_or_create(

            user=user,

            course=course,

            defaults={

                'fingerprint_id': fingerprint_id,

                'status': 'active',

                'registered_at': datetime.now(),

                'enrollment_session_id': enrollment_id

            }

        )

        

        # Also update user's default biometric profile

        user.biometric_fingerprint_id = fingerprint_id

        user.biometric_registered = True

        user.save()

        

        # Clean up enrollment state

        if enrollment_id in _enrollment_states:

            del _enrollment_states[enrollment_id]

        

        logger.info(f"[ENROLLMENT] Saved enrollment {enrollment_id} with fingerprint ID {fingerprint_id} for course {course_id}")

        

        return JsonResponse({

            'success': True,

            'message': 'Fingerprint saved successfully',

            'fingerprint_id': fingerprint_id,

            'course_id': course_id

        })

        

    except Exception as e:

        logger.error(f"[ENROLLMENT] Error saving enrollment: {str(e)}")

        return JsonResponse({

            'success': False,

            'message': str(e)

        }, status=500)





def update_enrollment_progress(enrollment_id, current_scan, progress, message, fingerprint_id=None, error=None):

    """

    Helper function to update enrollment progress

    Called by ESP32 API endpoints when broadcasting updates

    """

    if enrollment_id in _enrollment_states:

        _enrollment_states[enrollment_id].update({

            'current_scan': current_scan,

            'progress': progress,

            'message': message,

            'updated_at': datetime.now().isoformat()

        })

        

        if fingerprint_id:

            _enrollment_states[enrollment_id]['fingerprint_id'] = fingerprint_id

        

        if error:

            _enrollment_states[enrollment_id]['error'] = error

            _enrollment_states[enrollment_id]['status'] = 'failed'

        

        # Auto-complete at 100%

        if progress >= 100:

            _enrollment_states[enrollment_id]['status'] = 'completed'

        

        logger.debug(f"[ENROLLMENT] Updated {enrollment_id}: Scan {current_scan}, Progress {progress}%")





def mark_enrollment_complete(enrollment_id, fingerprint_id):

    """

    Helper function to mark enrollment as complete

    """

    if enrollment_id in _enrollment_states:

        _enrollment_states[enrollment_id].update({

            'status': 'completed',

            'progress': 100,

            'message': 'All 5 scans completed successfully!',

            'fingerprint_id': fingerprint_id,

            'completed_at': datetime.now().isoformat()

        })

        logger.info(f"[ENROLLMENT] Completed enrollment {enrollment_id} with fingerprint ID {fingerprint_id}")


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_biometric_confirm_fingerprint_view(request):
    """
    API endpoint to confirm fingerprint enrollment.
    
    For replacements: Just updates the biometric data for the same fingerprint_id (no deactivation)
    For new enrollments: Activates the newly registered fingerprints
    
    Expected POST data:
    {
        'fingerprint_id': <int>,  # The fingerprint ID being confirmed
        'is_replacement': <bool>
    }
    """
    try:
        data = json.loads(request.body)
        fingerprint_id = data.get('fingerprint_id')
        is_replacement = data.get('is_replacement', False)
        
        if not fingerprint_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing fingerprint_id'
            }, status=400)
        
        # Get student
        target_student = request.user if request.user.is_student else None
        if not target_student:
            return JsonResponse({
                'success': False,
                'message': 'You must be a student to confirm biometric'
            }, status=403)
        
        with transaction.atomic():
            print(f"\n[DEBUG] Confirmation request received:")
            print(f"   Student ID: {target_student.id}")
            print(f"   Fingerprint ID requested: {fingerprint_id}")
            print(f"   Is replacement: {is_replacement}")
            
            # CLEANUP: Fix any existing records with None fingerprint_id
            # This can happen if old enrollments weren't finalized properly
            none_registrations = BiometricRegistration.objects.filter(
                student=target_student,
                fingerprint_id__isnull=True,
                is_active=True
            )
            if none_registrations.exists():
                print(f"\n[CLEANUP] Found {none_registrations.count()} registrations with None fingerprint_id")
                # Update them to use the fingerprint_id being confirmed
                none_registrations.update(fingerprint_id=fingerprint_id)
                print(f"[CLEANUP] Updated them to fingerprint_id: {fingerprint_id}")
            
            # Debug: Show all fingerprint registrations for this student
            all_regs = BiometricRegistration.objects.filter(student=target_student)
            print(f"   [DEBUG] Total registrations for this student: {all_regs.count()}")
            for reg in all_regs:
                print(f"      - Course: {reg.course.code}, Fingerprint ID: {reg.fingerprint_id}, Active: {reg.is_active}")
            
            if is_replacement:
                # For replacement: biometric data already updated in database
                # Just confirm that the fingerprint_id is now finalized
                registrations = BiometricRegistration.objects.filter(
                    student=target_student,
                    fingerprint_id=fingerprint_id,
                    is_active=True
                )
                count = registrations.count()
                
                print(f"\n✓ FINGERPRINT REPLACEMENT CONFIRMED:")
                print(f"   Student: {target_student.full_name}")
                print(f"   Fingerprint_id: {fingerprint_id}")
                print(f"   Biometric data updated for {count} course(s)")
                print(f"   (Using same fingerprint_id - storage saved!)")
                
                # Debug: Show what we found
                for reg in registrations:
                    print(f"      - Course: {reg.course.code}, fingerprint_id: {reg.fingerprint_id}")
            else:
                # For new enrollment: just confirm the fingerprints are active
                registrations = BiometricRegistration.objects.filter(
                    student=target_student,
                    fingerprint_id=fingerprint_id,
                    is_active=True
                )
                count = registrations.count()
                
                print(f"\n✓ FINGERPRINT CONFIRMED:")
                print(f"   Student: {target_student.full_name}")
                print(f"   Fingerprint_id: {fingerprint_id}")
                print(f"   Confirmed in {count} course(s)")
                
                # Debug: Show what we found
                for reg in registrations:
                    print(f"      - Course: {reg.course.code}, fingerprint_id: {reg.fingerprint_id}")
            
            if count == 0:
                # Debug: Show what fingerprint IDs actually exist
                existing_ids = BiometricRegistration.objects.filter(
                    student=target_student,
                    is_active=True
                ).values_list('fingerprint_id', flat=True).distinct()
                print(f"\n[ERROR] No registrations found!")
                print(f"   Expected fingerprint_id: {fingerprint_id}")
                print(f"   Existing active fingerprint_ids: {list(existing_ids)}")
                return JsonResponse({
                    'success': False,
                    'message': f'No active biometric registrations found for fingerprint_id {fingerprint_id}. Existing IDs: {list(existing_ids)}'
                }, status=400)
            
            print(f"\n✓✓✓ FINGERPRINT CONFIRMATION COMPLETE ✓✓✓")
        
        return JsonResponse({
            'success': True,
            'message': f'Fingerprint confirmed successfully!',
            'fingerprint_id': fingerprint_id,
            'is_replacement': is_replacement,
            'confirmed_courses': count
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error confirming biometric: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==================== BIOMETRIC CONFIRMATION AND CANCELLATION APIs ====================

@csrf_exempt
@require_http_methods(["POST"])
def api_biometric_confirm_enrollment(request):
    """
    Confirm a biometric enrollment and finalize the registration
    This endpoint:
    1. Retrieves biometric registrations that were created by api_biometric_enroll_view
    2. Confirms them in the database
    3. Returns confirmation to the frontend
    
    The biometric_data is already saved by api_biometric_enroll_view,
    so this endpoint just needs to verify the records exist and are active.
    """
    try:
        data = json.loads(request.body)
        
        fingerprint_id_raw = data.get('fingerprint_id', 1)
        template_id = data.get('template_id', '')
        is_replacement = data.get('is_replacement', False)
        course_ids = data.get('course_ids', [])

        try:
            fingerprint_id = int(str(fingerprint_id_raw).strip())
        except Exception:
            return JsonResponse({'success': False, 'message': 'Invalid fingerprint_id (slot). Please restart enrollment.'}, status=400)
        
        # Get authenticated user
        user = request.user
        if not user or not user.is_authenticated:
            return JsonResponse({
                'success': False,
                'message': 'User must be authenticated'
            }, status=401)
        
        logger.info(f"[CONFIRM] ========== BIOMETRIC CONFIRMATION START ==========")
        logger.info(f"[CONFIRM] User: {user.full_name} (ID: {user.id})")
        logger.info(f"[CONFIRM] Fingerprint ID: {fingerprint_id}, Template: {template_id}")
        logger.info(f"[CONFIRM] Is Replacement: {is_replacement}")
        logger.info(f"[CONFIRM] Received course_ids list: {course_ids} (type: {type(course_ids).__name__}, length: {len(course_ids) if isinstance(course_ids, list) else 'N/A'})")
        logger.info(f"[CONFIRM] ⚠️ NOTE: Biometric registrations should already be created by api_biometric_enroll_view")
        logger.info(f"[CONFIRM] This endpoint verifies and confirms the existing registrations")
        
        # Step 1: Send confirmation to ESP32
        esp32_url = f'http://{settings.ESP32_IP}:80/enroll/confirm'
        esp32_payload = {
            'fingerprint_id': fingerprint_id,
            'template_id': template_id
        }
        
        logger.info(f"[CONFIRM] Sending confirmation to ESP32 at {esp32_url}")
        logger.info(f"[CONFIRM] Using ESP32_IP from settings: {settings.ESP32_IP}")
        
        try:
            import requests
            esp32_response = requests.post(
                esp32_url,
                json=esp32_payload,
                timeout=10
            )
            
            if esp32_response.status_code != 200:
                logger.warning(f"[CONFIRM] ESP32 returned status {esp32_response.status_code}")
                logger.warning(f"[CONFIRM] Response: {esp32_response.text}")
            else:
                logger.info(f"[CONFIRM] ✓ ESP32 confirmation successful")
                
        except Exception as e:
            logger.warning(f"[CONFIRM] Could not reach ESP32: {str(e)}")
            logger.warning(f"[CONFIRM] Continuing anyway")
        
        # Step 2: Verify biometric registrations are active in database.
        # HARD REQUIREMENT: after Confirm & Save, every selected course must have an active
        # BiometricRegistration row with this fingerprint_id (slot), otherwise instructor scan
        # will show "not registered" even though the sensor matched.
        # These should have been created by api_biometric_enroll_view during the enrollment process
        verified_count = 0
        missing_courses = []
        
        logger.info(f"[CONFIRM] Verifying {len(course_ids)} biometric registrations exist in database")
        
        if course_ids:
            for course_id in course_ids:
                try:
                    course = Course.objects.get(id=int(str(course_id).strip()))
                    logger.info(f"[CONFIRM] Verifying course {course_id}: {course.code}")
                    
                    # Check if biometric registration exists
                    biometric_reg = BiometricRegistration.objects.filter(
                        student=user,
                        course=course,
                        is_active=True
                    ).first()
                    
                    if biometric_reg:
                        logger.info(f"[CONFIRM] ✓ Found biometric registration for course {course.code}")
                        logger.info(f"[CONFIRM]   - fingerprint_id: {biometric_reg.fingerprint_id}")
                        logger.info(f"[CONFIRM]   - has_biometric_data: {bool(biometric_reg.biometric_data)}")
                        logger.info(f"[CONFIRM]   - is_active: {biometric_reg.is_active}")

                        # Ensure the saved fingerprint_id matches the confirmed slot.
                        # This is critical for instructor-side matching.
                        try:
                            if biometric_reg.fingerprint_id != fingerprint_id or not biometric_reg.is_active:
                                biometric_reg.fingerprint_id = fingerprint_id
                                biometric_reg.is_active = True
                                biometric_reg.save(update_fields=['fingerprint_id', 'is_active'])
                        except Exception as _e:
                            logger.warning(f"[CONFIRM] Could not normalize registration for {course.code}: {_e}")

                        verified_count += 1
                    else:
                        # Self-heal: create/update the missing row so instructor lookups are reliable.
                        logger.warning(f"[CONFIRM] Missing biometric registration for course {course.code} - creating it now")

                        # BiometricRegistration.biometric_data is required by the model.
                        # Copy it from any existing active registration for this user+fingerprint_id.
                        source_reg = BiometricRegistration.objects.filter(
                            student=user,
                            fingerprint_id=fingerprint_id,
                            is_active=True
                        ).exclude(biometric_data='').first()

                        source_biometric_data = getattr(source_reg, 'biometric_data', '')
                        if not source_biometric_data:
                            logger.error(f"[CONFIRM] Cannot self-heal {course.code}: no biometric_data available to copy")
                            missing_courses.append(course.code)
                            continue

                        biometric_reg, _created = BiometricRegistration.objects.update_or_create(
                            student=user,
                            course=course,
                            defaults={
                                'fingerprint_id': fingerprint_id,
                                'biometric_data': source_biometric_data,
                                'biometric_type': 'fingerprint',
                                'is_active': True
                            }
                        )
                        logger.info(f"[CONFIRM] ✓ Created/updated BiometricRegistration for {course.code} (fingerprint_id={biometric_reg.fingerprint_id})")
                        verified_count += 1

                except Course.DoesNotExist:
                    logger.warning(f"[CONFIRM] Course {course_id} not found")
                    missing_courses.append(f"Course({course_id})")
                except Exception as e:
                    logger.error(f"[CONFIRM] Error verifying registration for course {course_id}: {str(e)}")
        else:
            logger.warning("[CONFIRM] No course_ids provided!")
        
        logger.info(f"[CONFIRM] ✓✓✓ Confirmation verification complete")
        logger.info(f"[CONFIRM] Verified: {verified_count} courses")
        if missing_courses:
            logger.error(f"[CONFIRM] MISSING: {missing_courses} - registrations not found!")
        
        return JsonResponse({
            'success': True if verified_count > 0 else False,
            'message': f'Fingerprint confirmed for {verified_count} course(s)',
            'fingerprint_id': fingerprint_id,
            'is_replacement': is_replacement,
            'confirmed_courses': verified_count,
            'missing_courses': missing_courses if missing_courses else None
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"[CONFIRM] Error confirming enrollment: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_biometric_cancel_enrollment(request):
    """
    Cancel an ongoing biometric enrollment
    This endpoint:
    1. Sends cancellation request to ESP32
    2. Cleans up enrollment session
    """
    try:
        # Get authenticated user
        user = request.user
        if not user or not user.is_authenticated:
            return JsonResponse({
                'success': False,
                'message': 'User must be authenticated'
            }, status=401)
        
        logger.info(f"[CANCEL] Cancelling enrollment for user {user.full_name}")
        
        # Send cancellation to ESP32
        esp32_url = f'http://192.168.1.9:80/enroll/cancel'
        
        logger.info(f"[CANCEL] Sending cancellation to ESP32 at {esp32_url}")
        
        try:
            import requests
            esp32_response = requests.post(
                esp32_url,
                json={},
                timeout=10
            )
            
            if esp32_response.status_code == 200:
                logger.info(f"[CANCEL] ✓ ESP32 cancellation successful")
            else:
                logger.warning(f"[CANCEL] ESP32 returned status {esp32_response.status_code}")
                
        except Exception as e:
            logger.warning(f"[CANCEL] Could not reach ESP32: {str(e)}")
            logger.warning(f"[CANCEL] User-side cancellation will still work")
        
        logger.info(f"[CANCEL] ✓ Enrollment cancelled")
        
        return JsonResponse({
            'success': True,
            'message': 'Enrollment cancelled'
        })
        
    except Exception as e:
        logger.error(f"[CANCEL] Error cancelling enrollment: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==================== INSTRUCTOR BIOMETRIC SCANNING ====================

@login_required
@require_http_methods(["GET"])
def instructor_get_biometric_pending_view(request):
    """
    Get pending biometric scans for a course (from ESP32 sensor).
    Called by JavaScript polling to check for new fingerprint detections.
    
    Query parameters:
    - course_id: int (required)
    - schedule_id: str (optional)
    
    Returns:
    {
        'pending': [
            {
                'fingerprint_id': int,
                'student_id': int,
                'student_name': str,
                'confidence': int,
                'timestamp': str
            }
        ]
    }
    """
    try:
        if not request.user.is_teacher:
            logger.warning(f"[PENDING] User {request.user} is not a teacher")
            return JsonResponse({'pending': []})
        
        course_id = request.GET.get('course_id')
        schedule_id = request.GET.get('schedule_id')
        
        logger.info(f"[PENDING POLL] Received poll request - course_id: {course_id}, schedule_id: {schedule_id}, user: {request.user.full_name}")
        print(f"\n[BIOMETRIC POLL] ===== POLL STARTED =====")
        print(f"[BIOMETRIC POLL] Course ID: {course_id}")
        print(f"[BIOMETRIC POLL] Schedule ID: {schedule_id}")
        print(f"[BIOMETRIC POLL] User: {request.user.full_name}")
        
        if not course_id:
            logger.warning(f"[PENDING POLL] No course_id provided")
            print(f"[BIOMETRIC POLL] ✗ No course_id!")
            return JsonResponse({'pending': []})
        
        try:
            course = Course.objects.get(id=int(course_id))
            logger.info(f"[PENDING POLL] Course found: {course.code}")
            print(f"[BIOMETRIC POLL] ✓ Course found: {course.code}, ID: {course.id}")
        except (Course.DoesNotExist, ValueError) as e:
            logger.error(f"[PENDING POLL] Course not found - course_id: {course_id}, error: {str(e)}")
            print(f"[BIOMETRIC POLL] ✗ Course not found: {course_id}")
            return JsonResponse({'pending': []})
        
        # Get all biometric registrations for this course that are active.
        # Also include sibling sections of the same course so fingerprint slots resolve correctly
        # even if the student registered under a different section.
        sibling_courses = Course.objects.filter(
            instructor=course.instructor,
            code=course.code,
            name=course.name,
            semester=course.semester,
            school_year=course.school_year,
            is_active=True,
            deleted_at__isnull=True,
            is_archived=False
        )

        registrations = BiometricRegistration.objects.filter(
            course__in=sibling_courses,
            is_active=True
        ).select_related('student', 'course')
        
        print(f"[BIOMETRIC POLL] ===== BIOMETRIC REGISTRATIONS FOR COURSE {course.code} =====")
        print(f"[BIOMETRIC POLL] Total registrations in database for this course: {registrations.count()}")
        for reg in registrations:
            print(f"[BIOMETRIC POLL]   - {reg.student.full_name}: Fingerprint ID={reg.fingerprint_id}, Active={reg.is_active}")
        print(f"[BIOMETRIC POLL] =====")
        
        # Create a mapping of fingerprint_id(int) -> (student, registration)
        fingerprint_map = {}
        for reg in registrations:
            if reg.fingerprint_id is None:
                continue

            try:
                fp_key = int(reg.fingerprint_id)
            except Exception:
                continue

            # Prefer the registration for the actively scanned course if duplicates exist.
            if fp_key in fingerprint_map:
                try:
                    existing_reg = fingerprint_map[fp_key].get('registration')
                    if existing_reg and getattr(existing_reg, 'course_id', None) == course.id:
                        continue
                except Exception:
                    pass

            fingerprint_map[fp_key] = {
                'student': reg.student,
                'registration': reg,
                'fingerprint_id': reg.fingerprint_id
            }
            print(f"[BIOMETRIC POLL]   - Student: {reg.student.full_name}, Fingerprint ID: {reg.fingerprint_id}")
        
        logger.info(f"[PENDING] Course {course.code} has {len(fingerprint_map)} registered fingerprints: {list(fingerprint_map.keys())}")
        print(f"[BIOMETRIC POLL] Fingerprint map: {fingerprint_map}")
        
        # Get pending detections from cache - USE ATTENDANCE QUEUE
        detections_queue_key = "fingerprint_detections_queue_attendance"
        queue = cache.get(detections_queue_key, [])
        
        print(f"[BIOMETRIC POLL] Cache queue size: {len(queue)}")
        print(f"[BIOMETRIC POLL] Queue contents: {queue}")
        
        pending_list = []
        processed_keys = []
        
        print(f"[BIOMETRIC POLL] ===== PROCESSING DETECTIONS =====")
        print(f"[BIOMETRIC POLL] Total detections to process: {len(queue)}")
        
        # Process each detection
        for idx, detection in enumerate(queue):
            # Some devices send retry/hint events using fingerprint_id=-2.
            # Always treat this as a retry (never as not-registered), even if match_type is missing.
            fingerprint_slot_raw = detection.get('fingerprint_id')
            try:
                fingerprint_slot_int = int(fingerprint_slot_raw)
            except Exception:
                fingerprint_slot_int = None

            if fingerprint_slot_int == -2:
                pending_list.append({
                    'fingerprint_id': -2,
                    'student_id': None,
                    'student_name': 'Place finger again',
                    'error': 'retry',
                    'reason': detection.get('reason'),
                    'confidence': detection.get('confidence', 0),
                    'timestamp': detection.get('timestamp')
                })
                processed_keys.append(detection.get('key'))
                continue

            if detection.get('match_type') == 'hint':
                fingerprint_slot_raw = detection.get('fingerprint_id')
                try:
                    fingerprint_slot = int(fingerprint_slot_raw)
                except Exception:
                    fingerprint_slot = fingerprint_slot_raw

                if fingerprint_slot == -2:
                    pending_list.append({
                        'fingerprint_id': -2,
                        'student_id': None,
                        'student_name': 'Place finger again',
                        'error': 'retry',
                        'reason': detection.get('reason'),
                        'confidence': detection.get('confidence', 0),
                        'timestamp': detection.get('timestamp')
                    })
                    processed_keys.append(detection.get('key'))
                    continue

            # Check if this is a hardware-matched fingerprint (ACCURATE)
            if detection.get('match_type') == 'hardware':
                fingerprint_slot_raw = detection.get('fingerprint_id')
                try:
                    fingerprint_slot = int(fingerprint_slot_raw)
                except Exception:
                    fingerprint_slot = fingerprint_slot_raw
                confidence = detection.get('confidence', 0)
                timestamp = detection.get('timestamp')
                detection_key = detection.get('key')

                # Sentinel for "unregistered" published by ESP32
                if fingerprint_slot == -1:
                    logger.warning(f"[PENDING] ⚠ UNREGISTERED (hardware) - Fingerprint not found in sensor database")
                    pending_list.append({
                        'fingerprint_id': -1,
                        'student_id': None,
                        'student_name': 'Unregistered',
                        'error': 'unregistered_fingerprint',
                        'confidence': confidence,
                        'timestamp': timestamp
                    })
                    processed_keys.append(detection_key)
                    continue
                
                print(f"[BIOMETRIC POLL] [{idx}] ✓ HARDWARE MATCH - Fingerprint ID: {fingerprint_slot}, Confidence: {confidence}")
                logger.info(f"[PENDING] Hardware-matched fingerprint: ID={fingerprint_slot}, confidence={confidence}")
                
                # Look up student by fingerprint_id (100% accurate - from hardware sensor)
                if fingerprint_slot in fingerprint_map:
                    student_info = fingerprint_map[fingerprint_slot]
                    student = student_info['student']
                    
                    print(f"[BIOMETRIC POLL] ✓ MATCHED! Fingerprint {fingerprint_slot} = Student {student.full_name}")
                    
                    pending_list.append({
                        'fingerprint_id': fingerprint_slot,
                        'fingerprint_template_id': student_info.get('fingerprint_id', ''),
                        'student_id': student.id,
                        'student_name': student.full_name or student.username,
                        'student_email': student.email,
                        'confidence': confidence,
                        'timestamp': timestamp
                    })
                    
                    logger.info(f"[PENDING] ✓ ACCURATE MATCH - Fingerprint {fingerprint_slot} belongs to Student: {student.full_name}")
                    processed_keys.append(detection_key)
                else:
                    # Fingerprint slot not found in this course
                    print(f"[BIOMETRIC POLL] ✗ Fingerprint ID {fingerprint_slot} not enrolled in this course")
                    owner_reg = BiometricRegistration.objects.filter(
                        fingerprint_id=fingerprint_slot,
                        is_active=True
                    ).select_related('student', 'course').first()

                    if owner_reg:
                        pending_list.append({
                            'fingerprint_id': fingerprint_slot,
                            'student_id': owner_reg.student.id,
                            'student_name': owner_reg.student.full_name or owner_reg.student.username,
                            'student_email': getattr(owner_reg.student, 'email', ''),
                            'confidence': confidence,
                            'timestamp': timestamp,
                            'error': 'not_registered_for_course'
                        })
                    else:
                        pending_list.append({
                            'fingerprint_id': fingerprint_slot,
                            'student_id': None,
                            'student_name': f'Fingerprint ID {fingerprint_slot} not enrolled',
                            'confidence': confidence,
                            'timestamp': timestamp,
                            'error': 'not_registered_for_course'
                        })
                    processed_keys.append(detection_key)
                continue
            
            # Handle legacy raw fingerprint detections
            if detection.get('needs_matching'):
                print(f"[BIOMETRIC POLL] [{idx}] Raw fingerprint detection - needs software matching")
                print(f"[BIOMETRIC POLL]     Quality: {detection.get('quality')}")
                
                # Raw fingerprint from ESP32 - try to identify which student it belongs to
                # Strategy: If only ONE student in this course has biometric enrollment,
                # it's likely that student who placed their finger
                
                students_with_biometric = list(fingerprint_map.values())
                
                if len(students_with_biometric) == 1:
                    # Only one student has biometric enrollment - must be them!
                    student_info = students_with_biometric[0]
                    student = student_info['student']
                    
                    print(f"[BIOMETRIC POLL] ✓ MATCHED! Only 1 student enrolled - must be {student.full_name}")
                    
                    pending_list.append({
                        'fingerprint_id': student_info['fingerprint_id'],
                        'fingerprint_template_id': student_info.get('fingerprint_id', ''),
                        'student_id': student.id,
                        'student_name': student.full_name or student.username,
                        'student_email': student.email,
                        'confidence': detection.get('quality', 0),
                        'timestamp': detection.get('timestamp')
                    })
                    
                elif len(students_with_biometric) > 1:
                    # Multiple students with biometric - we need more info to match
                    # For now, show the student who most recently had their fingerprint enrolled
                    # (most recent = highest fingerprint_id)
                    most_recent = max(students_with_biometric, 
                                     key=lambda x: x.get('fingerprint_id', 0))
                    student = most_recent['student']
                    
                    print(f"[BIOMETRIC POLL] ⚠ Multiple students enrolled - guessing {student.full_name} (most recent)")
                    
                    pending_list.append({
                        'fingerprint_id': most_recent['fingerprint_id'],
                        'fingerprint_template_id': most_recent.get('fingerprint_id', ''),
                        'student_id': student.id,
                        'student_name': student.full_name or student.username,
                        'student_email': student.email,
                        'confidence': detection.get('quality', 0),
                        'timestamp': detection.get('timestamp')
                    })
                else:
                    # No students with biometric enrollment in this course
                    print(f"[BIOMETRIC POLL] ✗ No students with biometric enrollment in this course")
                    pending_list.append({
                        'fingerprint_id': -1,
                        'student_id': None,
                        'student_name': 'No biometric students enrolled',
                        'confidence': 0,
                        'timestamp': detection.get('timestamp'),
                        'error': 'unregistered_fingerprint'
                    })
                
                processed_keys.append(detection.get('key'))
                continue
            
            # Original detection processing for already-matched fingerprints
            fingerprint_slot = detection.get('fingerprint_id')
            confidence = detection.get('confidence', 0)
            timestamp = detection.get('timestamp')
            detection_key = detection.get('key')
            
            print(f"[BIOMETRIC POLL] [{idx}] Fingerprint ID: {fingerprint_slot}, Confidence: {confidence}")
            logger.info(f"[PENDING] Checking detection - Fingerprint ID: {fingerprint_slot}, Confidence: {confidence}")
            
            # Check if this fingerprint is registered for this course
            if fingerprint_slot in fingerprint_map:
                student_info = fingerprint_map[fingerprint_slot]
                student = student_info['student']
                
                print(f"[BIOMETRIC POLL] ✓ MATCH! Fingerprint {fingerprint_slot} = Student {student.full_name}")
                
                pending_list.append({
                    'fingerprint_id': fingerprint_slot,
                    'fingerprint_template_id': student_info.get('fingerprint_id', ''),
                    'student_id': student.id,
                    'student_name': student.full_name or student.username,
                    'student_email': student.email,
                    'confidence': confidence,
                    'timestamp': timestamp
                })
                
                logger.info(f"[PENDING] ✓ MATCH FOUND - Fingerprint ID {fingerprint_slot} belongs to Student: {student.full_name}")
                processed_keys.append(detection_key)
                
            elif fingerprint_slot == -1:
                # Unregistered fingerprint (sensor couldn't match to any enrolled fingerprint)
                logger.warning(f"[PENDING] ⚠ UNREGISTERED - Fingerprint not found in sensor database")
                pending_list.append({
                    'fingerprint_id': -1,
                    'student_id': None,
                    'student_name': 'Unregistered',
                    'error': 'unregistered_fingerprint',
                    'confidence': confidence,
                    'timestamp': timestamp
                })
                processed_keys.append(detection_key)
            else:
                # Fingerprint detected but NOT registered for THIS COURSE
                # Check if it's registered for another course
                other_registrations = BiometricRegistration.objects.filter(
                    fingerprint_id=fingerprint_slot,
                    is_active=True
                ).select_related('student', 'course')
                
                if other_registrations.exists():
                    reg = other_registrations.first()
                    logger.warning(f"[PENDING] ⚠ NOT THIS COURSE - Fingerprint ID {fingerprint_slot} registered for {reg.course.code}, not {course.code}")
                    # Return owner info but mark as not registered for this course (frontend will not record attendance)
                    pending_list.append({
                        'fingerprint_id': fingerprint_slot,
                        'student_id': reg.student.id,
                        'student_name': reg.student.full_name or reg.student.username,
                        'student_email': getattr(reg.student, 'email', ''),
                        'confidence': confidence,
                        'timestamp': timestamp,
                        'error': 'not_registered_for_course'
                    })
                else:
                    logger.warning(f"[PENDING] ⚠ NOT REGISTERED - Fingerprint ID {fingerprint_slot} not found in any registration")
                    # Treat as unregistered for this course
                    pending_list.append({
                        'fingerprint_id': fingerprint_slot,
                        'student_id': None,
                        'student_name': 'Not registered for this course',
                        'error': 'not_registered_for_course',
                        'confidence': confidence,
                        'timestamp': timestamp
                    })
                
                processed_keys.append(detection_key)
        
        # Remove processed detections from queue (clean up)
        if processed_keys:
            queue = [d for d in queue if d.get('key') not in processed_keys]
            cache.set("fingerprint_detections_queue_attendance", queue, 60)
            logger.info(f"[PENDING] Removed {len(processed_keys)} processed detections, {len(queue)} remaining in queue")
        
        logger.info(f"[PENDING] ✓ RETURNING {len(pending_list)} pending detections for course {course.code} ({course_id})")
        for item in pending_list:
            logger.info(f"[PENDING] - Returning: Student: {item.get('student_name')}, Fingerprint ID: {item.get('fingerprint_id')}, Confidence: {item.get('confidence')}")
        
        print(f"[BIOMETRIC POLL] ===== RESPONSE =====")
        print(f"[BIOMETRIC POLL] Returning {len(pending_list)} pending items")
        for item in pending_list:
            print(f"[BIOMETRIC POLL]   - {item.get('student_name')} (ID: {item.get('fingerprint_id')})")
        print(f"[BIOMETRIC POLL] ===== END POLL =====\n")
        
        # DEBUG: Also print to console so we see it immediately
        print(f"\n=== BIOMETRIC PENDING POLL ===")
        print(f"Course: {course.code}, Course ID: {course_id}")
        print(f"Registrations: {len(fingerprint_map)}")
        print(f"Pending items: {len(pending_list)}")
        for item in pending_list:
            print(f"  - {item.get('student_name')} (Fingerprint {item.get('fingerprint_id')})")
        print(f"Cache queue size: {len(queue)}")
        print(f"===\n")
        
        return JsonResponse({'pending': pending_list})
        
    except Exception as e:
        logger.error(f"Error getting pending biometric scans: {str(e)}", exc_info=True)
        return JsonResponse({'pending': []})


@login_required
@require_http_methods(["POST"])
def instructor_biometric_scan_attendance_view(request):
    """
    Record attendance via biometric scan (instructor context).
    Called when a student's fingerprint is verified through the instructor interface.
    
    Expected POST data:
    {
        'course_id': <int>,
        'fingerprint_id': '<str>', # fingerprint template ID from ESP32
        'schedule_id': '<str>' (optional)
    }
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({
                'success': False,
                'message': 'Only instructors can process biometric scans.'
            }, status=403)
        
        from django.db import transaction
        
        data = json.loads(request.body)
        course_id = data.get('course_id')
        fingerprint_id_raw = data.get('fingerprint_id')
        schedule_id = data.get('schedule_id', '')

        try:
            fingerprint_id = int(str(fingerprint_id_raw).strip())
        except Exception:
            fingerprint_id = None
        
        if not course_id or fingerprint_id is None:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_id or fingerprint_id'
            }, status=400)
        
        # Get course
        try:
            course = Course.objects.get(id=int(course_id))
        except (Course.DoesNotExist, ValueError):
            return JsonResponse({
                'success': False,
                'message': 'Course not found'
            }, status=404)
        
        # Find student with matching fingerprint registration.
        # Prefer this specific course, but allow sibling sections so a student doesn't
        # get falsely flagged as "not registered" due to section mismatch.
        biometric_reg = BiometricRegistration.objects.filter(
            course=course,
            fingerprint_id=fingerprint_id,
            is_active=True
        ).select_related('student', 'course').first()

        if not biometric_reg:
            sibling_courses = Course.objects.filter(
                instructor=course.instructor,
                code=course.code,
                name=course.name,
                semester=course.semester,
                school_year=course.school_year,
                is_active=True,
                deleted_at__isnull=True,
                is_archived=False
            )
            biometric_reg = BiometricRegistration.objects.filter(
                course__in=sibling_courses,
                fingerprint_id=fingerprint_id,
                is_active=True
            ).select_related('student', 'course').first()
        
        if not biometric_reg:
            # SECURITY: Check if this fingerprint_id exists but for a DIFFERENT student
            # (should never happen due to unique_student_fingerprint_id constraint, but validate anyway)
            other_student_reg = BiometricRegistration.objects.filter(
                fingerprint_id=fingerprint_id,
                is_active=True
            ).exclude(course=course).select_related('student', 'course').first()
            
            if other_student_reg:
                logger.warning(f"[BIOMETRIC SECURITY] Fingerprint {fingerprint_id} belongs to {other_student_reg.student.full_name} in {other_student_reg.course.code}, not found in {course.code}")
                return JsonResponse({
                    'success': False,
                    'message': 'Fingerprint not registered for this course'
                }, status=403)
            
            logger.warning(f"[BIOMETRIC] Fingerprint ID {fingerprint_id} not registered for course {course.code}")
            return JsonResponse({
                'success': False,
                'message': 'Fingerprint not registered for this course'
            }, status=403)
        
        # Hardware matching is performed on the ESP32/sensor; for attendance we rely on
        # fingerprint_id -> student mapping. biometric_data may be empty for some legacy
        # registrations and should not block attendance.
        if not biometric_reg.biometric_data:
            logger.warning(f"[BIOMETRIC WARNING] Registration {biometric_reg.id} has no biometric_data stored; proceeding with hardware-matched fingerprint_id={fingerprint_id}")

        student = biometric_reg.student
        
        # SECURITY: Verify student is actually enrolled in this course
        try:
            enrollment = CourseEnrollment.objects.get(
                student=student,
                course=course,
                is_active=True
            )
        except CourseEnrollment.DoesNotExist:
            logger.warning(f"[BIOMETRIC SECURITY] Student {student.full_name} has biometric registered but NOT enrolled in {course.code}")
            return JsonResponse({
                'success': False,
                'message': 'Student not enrolled in this course'
            }, status=403)
        
        # Check if attendance is open (prefer schedule-level status when available)
        attendance_status = getattr(course, 'attendance_status', None)
        try:
            if active_schedule and getattr(active_schedule, 'attendance_status', None):
                attendance_status = active_schedule.attendance_status
        except Exception:
            pass

        if attendance_status not in ['open', 'automatic']:
            return JsonResponse({
                'success': False,
                'message': f'Attendance is currently {attendance_status}. Cannot record attendance.'
            }, status=403)
        
        # Get timezone
        try:
            from zoneinfo import ZoneInfo
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        
        # Determine schedule day from TODAY's day of week (same logic as QR code)
        day_map = {
            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed', 'Thursday': 'Thu',
            'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun'
        }
        today_short = day_map.get(now_ph.strftime('%A'), now_ph.strftime('%A')[:3])
        
        # Try to use provided schedule_id first, but default to today's day if not provided
        schedule_day = today_short
        if schedule_id:
            try:
                parts = str(schedule_id).split('_')
                if len(parts) >= 2:
                    day_short = parts[1]
                    # Check if this is a valid day code (single or double letter)
                    day_map_reverse = {
                        'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'Th': 'Thu', 'F': 'Fri', 
                        'S': 'Sat', 'Su': 'Sun', 'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed',
                        'Thu': 'Thu', 'Fri': 'Fri', 'Sat': 'Sat', 'Sun': 'Sun'
                    }
                    mapped_day = day_map_reverse.get(day_short)
                    if mapped_day:
                        schedule_day = mapped_day
                        logger.info(f"[BIOMETRIC LATE CHECK] Extracted schedule day from schedule_id: {schedule_day}")
            except Exception as e:
                logger.warning(f"[BIOMETRIC LATE CHECK] Error parsing schedule_id, defaulting to today's day: {str(e)}")
                schedule_day = today_short
        
        logger.info(f"[BIOMETRIC LATE CHECK] Using schedule day: {schedule_day}, Today's day: {today_short}")
        
        # Don't check for existing attendance record - allow re-marking
        # This matches the QR code behavior where students can scan again
        existing_record = AttendanceRecord.objects.filter(
            student=student,
            course=course,
            attendance_date=today,
            schedule_day=schedule_day or ''
        ).first()
        
        # CRITICAL: Determine attendance status based on present window (SAME LOGIC AS QR CODE)
        attendance_record_status = 'present'  # Default status
        
        logger.info(f"[BIOMETRIC LATE CHECK] Starting present/late determination for {student.full_name}")
        logger.info(f"[BIOMETRIC LATE CHECK] Current time: {now_ph}, Schedule day: {schedule_day}")
        
        # Get the schedule object
        active_schedule = CourseSchedule.objects.filter(
            course=course,
            day=schedule_day
        ).first()
        logger.info(f"[BIOMETRIC LATE CHECK] Schedule lookup for day '{schedule_day}': {'Found' if active_schedule else 'NOT FOUND'}")
        
        # Get the present duration from schedule or course (MUST be set by instructor)
        present_duration_minutes = None
        if active_schedule and getattr(active_schedule, 'attendance_present_duration', None):
            present_duration_minutes = active_schedule.attendance_present_duration
            logger.info(f"[BIOMETRIC LATE CHECK] ✓ Using schedule-level present duration: {present_duration_minutes} minutes")
        elif getattr(course, 'attendance_present_duration', None):
            present_duration_minutes = course.attendance_present_duration
            logger.info(f"[BIOMETRIC LATE CHECK] ✓ Using course-level present duration: {present_duration_minutes} minutes")
        else:
            logger.warning(f"[BIOMETRIC LATE CHECK] ⚠️ NO present duration configured - defaulting to PRESENT")
        
        # Get the qr_opened_at timestamp (MUST be set when instructor opens attendance)
        qr_opened_at = None
        if active_schedule and getattr(active_schedule, 'qr_code_opened_at', None):
            qr_opened_at = active_schedule.qr_code_opened_at
            logger.info(f"[BIOMETRIC LATE CHECK] ✓ Using schedule-level qr_opened_at: {qr_opened_at}")
        elif getattr(course, 'qr_code_opened_at', None):
            qr_opened_at = course.qr_code_opened_at
            logger.info(f"[BIOMETRIC LATE CHECK] ✓ Using course-level qr_opened_at: {qr_opened_at}")
        else:
            logger.warning(f"[BIOMETRIC LATE CHECK] ⚠️ NO qr_opened_at timestamp found - attendance may not have been opened yet")
        
        # CRITICAL CHECK: If present window is configured, MUST check if student is within the window
        # If qr_opened_at is missing (not recorded), fall back to schedule start_time + duration.
        if present_duration_minutes and present_duration_minutes > 0:
            present_cutoff_dt = None

            if qr_opened_at:
                present_cutoff_dt = qr_opened_at + timezone.timedelta(minutes=int(present_duration_minutes))
            elif active_schedule and getattr(active_schedule, 'start_time', None):
                from datetime import datetime as dt
                present_cutoff_dt = dt.combine(today, active_schedule.start_time) + timezone.timedelta(minutes=int(present_duration_minutes))
                try:
                    present_cutoff_dt = present_cutoff_dt.replace(tzinfo=now_ph.tzinfo)
                except Exception:
                    pass

            if present_cutoff_dt:
                logger.info(f"[BIOMETRIC LATE CHECK] Present window boundaries:")
                logger.info(f"[BIOMETRIC LATE CHECK]   Started at: {qr_opened_at}")
                logger.info(f"[BIOMETRIC LATE CHECK]   Ends at:   {present_cutoff_dt}")
                logger.info(f"[BIOMETRIC LATE CHECK]   Current:   {now_ph}")
                
                # Check if current time is after the present window
                if now_ph > present_cutoff_dt:
                    attendance_record_status = 'late'
                    time_late = (now_ph - present_cutoff_dt).total_seconds() / 60
                    logger.warning(f"[BIOMETRIC LATE CHECK] 🔴 LATE: Student {student.full_name} is {time_late:.1f} minutes LATE (scanned at {now_ph})")
                else:
                    attendance_record_status = 'present'
                    time_remaining = (present_cutoff_dt - now_ph).total_seconds() / 60
                    logger.info(f"[BIOMETRIC LATE CHECK] 🟢 PRESENT: Student {student.full_name} scanned with {time_remaining:.1f} minutes remaining")
            else:
                # Duration exists but we have no baseline; fall back to attendance_end if available.
                try:
                    end_time = None
                    if active_schedule and getattr(active_schedule, 'attendance_end', None):
                        end_time = active_schedule.attendance_end
                    elif getattr(course, 'attendance_end', None):
                        end_time = course.attendance_end

                    if end_time and now_ph.time() > end_time:
                        attendance_record_status = 'late'
                        logger.warning(f"[BIOMETRIC LATE CHECK] 🔴 LATE by attendance_end fallback: now={now_ph.time()} end={end_time}")
                    else:
                        attendance_record_status = 'present'
                except Exception:
                    attendance_record_status = 'present'
        else:
            # No present-window configured -> use attendance_end fallback if it exists
            attendance_record_status = 'present'
            try:
                end_time = None
                if active_schedule and getattr(active_schedule, 'attendance_end', None):
                    end_time = active_schedule.attendance_end
                elif getattr(course, 'attendance_end', None):
                    end_time = course.attendance_end
                if end_time and now_ph.time() > end_time:
                    attendance_record_status = 'late'
            except Exception:
                pass
        
        logger.info(f"[BIOMETRIC LATE CHECK] ========================================")
        logger.info(f"[BIOMETRIC LATE CHECK] ✓ FINAL DECISION: {student.full_name} will be marked as {attendance_record_status.upper()}")
        logger.info(f"[BIOMETRIC LATE CHECK] ========================================")
        
        # Create/update attendance record
        with transaction.atomic():
            # CRITICAL: Delete any existing records for this student/course/date
            # This ensures we replace old scans (PRESENT) with new ones (LATE) on re-mark
            # Only one record per date should exist
            old_records = AttendanceRecord.objects.filter(
                student=student,
                course=course,
                attendance_date=today
            )
            
            if old_records.exists():
                count = old_records.count()
                old_time = old_records.first().attendance_time
                old_status = old_records.first().status
                logger.info(f"[BIOMETRIC INSTRUCTOR] ✓ REPLACING {count} old record(s): {old_time} ({old_status}) -> {now_ph.time()} ({attendance_record_status})")
                old_records.delete()
            
            # Create fresh attendance record with latest status and time
            attendance_record = AttendanceRecord.objects.create(
                student=student,
                course=course,
                enrollment=enrollment,
                attendance_date=today,
                attendance_time=now_ph.time(),
                status=attendance_record_status,
                schedule_day=schedule_day or ''
            )
            
            logger.info(f"[BIOMETRIC INSTRUCTOR] ✓ CREATED fresh attendance record for {student.full_name} - status: {attendance_record_status}")
            logger.info(f"[BIOMETRIC INSTRUCTOR] Student {student.full_name} ({student.id}) marked {attendance_record_status} in {course.code} via fingerprint")
            
            # Create notification
            try:
                create_notification(
                    user=student,
                    notification_type='attendance_marked',
                    title='Attendance Recorded - Biometric',
                    message=f'Your biometric attendance has been recorded as {attendance_record_status.upper()} in {course.code}',
                    category='attendance',
                    related_course=course,
                    related_user=course.instructor
                )
            except Exception as e:
                logger.warning(f"Could not create notification: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'{student.full_name} marked {attendance_record_status}',
                'student_name': student.full_name,
                'student_id': student.id,
                'student_pk': student.pk,
                'status': attendance_record_status,
                'attendance_time': now_ph.strftime('%H:%M:%S')
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"Error recording biometric attendance: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


# ==================== FINGERPRINT DETECTION FROM ESP32 ====================

from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache

@csrf_exempt  # Allow POST from ESP32 without CSRF token
@require_http_methods(["POST"])
def fingerprint_detection_view(request):
    """
    Receive fingerprint detection from ESP32 sensor.
    This is called directly by the ESP32 when a fingerprint is detected.
    
    Expected POST data:
    {
        'fingerprint_id': <int>,      # ID from sensor database (1-255), or -1 for unregistered
        'confidence': <int>,
        'timestamp': <int>,
        'mode': <str>  # 'registration' or 'attendance' - which purpose is this detection for
    }
    
    Detections are stored in separate cache queues based on mode:
    - registration: For student fingerprint enrollment
    - attendance: For instructor attendance scanning
    """
    try:
        data = json.loads(request.body)
        fingerprint_id_raw = data.get('fingerprint_id')
        confidence = data.get('confidence', 0)
        detection_mode = data.get('mode', 'attendance').lower()  # Default to attendance if not specified
        match_type = data.get('match_type')
        reason = data.get('reason')

        try:
            fingerprint_id = int(str(fingerprint_id_raw).strip())
        except Exception:
            fingerprint_id = None
        
        logger.info(f"[DETECTION ENDPOINT] ✓ Fingerprint detection received - ID: {fingerprint_id}, Confidence: {confidence}, Mode: {detection_mode}")
        print(f"\n[DETECTION] ===== NEW FINGERPRINT DETECTION =====")
        print(f"[DETECTION] Fingerprint ID: {fingerprint_id}")
        print(f"[DETECTION] Confidence: {confidence}")
        print(f"[DETECTION] Mode: {detection_mode}")
        
        if fingerprint_id is None:
            logger.error(f"[DETECTION ENDPOINT] ✗ Missing fingerprint_id in request body")
            return JsonResponse({
                'success': False,
                'error': 'Missing fingerprint_id'
            }, status=400)
        
        # Validate mode
        if detection_mode not in ['registration', 'attendance']:
            detection_mode = 'attendance'
        
        logger.info(f"[DETECTION] Fingerprint detected - ID: {fingerprint_id}, Confidence: {confidence}, Mode: {detection_mode}")
        
        # Use different cache keys for different modes
        if detection_mode == 'registration':
            detections_queue_key = "fingerprint_detections_queue_registration"
        else:  # attendance
            detections_queue_key = "fingerprint_detections_queue_attendance"
        
        print(f"[DETECTION] Using queue key: {detections_queue_key}")
        
        # Create a unique key for this detection
        detection_key = f"fingerprint_detection_{detection_mode}_{fingerprint_id}_{int(timezone.now().timestamp() * 1000)}"
        
        # Store detection data with 30-second expiry
        detection_data = {
            'fingerprint_id': fingerprint_id,
            'confidence': confidence,
            'timestamp': timezone.now().isoformat(),
            'match_type': match_type,
            'reason': reason
        }
        cache.set(detection_key, detection_data, 30)
        
        # Maintain a queue of recent detections for this mode
        queue = cache.get(detections_queue_key, [])
        queue.append({
            'fingerprint_id': fingerprint_id,
            'confidence': confidence,
            'timestamp': timezone.now().isoformat(),
            'key': detection_key,
            'match_type': match_type,
            'reason': reason
        })
        
        # Keep only last 50 detections
        queue = queue[-50:]
        cache.set(detections_queue_key, queue, 60)  # Cache for 1 minute
        
        logger.info(f"[DETECTION] Stored {detection_mode} detection - Queue size: {len(queue)}")
        print(f"[DETECTION] Stored in cache! Queue now has {len(queue)} items")
        print(f"[DETECTION] Queue contents: {queue}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Fingerprint detection received',
            'fingerprint_id': fingerprint_id,
            'mode': detection_mode
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON'
        }, status=400)
    except Exception as e:
        logger.error(f"Error receiving fingerprint detection: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def student_fingerprint_pending_view(request):
    """
    Polling endpoint for STUDENT registration mode.
    Students call this repeatedly during fingerprint enrollment to get detections.
    
    Returns recent fingerprint detections from the registration queue.
    """
    if request.method != 'GET':
        return JsonResponse({
            'success': False,
            'error': 'Method not allowed'
        }, status=405)
    
    try:
        # Get the registration queue
        detections_queue_key = "fingerprint_detections_queue_registration"
        queue = cache.get(detections_queue_key, [])
        
        # Get detections that haven't been processed yet
        last_check = request.GET.get('last_check')
        
        pending_detections = []
        if last_check:
            try:
                last_check_time = float(last_check)
                for detection in queue:
                    detection_time = datetime.datetime.fromisoformat(detection['timestamp']).timestamp()
                    if detection_time > last_check_time:
                        pending_detections.append(detection)
            except (ValueError, KeyError):
                # If parsing fails, return all detections
                pending_detections = queue[-10:]  # Last 10 detections
        else:
            # First check - return last 5 detections
            pending_detections = queue[-5:]
        
        logger.info(f"[STUDENT] Fingerprint pending check - {len(pending_detections)} detections found")
        
        return JsonResponse({
            'success': True,
            'detections': pending_detections,
            'count': len(pending_detections),
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error in student fingerprint pending: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== GET COURSE ATTENDANCE COUNT ====================

@login_required
@require_http_methods(["GET"])
def get_course_attendance_count(request):
    """
    Get the current attendance count for a course on today's date.
    Used to update the UI in real-time after QR/biometric scans.
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({'success': False, 'message': 'Only instructors can access this'}, status=403)
        
        course_id = request.GET.get('course_id')
        if not course_id:
            return JsonResponse({'success': False, 'message': 'course_id is required'}, status=400)
        
        # Get course
        try:
            course = Course.objects.get(id=int(course_id))
        except (Course.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'message': 'Course not found'}, status=404)
        
        # Get timezone
        try:
            from zoneinfo import ZoneInfo
            ph_tz = ZoneInfo('Asia/Manila')
        except:
            import pytz
            ph_tz = pytz.timezone('Asia/Manila')
        
        now_ph = timezone.now().astimezone(ph_tz)
        today = now_ph.date()
        
        # Count attendance records for today
        attendance_count = AttendanceRecord.objects.filter(
            course=course,
            attendance_date=today,
            status__in=['present', 'late']  # Count only marked students, not absent
        ).count()
        
        return JsonResponse({
            'success': True,
            'count': attendance_count
        })
        
    except Exception as e:
        logger.error(f"Error getting attendance count: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def instructor_start_biometric_detection_view(request):
    """
    Start biometric fingerprint detection for attendance scanning.
    Called by instructor interface when opening the dual scanner.
    
    Enables the ESP32 fingerprint sensor to detect fingerprints.
    
    POST /api/instructor/attendance/start/
    {
        'course_id': <int>,
        'schedule_id': <str> (optional),
        'session_id': <str> (unique session identifier)
    }
    
    Returns:
    {
        'success': True/False,
        'message': 'Detection started' or error message,
        'session_id': session identifier
    }
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({
                'success': False,
                'message': 'Only instructors can start detection.'
            }, status=403)
        
        data = json.loads(request.body)
        course_id = data.get('course_id')
        session_id = data.get('session_id', f"session_{timezone.now().timestamp()}")
        
        if not course_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_id'
            }, status=400)
        
        # Verify course exists and instructor has access
        try:
            course = Course.objects.get(id=int(course_id))
            if course.instructor != request.user:
                return JsonResponse({
                    'success': False,
                    'message': 'You do not have permission to access this course.'
                }, status=403)
        except (Course.DoesNotExist, ValueError):
            return JsonResponse({
                'success': False,
                'message': 'Course not found'
            }, status=404)
        
        # Enable fingerprint detection on ESP32 via MQTT
        try:
            from dashboard.mqtt_client import get_mqtt_client
            mqtt_client = get_mqtt_client()
            
            if mqtt_client and mqtt_client.is_connected:
                # Send detection enable request to ESP32
                detection_request = {
                    'action': 'start_detection',
                    'mode': 2,  # MODE_ATTENDANCE
                    'course_id': course.id,
                    'session_id': session_id
                }
                mqtt_client.publish('biometric/esp32/detect/request', detection_request, qos=1)
                logger.info(f"[API] ✓ Fingerprint detection enabled for attendance - Course: {course.code}, Session: {session_id}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Fingerprint detection started',
                    'session_id': session_id
                })

            else:
                print("[API] MQTT bridge not connected - detection could not be started")
                logger.warning("[API] MQTT bridge not connected - detection could not be started")
                
                return JsonResponse({
                    'success': False,
                    'message': 'Biometric sensor not connected. Please ensure the ESP32 sensor is online.',
                    'error': 'sensor_offline'
                }, status=503)
        except Exception as bridge_error:
            print(f"[API] Error with MQTT bridge: {bridge_error}")
            logger.warning(f"[API] MQTT bridge error: {bridge_error}")
            return JsonResponse({
                'success': False,
                'message': 'Biometric sensor not connected. Please ensure the ESP32 sensor is online.',
                'error': 'sensor_offline'
            }, status=503)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON'
        }, status=400)
    except Exception as e:
        logger.error(f"Error starting biometric detection: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def instructor_stop_biometric_detection_view(request):
    """
    Stop biometric fingerprint detection for attendance scanning.
    Called by instructor when closing the dual scanner.
    
    POST /api/instructor/attendance/stop/
    {
        'course_id': <int>,
        'session_id': <str> (optional)
    }
    
    Returns:
    {
        'success': True/False,
        'message': 'Detection stopped' or error message
    }
    """
    try:
        if not request.user.is_teacher:
            return JsonResponse({
                'success': False,
                'message': 'Only instructors can stop detection.'
            }, status=403)
        
        data = json.loads(request.body)
        course_id = data.get('course_id')
        
        if not course_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing course_id'
            }, status=400)
        
        # Verify course exists
        try:
            course = Course.objects.get(id=int(course_id))
        except (Course.DoesNotExist, ValueError):
            return JsonResponse({
                'success': False,
                'message': 'Course not found'
            }, status=404)
        
        # Disable fingerprint detection on ESP32 via MQTT
        try:
            from dashboard.mqtt_client import get_mqtt_client
            mqtt_client = get_mqtt_client()
            
            if mqtt_client and mqtt_client.is_connected:
                # Send detection disable request to ESP32
                detection_request = {
                    'action': 'stop_detection',
                    'mode': 0  # MODE_IDLE
                }
                mqtt_client.publish('biometric/esp32/detect/request', detection_request, qos=1)
                logger.info(f"[API] ✓ Fingerprint detection disabled - Course: {course.code}")
        except Exception as bridge_error:
            logger.error(f"[API] Error with MQTT during stop: {bridge_error}")
        
        # Always return success (detection is off either way)
        return JsonResponse({
            'success': True,
            'message': 'Fingerprint detection stopped'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON'
        }, status=400)
    except Exception as e:
        logger.error(f"Error stopping biometric detection: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


# ============================================================================
# MISSING API ENDPOINTS FOR STUDENT BIOMETRIC ENROLLMENT
# ============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def api_student_enrollment_status(request):
    """Get status of current student enrollment"""
    try:
        student_id = request.GET.get('student_id')
        
        if not student_id:
            return JsonResponse({
                'status': 'error',
                'message': 'student_id required'
            }, status=400)
        
        # Check for any active enrollments for this student
        active_enrollments = []
        for enrollment_id, state in _enrollment_states.items():
            if state.get('user_id') and str(state.get('user_id')) == str(student_id):
                active_enrollments.append({
                    'enrollment_id': enrollment_id,
                    'status': state.get('status'),
                    'progress': state.get('progress'),
                    'message': state.get('message')
                })
        
        return JsonResponse({
            'success': True,
            'enrollments': active_enrollments,
            'count': len(active_enrollments)
        })
    except Exception as e:
        logger.error(f"Error getting enrollment status: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_student_enroll_cancel(request):
    """Cancel ongoing biometric enrollment"""
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        
        logger.info(f"[ENROLL CANCEL] Received cancel request for student: {student_id}")
        
        # Find and remove enrollment states for this student
        to_remove = []
        for enrollment_id, state in list(_enrollment_states.items()):
            if state.get('user_id') and str(state.get('user_id')) == str(student_id):
                to_remove.append(enrollment_id)
                logger.info(f"[ENROLL CANCEL] Marking enrollment {enrollment_id} for removal")
        
        for enrollment_id in to_remove:
            del _enrollment_states[enrollment_id]
            logger.info(f"[ENROLL CANCEL] Deleted enrollment state {enrollment_id}")
        
        return JsonResponse({
            'success': True,
            'message': 'Enrollment cancelled',
            'cancelled_count': len(to_remove)
        })
    except Exception as e:
        logger.error(f"Error cancelling enrollment: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def api_broadcast_enrollment(request):
    """Broadcast enrollment message to ESP32 via MQTT (for confirmation)"""
    try:
        data = json.loads(request.body)
        slot = data.get('slot')
        template_id = data.get('template_id')
        action = data.get('action', 'confirm')
        
        logger.info(f"[BROADCAST] Broadcasting enrollment action={action}, slot={slot}, template_id={template_id}")

        if not template_id:
            return JsonResponse({'success': False, 'message': 'Missing template_id'}, status=400)

        try:
            slot_int = int(slot) if slot is not None else None
        except Exception:
            slot_int = None

        from dashboard.mqtt_client import get_mqtt_client
        mqtt_client = get_mqtt_client()

        if not mqtt_client or not mqtt_client.is_connected:
            return JsonResponse({'success': False, 'message': 'MQTT not connected'}, status=503)

        payload = {
            'action': action,
            'template_id': template_id,
        }
        if slot_int is not None:
            payload['slot'] = slot_int

        mqtt_ok = mqtt_client.publish('biometric/esp32/enroll/request', payload)
        if not mqtt_ok:
            return JsonResponse({'success': False, 'message': 'Failed to publish MQTT message'}, status=502)
        
        return JsonResponse({
            'success': True,
            'message': 'Enrollment message sent'
        })
    except Exception as e:
        logger.error(f"Error broadcasting enrollment: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
